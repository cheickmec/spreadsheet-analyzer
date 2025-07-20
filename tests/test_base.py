"""
Base test classes and utilities for spreadsheet analyzer tests.

This module provides foundational test infrastructure following best practices:
- Clear test documentation and naming conventions
- Comprehensive fixtures and test data generators
- Type-safe test utilities
- Consistent error handling patterns

CLAUDE-KNOWLEDGE: Tests should be as readable as production code. Each test
tells a story about expected behavior.
"""

from abc import ABC, abstractmethod
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path
from typing import Any, TypeVar, final

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from spreadsheet_analyzer.pipeline.types import Err, Ok

# Type variables for generic test utilities
T = TypeVar("T")
E = TypeVar("E")


@dataclass(frozen=True)
class TestContext:
    """
    Immutable context object for sharing test state.

    Provides a clean way to pass test configuration and state between
    test methods without relying on mutable class attributes.
    """

    test_name: str
    test_data_dir: Path
    temp_dir: Path
    config: dict[str, Any]


class BaseSpreadsheetTest(ABC):
    """
    Abstract base class for all spreadsheet analyzer tests.

    Provides common test infrastructure and enforces consistent patterns:
    - Proper setup and teardown
    - Test data management
    - Result assertion helpers
    - Documentation standards

    CLAUDE-COMPLEX: This base class uses the Template Method pattern to ensure
    all tests follow the same lifecycle while allowing customization.
    """

    @pytest.fixture(autouse=True)
    def _setup_test_context(self, tmp_path: Path, test_data_dir: Path) -> None:
        """
        Automatically set up test context before each test.

        This fixture runs before every test method and ensures proper
        initialization of test environment.
        """
        self.context = TestContext(
            test_name=self._testMethodName if hasattr(self, "_testMethodName") else "unknown",
            test_data_dir=test_data_dir,
            temp_dir=tmp_path,
            config=self._get_test_config(),
        )
        self._custom_setup()

    @abstractmethod
    def _get_test_config(self) -> dict[str, Any]:
        """
        Define test-specific configuration.

        Subclasses must implement this to provide their configuration.
        This enforces explicit configuration over implicit defaults.
        """
        pass

    def _custom_setup(self) -> None:
        """
        Hook for test-specific setup logic.

        Override this method in subclasses to add custom setup logic
        that runs after the base setup.
        """
        pass

    # Result assertion helpers that make tests more readable

    def assert_ok(self, result: Ok[T] | Err) -> T:
        """
        Assert that a Result is Ok and return its value.

        This helper makes test assertions more readable:
            value = self.assert_ok(parse_formula("=A1+B1"))

        Instead of:
            result = parse_formula("=A1+B1")
            assert isinstance(result, Ok)
            value = result.value
        """
        assert isinstance(result, Ok), (
            f"Expected Ok but got Err: {result.error if isinstance(result, Err) else 'unknown'}"
        )
        return result.value

    def assert_err(self, result: Ok[T] | Err, expected_error_pattern: str | None = None) -> str:
        """
        Assert that a Result is Err and optionally check error message.

        Args:
            result: The Result to check
            expected_error_pattern: Optional substring that should appear in error

        Returns:
            The error value for further assertions

        Example:
            error = self.assert_err(parse_formula("=INVALID("), "INVALID")
            assert "function" in error.lower()
        """
        assert isinstance(result, Err), (
            f"Expected Err but got Ok: {result.value if isinstance(result, Ok) else 'unknown'}"
        )

        if expected_error_pattern is not None:
            assert expected_error_pattern in str(result.error), (
                f"Expected error containing '{expected_error_pattern}' but got: {result.error}"
            )

        return result.error

    def assert_result_matches(
        self, result: Ok[T] | Err, predicate: Callable[[T], bool], description: str = "predicate check"
    ) -> None:
        """
        Assert that an Ok result's value matches a predicate.

        This is useful for complex assertions that would be hard to read inline:
            self.assert_result_matches(
                analyze_sheet(ws),
                lambda analysis: analysis.formula_count > 10,
                "sheet should have more than 10 formulas"
            )
        """
        value = self.assert_ok(result)
        assert predicate(value), f"Result failed {description}: {value}"


class ExcelTestDataBuilder:
    """
    Fluent builder for creating test Excel files with clear intent.

    This builder makes test data creation more readable and maintainable:

        test_file = (ExcelTestDataBuilder()
            .with_sheet("Data")
            .add_headers(["Name", "Value", "Total"])
            .add_row(["Item1", 10, "=B2*2"])
            .add_row(["Item2", 20, "=B3*2"])
            .with_sheet("Summary")
            .add_cell("A1", "Total:")
            .add_cell("B1", "=SUM(Data!C:C)")
            .build(tmp_path / "test.xlsx"))

    CLAUDE-KNOWLEDGE: Builder pattern provides clear, self-documenting test data
    creation that's easy to modify and understand.
    """

    def __init__(self) -> None:
        """Initialize a new workbook builder."""
        self.workbook = Workbook()
        self.current_sheet: Worksheet | None = None
        self._remove_default_sheet = True

    def with_sheet(self, name: str, make_active: bool = False) -> "ExcelTestDataBuilder":
        """
        Add or switch to a sheet with the given name.

        Args:
            name: Sheet name
            make_active: Whether to make this the active sheet

        Returns:
            Self for method chaining
        """
        if self._remove_default_sheet and "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
            self._remove_default_sheet = False

        if name in self.workbook.sheetnames:
            self.current_sheet = self.workbook[name]
        else:
            self.current_sheet = self.workbook.create_sheet(name)

        if make_active:
            self.workbook.active = self.current_sheet

        return self

    def add_headers(self, headers: list[str]) -> "ExcelTestDataBuilder":
        """
        Add header row to current sheet.

        Args:
            headers: List of header values

        Returns:
            Self for method chaining
        """
        if not self.current_sheet:
            raise ValueError("No active sheet. Use with_sheet() first.")

        self.current_sheet.append(headers)
        return self

    def add_row(self, values: list[Any]) -> "ExcelTestDataBuilder":
        """
        Add data row to current sheet.

        Args:
            values: List of cell values (can include formulas)

        Returns:
            Self for method chaining
        """
        if not self.current_sheet:
            raise ValueError("No active sheet. Use with_sheet() first.")

        self.current_sheet.append(values)
        return self

    def add_cell(self, cell_ref: str, value: Any) -> "ExcelTestDataBuilder":
        """
        Set a specific cell's value.

        Args:
            cell_ref: Cell reference (e.g., "A1")
            value: Cell value or formula

        Returns:
            Self for method chaining
        """
        if not self.current_sheet:
            raise ValueError("No active sheet. Use with_sheet() first.")

        self.current_sheet[cell_ref] = value
        return self

    def add_named_range(self, name: str, reference: str) -> "ExcelTestDataBuilder":
        """
        Add a named range to the workbook.

        Args:
            name: Range name
            reference: Range reference (e.g., "Sheet1!A1:B10")

        Returns:
            Self for method chaining
        """
        from openpyxl.workbook.defined_name import DefinedName

        defn = DefinedName(name=name, attr_text=reference)
        self.workbook.defined_names[name] = defn
        return self

    def with_data_validation(
        self, cell_range: str, validation_type: str = "list", formula1: str | None = None
    ) -> "ExcelTestDataBuilder":
        """
        Add data validation to cells.

        Args:
            cell_range: Range to apply validation to
            validation_type: Type of validation
            formula1: Validation formula or list

        Returns:
            Self for method chaining
        """
        from openpyxl.worksheet.datavalidation import DataValidation

        if not self.current_sheet:
            raise ValueError("No active sheet. Use with_sheet() first.")

        dv = DataValidation(type=validation_type, formula1=formula1)
        dv.add(cell_range)
        self.current_sheet.add_data_validation(dv)
        return self

    @final
    def build(self, output_path: Path) -> Path:
        """
        Save the workbook and return the path.

        Args:
            output_path: Where to save the Excel file

        Returns:
            Path to the created file

        CLAUDE-GOTCHA: Always close the workbook after saving to prevent
        file handle leaks in tests.
        """
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(output_path)
        self.workbook.close()
        return output_path


class TestDataPatterns:
    """
    Common test data patterns for spreadsheet testing.

    This class provides factory methods for creating common test scenarios,
    making tests more consistent and reducing duplication.
    """

    @staticmethod
    def create_simple_data_sheet(path: Path, rows: int = 10) -> Path:
        """Create a basic data sheet with headers and numeric data."""
        return (
            ExcelTestDataBuilder()
            .with_sheet("Data", make_active=True)
            .add_headers(["ID", "Name", "Value", "Status"])
            .add_row([1, "Item A", 100, "Active"])
            .add_row([2, "Item B", 200, "Active"])
            .add_row([3, "Item C", 300, "Inactive"])
            .build(path)
        )

    @staticmethod
    def create_formula_heavy_sheet(path: Path) -> Path:
        """Create a sheet with various formula patterns for testing."""
        return (
            ExcelTestDataBuilder()
            .with_sheet("Data")
            .add_headers(["Base", "Multiplier", "Result"])
            .add_row([10, 2, "=A2*B2"])
            .add_row([20, 3, "=A3*B3"])
            .add_row([30, 4, "=A4*B4"])
            .with_sheet("Summary")
            .add_cell("A1", "Total Base:")
            .add_cell("B1", "=SUM(Data!A:A)")
            .add_cell("A2", "Total Result:")
            .add_cell("B2", "=SUM(Data!C:C)")
            .add_cell("A3", "Average:")
            .add_cell("B3", "=AVERAGE(Data!C:C)")
            .build(path)
        )

    @staticmethod
    def create_circular_reference_sheet(path: Path) -> Path:
        """Create a sheet with circular references for edge case testing."""
        return (
            ExcelTestDataBuilder()
            .with_sheet("Circular")
            .add_cell("A1", "=B1+1")
            .add_cell("B1", "=C1+1")
            .add_cell("C1", "=A1+1")  # Creates circular reference
            .build(path)
        )

    @staticmethod
    def create_cross_sheet_dependencies(path: Path) -> Path:
        """Create complex cross-sheet formula dependencies."""
        return (
            ExcelTestDataBuilder()
            .with_sheet("Input")
            .add_headers(["Month", "Sales", "Costs"])
            .add_row(["Jan", 1000, 600])
            .add_row(["Feb", 1200, 700])
            .add_row(["Mar", 1100, 650])
            .with_sheet("Calculations")
            .add_cell("A1", "Profit Margin")
            .add_cell("B1", "=(SUM(Input!B:B)-SUM(Input!C:C))/SUM(Input!B:B)")
            .add_cell("A2", "Average Sales")
            .add_cell("B2", "=AVERAGE(Input!B:B)")
            .with_sheet("Report")
            .add_cell("A1", "Total Revenue:")
            .add_cell("B1", "=SUM(Input!B:B)")
            .add_cell("A2", "Profit Margin:")
            .add_cell("B2", "=Calculations!B1")
            .add_cell("A3", "Status:")
            .add_cell("B3", '=IF(B2>0.3,"Good","Needs Improvement")')
            .build(path)
        )
