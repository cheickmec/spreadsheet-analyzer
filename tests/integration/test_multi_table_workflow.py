#!/usr/bin/env python3
"""Integration tests for multi-table detection and analysis workflow.

These tests verify the complete end-to-end workflow including:
- Table detection with LLM
- Handoff between agents
- Analysis with detected boundaries
- Notebook generation with prompt hash tracking
"""

from pathlib import Path

import openpyxl
import pytest

from spreadsheet_analyzer.cli.notebook_analysis import AnalysisConfig
from spreadsheet_analyzer.workflows.multi_table_workflow import (
    SpreadsheetAnalysisState,
    create_multi_table_workflow,
    run_multi_table_analysis,
)


@pytest.fixture
def multi_table_excel(tmp_path: Path) -> Path:
    """Create Excel file with multiple distinct tables."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Table 1: Sales Data (rows 1-20, cols A-D)
    ws["A1"] = "Product ID"
    ws["B1"] = "Product Name"
    ws["C1"] = "Units Sold"
    ws["D1"] = "Revenue"

    for i in range(2, 21):
        ws[f"A{i}"] = f"PROD-{i:03d}"
        ws[f"B{i}"] = f"Product {i}"
        ws[f"C{i}"] = 100 + i * 5
        ws[f"D{i}"] = (100 + i * 5) * 29.99

    # Empty rows (21-23)

    # Table 2: Regional Summary (rows 24-30, cols A-C)
    ws["A24"] = "Region"
    ws["B24"] = "Total Sales"
    ws["C24"] = "Growth %"

    regions = ["North", "South", "East", "West", "Central"]
    for i, region in enumerate(regions, start=25):
        ws[f"A{i}"] = region
        ws[f"B{i}"] = 150000 + i * 10000
        ws[f"C{i}"] = 5.5 + i * 0.5

    # Side-by-side table: Employee Data (rows 1-15, cols F-I)
    ws["F1"] = "Employee ID"
    ws["G1"] = "Name"
    ws["H1"] = "Department"
    ws["I1"] = "Salary"

    for i in range(2, 16):
        ws[f"F{i}"] = f"EMP-{i:03d}"
        ws[f"G{i}"] = f"Employee {i}"
        ws[f"H{i}"] = ["Sales", "Marketing", "IT", "HR"][i % 4]
        ws[f"I{i}"] = 50000 + i * 2000

    file_path = tmp_path / "multi_table_test.xlsx"
    wb.save(file_path)
    return file_path


@pytest.fixture
def analysis_config(tmp_path: Path) -> AnalysisConfig:
    """Create test analysis configuration."""
    return AnalysisConfig(
        excel_path=tmp_path / "test.xlsx",
        sheet_index=0,
        model="gpt-4o-mini",  # Use a cheaper model for tests
        output_dir=tmp_path / "outputs",
        max_rounds=2,
        detector_max_rounds=2,
        track_costs=False,  # Disable cost tracking for tests
    )


class TestMultiTableWorkflow:
    """Integration tests for multi-table workflow."""

    @pytest.mark.asyncio
    async def test_workflow_creation(self):
        """Test that workflow graph is created correctly."""
        workflow = create_multi_table_workflow()

        # Verify workflow structure
        assert workflow is not None
        # Check that key nodes exist by trying to get the graph structure
        graph_def = workflow.get_graph()
        nodes = graph_def.nodes

        assert "supervisor" in nodes
        assert "detector" in nodes
        assert "analyst" in nodes

    @pytest.mark.asyncio
    async def test_state_initialization(self, multi_table_excel, analysis_config):
        """Test proper state initialization."""
        initial_state = SpreadsheetAnalysisState(
            excel_file_path=str(multi_table_excel),
            sheet_index=0,
            sheet_name=None,
            config=analysis_config,
            table_boundaries=None,
            detection_notebook_path=None,
            detection_error=None,
            analysis_notebook_path=None,
            analysis_error=None,
            messages=[],
            current_agent="",
            workflow_complete=False,
        )

        assert initial_state["excel_file_path"] == str(multi_table_excel)
        assert initial_state["table_boundaries"] is None
        assert initial_state["workflow_complete"] is False

    @pytest.mark.asyncio
    @pytest.mark.skip(reason="Requires LLM API key")
    async def test_detector_only_mode(self, multi_table_excel, tmp_path):
        """Test detector-only mode without running analyst."""
        config = AnalysisConfig(
            excel_path=multi_table_excel,
            sheet_index=0,
            model="gpt-4o-mini",
            output_dir=tmp_path / "outputs",
            detector_max_rounds=2,
            detector_only=True,  # Only run detector
            track_costs=False,
        )

        result = await run_multi_table_analysis(multi_table_excel, sheet_index=0, config=config)

        assert result.is_ok()
        output = result.unwrap()

        # Should have detection results but no analysis
        assert output["detection_notebook"] is not None
        assert output["analysis_notebook"] is None
        assert output["tables_found"] > 0

        # Verify detection notebook was created
        detection_path = Path(output["detection_notebook"])
        assert detection_path.exists()
        assert "_detection_" in detection_path.name

        # Verify prompt hash is in filename
        # Both detector and analyst use different prompts with different hashes
        assert any(char.isdigit() and char.isalpha() for char in detection_path.stem)

    @pytest.mark.asyncio
    @pytest.mark.skip(reason="Requires LLM API key")
    async def test_full_workflow_with_boundaries(self, multi_table_excel, tmp_path):
        """Test complete workflow from detection to analysis."""
        config = AnalysisConfig(
            excel_path=multi_table_excel,
            sheet_index=0,
            model="gpt-4o-mini",
            output_dir=tmp_path / "outputs",
            max_rounds=2,
            detector_max_rounds=2,
            track_costs=False,
        )

        result = await run_multi_table_analysis(multi_table_excel, sheet_index=0, config=config)

        assert result.is_ok()
        output = result.unwrap()

        # Should have both detection and analysis results
        assert output["detection_notebook"] is not None
        assert output["analysis_notebook"] is not None
        assert output["tables_found"] > 0

        # Verify both notebooks were created
        detection_path = Path(output["detection_notebook"])
        analysis_path = Path(output["analysis_notebook"])

        assert detection_path.exists()
        assert analysis_path.exists()

        # Verify different directories
        assert "detector" in str(detection_path)
        assert "analyst" in str(analysis_path)

        # Verify prompt hashes in filenames
        # Detection uses table_detector_system prompt
        # Analysis uses table_aware_analyst_system prompt
        assert "_detection_" in detection_path.name
        assert any(char.isdigit() and char.isalpha() for char in detection_path.stem)
        assert any(char.isdigit() and char.isalpha() for char in analysis_path.stem)

    @pytest.mark.asyncio
    async def test_workflow_error_handling(self, tmp_path):
        """Test workflow handles missing files gracefully."""
        non_existent = tmp_path / "does_not_exist.xlsx"

        config = AnalysisConfig(
            excel_path=non_existent,
            sheet_index=0,
            model="gpt-4o-mini",
            output_dir=tmp_path / "outputs",
            track_costs=False,
        )

        result = await run_multi_table_analysis(non_existent, sheet_index=0, config=config)

        # Should return error result
        assert result.is_err()
        error = result.unwrap_err()
        assert "Workflow failed" in error or "Workflow error" in error

    @pytest.mark.asyncio
    async def test_supervisor_routing(self, multi_table_excel, analysis_config):
        """Test supervisor correctly routes between agents."""
        workflow = create_multi_table_workflow()

        # Test initial routing - should go to detector first
        initial_state = SpreadsheetAnalysisState(
            excel_file_path=str(multi_table_excel),
            sheet_index=0,
            sheet_name=None,
            config=analysis_config,
            table_boundaries=None,
            detection_notebook_path=None,
            detection_error=None,
            analysis_notebook_path=None,
            analysis_error=None,
            messages=[],
            current_agent="",
            workflow_complete=False,
        )

        # Run supervisor node
        from spreadsheet_analyzer.workflows.multi_table_workflow import supervisor_node

        supervisor_result = await supervisor_node(initial_state)

        # Should route to detector first
        assert supervisor_result["current_agent"] == "detector"

        # Simulate detector completion
        from spreadsheet_analyzer.agents.table_detection_types import (
            TableBoundary,
            TableDetectionResult,
            TableType,
        )

        state_after_detection = {
            **initial_state,
            "table_boundaries": TableDetectionResult(
                sheet_name="Sheet1",
                tables=(
                    TableBoundary(
                        table_id="table_1",
                        description="Test table",
                        start_row=0,
                        end_row=10,
                        start_col=0,
                        end_col=3,
                        confidence=0.9,
                        table_type=TableType.DETAIL,
                        entity_type="sales",
                    ),
                ),
                metadata={},
            ),
            "detection_notebook_path": "/path/to/detection.ipynb",
        }

        supervisor_result = await supervisor_node(state_after_detection)

        # Should route to analyst after detection
        assert supervisor_result["current_agent"] == "analyst"

        # Simulate analyst completion
        state_after_analysis = {
            **state_after_detection,
            "analysis_notebook_path": "/path/to/analysis.ipynb",
        }

        supervisor_result = await supervisor_node(state_after_analysis)

        # Should mark workflow complete
        assert supervisor_result["current_agent"] == "end"
        assert supervisor_result["workflow_complete"] is True

    @pytest.mark.asyncio
    async def test_detector_only_supervisor_routing(self, multi_table_excel, tmp_path):
        """Test supervisor skips analyst in detector-only mode."""
        from spreadsheet_analyzer.agents.table_detection_types import (
            TableBoundary,
            TableDetectionResult,
            TableType,
        )
        from spreadsheet_analyzer.workflows.multi_table_workflow import supervisor_node

        # Create config with detector_only=True
        config = AnalysisConfig(
            excel_path=multi_table_excel,
            sheet_index=0,
            model="gpt-4o-mini",
            output_dir=tmp_path / "outputs",
            detector_only=True,
            track_costs=False,
        )

        # State after detection with detector_only flag
        state_after_detection = SpreadsheetAnalysisState(
            excel_file_path=str(multi_table_excel),
            sheet_index=0,
            sheet_name=None,
            config=config,
            table_boundaries=TableDetectionResult(
                sheet_name="Sheet1",
                tables=(
                    TableBoundary(
                        table_id="table_1",
                        description="Test table",
                        start_row=0,
                        end_row=10,
                        start_col=0,
                        end_col=3,
                        confidence=0.9,
                        table_type=TableType.DETAIL,
                        entity_type="sales",
                    ),
                ),
                metadata={},
            ),
            detection_notebook_path="/path/to/detection.ipynb",
            detection_error=None,
            analysis_notebook_path=None,
            analysis_error=None,
            messages=[],
            current_agent="detector",
            workflow_complete=False,
        )

        supervisor_result = await supervisor_node(state_after_detection)

        # Should skip analyst and mark complete
        assert supervisor_result["current_agent"] == "end"
        assert supervisor_result["workflow_complete"] is True
        assert "Detection complete" in supervisor_result["messages"][0].content


@pytest.mark.asyncio
async def test_prompt_hash_in_generated_notebooks(tmp_path):
    """Test that generated notebooks include prompt hash in their filenames."""
    from spreadsheet_analyzer.cli.utils.naming import (
        FileNameConfig,
        generate_notebook_name,
        get_short_hash,
    )
    from spreadsheet_analyzer.prompts import get_prompt_definition

    # Get actual prompt hashes
    detector_prompt = get_prompt_definition("table_detector_system")
    analyst_prompt = get_prompt_definition("table_aware_analyst_system")

    assert detector_prompt is not None
    assert analyst_prompt is not None

    detector_hash = get_short_hash(detector_prompt.content_hash)
    analyst_hash = get_short_hash(analyst_prompt.content_hash)

    # Create file configs with hashes
    detector_config = FileNameConfig(
        excel_file=tmp_path / "test.xlsx",
        model="gpt-4o-mini",
        sheet_index=0,
        sheet_name="Sheet1",
        max_rounds=3,
        prompt_hash=detector_hash,
    )

    analyst_config = FileNameConfig(
        excel_file=tmp_path / "test.xlsx",
        model="gpt-4o-mini",
        sheet_index=0,
        sheet_name="Sheet1",
        max_rounds=5,
        prompt_hash=analyst_hash,
    )

    # Generate names
    detector_name = generate_notebook_name(detector_config, include_timestamp=False)
    analyst_name = generate_notebook_name(analyst_config, include_timestamp=False)

    # Verify hashes are included
    assert detector_hash in detector_name
    assert analyst_hash in analyst_name

    # Verify they're different (different prompts = different hashes)
    assert detector_hash != analyst_hash
    assert detector_name != analyst_name
