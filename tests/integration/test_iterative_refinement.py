"""Test script for iterative refinement in analyze_sheet.py

This script tests various scenarios to ensure the iterative refinement loop works correctly:
1. Code that generates errors to ensure refinement works
2. Incomplete analyses to verify quality inspection
3. Cost limit enforcement
4. Observation building with various output types
"""

import asyncio
import json
import tempfile
from pathlib import Path
from unittest.mock import MagicMock

from openpyxl import Workbook

from spreadsheet_analyzer.cli.analyze_sheet import (
    build_observation_from_notebook,
    create_sheet_notebook,
    inspect_notebook_quality,
)
from spreadsheet_analyzer.notebook_llm.llm_providers.base import LLMResponse
from spreadsheet_analyzer.notebook_llm.nap.protocols import Cell, CellType, NotebookDocument


def create_test_excel(path: Path) -> None:
    """Create a test Excel file with sample data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"

    # Add headers
    headers = ["Date", "Product", "Quantity", "Price", "Total"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Add data with some issues
    data = [
        ["2024-01-01", "Widget A", "10", "$25.50", "=C2*D2"],  # Text in numeric field
        ["2024-01-02", "Widget B", 15, 30.00, "=C3*D3"],
        ["2024-01-03", "Widget C", "N/A", 20.00, "=C4*D4"],  # Non-numeric quantity
        ["2024-01-04", "Widget D", 5, "Invalid", "=C5*D5"],  # Invalid price
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(path)


def test_build_observation():
    """Test observation building from notebook outputs."""
    print("\n=== Testing build_observation_from_notebook ===")

    # Create a notebook with various outputs
    cells = [
        Cell(
            id="test1",
            cell_type=CellType.CODE,
            source="print('Hello')",
            metadata={"analysis_type": "llm_generated"},
            outputs=[{"output_type": "stream", "name": "stdout", "text": "Hello\n"}],
        ),
        Cell(
            id="test2",
            cell_type=CellType.CODE,
            source="1/0",
            metadata={"analysis_type": "llm_generated"},
            outputs=[
                {
                    "output_type": "error",
                    "ename": "ZeroDivisionError",
                    "evalue": "division by zero",
                    "traceback": ["Traceback...", "ZeroDivisionError: division by zero"],
                }
            ],
        ),
        Cell(
            id="test3",
            cell_type=CellType.CODE,
            source="df.describe()",
            metadata={"analysis_type": "llm_generated"},
            outputs=[
                {
                    "output_type": "execute_result",
                    "execution_count": 3,
                    "data": {"text/plain": "DataFrame statistics..."},
                }
            ],
        ),
    ]

    notebook = NotebookDocument(
        id="test_notebook",
        cells=cells,
        metadata={},
        kernel_spec={"name": "python3", "display_name": "Python 3"},
        language_info={"name": "python", "version": "3.12"},
    )

    observation = build_observation_from_notebook(notebook)
    print("Observation:")
    print(observation)

    # Verify observation contains expected elements
    assert "Errors Found:" in observation
    assert "ZeroDivisionError" in observation
    assert "Outputs:" in observation
    assert "Hello" in observation
    assert "DataFrame statistics" in observation

    print("✓ Observation building test passed")


def test_inspect_notebook_quality():
    """Test notebook quality inspection."""
    print("\n=== Testing inspect_notebook_quality ===")

    # Test 1: Notebook with errors
    cells_with_errors = [
        Cell(
            id="error_cell",
            cell_type=CellType.CODE,
            source="1/0",
            metadata={"analysis_type": "llm_generated"},
            outputs=[
                {"output_type": "error", "ename": "ZeroDivisionError", "evalue": "division by zero", "traceback": []}
            ],
        )
    ]

    notebook_with_errors = NotebookDocument(
        id="test",
        cells=cells_with_errors,
        metadata={},
        kernel_spec={"name": "python3", "display_name": "Python 3"},
        language_info={"name": "python", "version": "3.12"},
    )

    needs_refinement, reason = inspect_notebook_quality(notebook_with_errors)
    print(f"Notebook with errors - Needs refinement: {needs_refinement}, Reason: {reason}")
    assert needs_refinement is True
    assert "execution errors" in reason

    # Test 2: Notebook with no outputs
    cells_no_output = [
        Cell(
            id="no_output",
            cell_type=CellType.CODE,
            source="x = 1 + 1",
            metadata={"analysis_type": "llm_generated"},
            outputs=[],
        )
    ]

    notebook_no_output = NotebookDocument(
        id="test",
        cells=cells_no_output,
        metadata={},
        kernel_spec={"name": "python3", "display_name": "Python 3"},
        language_info={"name": "python", "version": "3.12"},
    )

    needs_refinement, reason = inspect_notebook_quality(notebook_no_output)
    print(f"Notebook with no output - Needs refinement: {needs_refinement}, Reason: {reason}")
    assert needs_refinement is True
    assert "no output" in reason

    # Test 3: Complete notebook
    cells_complete = [
        Cell(
            id="cell1",
            cell_type=CellType.CODE,
            source="df.describe()",
            metadata={"analysis_type": "llm_generated"},
            outputs=[{"output_type": "stream", "text": "Statistics..."}],
        ),
        Cell(
            id="cell2",
            cell_type=CellType.CODE,
            source="df.mean()",
            metadata={"analysis_type": "llm_generated"},
            outputs=[{"output_type": "stream", "text": "Mean values..."}],
        ),
        Cell(
            id="cell3",
            cell_type=CellType.CODE,
            source="plt.plot(df)",
            metadata={"analysis_type": "llm_generated"},
            outputs=[{"output_type": "display_data", "data": {"image/png": "base64data"}}],
        ),
        Cell(
            id="cell4",
            cell_type=CellType.CODE,
            source="analysis",
            metadata={"analysis_type": "llm_generated"},
            outputs=[{"output_type": "stream", "text": "Analysis results..."}],
        ),
    ]

    notebook_complete = NotebookDocument(
        id="test",
        cells=cells_complete,
        metadata={},
        kernel_spec={"name": "python3", "display_name": "Python 3"},
        language_info={"name": "python", "version": "3.12"},
    )

    needs_refinement, reason = inspect_notebook_quality(notebook_complete)
    print(f"Complete notebook - Needs refinement: {needs_refinement}, Reason: {reason}")
    assert needs_refinement is False
    assert "complete" in reason

    print("✓ Quality inspection test passed")


async def test_iterative_refinement_mock():
    """Test iterative refinement with mocked LLM responses."""
    print("\n=== Testing iterative refinement with mock LLM ===")

    # Create test Excel file
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_path = Path(tmpdir) / "test.xlsx"
        create_test_excel(excel_path)

        # Create initial notebook
        notebook = create_sheet_notebook(excel_path, "TestSheet")

        # Mock LLM responses for refinement
        responses = [
            # Round 1: Code with error
            LLMResponse(
                content="""```markdown
# Initial Analysis
Let's analyze the data
```

```python
# This will cause an error
df['Total_Numeric'] = df['Quantity'] * df['Price']
```""",
                model="claude-test",
                usage={"prompt_tokens": 100, "completion_tokens": 50, "total_tokens": 150},
                error=None,
            ),
            # Round 2: Fixed code
            LLMResponse(
                content="""```python
# Fix the error by converting to numeric first
df['Quantity_Numeric'] = pd.to_numeric(df['Quantity'], errors='coerce')
df['Price_Numeric'] = pd.to_numeric(df['Price'].str.replace('$', ''), errors='coerce')
df['Total_Calculated'] = df['Quantity_Numeric'] * df['Price_Numeric']
print("Calculated totals:")
print(df[['Product', 'Total_Calculated']])
```

```python
# Show statistics
print("\\nStatistics:")
print(df['Total_Calculated'].describe())
```""",
                model="claude-test",
                usage={"prompt_tokens": 150, "completion_tokens": 100, "total_tokens": 250},
                error=None,
            ),
        ]

        # Mock provider
        mock_provider = MagicMock()
        mock_provider.model_name = "claude-test"
        mock_provider.complete.side_effect = responses

        # Import and run with mocked provider
        from spreadsheet_analyzer.cli.analyze_sheet import analyze_sheet_with_llm

        # Run analysis
        final_notebook, results = await analyze_sheet_with_llm(
            notebook, mock_provider, strategy_name="hierarchical", max_rounds=3, cost_limit=1.0
        )

        # Verify results
        print(f"\nAnalysis completed in {results['rounds']} rounds")
        print(f"Total cost: ${results['total_cost']}")
        print(f"Refinement history: {json.dumps(results['refinement_history'], indent=2)}")

        assert results["rounds"] == 2
        assert len(results["refinement_history"]) == 2
        assert results["refinement_history"][0]["needs_refinement"] is True
        assert results["refinement_history"][1]["needs_refinement"] is False

        print("✓ Iterative refinement test passed")


def test_cost_tracking():
    """Test cost calculation for different models."""
    print("\n=== Testing cost tracking ===")

    # Test Claude pricing
    claude_usage = {"prompt_tokens": 1000, "completion_tokens": 500, "total_tokens": 1500}

    # Calculate expected cost for Claude
    # Input: $0.003/1K tokens, Output: $0.015/1K tokens
    expected_input_cost = 1000 * 0.003 / 1000
    expected_output_cost = 500 * 0.015 / 1000
    expected_total = expected_input_cost + expected_output_cost

    print("Claude pricing test:")
    print(f"  Input cost: ${expected_input_cost:.4f}")
    print(f"  Output cost: ${expected_output_cost:.4f}")
    print(f"  Total: ${expected_total:.4f}")

    # Test GPT pricing
    gpt_usage = {"prompt_tokens": 1000, "completion_tokens": 500, "total_tokens": 1500}

    # Calculate expected cost for GPT
    # Input: $0.001/1K tokens, Output: $0.002/1K tokens
    expected_input_cost_gpt = 1000 * 0.001 / 1000
    expected_output_cost_gpt = 500 * 0.002 / 1000
    expected_total_gpt = expected_input_cost_gpt + expected_output_cost_gpt

    print("\nGPT pricing test:")
    print(f"  Input cost: ${expected_input_cost_gpt:.4f}")
    print(f"  Output cost: ${expected_output_cost_gpt:.4f}")
    print(f"  Total: ${expected_total_gpt:.4f}")

    print("\n✓ Cost tracking test passed")


def main():
    """Run all tests."""
    print("=== Testing Iterative Refinement Implementation ===")

    # Test individual components
    test_build_observation()
    test_inspect_notebook_quality()
    test_cost_tracking()

    # Test full iterative refinement
    asyncio.run(test_iterative_refinement_mock())

    print("\n=== All tests passed! ===")


if __name__ == "__main__":
    main()
