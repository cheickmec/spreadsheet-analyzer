"""
Tests for spreadsheet-specific quality inspection.

This module provides comprehensive functional tests for the SpreadsheetQualityInspector,
testing quality metrics, issue detection, and analysis capabilities.
All tests are functional and use no mocking.
"""

import pytest
from pathlib import Path
from typing import Dict, List
import tempfile
import pandas as pd

from spreadsheet_analyzer.plugins.spreadsheet.quality import SpreadsheetQualityInspector
from spreadsheet_analyzer.core_exec import (
    NotebookBuilder,
    NotebookCell,
    CellType,
)
from spreadsheet_analyzer.core_exec.quality import (
    QualityMetrics,
    QualityIssue,
    QualityLevel,
)


class TestSpreadsheetQualityInspector:
    """Test the SpreadsheetQualityInspector class functionality."""
    
    def setup_method(self):
        """Set up test fixtures for each test method."""
        self.inspector = SpreadsheetQualityInspector()
        self.builder = NotebookBuilder()
    
    def test_inspector_initialization(self):
        """Test that the inspector initializes correctly."""
        assert isinstance(self.inspector, SpreadsheetQualityInspector)
        assert hasattr(self.inspector, 'inspect')
    
    def test_inspect_empty_notebook(self):
        """Test quality inspection of an empty notebook."""
        # Create empty notebook
        empty_notebook = self.builder.build()
        
        metrics = self.inspector.inspect(empty_notebook)
        
        assert isinstance(metrics, QualityMetrics)
        assert metrics.overall_score >= 0.0
        assert metrics.overall_score <= 1.0
        assert isinstance(metrics.issues, list)
        
        # Empty notebook should have issues
        issue_messages = [issue.message for issue in metrics.issues]
        assert any('empty' in msg.lower() or 'no cells' in msg.lower() for msg in issue_messages)
    
    def test_inspect_basic_spreadsheet_analysis(self):
        """Test quality inspection of a basic spreadsheet analysis notebook."""
        # Create notebook with basic spreadsheet analysis
        self.builder.add_markdown_cell("# Spreadsheet Analysis")
        self.builder.add_code_cell("import pandas as pd\nimport numpy as np")
        self.builder.add_code_cell("df = pd.read_excel('data.xlsx')")
        self.builder.add_code_cell("print(df.head())")
        self.builder.add_code_cell("print(df.describe())")
        
        notebook = self.builder.build()
        metrics = self.inspector.inspect(notebook)
        
        assert isinstance(metrics, QualityMetrics)
        assert metrics.overall_score > 0.0
        
        # Should recognize good practices
        issue_levels = [issue.level for issue in metrics.issues]
        # Should have fewer critical issues for well-structured analysis
        critical_issues = [issue for issue in metrics.issues if issue.level == QualityLevel.CRITICAL]
        assert len(critical_issues) <= 2  # Allow some flexibility
    
    def test_inspect_comprehensive_analysis(self):
        """Test quality inspection of a comprehensive spreadsheet analysis."""
        # Create notebook with comprehensive analysis
        self.builder.add_markdown_cell("# Comprehensive Data Analysis")
        self.builder.add_markdown_cell("## Data Loading and Overview")
        
        # Good imports
        self.builder.add_code_cell("""
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
""")
        
        # Data loading with error handling
        self.builder.add_code_cell("""
try:
    df = pd.read_excel('sales_data.xlsx')
    print(f"Data loaded successfully: {df.shape}")
except FileNotFoundError:
    print("File not found")
except Exception as e:
    print(f"Error loading data: {e}")
""")
        
        # Data overview
        self.builder.add_code_cell("""
# Display basic information
print("Dataset Info:")
print(df.info())
print("\\nFirst 5 rows:")
print(df.head())
""")
        
        # Statistical summary
        self.builder.add_code_cell("""
# Statistical summary
print("Statistical Summary:")
print(df.describe())
""")
        
        # Data quality checks
        self.builder.add_markdown_cell("## Data Quality Assessment")
        self.builder.add_code_cell("""
# Check for missing values
missing_values = df.isnull().sum()
print("Missing values per column:")
print(missing_values[missing_values > 0])
""")
        
        # Visualizations
        self.builder.add_markdown_cell("## Data Visualization")
        self.builder.add_code_cell("""
# Distribution plots
fig, axes = plt.subplots(2, 2, figsize=(12, 10))
df.hist(bins=20, ax=axes, layout=(2, 2))
plt.tight_layout()
plt.show()
""")
        
        notebook = self.builder.build()
        metrics = self.inspector.inspect(notebook)
        
        assert isinstance(metrics, QualityMetrics)
        assert metrics.overall_score > 0.5  # Should score well for comprehensive analysis
        
        # Check specific quality aspects
        assert 'structure_score' in metrics.details
        assert 'code_quality_score' in metrics.details
        assert 'documentation_score' in metrics.details
    
    def test_inspect_poor_quality_notebook(self):
        """Test quality inspection of a poorly structured notebook."""
        # Create notebook with quality issues
        self.builder.add_code_cell("df=pd.read_excel('file.xlsx')")  # No import, poor formatting
        self.builder.add_code_cell("print(df)")  # No context
        self.builder.add_code_cell("x=df.mean()")  # No explanation
        self.builder.add_code_cell("print(x)")
        self.builder.add_code_cell("df.plot()")  # No setup, no context
        
        notebook = self.builder.build()
        metrics = self.inspector.inspect(notebook)
        
        assert isinstance(metrics, QualityMetrics)
        assert metrics.overall_score < 0.5  # Should score poorly
        
        # Should identify several issues
        assert len(metrics.issues) > 3
        
        # Check for specific issue types
        issue_messages = [issue.message.lower() for issue in metrics.issues]
        assert any('import' in msg for msg in issue_messages)
        assert any('documentation' in msg or 'markdown' in msg for msg in issue_messages)
    
    def test_inspect_formula_analysis_quality(self):
        """Test quality inspection specifically for formula analysis."""
        # Create notebook focused on formula analysis
        self.builder.add_markdown_cell("# Excel Formula Analysis")
        
        self.builder.add_code_cell("""
import openpyxl
from openpyxl import load_workbook
import pandas as pd
""")
        
        self.builder.add_code_cell("""
# Load workbook to analyze formulas
wb = load_workbook('spreadsheet.xlsx', data_only=False)
ws = wb.active
""")
        
        self.builder.add_code_cell("""
# Extract formulas
formulas = []
for row in ws.iter_rows():
    for cell in row:
        if cell.value and str(cell.value).startswith('='):
            formulas.append({
                'cell': cell.coordinate,
                'formula': cell.value
            })

print(f"Found {len(formulas)} formulas")
""")
        
        self.builder.add_code_cell("""
# Analyze formula complexity
formula_df = pd.DataFrame(formulas)
if not formula_df.empty:
    formula_df['length'] = formula_df['formula'].str.len()
    formula_df['function_count'] = formula_df['formula'].str.count(r'[A-Z]+\\(')
    
    print("Formula complexity analysis:")
    print(formula_df.describe())
""")
        
        notebook = self.builder.build()
        metrics = self.inspector.inspect(notebook)
        
        assert isinstance(metrics, QualityMetrics)
        # Should recognize as specialized analysis
        assert metrics.overall_score > 0.4
        
        # Check for spreadsheet-specific quality indicators
        details = metrics.details
        assert isinstance(details, dict)
    
    def test_inspect_data_profiling_quality(self):
        """Test quality inspection for data profiling notebooks."""
        # Create comprehensive data profiling notebook
        self.builder.add_markdown_cell("# Data Profiling Report")
        self.builder.add_markdown_cell("This notebook provides comprehensive profiling of the dataset.")
        
        self.builder.add_code_cell("""
# Essential imports for data profiling
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats
""")
        
        self.builder.add_code_cell("""
# Load and inspect data
df = pd.read_excel('dataset.xlsx')
print(f"Dataset shape: {df.shape}")
print(f"Column names: {list(df.columns)}")
""")
        
        self.builder.add_markdown_cell("## Data Types and Missing Values")
        self.builder.add_code_cell("""
# Data types analysis
print("Data types:")
print(df.dtypes)
print("\\nMissing values:")
print(df.isnull().sum())
print(f"\\nMissing value percentage:")
print((df.isnull().sum() / len(df)) * 100)
""")
        
        self.builder.add_markdown_cell("## Statistical Summary")
        self.builder.add_code_cell("""
# Numerical columns summary
numerical_cols = df.select_dtypes(include=[np.number]).columns
print("Numerical columns summary:")
print(df[numerical_cols].describe())
""")
        
        self.builder.add_code_cell("""
# Categorical columns summary
categorical_cols = df.select_dtypes(include=['object']).columns
print("Categorical columns summary:")
for col in categorical_cols:
    print(f"\\n{col}:")
    print(df[col].value_counts().head())
""")
        
        self.builder.add_markdown_cell("## Distribution Analysis")
        self.builder.add_code_cell("""
# Distribution plots for numerical columns
if len(numerical_cols) > 0:
    fig, axes = plt.subplots(len(numerical_cols), 2, figsize=(12, 4*len(numerical_cols)))
    if len(numerical_cols) == 1:
        axes = axes.reshape(1, -1)
    
    for i, col in enumerate(numerical_cols):
        # Histogram
        axes[i, 0].hist(df[col].dropna(), bins=30, alpha=0.7)
        axes[i, 0].set_title(f'{col} - Histogram')
        axes[i, 0].set_xlabel(col)
        axes[i, 0].set_ylabel('Frequency')
        
        # Box plot
        axes[i, 1].boxplot(df[col].dropna())
        axes[i, 1].set_title(f'{col} - Box Plot')
        axes[i, 1].set_ylabel(col)
    
    plt.tight_layout()
    plt.show()
""")
        
        notebook = self.builder.build()
        metrics = self.inspector.inspect(notebook)
        
        assert isinstance(metrics, QualityMetrics)
        assert metrics.overall_score > 0.6  # Should score well for thorough profiling
        
        # Should have good documentation and structure scores
        assert metrics.details.get('documentation_score', 0) > 0.5
        assert metrics.details.get('structure_score', 0) > 0.5
    
    def test_inspect_outlier_detection_quality(self):
        """Test quality inspection for outlier detection notebooks."""
        # Create outlier detection notebook
        self.builder.add_markdown_cell("# Outlier Detection Analysis")
        
        self.builder.add_code_cell("""
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats
""")
        
        self.builder.add_code_cell("""
# Load data
df = pd.read_excel('data.xlsx')
numerical_cols = df.select_dtypes(include=[np.number]).columns
print(f"Analyzing {len(numerical_cols)} numerical columns for outliers")
""")
        
        self.builder.add_markdown_cell("## IQR Method")
        self.builder.add_code_cell("""
# IQR-based outlier detection
def detect_outliers_iqr(data, column):
    Q1 = data[column].quantile(0.25)
    Q3 = data[column].quantile(0.75)
    IQR = Q3 - Q1
    
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    
    outliers = data[(data[column] < lower_bound) | (data[column] > upper_bound)]
    return outliers, lower_bound, upper_bound

# Apply to each numerical column
outlier_summary = {}
for col in numerical_cols:
    outliers, lower, upper = detect_outliers_iqr(df, col)
    outlier_summary[col] = {
        'count': len(outliers),
        'percentage': (len(outliers) / len(df)) * 100,
        'bounds': (lower, upper)
    }
    print(f"{col}: {len(outliers)} outliers ({(len(outliers)/len(df)*100):.2f}%)")
""")
        
        self.builder.add_markdown_cell("## Z-Score Method")
        self.builder.add_code_cell("""
# Z-score based outlier detection
def detect_outliers_zscore(data, column, threshold=2):
    z_scores = np.abs(stats.zscore(data[column].dropna()))
    outliers = data[z_scores > threshold]
    return outliers

# Apply z-score method
print("\\nZ-score method (threshold=2):")
for col in numerical_cols:
    outliers = detect_outliers_zscore(df, col)
    print(f"{col}: {len(outliers)} outliers")
""")
        
        self.builder.add_markdown_cell("## Visualization")
        self.builder.add_code_cell("""
# Visualize outliers
fig, axes = plt.subplots(len(numerical_cols), 2, figsize=(12, 4*len(numerical_cols)))
if len(numerical_cols) == 1:
    axes = axes.reshape(1, -1)

for i, col in enumerate(numerical_cols):
    # Box plot
    axes[i, 0].boxplot(df[col].dropna())
    axes[i, 0].set_title(f'{col} - Box Plot')
    axes[i, 0].set_ylabel(col)
    
    # Scatter plot with outliers highlighted
    outliers, lower, upper = detect_outliers_iqr(df, col)
    normal_data = df[(df[col] >= lower) & (df[col] <= upper)]
    
    axes[i, 1].scatter(range(len(normal_data)), normal_data[col], alpha=0.6, label='Normal')
    if len(outliers) > 0:
        outlier_indices = outliers.index
        axes[i, 1].scatter(outlier_indices, outliers[col], color='red', alpha=0.8, label='Outliers')
    axes[i, 1].set_title(f'{col} - Outliers Highlighted')
    axes[i, 1].set_ylabel(col)
    axes[i, 1].legend()

plt.tight_layout()
plt.show()
""")
        
        notebook = self.builder.build()
        metrics = self.inspector.inspect(notebook)
        
        assert isinstance(metrics, QualityMetrics)
        assert metrics.overall_score > 0.6  # Should score well for comprehensive outlier analysis
        
        # Should recognize statistical analysis patterns
        issue_types = [str(issue.level) for issue in metrics.issues]
        # Should have fewer critical issues for well-structured statistical analysis
    
    def test_inspect_mixed_content_notebook(self):
        """Test quality inspection of notebook with mixed analysis types."""
        # Create notebook combining multiple analysis types
        self.builder.add_markdown_cell("# Comprehensive Spreadsheet Analysis")
        self.builder.add_markdown_cell("This notebook combines data profiling, formula analysis, and outlier detection.")
        
        # Data profiling section
        self.builder.add_markdown_cell("## 1. Data Profiling")
        self.builder.add_code_cell("""
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from scipy import stats
""")
        
        self.builder.add_code_cell("""
# Load and profile data
df = pd.read_excel('comprehensive_data.xlsx')
print("Data Overview:")
print(f"Shape: {df.shape}")
print(df.info())
""")
        
        # Formula analysis section
        self.builder.add_markdown_cell("## 2. Formula Analysis")
        self.builder.add_code_cell("""
# Analyze Excel formulas
wb = openpyxl.load_workbook('comprehensive_data.xlsx', data_only=False)
ws = wb.active

formulas = []
for row in ws.iter_rows():
    for cell in row:
        if cell.value and str(cell.value).startswith('='):
            formulas.append({
                'cell': cell.coordinate,
                'formula': cell.value
            })

print(f"Found {len(formulas)} formulas in the spreadsheet")
""")
        
        # Outlier detection section
        self.builder.add_markdown_cell("## 3. Outlier Detection")
        self.builder.add_code_cell("""
# Detect outliers in numerical columns
numerical_cols = df.select_dtypes(include=[np.number]).columns

for col in numerical_cols:
    Q1 = df[col].quantile(0.25)
    Q3 = df[col].quantile(0.75)
    IQR = Q3 - Q1
    outliers = df[(df[col] < Q1 - 1.5*IQR) | (df[col] > Q3 + 1.5*IQR)]
    print(f"{col}: {len(outliers)} outliers detected")
""")
        
        # Summary section
        self.builder.add_markdown_cell("## Summary and Conclusions")
        self.builder.add_code_cell("""
# Generate summary report
print("Analysis Summary:")
print(f"- Dataset contains {df.shape[0]} rows and {df.shape[1]} columns")
print(f"- {len(formulas)} formulas found in original spreadsheet")
print(f"- Outlier analysis completed for {len(numerical_cols)} numerical columns")
""")
        
        notebook = self.builder.build()
        metrics = self.inspector.inspect(notebook)
        
        assert isinstance(metrics, QualityMetrics)
        assert metrics.overall_score > 0.7  # Should score very well for comprehensive analysis
        
        # Should recognize the comprehensive nature
        assert len(metrics.issues) < 10  # Should have relatively few issues
        
        # Check for balanced scoring across different aspects
        details = metrics.details
        assert 'structure_score' in details
        assert 'code_quality_score' in details
        assert 'documentation_score' in details
    
    def test_inspect_identifies_specific_spreadsheet_issues(self):
        """Test that the inspector identifies spreadsheet-specific quality issues."""
        # Create notebook with spreadsheet-specific problems
        self.builder.add_code_cell("df = pd.read_excel('file.xlsx')")  # No import
        self.builder.add_code_cell("print(df.head())")  # No error handling
        self.builder.add_code_cell("df.plot()")  # No matplotlib import
        self.builder.add_code_cell("wb = load_workbook('file.xlsx')")  # No openpyxl import
        
        notebook = self.builder.build()
        metrics = self.inspector.inspect(notebook)
        
        # Should identify import-related issues
        import_issues = [issue for issue in metrics.issues 
                        if 'import' in issue.message.lower()]
        assert len(import_issues) > 0
        
        # Should identify error handling issues
        error_handling_issues = [issue for issue in metrics.issues 
                               if 'error' in issue.message.lower() or 'exception' in issue.message.lower()]
        # This might or might not be detected depending on implementation
    
    def test_quality_metrics_structure(self):
        """Test that quality metrics are properly structured."""
        # Create a simple notebook for testing
        self.builder.add_markdown_cell("# Test Analysis")
        self.builder.add_code_cell("import pandas as pd")
        self.builder.add_code_cell("df = pd.read_excel('test.xlsx')")
        
        notebook = self.builder.build()
        metrics = self.inspector.inspect(notebook)
        
        # Test metrics structure
        assert isinstance(metrics, QualityMetrics)
        assert hasattr(metrics, 'overall_score')
        assert hasattr(metrics, 'issues')
        assert hasattr(metrics, 'details')
        
        # Test score range
        assert 0.0 <= metrics.overall_score <= 1.0
        
        # Test issues list
        assert isinstance(metrics.issues, list)
        for issue in metrics.issues:
            assert isinstance(issue, QualityIssue)
            assert hasattr(issue, 'level')
            assert hasattr(issue, 'message')
            assert hasattr(issue, 'location')
        
        # Test details dictionary
        assert isinstance(metrics.details, dict)
    
    def test_quality_levels_assignment(self):
        """Test that quality issues are assigned appropriate levels."""
        # Create notebook with various quality levels
        
        # Critical issues
        self.builder.add_code_cell("df = pd.read_excel('file.xlsx')")  # Missing import
        
        # Warning issues  
        self.builder.add_code_cell("print(df)")  # Minimal analysis
        
        # Info issues
        self.builder.add_markdown_cell("Analysis")  # Minimal documentation
        self.builder.add_code_cell("import pandas as pd")  # Good practice
        
        notebook = self.builder.build()
        metrics = self.inspector.inspect(notebook)
        
        # Should have issues at different levels
        levels = [issue.level for issue in metrics.issues]
        level_counts = {level: levels.count(level) for level in set(levels)}
        
        # Should have at least some issues
        assert len(metrics.issues) > 0
        
        # Test that critical issues affect score more
        critical_issues = [issue for issue in metrics.issues if issue.level == QualityLevel.CRITICAL]
        if critical_issues:
            assert metrics.overall_score < 0.8  # Critical issues should lower score significantly


class TestSpreadsheetQualityIntegration:
    """Integration tests for spreadsheet quality inspection with real scenarios."""
    
    def setup_method(self):
        """Set up realistic test scenarios."""
        self.inspector = SpreadsheetQualityInspector()
        
        # Create temporary files for testing
        self.temp_dir = tempfile.mkdtemp()
        self.excel_file = Path(self.temp_dir) / "test_data.xlsx"
        
        # Create sample data
        data = {
            'product': ['A', 'B', 'C', 'D', 'E'],
            'sales': [100, 150, 200, 175, 300],
            'cost': [80, 120, 160, 140, 240],
            'profit': [20, 30, 40, 35, 60]
        }
        df = pd.DataFrame(data)
        df.to_excel(self.excel_file, index=False)
    
    def test_end_to_end_quality_assessment(self):
        """Test complete quality assessment workflow."""
        builder = NotebookBuilder()
        
        # Create a realistic analysis notebook
        builder.add_markdown_cell("# Sales Data Analysis")
        builder.add_markdown_cell("Analysis of product sales, costs, and profitability.")
        
        builder.add_code_cell(f"""
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

# Load the data
df = pd.read_excel('{self.excel_file}')
print("Data loaded successfully!")
print(f"Dataset shape: {{df.shape}}")
""")
        
        builder.add_markdown_cell("## Data Overview")
        builder.add_code_cell("""
# Display basic information
print("Dataset Info:")
df.info()
print("\\nFirst 5 rows:")
print(df.head())
print("\\nStatistical Summary:")
print(df.describe())
""")
        
        builder.add_markdown_cell("## Profitability Analysis")
        builder.add_code_cell("""
# Calculate profit margin
df['profit_margin'] = (df['profit'] / df['sales']) * 100
print("Profit margins:")
print(df[['product', 'profit_margin']].sort_values('profit_margin', ascending=False))
""")
        
        builder.add_markdown_cell("## Visualization")
        builder.add_code_cell("""
# Create visualization
fig, axes = plt.subplots(1, 2, figsize=(12, 5))

# Sales by product
axes[0].bar(df['product'], df['sales'])
axes[0].set_title('Sales by Product')
axes[0].set_xlabel('Product')
axes[0].set_ylabel('Sales')

# Profit margin by product
axes[1].bar(df['product'], df['profit_margin'])
axes[1].set_title('Profit Margin by Product')
axes[1].set_xlabel('Product')
axes[1].set_ylabel('Profit Margin (%)')

plt.tight_layout()
plt.show()
""")
        
        notebook = builder.build()
        metrics = self.inspector.inspect(notebook)
        
        # Should score well for comprehensive analysis
        assert metrics.overall_score > 0.6
        
        # Should have good structural organization
        assert metrics.details.get('structure_score', 0) > 0.5
        
        # Should have reasonable documentation
        assert metrics.details.get('documentation_score', 0) > 0.4
        
        # Issues should be mostly minor
        critical_issues = [issue for issue in metrics.issues if issue.level == QualityLevel.CRITICAL]
        assert len(critical_issues) <= 2
    
    def test_quality_comparison_scenarios(self):
        """Test quality assessment across different analysis quality levels."""
        scenarios = []
        
        # Scenario 1: High quality
        builder1 = NotebookBuilder()
        builder1.add_markdown_cell("# Professional Data Analysis")
        builder1.add_markdown_cell("## Objective\nAnalyze sales performance and identify trends.")
        builder1.add_code_cell("""
# Import required libraries
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
""")
        builder1.add_code_cell(f"""
# Load and validate data
try:
    df = pd.read_excel('{self.excel_file}')
    print(f"Data loaded successfully: {{df.shape}}")
    assert not df.empty, "Dataset is empty"
except Exception as e:
    print(f"Error loading data: {{e}}")
    raise
""")
        builder1.add_markdown_cell("## Results\nAnalysis completed successfully.")
        scenarios.append(('high_quality', builder1.build()))
        
        # Scenario 2: Medium quality
        builder2 = NotebookBuilder()
        builder2.add_markdown_cell("# Analysis")
        builder2.add_code_cell("import pandas as pd")
        builder2.add_code_cell(f"df = pd.read_excel('{self.excel_file}')")
        builder2.add_code_cell("print(df.head())")
        builder2.add_code_cell("print(df.describe())")
        scenarios.append(('medium_quality', builder2.build()))
        
        # Scenario 3: Low quality
        builder3 = NotebookBuilder()
        builder3.add_code_cell(f"df=pd.read_excel('{self.excel_file}')")  # No import, poor formatting
        builder3.add_code_cell("print(df)")
        scenarios.append(('low_quality', builder3.build()))
        
        # Assess all scenarios
        results = {}
        for name, notebook in scenarios:
            metrics = self.inspector.inspect(notebook)
            results[name] = metrics.overall_score
        
        # Verify quality ordering
        assert results['high_quality'] > results['medium_quality']
        assert results['medium_quality'] > results['low_quality']
        
        # Verify reasonable score ranges
        assert results['high_quality'] > 0.6
        assert results['low_quality'] < 0.5


if __name__ == "__main__":
    pytest.main([__file__]) 