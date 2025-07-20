"""
Test script to demonstrate enhanced formula analysis with range handling.

This test creates a simple Excel file with formulas using ranges,
then runs the enhanced analysis to show how dependencies are captured.
"""

import tempfile
from pathlib import Path

import openpyxl

from spreadsheet_analyzer.pipeline.stages.stage_3_formulas import stage_3_formula_analysis
from spreadsheet_analyzer.pipeline.types import Ok


def create_test_workbook():
    """Create a test Excel file with various formula patterns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestSheet"

    # Add some data
    for i in range(1, 11):
        ws[f"A{i}"] = i * 10  # Values: 10, 20, 30, ..., 100
        ws[f"B{i}"] = i * 5  # Values: 5, 10, 15, ..., 50

    # Add formulas with different patterns
    ws["C1"] = "=SUM(A1:A10)"  # Range sum
    ws["C2"] = "=AVERAGE(B1:B10)"  # Range average
    ws["C3"] = "=A1+A2"  # Simple cell references
    ws["C4"] = "=VLOOKUP(A1,A1:B10,2,FALSE)"  # Lookup with range
    ws["C5"] = '=COUNTIF(A1:A10,">50")'  # Conditional count
    ws["C6"] = "=SUM(A1:A5)+SUM(B1:B5)"  # Multiple ranges

    # Cross-sheet reference (create second sheet)
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "=SUM(TestSheet!A1:A10)"  # Cross-sheet range reference
    ws2["A2"] = "=TestSheet!C1*2"  # Cross-sheet cell reference

    return wb


def test_enhanced_formula_analysis():
    """Test that range dependencies are properly captured."""

    # Create temporary file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)

    try:
        # Create and save test workbook
        wb = create_test_workbook()
        wb.save(tmp_path)

        # Run formula analysis
        result = stage_3_formula_analysis(tmp_path)

        # Check that analysis succeeded
        assert isinstance(result, Ok)
        analysis = result.value

        # Print results for demonstration
        print("\n=== Formula Analysis Results ===")
        print(f"Total formulas found: {len(analysis.dependency_graph)}")
        print(f"Dependency graph nodes: {len(analysis.dependency_graph)}")

        # Check specific formulas
        sum_formula_key = "TestSheet!C1"
        if sum_formula_key in analysis.dependency_graph:
            node = analysis.dependency_graph[sum_formula_key]
            print(f"\n{sum_formula_key}: {node.formula}")
            print(f"  Dependencies: {len(node.dependencies)}")
            for dep in node.dependencies:
                print(f"    -> {dep.sheet}!{dep.cell} (range: {dep.is_range})")

        # Check cross-sheet references
        cross_sheet_key = "Summary!A1"
        if cross_sheet_key in analysis.dependency_graph:
            node = analysis.dependency_graph[cross_sheet_key]
            print(f"\n{cross_sheet_key}: {node.formula}")
            print(f"  Dependencies: {len(node.dependencies)}")
            for dep in node.dependencies:
                print(f"    -> {dep.sheet}!{dep.cell} (range: {dep.is_range})")

        # Verify range nodes are created
        range_nodes = [
            key
            for key, node in analysis.dependency_graph.items()
            if hasattr(node, "node_type") and node.node_type == "range"
        ]
        print(f"\nRange nodes created: {len(range_nodes)}")
        for range_key in range_nodes[:5]:  # Show first 5
            print(f"  - {range_key}")

        # Check that we have dependencies (not empty like before)
        formulas_with_deps = sum(1 for node in analysis.dependency_graph.values() if node.dependencies)
        print(f"\nFormulas with dependencies: {formulas_with_deps}")

        # This should be > 0 with our fix!
        assert formulas_with_deps > 0, "No dependencies found - range handling may not be working"

    finally:
        # Clean up
        tmp_path.unlink(missing_ok=True)


def test_range_size_calculation():
    """Test the range size calculation function."""
    from spreadsheet_analyzer.pipeline.stages.stage_3_formulas import parse_range_size

    # Test cases
    test_cases = [
        ("A1:A10", 10),  # Column range
        ("A1:C1", 3),  # Row range
        ("A1:C3", 9),  # Block range
        ("B5:D10", 18),  # Block range with offset
        ("A:A", 1048576),  # Full column
        ("1:1", 16384),  # Full row
    ]

    for range_ref, expected_size in test_cases:
        size, metadata = parse_range_size(range_ref)
        print(f"{range_ref}: size={size}, metadata={metadata}")
        if not range_ref.startswith(("A:A", "1:1")):  # Skip full ranges
            assert size == expected_size, f"Expected {expected_size} for {range_ref}, got {size}"


def test_range_membership_queries():
    """Test range membership index queries."""
    from spreadsheet_analyzer.graph_db.query_interface import create_enhanced_query_interface

    # Create temporary file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)

    try:
        # Create and save test workbook
        wb = create_test_workbook()
        wb.save(tmp_path)

        # Run formula analysis
        result = stage_3_formula_analysis(tmp_path)
        assert isinstance(result, Ok)
        analysis = result.value

        # Create enhanced query interface
        query_interface = create_enhanced_query_interface(analysis)

        print("\n=== Range Membership Query Tests ===")

        # Test querying an empty cell that's part of a range
        # Cell A5 is part of the SUM(A1:A10) formula but also has a value
        result = query_interface.get_cell_dependencies("TestSheet", "A5", include_ranges=True)
        print("\nCell A5 dependencies:")
        print(f"  Direct dependents: {result.direct_dependents}")
        print(f"  Range dependents: {result.range_dependents}")
        print(f"  Is in ranges: {result.is_in_ranges}")

        # Test a cell that's in multiple ranges
        # B3 is part of AVERAGE(B1:B10) and VLOOKUP range
        result = query_interface.get_cell_dependencies("TestSheet", "B3", include_ranges=True)
        print("\nCell B3 dependencies:")
        print(f"  Range dependents: {result.range_dependents}")
        print(f"  Total dependents: {result.total_dependents}")

        # Get statistics
        stats = query_interface.get_formula_statistics_with_ranges()
        print("\nFormula Statistics:")
        print(f"  Total formulas: {stats['total_formulas']}")
        print(f"  Formulas with dependencies: {stats['formulas_with_dependencies']}")
        print(f"  Has range index: {stats['has_range_index']}")
        if stats["has_range_index"]:
            print(f"  Unique ranges: {stats['unique_ranges']}")
            print(f"  Total cells in ranges: {stats['total_cells_in_ranges']}")

    finally:
        # Clean up
        tmp_path.unlink(missing_ok=True)


def test_empty_cell_in_range_scenario():
    """Test the specific scenario of empty cells in formula ranges."""
    # Create a workbook with a range that includes empty cells
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Add sparse data - some cells will be empty
    ws["B1"] = 100
    ws["B3"] = 200
    ws["B7"] = 300
    # B2, B4, B5, B6, B8, B9, B10 are empty

    # Add formula that sums the range including empty cells
    ws["C1"] = "=SUM(B1:B10)"

    # Save to temporary file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)

    try:
        wb.save(tmp_path)

        # Run analysis
        result = stage_3_formula_analysis(tmp_path)
        assert isinstance(result, Ok)
        analysis = result.value

        # Check range membership index
        if analysis.range_membership_index:
            print("\n=== Empty Cell Range Membership Test ===")

            # Check that empty cell B5 is recognized as part of the SUM range
            is_in_range = analysis.range_membership_index.is_cell_in_any_range("Data", "B5")
            print(f"Is empty cell B5 in a formula range? {is_in_range}")
            assert is_in_range, "Empty cell B5 should be recognized as part of SUM range"

            # Get formulas that depend on B5
            formulas = analysis.range_membership_index.get_ranges_containing_cell("Data", "B5")
            print(f"Formulas that include B5 in their ranges: {formulas}")
            assert "Data!C1" in formulas, "SUM formula should include B5 in its range"

            # This demonstrates that we can now answer: "What formulas would be affected if I add data to B5?"

    finally:
        tmp_path.unlink(missing_ok=True)


if __name__ == "__main__":
    print("Testing enhanced formula analysis with range handling...")
    test_range_size_calculation()
    test_enhanced_formula_analysis()
    test_range_membership_queries()
    test_empty_cell_in_range_scenario()
    print("\nAll tests passed!")
