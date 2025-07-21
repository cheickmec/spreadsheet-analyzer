"""Pytest configuration and fixtures."""

import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import pytest

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))


@pytest.fixture
def sample_excel_file(tmp_path):
    """Create a sample Excel file for testing."""
    file_path = tmp_path / "sample.xlsx"

    # Create workbook with openpyxl directly
    wb = openpyxl.Workbook()

    # Simple data sheet
    ws1 = wb.active
    ws1.title = "Employees"
    ws1.append(["Name", "Age", "Salary"])
    ws1.append(["Alice", 25, 50000])
    ws1.append(["Bob", 30, 60000])
    ws1.append(["Charlie", 35, 70000])

    # Sheet with formulas
    ws2 = wb.create_sheet("Sales")
    ws2.append(["Product", "Price", "Quantity", "Total"])
    ws2.append(["A", 10, 5, "=B2*C2"])
    ws2.append(["B", 20, 3, "=B3*C3"])
    ws2.append(["C", 30, 2, "=B4*C4"])

    # Summary sheet
    ws3 = wb.create_sheet("Summary")
    ws3.append(["Metric", "Value"])
    ws3.append(["Total Employees", "=COUNTA(Employees!A:A)-1"])
    ws3.append(["Total Sales", "=SUM(Sales!D:D)"])

    # Save the workbook
    wb.save(file_path)
    wb.close()

    return file_path


@pytest.fixture
def mock_llm_response():
    """Mock LLM response for testing AI components."""

    def _mock_response(prompt: str) -> str:
        responses = {
            "analyze": "This spreadsheet contains employee data with salary information.",
            "pattern": "Found repeating pattern in column formulas.",
            "validate": "All formulas are correctly structured.",
        }

        for key, response in responses.items():
            if key in prompt.lower():
                return response

        return "Generic analysis response."

    return _mock_response


@pytest.fixture(scope="session")
def test_data_dir():
    """Path to test data directory."""
    return Path(__file__).parent / "fixtures" / "inputs"


# Clean up any test files after test run
@pytest.fixture(autouse=True)
def cleanup_test_files(request):
    """Clean up temporary test files after test completion."""
    return
    # Cleanup happens after test
    # Add any cleanup logic here if needed


@pytest.fixture
def performance_tracker() -> dict[str, list[float]]:
    """Track performance metrics across tests."""
    return defaultdict(list)


@pytest.fixture
def excel_builder():
    """Create ExcelTestDataBuilder instance for tests."""
    from tests.test_base import ExcelTestDataBuilder

    return ExcelTestDataBuilder()
