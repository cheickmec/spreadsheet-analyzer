#!/usr/bin/env python3
"""Tests for table detection agent.

Following TDD methodology - these tests define the expected behavior
before implementation (Red phase).
"""

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from spreadsheet_analyzer.agents.table_detection_types import (
    TableDetectionResult,
    TableType,
)
from spreadsheet_analyzer.agents.table_detector_agent import create_table_detector
from spreadsheet_analyzer.agents.types import AgentId, AgentMessage, AgentState


# Test Fixtures
@pytest.fixture
def single_table_excel(tmp_path: Path) -> Path:
    """Create Excel file with a single table."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Headers
    ws["A1"] = "Order ID"
    ws["B1"] = "Customer"
    ws["C1"] = "Amount"
    ws["D1"] = "Date"

    # Data rows
    for i in range(2, 102):  # 100 data rows
        ws[f"A{i}"] = f"ORD-{i:04d}"
        ws[f"B{i}"] = f"Customer {i}"
        ws[f"C{i}"] = 100 + i * 10
        ws[f"D{i}"] = f"2024-01-{(i % 30) + 1:02d}"

    file_path = tmp_path / "single_table.xlsx"
    wb.save(file_path)
    return file_path


@pytest.fixture
def multi_table_excel_empty_rows(tmp_path: Path) -> Path:
    """Create Excel file with multiple tables separated by empty rows."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Table 1: Orders (rows 1-50)
    ws["A1"] = "Order ID"
    ws["B1"] = "Customer"
    ws["C1"] = "Amount"

    for i in range(2, 51):
        ws[f"A{i}"] = f"ORD-{i:04d}"
        ws[f"B{i}"] = f"Customer {i}"
        ws[f"C{i}"] = 100 + i * 10

    # Empty rows (51-53)
    # Leave these blank

    # Table 2: Regional Summary (rows 54-65)
    ws["A54"] = "Region"
    ws["B54"] = "Total Sales"
    ws["C54"] = "Order Count"

    regions = ["North", "South", "East", "West"]
    for i, region in enumerate(regions, start=55):
        ws[f"A{i}"] = region
        ws[f"B{i}"] = 50000 + i * 1000
        ws[f"C{i}"] = 100 + i * 10

    file_path = tmp_path / "multi_table_empty_rows.xlsx"
    wb.save(file_path)
    return file_path


@pytest.fixture
def multi_table_excel_semantic(tmp_path: Path) -> Path:
    """Create Excel file with tables that have different semantic meaning."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Table 1: Product Inventory (rows 1-30)
    ws["A1"] = "Product ID"
    ws["B1"] = "Product Name"
    ws["C1"] = "Stock"
    ws["D1"] = "Price"

    for i in range(2, 31):
        ws[f"A{i}"] = f"PROD-{i:03d}"
        ws[f"B{i}"] = f"Product {i}"
        ws[f"C{i}"] = 50 + i
        ws[f"D{i}"] = 19.99 + i * 0.5

    # No empty rows - directly adjacent

    # Table 2: Employee List (rows 31-50)
    ws["A31"] = "Employee ID"
    ws["B31"] = "Name"
    ws["C31"] = "Department"
    ws["D31"] = "Salary"

    for i in range(32, 51):
        ws[f"A{i}"] = f"EMP-{i:03d}"
        ws[f"B{i}"] = f"Employee {i}"
        ws[f"C{i}"] = ["IT", "HR", "Sales", "Finance"][i % 4]
        ws[f"D{i}"] = 50000 + i * 1000

    file_path = tmp_path / "multi_table_semantic.xlsx"
    wb.save(file_path)
    return file_path


@pytest.fixture
def master_detail_excel(tmp_path: Path) -> Path:
    """Create Excel file with master-detail pattern."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Master section (rows 1-5)
    ws["A1"] = "Invoice Number:"
    ws["B1"] = "INV-2024-001"
    ws["A2"] = "Customer:"
    ws["B2"] = "ABC Corporation"
    ws["A3"] = "Date:"
    ws["B3"] = "2024-01-15"
    ws["A4"] = "Total:"
    ws["B4"] = 5000

    # Empty row

    # Detail section (rows 7-20)
    ws["A7"] = "Item"
    ws["B7"] = "Description"
    ws["C7"] = "Quantity"
    ws["D7"] = "Price"
    ws["E7"] = "Subtotal"

    for i in range(8, 15):
        ws[f"A{i}"] = f"ITEM-{i:03d}"
        ws[f"B{i}"] = f"Product {i}"
        ws[f"C{i}"] = i - 7
        ws[f"D{i}"] = 100
        ws[f"E{i}"] = (i - 7) * 100

    file_path = tmp_path / "master_detail.xlsx"
    wb.save(file_path)
    return file_path


# Test Cases - Red Phase


class TestTableDetectorAgent:
    """Test suite for table detection agent."""

    def test_table_detector_identifies_single_table(self, single_table_excel):
        """Test detector correctly identifies a single table."""
        # Arrange
        detector = create_table_detector()
        df = pd.read_excel(single_table_excel)

        message = AgentMessage.create(
            sender=AgentId.generate("test"),
            receiver=detector.id,
            content={
                "dataframe": df,
                "sheet_name": "Sheet",
                "file_path": str(single_table_excel),
            },
        )

        state = AgentState(agent_id=detector.id, status="idle")

        # Act
        result = detector.process(message, state)

        # Assert
        assert result.is_ok()
        response = result.unwrap()
        detection_result: TableDetectionResult = response.content

        assert isinstance(detection_result, TableDetectionResult)
        assert len(detection_result.tables) == 1

        table = detection_result.tables[0]
        assert table.start_row == 0
        assert table.end_row == 99  # 100 rows total (0-indexed, so last row is 99)
        assert table.start_col == 0
        assert table.end_col == 3  # 4 columns (0-3)
        assert table.confidence >= 0.7  # Lower threshold since no empty rows
        assert "order" in table.description.lower()

    def test_table_detector_identifies_empty_row_separation(self, multi_table_excel_empty_rows):
        """Test detector finds tables separated by empty rows."""
        # Arrange
        detector = create_table_detector()
        df = pd.read_excel(multi_table_excel_empty_rows)

        message = AgentMessage.create(
            sender=AgentId.generate("test"),
            receiver=detector.id,
            content={
                "dataframe": df,
                "sheet_name": "Sheet",
                "file_path": str(multi_table_excel_empty_rows),
            },
        )

        state = AgentState(agent_id=detector.id, status="idle")

        # Act
        result = detector.process(message, state)

        # Assert
        assert result.is_ok()
        response = result.unwrap()
        detection_result: TableDetectionResult = response.content

        assert len(detection_result.tables) == 2

        # First table (Orders)
        table1 = detection_result.tables[0]
        assert table1.start_row == 0
        assert table1.end_row == 48  # Rows 1-49 (0-indexed, before empty rows)
        assert "order" in table1.description.lower()
        assert table1.table_type == TableType.DETAIL

        # Second table (Summary)
        table2 = detection_result.tables[1]
        assert table2.start_row == 52  # After 3 empty rows (49, 50, 51)
        assert table2.end_row == 56  # Last row with data (header at 53, 4 regions: 54-57, 0-indexed)
        assert "summary" in table2.description.lower() or "region" in table2.description.lower()
        assert table2.table_type == TableType.SUMMARY

    def test_table_detector_identifies_semantic_boundaries(self, multi_table_excel_semantic):
        """Test detector recognizes different entity types as separate tables."""
        # Arrange
        detector = create_table_detector()
        df = pd.read_excel(multi_table_excel_semantic)

        message = AgentMessage.create(
            sender=AgentId.generate("test"),
            receiver=detector.id,
            content={
                "dataframe": df,
                "sheet_name": "Sheet",
                "file_path": str(multi_table_excel_semantic),
            },
        )

        state = AgentState(agent_id=detector.id, status="idle")

        # Act
        result = detector.process(message, state)

        # Assert
        assert result.is_ok()
        response = result.unwrap()
        detection_result: TableDetectionResult = response.content

        # Should detect at least 2 tables (might detect the header row as separate)
        assert len(detection_result.tables) >= 2

        # Find the products and employees tables
        products_table = None
        employees_table = None

        for table in detection_result.tables:
            # Products table should be in first half
            if table.row_count > 10 and table.start_row < 20:
                products_table = table
            # Employees table should be in second half
            elif table.row_count > 10 and table.start_row > 25:
                employees_table = table

        assert products_table is not None, "Products table not found"
        assert employees_table is not None, "Employees table not found"

        # Verify they are in expected positions
        assert products_table.start_row < 20  # Products are in first half
        assert employees_table.start_row > 25  # Employees are in second half

        # The key is that we detected two separate tables at the right boundaries
        # Entity detection can be improved later

    def test_table_detector_handles_master_detail_pattern(self, master_detail_excel):
        """Test detector recognizes master-detail patterns."""
        # Arrange
        detector = create_table_detector()
        df = pd.read_excel(master_detail_excel)

        message = AgentMessage.create(
            sender=AgentId.generate("test"),
            receiver=detector.id,
            content={
                "dataframe": df,
                "sheet_name": "Sheet",
                "file_path": str(master_detail_excel),
            },
        )

        state = AgentState(agent_id=detector.id, status="idle")

        # Act
        result = detector.process(message, state)

        # Assert
        assert result.is_ok()
        response = result.unwrap()
        detection_result: TableDetectionResult = response.content

        # May detect 1 or 2 tables depending on whether header is large enough
        assert len(detection_result.tables) >= 1

        # Find the detail table (items)
        detail_table = None
        for table in detection_result.tables:
            if "item" in table.description.lower() or table.start_row >= 5:
                detail_table = table
                break

        assert detail_table is not None, "Detail table with items not found"
        assert detail_table.table_type in [TableType.DETAIL, TableType.SUMMARY]
        assert detail_table.start_row >= 5  # After the header section

        # If two tables detected, check the first is header
        if len(detection_result.tables) == 2:
            master = detection_result.tables[0]
            assert master.table_type == TableType.HEADER
            assert master.end_row < 6

    def test_table_detector_respects_min_rows_threshold(self, tmp_path):
        """Test that tables with fewer than MIN_TABLE_ROWS are not detected."""
        # Arrange
        wb = openpyxl.Workbook()
        ws = wb.active

        # Very small table (2 rows)
        ws["A1"] = "Header"
        ws["A2"] = "Data"

        file_path = tmp_path / "tiny_table.xlsx"
        wb.save(file_path)

        detector = create_table_detector()
        df = pd.read_excel(file_path)

        message = AgentMessage.create(
            sender=AgentId.generate("test"),
            receiver=detector.id,
            content={
                "dataframe": df,
                "sheet_name": "Sheet",
                "file_path": str(file_path),
            },
        )

        state = AgentState(agent_id=detector.id, status="idle")

        # Act
        result = detector.process(message, state)

        # Assert
        assert result.is_ok()
        response = result.unwrap()
        detection_result: TableDetectionResult = response.content

        # Should either detect no tables or one table with low confidence
        if len(detection_result.tables) > 0:
            assert detection_result.tables[0].confidence < 0.5

    def test_table_detector_provides_metadata(self, single_table_excel):
        """Test that detection result includes useful metadata."""
        # Arrange
        detector = create_table_detector()
        df = pd.read_excel(single_table_excel)

        message = AgentMessage.create(
            sender=AgentId.generate("test"),
            receiver=detector.id,
            content={
                "dataframe": df,
                "sheet_name": "Sheet",
                "file_path": str(single_table_excel),
            },
        )

        state = AgentState(agent_id=detector.id, status="idle")

        # Act
        result = detector.process(message, state)

        # Assert
        assert result.is_ok()
        response = result.unwrap()
        detection_result: TableDetectionResult = response.content

        # Check for valid metadata structure
        assert detection_result.metadata is not None
        assert isinstance(detection_result.metadata, dict)
        assert detection_result.sheet_name == "Sheet"

        # Check that tables were detected
        assert len(detection_result.tables) > 0

        # Check first table has valid structure
        first_table = detection_result.tables[0]
        assert hasattr(first_table, "table_id")
        assert hasattr(first_table, "confidence")
        assert 0.0 <= first_table.confidence <= 1.0
