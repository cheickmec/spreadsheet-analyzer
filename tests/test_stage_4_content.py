"""
Tests for Stage 4: Content Intelligence.

Tests the actual functionality that was broken by read_only mode.
"""

from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl

from spreadsheet_analyzer.pipeline.stages.stage_4_content import (
    analyze_sheet_content,
    stage_4_content_intelligence,
)
from spreadsheet_analyzer.pipeline.types import Ok


def test_analyze_sheet_content_with_mock():
    """Test that analyze_sheet_content works with regular worksheet (not read-only)."""
    # Create a mock worksheet
    mock_sheet = MagicMock()
    mock_sheet.max_row = 10
    mock_sheet.max_column = 3

    # Create mock cells with column_letter attribute
    mock_cells = []
    for col_idx, col_letter in enumerate(["A", "B", "C"], 1):
        column_cells = []
        for row_idx in range(1, 11):
            cell = MagicMock()
            cell.value = f"{col_letter}{row_idx}" if row_idx == 1 else row_idx * col_idx
            cell.column_letter = col_letter
            column_cells.append(cell)
        mock_cells.append(column_cells)

    # Mock iter_cols to return our cells
    mock_sheet.iter_cols.return_value = mock_cells

    # This should not raise AttributeError
    column_analyses, patterns = analyze_sheet_content(mock_sheet, "TestSheet", sample_size=10)

    # Verify the function could access column_letter
    assert len(column_analyses) == 3
    assert "A1" in column_analyses
    assert "B1" in column_analyses
    assert "C1" in column_analyses


def test_stage_4_handles_readonly_worksheet_error():
    """Test that stage 4 properly handles worksheets and doesn't use read_only mode."""
    test_file = Path("test-files/test_content.xlsx")

    # Create a simple test Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestSheet"

    # Add some data
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Test"
    ws["B2"] = 123

    # Save it
    test_file.parent.mkdir(exist_ok=True)
    wb.save(test_file)
    wb.close()

    try:
        # Run stage 4
        result = stage_4_content_intelligence(test_file, sample_size=100)

        # Should succeed
        assert isinstance(result, Ok)

        # Verify it analyzed the content
        content_analysis = result.value
        assert content_analysis.data_quality_score >= 0
        assert "total_sheets_analyzed" in content_analysis.key_metrics
        assert content_analysis.key_metrics["total_sheets_analyzed"] == 1

    finally:
        # Clean up
        if test_file.exists():
            test_file.unlink()


@patch("openpyxl.load_workbook")
def test_read_only_mode_not_used(mock_load_workbook):
    """Verify that stage_4 doesn't use read_only=True."""
    # Mock the workbook
    mock_wb = MagicMock()
    mock_wb.sheetnames = ["Sheet1"]
    mock_ws = MagicMock()
    mock_ws.max_row = 1
    mock_ws.max_column = 1
    mock_ws.iter_cols.return_value = []
    mock_wb.__getitem__.return_value = mock_ws
    mock_wb.close = MagicMock()

    mock_load_workbook.return_value = mock_wb

    # Run stage 4
    _ = stage_4_content_intelligence(Path("dummy.xlsx"))

    # Verify load_workbook was called with read_only=False
    mock_load_workbook.assert_called_once()
    call_args = mock_load_workbook.call_args
    assert call_args[1]["read_only"] is False
    assert call_args[1]["data_only"] is True
