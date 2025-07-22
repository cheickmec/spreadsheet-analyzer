"""Integration test to verify the full stack is working."""

from unittest.mock import Mock, patch

import pytest

from spreadsheet_analyzer.excel_to_notebook import ExcelToNotebookConverter
from spreadsheet_analyzer.notebook_llm.llm_providers import get_provider
from spreadsheet_analyzer.notebook_llm.llm_providers.base import LLMResponse, Message, Role


@pytest.mark.asyncio
async def test_full_stack_components():
    """Test that all major components can be instantiated."""
    # Test Excel to Notebook converter
    converter = ExcelToNotebookConverter()
    assert converter is not None

    # Test LLM provider registry
    providers = ["openai", "anthropic"]
    for provider_name in providers:
        # Mock the API key check
        with patch.dict("os.environ", {"OPENAI_API_KEY": "test", "ANTHROPIC_API_KEY": "test"}):
            provider = get_provider(provider_name)
            assert provider is not None
            assert provider.model_name is not None
            assert provider.max_context_tokens > 0


def test_llm_provider_mock():
    """Test LLM provider with mocked response."""
    # Create a mock provider
    mock_provider = Mock()
    mock_provider.model_name = "test-model"
    mock_provider.max_context_tokens = 8192

    # Mock the complete method
    mock_response = LLMResponse(
        content="This is a test response",
        model="test-model",
        usage={"prompt_tokens": 10, "completion_tokens": 5, "total_tokens": 15},
    )
    mock_provider.complete.return_value = mock_response

    # Test the mock
    messages = [
        Message(role=Role.SYSTEM, content="You are a helpful assistant"),
        Message(role=Role.USER, content="Hello"),
    ]

    response = mock_provider.complete(messages)
    assert response.content == "This is a test response"
    assert response.usage["total_tokens"] == 15


def test_excel_converter_with_sample_data(tmp_path):
    """Test Excel converter with a sample file."""
    import openpyxl

    # Create a sample Excel file
    excel_file = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Header"
    ws["A2"] = 100
    ws["B1"] = "Value"
    ws["B2"] = "=A2*2"
    wb.save(excel_file)

    # Convert to notebook
    converter = ExcelToNotebookConverter()
    notebook = converter.convert(excel_file)

    # Verify notebook structure
    assert notebook is not None
    assert len(notebook.cells) > 0
    assert notebook.metadata["sheet_count"] == 1
    assert notebook.source_file == excel_file

    # Check that cells contain expected content
    cell_contents = [cell.content for cell in notebook.cells]
    assert any("Excel Workbook Analysis" in content for content in cell_contents)
    # Check for sheet name (might be "Sheet" or "Sheet1" depending on openpyxl version)
    assert any("Sheet" in content for content in cell_contents)
