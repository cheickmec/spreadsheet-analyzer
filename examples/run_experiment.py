"""Example script demonstrating the full LLM Excel analysis stack.

This script shows how to use the complete system to analyze Excel files
with different LLMs and strategies.
"""

import asyncio
import logging
from pathlib import Path

from spreadsheet_analyzer.experiments import ExperimentConfig, ExperimentRunner
from spreadsheet_analyzer.notebook_llm.strategies.base import AnalysisFocus

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)


async def main():
    """Run a sample experiment comparing LLMs on Excel analysis."""
    # Create sample Excel file path (you'll need to provide your own)
    excel_files = [
        Path("examples/data/sample_financial.xlsx"),  # Replace with your file
    ]

    # Skip files that don't exist
    excel_files = [f for f in excel_files if f.exists()]

    if not excel_files:
        print("No Excel files found. Please add Excel files to examples/data/")
        print("Creating a minimal example...")

        # Create a minimal Excel file for demonstration
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Data"

        # Add headers
        headers = ["Month", "Product A", "Product B", "Total"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Add data
        months = ["January", "February", "March", "April"]
        sales_a = [1000, 1200, 1100, 1300]
        sales_b = [800, 900, 950, 1000]

        for row, (month, a, b) in enumerate(zip(months, sales_a, sales_b, strict=False), 2):
            ws.cell(row=row, column=1, value=month)
            ws.cell(row=row, column=2, value=a)
            ws.cell(row=row, column=3, value=b)
            # Add formula for total
            ws.cell(row=row, column=4, value=f"=B{row}+C{row}")

        # Save the file
        Path("examples/data").mkdir(parents=True, exist_ok=True)
        sample_file = Path("examples/data/sample_sales.xlsx")
        wb.save(sample_file)
        excel_files = [sample_file]
        print(f"Created sample file: {sample_file}")

    # Configure experiment
    config = ExperimentConfig(
        name="llm_excel_comparison",
        excel_files=excel_files,
        llm_providers=["openai", "anthropic"],
        llm_models={
            "openai": "gpt-4",
            "anthropic": "claude-3-sonnet-20240229",
        },
        strategies=["hierarchical"],  # Start with one strategy
        analysis_focus=AnalysisFocus.GENERAL,
        output_dir=Path("experiments/results"),
        max_parallel_kernels=1,  # Run sequentially for demo
        timeout_seconds=120,  # 2 minute timeout
    )

    # Create and run experiment
    runner = ExperimentRunner(config)

    print(f"\nRunning experiment: {config.name}")
    print(f"Excel files: {[f.name for f in config.excel_files]}")
    print(f"Providers: {config.llm_providers}")
    print(f"Strategies: {config.strategies}")
    print("\nThis may take a few minutes...\n")

    try:
        results = await runner.run()

        # Print summary
        print("\n=== Experiment Results ===")
        for result in results:
            print(f"\nProvider: {result.provider}/{result.model}")
            print(f"Strategy: {result.strategy}")
            print(f"File: {result.excel_file.name}")
            print(f"Success: {result.success}")
            print(f"Duration: {result.duration_seconds:.2f}s")
            if result.tokens_used:
                print(f"Tokens: {result.tokens_used.get('total', 'N/A')}")

            if result.insights:
                print("Insights:")
                for insight in result.insights[:3]:
                    print(f"  - {insight}")

            if result.errors:
                print("Errors:")
                for error in result.errors:
                    print(f"  - {error}")

        print(f"\nResults saved to: {config.output_dir}")

    except Exception as e:
        print(f"Experiment failed: {e}")
        import traceback

        traceback.print_exc()


if __name__ == "__main__":
    # Run the async main function
    asyncio.run(main())
