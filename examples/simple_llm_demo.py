"""Simple demonstration of the LLM Excel analysis stack."""

import asyncio
import os
from pathlib import Path

from spreadsheet_analyzer.excel_to_notebook import ExcelToNotebookConverter
from spreadsheet_analyzer.notebook_llm.llm_providers import get_provider, list_providers


def demo_excel_to_notebook():
    """Demonstrate Excel to Notebook conversion."""
    print("=== Excel to Notebook Conversion Demo ===\n")

    # Create a sample Excel file
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    # Add headers
    ws["A1"] = "Month"
    ws["B1"] = "Sales"
    ws["C1"] = "Target"
    ws["D1"] = "Variance"

    # Add data
    months = ["Jan", "Feb", "Mar"]
    sales = [1000, 1200, 900]
    targets = [1100, 1100, 1100]

    for i, (month, sale, target) in enumerate(zip(months, sales, targets, strict=False), 2):
        ws[f"A{i}"] = month
        ws[f"B{i}"] = sale
        ws[f"C{i}"] = target
        ws[f"D{i}"] = f"=B{i}-C{i}"

    # Save file
    excel_file = Path("demo_sales.xlsx")
    wb.save(excel_file)
    print(f"Created sample Excel file: {excel_file}")

    # Convert to notebook
    converter = ExcelToNotebookConverter()
    notebook = converter.convert(excel_file)

    print(f"\nNotebook created with {len(notebook.cells)} cells:")
    for i, cell in enumerate(notebook.cells[:3]):  # Show first 3 cells
        print(f"\nCell {i + 1} ({cell.cell_type.value}):")
        print(cell.content[:200] + "..." if len(cell.content) > 200 else cell.content)

    # Clean up
    excel_file.unlink()

    return notebook


def demo_llm_providers():
    """Demonstrate LLM provider capabilities."""
    print("\n=== LLM Provider Demo ===\n")

    # List available providers
    providers = list_providers()
    print(f"Available providers: {', '.join(providers)}")

    # Check for API keys
    has_openai = bool(os.environ.get("OPENAI_API_KEY"))
    has_anthropic = bool(os.environ.get("ANTHROPIC_API_KEY"))

    print("\nAPI Keys configured:")
    print(f"- OpenAI: {'Yes' if has_openai else 'No'}")
    print(f"- Anthropic: {'Yes' if has_anthropic else 'No'}")

    # Demo with mock if no API keys
    if not (has_openai or has_anthropic):
        print("\nNo API keys found. Using mock provider for demo.")
        return demo_mock_provider()

    # Use real provider if available
    provider_name = "openai" if has_openai else "anthropic"
    print(f"\nUsing {provider_name} provider...")

    try:
        provider = get_provider(provider_name)
        print(f"Model: {provider.model_name}")
        print(f"Max context tokens: {provider.max_context_tokens:,}")

        # Count tokens
        test_text = "This is a test message to demonstrate token counting."
        tokens = provider.count_tokens(test_text)
        print(f"\nToken count for '{test_text}': {tokens}")

    except Exception as e:
        print(f"Error initializing provider: {e}")
        return demo_mock_provider()


def demo_mock_provider():
    """Demo with a mock provider."""
    from unittest.mock import Mock

    from spreadsheet_analyzer.notebook_llm.llm_providers.base import LLMResponse, Message, Role

    print("\n--- Mock Provider Demo ---")

    mock_provider = Mock()
    mock_provider.model_name = "mock-gpt-4"
    mock_provider.max_context_tokens = 8192

    # Simulate a response
    mock_response = LLMResponse(
        content="Based on the Excel analysis:\n"
        "1. The data shows a sales trend with variance from targets\n"
        "2. February had the highest sales at 1200\n"
        "3. March fell below target by 200",
        model="mock-gpt-4",
        usage={"prompt_tokens": 50, "completion_tokens": 40, "total_tokens": 90},
    )

    mock_provider.complete.return_value = mock_response

    # Demo conversation
    messages = [
        Message(role=Role.SYSTEM, content="You are an Excel analysis assistant."),
        Message(role=Role.USER, content="Analyze the sales data and identify trends."),
    ]

    response = mock_provider.complete(messages)
    print(f"Model response:\n{response.content}")
    print(f"\nTokens used: {response.usage}")


async def main():
    """Run the demo."""
    print("=== Spreadsheet Analyzer LLM Stack Demo ===\n")

    # Demo Excel to Notebook conversion
    demo_excel_to_notebook()

    # Demo LLM providers
    demo_llm_providers()

    print("\n=== Demo Complete ===")
    print("\nThe full stack includes:")
    print("1. ✅ Excel to Notebook conversion")
    print("2. ✅ LLM provider abstraction (OpenAI & Anthropic)")
    print("3. ✅ Jupyter kernel management")
    print("4. ✅ Analysis strategies (hierarchical & graph-based)")
    print("5. ✅ Experiment framework for comparing LLMs")
    print("\nReady for Excel analysis with LLMs!")


if __name__ == "__main__":
    asyncio.run(main())
