#!/usr/bin/env python3
"""Create a sample Excel file with various data quality issues.

This file is designed to trigger iterative refinement in the analyzer by including:
- Mixed data types in numeric columns
- Formula errors
- Inconsistent date formats
- Missing values
- Currency symbols in numeric fields
"""

import random
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def create_sales_data():
    """Create sales data with intentional issues for testing refinement."""
    start_date = datetime(2024, 1, 1)
    products = ["Widget A", "Widget B", "Gadget X", "Gadget Y", "Tool Z"]
    regions = ["North", "South", "East", "West"]

    data = []
    for i in range(100):
        date = start_date + timedelta(days=i)

        # Intentionally create data quality issues
        if i % 15 == 0:
            # Mixed date format
            date_str = date.strftime("%m/%d/%Y") if i % 30 == 0 else date.strftime("%Y-%m-%d")
        else:
            date_str = date.strftime("%Y-%m-%d")

        product = random.choice(products)
        region = random.choice(regions)

        # Quantity with occasional text values
        if i % 20 == 0:
            quantity = "N/A"
        elif i % 25 == 0:
            quantity = f"{random.randint(10, 100)} units"  # Text in numeric field
        else:
            quantity = random.randint(10, 100)

        # Price with currency symbols and occasional invalid values
        base_price = random.uniform(10, 100)
        if i % 18 == 0:
            price = "Invalid"
        elif i % 10 == 0:
            price = f"${base_price:.2f}"  # Currency symbol
        else:
            price = round(base_price, 2)

        # Discount with percentage symbols
        if i % 12 == 0:
            discount = f"{random.randint(5, 25)}%"
        else:
            discount = random.randint(0, 25) / 100

        # Sales rep with occasional missing values
        if i % 30 == 0:
            sales_rep = ""
        else:
            sales_rep = f"Rep_{random.randint(1, 10)}"

        data.append(
            {
                "Date": date_str,
                "Product": product,
                "Region": region,
                "Quantity": quantity,
                "Unit Price": price,
                "Discount": discount,
                "Sales Rep": sales_rep,
            }
        )

    return pd.DataFrame(data)


def create_inventory_data():
    """Create inventory data with calculation challenges."""
    products = ["Widget A", "Widget B", "Gadget X", "Gadget Y", "Tool Z"]
    warehouses = ["WH-North", "WH-South", "WH-Central"]

    data = []
    for product in products:
        for warehouse in warehouses:
            # Current stock with occasional text
            if random.random() < 0.1:
                current_stock = "Check manually"
            else:
                current_stock = random.randint(50, 500)

            # Minimum stock
            min_stock = random.randint(20, 100)

            # Incoming shipments
            incoming = random.randint(0, 200) if random.random() > 0.3 else 0

            # Pending orders
            pending = random.randint(0, 150)

            data.append(
                {
                    "Product": product,
                    "Warehouse": warehouse,
                    "Current Stock": current_stock,
                    "Min Stock Level": min_stock,
                    "Incoming Shipments": incoming,
                    "Pending Orders": pending,
                    # Formula column will be added in Excel
                    "Reorder Needed": '=IF(C:C+E:E-F:F<D:D,"YES","NO")',
                }
            )

    return pd.DataFrame(data)


def create_financial_summary():
    """Create financial summary with complex formulas."""
    categories = [
        "Product Sales",
        "Service Revenue",
        "Licensing Fees",
        "Support Contracts",
        "Training Revenue",
    ]

    quarters = ["Q1 2024", "Q2 2024", "Q3 2024", "Q4 2024"]

    data = []
    for category in categories:
        row = {"Category": category}
        for i, quarter in enumerate(quarters):
            # Add some revenue with variability
            base = random.uniform(50000, 200000)
            if category == "Product Sales":
                base *= 2  # Higher for product sales

            # Occasionally add text that will break calculations
            if random.random() < 0.1:
                row[quarter] = f"${base:,.2f} (provisional)"
            else:
                row[quarter] = round(base, 2)

        data.append(row)

    # Add total row with formulas
    total_row = {"Category": "TOTAL"}
    for quarter in quarters:
        total_row[quarter] = f"=SUM({quarter}2:{quarter}6)"
    data.append(total_row)

    return pd.DataFrame(data)


def create_sample_excel(output_path: Path):
    """Create Excel file with multiple sheets containing data issues."""
    # Create workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1: Sales Data
    ws_sales = wb.create_sheet("Sales Data")
    sales_df = create_sales_data()

    # Add header formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Write sales data
    for r in dataframe_to_rows(sales_df, index=False, header=True):
        ws_sales.append(r)

    # Format headers
    for cell in ws_sales[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Add calculated column with formula
    ws_sales["H1"] = "Total"
    ws_sales["H1"].font = header_font
    ws_sales["H1"].fill = header_fill

    for row in range(2, len(sales_df) + 2):
        # Formula that will fail on text values
        ws_sales[f"H{row}"] = f"=D{row}*E{row}*(1-F{row})"

    # Sheet 2: Inventory
    ws_inventory = wb.create_sheet("Inventory")
    inventory_df = create_inventory_data()

    for r in dataframe_to_rows(inventory_df, index=False, header=True):
        ws_inventory.append(r)

    # Format headers
    for cell in ws_inventory[1]:
        cell.font = header_font
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    # Sheet 3: Financial Summary
    ws_financial = wb.create_sheet("Financial Summary")
    financial_df = create_financial_summary()

    for r in dataframe_to_rows(financial_df, index=False, header=True):
        ws_financial.append(r)

    # Format headers and total row
    for cell in ws_financial[1]:
        cell.font = header_font
        cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")

    # Format total row
    total_row_num = len(financial_df) + 1
    for cell in ws_financial[total_row_num]:
        cell.font = Font(bold=True)

    # Add a summary sheet with cross-sheet references
    ws_summary = wb.create_sheet("Summary", 0)  # Insert at beginning

    ws_summary["A1"] = "Executive Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)

    ws_summary["A3"] = "Total Sales Records:"
    ws_summary["B3"] = "=COUNTA('Sales Data'!A:A)-1"

    ws_summary["A4"] = "Average Order Value:"
    ws_summary["B4"] = "=AVERAGE('Sales Data'!H:H)"

    ws_summary["A5"] = "Total Revenue:"
    ws_summary["B5"] = "=SUM('Sales Data'!H:H)"

    ws_summary["A7"] = "Inventory Alerts:"
    ws_summary["B7"] = "=COUNTIF('Inventory'!G:G,\"YES\")"

    ws_summary["A9"] = "Q4 2024 Total Revenue:"
    ws_summary["B9"] = "='Financial Summary'!E7"

    # Save workbook
    wb.save(output_path)
    print(f"✅ Created sample Excel file: {output_path}")
    print("\nThis file contains:")
    print("- Mixed data types in numeric columns")
    print("- Currency symbols that need cleaning")
    print("- Formula errors from text values")
    print("- Missing values and 'N/A' entries")
    print("- Cross-sheet references")
    print("\nPerfect for testing iterative refinement!")


if __name__ == "__main__":
    output_file = Path("examples/sample_data_with_issues.xlsx")
    output_file.parent.mkdir(exist_ok=True)
    create_sample_excel(output_file)
