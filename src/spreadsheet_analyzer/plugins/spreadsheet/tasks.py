"""
Spreadsheet analysis tasks for notebook generation.

This module contains task implementations for spreadsheet-specific analysis:
- Data profiling and statistical analysis
- Formula validation and error detection
- Outlier detection and anomaly analysis
- Business logic validation

Tasks generate initial cells and can post-process results.
"""

from pathlib import Path
from typing import Any

from ...core_exec import CellType, NotebookCell
from ..base import BaseTask


class DataProfilingTask(BaseTask):
    """
    Task for generating comprehensive data profiling analysis.

    Creates cells that:
    - Load and explore spreadsheet data
    - Generate statistical summaries
    - Identify data types and patterns
    - Check for missing values and completeness
    - Create basic visualizations
    """

    def __init__(self):
        super().__init__(name="data_profiling", description="Comprehensive data profiling and statistical analysis")
        self.category = "analysis"

    def build_initial_cells(self, context: dict[str, Any]) -> list[NotebookCell]:
        """Generate data profiling cells."""
        cells = []

        # Get file path and sheet info from context
        file_path = context.get("file_path")
        sheet_name = context.get("sheet_name", "Sheet1")

        if not file_path:
            return cells

        # Header cell
        cells.append(
            NotebookCell(
                cell_type=CellType.MARKDOWN,
                source=self._format_source(
                    "## Data Profiling Analysis\n\nComprehensive statistical analysis and data exploration."
                ),
                metadata={"tags": ["data_profiling", "header"]},
            )
        )

        # Import and setup cell
        import_code = self._generate_import_code()
        cells.append(
            NotebookCell(
                cell_type=CellType.CODE,
                source=self._format_source(import_code),
                metadata={"tags": ["imports", "setup"]},
                execution_count=None,
                outputs=[],
            )
        )

        # Data loading cell
        load_code = self._generate_load_code(file_path, sheet_name)
        cells.append(
            NotebookCell(
                cell_type=CellType.CODE,
                source=self._format_source(load_code),
                metadata={"tags": ["data_loading"]},
                execution_count=None,
                outputs=[],
            )
        )

        # Basic profiling cell
        profile_code = self._generate_profiling_code(sheet_name)
        cells.append(
            NotebookCell(
                cell_type=CellType.CODE,
                source=self._format_source(profile_code),
                metadata={"tags": ["profiling", "statistics"]},
                execution_count=None,
                outputs=[],
            )
        )

        return cells

    def validate_context(self, context: dict[str, Any]) -> list[str]:
        """Validate that required context is provided."""
        issues = []

        if "file_path" not in context:
            issues.append("Missing required 'file_path' in context")
        elif not Path(context["file_path"]).exists():
            issues.append(f"File does not exist: {context['file_path']}")

        return issues

    def _generate_import_code(self) -> str:
        """Generate import statements for data profiling."""
        return """
# Data profiling imports
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

# Set up plotting
plt.style.use('default')
sns.set_palette("husl")
"""

    def _generate_load_code(self, file_path: str, sheet_name: str) -> str:
        """Generate code to load spreadsheet data."""
        file_ext = Path(file_path).suffix.lower()

        if file_ext in [".xlsx", ".xls", ".xlsm"]:
            return f"""
# Load Excel data
file_path = r"{file_path}"
sheet_name = "{sheet_name}"

try:
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    print(f"✅ Successfully loaded {{len(df)}} rows and {{len(df.columns)}} columns from {sheet_name}")
    print(f"📊 Data shape: {{df.shape}}")
except Exception as e:
    print(f"❌ Error loading data: {{e}}")
    df = pd.DataFrame()  # Empty fallback
"""
        elif file_ext == ".csv":
            return f"""
# Load CSV data
file_path = r"{file_path}"

try:
    df = pd.read_csv(file_path)
    print(f"✅ Successfully loaded {{len(df)}} rows and {{len(df.columns)}} columns")
    print(f"📊 Data shape: {{df.shape}}")
except Exception as e:
    print(f"❌ Error loading data: {{e}}")
    df = pd.DataFrame()  # Empty fallback
"""
        else:
            return f"""
# Unsupported file format: {file_ext}
print("❌ Unsupported file format: {file_ext}")
df = pd.DataFrame()  # Empty fallback
"""

    def _generate_profiling_code(self, sheet_name: str) -> str:
        """Generate comprehensive data profiling code."""
        return f"""
# Comprehensive Data Profiling for {sheet_name}

if not df.empty:
    print("=" * 60)
    print(f"📋 DATA PROFILING REPORT - {sheet_name}")
    print("=" * 60)

    # Basic Information
    print("\\n🔍 BASIC INFORMATION")
    print(f"Shape: {{df.shape}}")
    print(f"Memory usage: {{df.memory_usage(deep=True).sum() / 1024**2:.2f}} MB")

    # Column Information
    print("\\n📊 COLUMN ANALYSIS")
    for col in df.columns:
        dtype = df[col].dtype
        null_count = df[col].isnull().sum()
        null_pct = (null_count / len(df)) * 100
        unique_count = df[col].nunique()

        print(f"  {{col:<30}} | {{dtype!s:<12}} | Nulls: {{null_count:>6}} ({{null_pct:>5.1f}}%) | Unique: {{unique_count:>6}}")

    # Statistical Summary
    print("\\n📈 STATISTICAL SUMMARY")
    numeric_cols = df.select_dtypes(include=[np.number]).columns
    if len(numeric_cols) > 0:
        display(df[numeric_cols].describe())
    else:
        print("No numeric columns found for statistical analysis")

    # Missing Values Analysis
    print("\\n❓ MISSING VALUES ANALYSIS")
    missing_summary = pd.DataFrame({{
        "Column": df.columns,
        'Missing_Count': df.isnull().sum(),
        'Missing_Percentage': (df.isnull().sum() / len(df)) * 100
    }}).sort_values('Missing_Percentage', ascending=False)

    display(missing_summary[missing_summary['Missing_Count'] > 0])

    # Data Types Distribution
    print("\\n🏷️ DATA TYPES DISTRIBUTION")
    dtype_counts = df.dtypes.value_counts()
    display(dtype_counts)

    # Sample Data Preview
    print("\\n👀 SAMPLE DATA PREVIEW")
    print("First 5 rows:")
    display(df.head())

    if len(df) > 5:
        print("\\nLast 5 rows:")
        display(df.tail())

    # Quick Quality Checks
    print("\\n✅ QUICK QUALITY CHECKS")

    # Duplicate rows
    duplicate_count = df.duplicated().sum()
    print(f"Duplicate rows: {{duplicate_count}} ({{duplicate_count / len(df) * 100:.1f}}%)")

    # Empty rows
    empty_rows = df.isnull().all(axis=1).sum()
    print(f"Completely empty rows: {{empty_rows}}")

    # Potential ID columns
    potential_ids = [col for col in df.columns if df[col].nunique() == len(df) and not df[col].isnull().any()]
    if potential_ids:
        print(f"Potential ID columns: {{potential_ids}}")

else:
    print("❌ No data available for profiling")
"""

    def _format_source(self, content: str) -> list[str]:
        """Format content as list of lines for Jupyter format."""
        if not content:
            return [""]

        lines = content.strip().split("\n")
        formatted = []
        for i, line in enumerate(lines):
            if i < len(lines) - 1:
                formatted.append(line + "\n")
            else:
                formatted.append(line)
        return formatted


class FormulaAnalysisTask(BaseTask):
    """
    Task for analyzing Excel formulas and detecting errors.

    Creates cells that:
    - Extract and analyze formulas
    - Detect formula errors and circular references
    - Validate formula logic and dependencies
    - Check for common formula issues
    """

    def __init__(self):
        super().__init__(name="formula_analysis", description="Excel formula validation and error detection")
        self.category = "validation"

    def build_initial_cells(self, context: dict[str, Any]) -> list[NotebookCell]:
        """Generate formula analysis cells."""
        cells = []

        file_path = context.get("file_path")
        sheet_name = context.get("sheet_name", "Sheet1")

        if not file_path:
            return cells

        # Only create cells for Excel files
        if Path(file_path).suffix.lower() not in [".xlsx", ".xls", ".xlsm"]:
            return cells

        # Header cell
        cells.append(
            NotebookCell(
                cell_type=CellType.MARKDOWN,
                source=self._format_source(
                    "## Formula Analysis\n\nDetection and validation of Excel formulas and errors."
                ),
                metadata={"tags": ["formula_analysis", "header"]},
            )
        )

        # Import cell
        import_code = self._generate_import_code()
        cells.append(
            NotebookCell(
                cell_type=CellType.CODE,
                source=self._format_source(import_code),
                metadata={"tags": ["imports"]},
                execution_count=None,
                outputs=[],
            )
        )

        # Formula analysis cell
        analysis_code = self._generate_analysis_code(file_path, sheet_name)
        cells.append(
            NotebookCell(
                cell_type=CellType.CODE,
                source=self._format_source(analysis_code),
                metadata={"tags": ["formula_analysis", "validation"]},
                execution_count=None,
                outputs=[],
            )
        )

        return cells

    def _generate_import_code(self) -> str:
        """Generate imports for formula analysis."""
        return """
# Formula analysis imports
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
from collections import defaultdict, Counter
import re
"""

    def _generate_analysis_code(self, file_path: str, sheet_name: str) -> str:
        """Generate formula analysis code."""
        return f"""
# Excel Formula Analysis for {sheet_name}

file_path = r"{file_path}"
sheet_name = "{sheet_name}"

try:
    # Load workbook with formulas
    workbook = openpyxl.load_workbook(file_path, data_only=False)
    worksheet = workbook[sheet_name]

    print("=" * 60)
    print(f"🧮 FORMULA ANALYSIS REPORT - {sheet_name}")
    print("=" * 60)

    # Extract all formulas
    formulas = []
    formula_cells = []
    error_cells = []

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas.append(cell.value)
                    formula_cells.append(get_column_letter(cell.column) + str(cell.row))
                elif isinstance(cell.value, str) and any(err in str(cell.value) for err in ['#DIV/0!', '#VALUE!', '#REF!', '#N/A', '#NUM!', '#NAME?', '#NULL!']):
                    error_cells.append({{
                        "cell": f"{{get_column_letter(cell.column)}}{{cell.row}}",
                        'error': cell.value
                    }})

    print(f"\\n📊 FORMULA SUMMARY")
    print(f"Total cells with formulas: {{len(formulas)}}")
    print(f"Cells with errors: {{len(error_cells)}}")

    if formulas:
        # Analyze formula types
        print(f"\\n🔍 FORMULA BREAKDOWN")

        formula_functions = []
        for formula in formulas:
            # Extract function names (simple regex)
            functions = re.findall(r'([A-Z]+)\\(', formula.upper())
            formula_functions.extend(functions)

        if formula_functions:
            function_counts = Counter(formula_functions)
            print("Most common functions:")
            for func, count in function_counts.most_common(10):
                print(f"  {{func}}: {{count}} times")

        # Show sample formulas
        print(f"\\n📝 SAMPLE FORMULAS")
        for i, (cell, formula) in enumerate(zip(formula_cells[:5], formulas[:5])):
            print(f"  {{cell}}: {{formula}}")

        if len(formulas) > 5:
            print(f"  ... and {{len(formulas) - 5}} more")

    # Error Analysis
    if error_cells:
        print(f"\\n❌ FORMULA ERRORS DETECTED")
        error_types = Counter([err['error'] for err in error_cells])

        print("Error breakdown:")
        for error_type, count in error_types.items():
            print(f"  {{error_type}}: {{count}} occurrences")

        print(f"\\nFirst 10 error locations:")
        for error in error_cells[:10]:
            print(f"  {{error['cell']}}: {{error['error']}}")
    else:
        print(f"\\n✅ No formula errors detected!")

    # Formula complexity analysis
    if formulas:
        print(f"\\n📐 FORMULA COMPLEXITY")
        complex_formulas = [f for f in formulas if len(f) > 100 or f.count('(') > 3]

        if complex_formulas:
            print(f"Complex formulas (>100 chars or >3 nested functions): {{len(complex_formulas)}}")
            print("Examples:")
            for i, formula in enumerate(complex_formulas[:3]):
                print(f"  {{i + 1}}. {{formula[:80]}}..." if len(formula) > 80 else f"  {{i + 1}}. {{formula}}")
        else:
            print("No overly complex formulas detected")

    workbook.close()

except Exception as e:
    print(f"❌ Error analyzing formulas: {{e}}")
"""

    def _format_source(self, content: str) -> list[str]:
        """Format content as list of lines for Jupyter format."""
        if not content:
            return [""]

        lines = content.strip().split("\n")
        formatted = []
        for i, line in enumerate(lines):
            if i < len(lines) - 1:
                formatted.append(line + "\n")
            else:
                formatted.append(line)
        return formatted


class OutlierDetectionTask(BaseTask):
    """
    Task for detecting outliers and anomalies in data.

    Creates cells that:
    - Statistical outlier detection (Z-score, IQR)
    - Visual outlier identification
    - Anomaly pattern recognition
    - Data quality assessment
    """

    def __init__(self):
        super().__init__(name="outlier_detection", description="Statistical outlier detection and anomaly analysis")
        self.category = "analysis"

    def build_initial_cells(self, context: dict[str, Any]) -> list[NotebookCell]:
        """Generate outlier detection cells."""
        cells = []

        file_path = context.get("file_path")

        if not file_path:
            return cells

        # Header cell
        cells.append(
            NotebookCell(
                cell_type=CellType.MARKDOWN,
                source=self._format_source(
                    "## Outlier Detection\n\nStatistical analysis to identify outliers and anomalies in the data."
                ),
                metadata={"tags": ["outlier_detection", "header"]},
            )
        )

        # Outlier detection code
        detection_code = self._generate_detection_code()
        cells.append(
            NotebookCell(
                cell_type=CellType.CODE,
                source=self._format_source(detection_code),
                metadata={"tags": ["outlier_detection", "statistics"]},
                execution_count=None,
                outputs=[],
            )
        )

        return cells

    def _generate_detection_code(self) -> str:
        """Generate outlier detection code."""
        return """
# Statistical Outlier Detection

# Assuming 'df' DataFrame is available from previous cells
if 'df' in locals() and not df.empty:
    print("=" * 60)
    print("🔍 OUTLIER DETECTION ANALYSIS")
    print("=" * 60)

    # Select numeric columns only
    numeric_cols = df.select_dtypes(include=[np.number]).columns

    if len(numeric_cols) == 0:
        print("❌ No numeric columns found for outlier detection")
    else:
        print(f"\\n📊 Analyzing {{{{len(numeric_cols)}}}} numeric columns for outliers...")

        outlier_summary = []

        for col in numeric_cols:
            if df[col].isnull().all():
                continue

            # Calculate statistics
            Q1 = df[col].quantile(0.25)
            Q3 = df[col].quantile(0.75)
            IQR = Q3 - Q1

            # IQR method outliers
            lower_bound = Q1 - 1.5 * IQR
            upper_bound = Q3 + 1.5 * IQR
            iqr_outliers = df[(df[col] < lower_bound) | (df[col] > upper_bound)]

            # Z-score method outliers (>3 standard deviations)
            z_scores = np.abs((df[col] - df[col].mean()) / df[col].std())
            zscore_outliers = df[z_scores > 3]

            outlier_summary.append({{
                'Column': col,
                'IQR_Outliers': len(iqr_outliers),
                'ZScore_Outliers': len(zscore_outliers),
                'Min_Value': df[col].min(),
                'Max_Value': df[col].max(),
                'Mean': df[col].mean(),
                'Std': df[col].std()
            }})

            # Print detailed analysis for columns with outliers
            if len(iqr_outliers) > 0 or len(zscore_outliers) > 0:
                print(f"\\n🚨 OUTLIERS DETECTED in '{{{{col}}}}':")
                print(f"  IQR Method: {{{{len(iqr_outliers)}}}} outliers ({{{{len(iqr_outliers)/len(df)*100:.1f}}}}%)")
                print(f"  Z-Score Method: {{{{len(zscore_outliers)}}}} outliers ({{{{len(zscore_outliers)/len(df)*100:.1f}}}}%)")
                print(f"  Value range: {{{{df[col].min():.2f}}}} to {{{{df[col].max():.2f}}}}")
                print(f"  Expected range (IQR): {{{{lower_bound:.2f}}}} to {{{{upper_bound:.2f}}}}")

                # Show some outlier values
                if len(iqr_outliers) > 0:
                    outlier_values = iqr_outliers[col].values
                    print(f"  Sample outlier values: {{{{outlier_values[:5].tolist()}}}}")

        # Summary table
        if outlier_summary:
            print(f"\\n📋 OUTLIER SUMMARY TABLE")
            outlier_df = pd.DataFrame(outlier_summary)
            display(outlier_df)

            # Visualization for columns with outliers
            cols_with_outliers = outlier_df[(outlier_df['IQR_Outliers'] > 0) | (outlier_df['ZScore_Outliers'] > 0)]['Column'].tolist()

            if cols_with_outliers:
                print(f"\\n📈 OUTLIER VISUALIZATIONS")

                # Create box plots
                n_cols = min(3, len(cols_with_outliers))
                n_rows = (len(cols_with_outliers) + n_cols - 1) // n_cols

                fig, axes = plt.subplots(n_rows, n_cols, figsize=(15, 5*n_rows))
                if n_rows == 1 and n_cols == 1:
                    axes = [axes]
                elif n_rows == 1:
                    axes = axes
                else:
                    axes = axes.flatten()

                for i, col in enumerate(cols_with_outliers[:len(axes)]):
                    if i < len(axes):
                        df[col].plot(kind='box', ax=axes[i], title=f'Box Plot: {col}')
                        axes[i].grid(True, alpha=0.3)

                # Hide empty subplots
                for i in range(len(cols_with_outliers), len(axes)):
                    axes[i].set_visible(False)

                plt.tight_layout()
                plt.show()
            else:
                print(f"\\n✅ No significant outliers detected in any numeric columns!")
        else:
            print(f"\\n❌ No numeric data available for outlier analysis")

else:
    print("❌ No DataFrame 'df' available. Please run data loading cells first.")
"""

    def _format_source(self, content: str) -> list[str]:
        """Format content as list of lines for Jupyter format."""
        if not content:
            return [""]

        lines = content.strip().split("\n")
        formatted = []
        for i, line in enumerate(lines):
            if i < len(lines) - 1:
                formatted.append(line + "\n")
            else:
                formatted.append(line)
        return formatted
