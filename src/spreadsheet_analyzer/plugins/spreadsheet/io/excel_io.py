"""Excel I/O utilities with performance optimizations."""

import asyncio
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from structlog import get_logger

logger = get_logger(__name__)

# CLAUDE-PERFORMANCE: Threshold for considering a file "large"
LARGE_FILE_THRESHOLD_MB = 50
LARGE_FILE_ROW_LIMIT = 100000


async def read_sheet_async(
    path: Path, sheet: str, *, header: int | None = None, na_filter: bool = False, nrows: int | None = None
) -> pd.DataFrame:
    """Read Excel sheet asynchronously using thread pool.

    CLAUDE-PERFORMANCE: asyncio.to_thread keeps event loop responsive
    for concurrent operations.
    """
    return await asyncio.to_thread(
        pd.read_excel,
        path,
        sheet_name=sheet,
        header=header,
        na_filter=na_filter,  # CLAUDE-PERFORMANCE: Disable NA inference for speed
        engine="openpyxl",
        nrows=nrows,  # Limit rows for large files
    )


async def check_file_size(excel_path: Path) -> tuple[float, bool]:
    """Check Excel file size and determine if special handling is needed.

    Returns:
        Tuple of (file_size_mb, is_large_file)
    """

    def _get_size():
        size_bytes = excel_path.stat().st_size
        size_mb = size_bytes / (1024 * 1024)
        is_large = size_mb > LARGE_FILE_THRESHOLD_MB
        return size_mb, is_large

    return await asyncio.to_thread(_get_size)


def list_sheets(excel_path: Path) -> list[str]:
    """List all sheet names in an Excel file.

    CLAUDE-PERFORMANCE: Uses read_only mode to avoid loading entire workbook.
    """
    try:
        wb = load_workbook(excel_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except Exception:
        logger.exception("Failed to list sheets", excel_path=str(excel_path))
        return []


def get_sheet_dimensions(excel_path: Path, sheet_name: str) -> tuple[int, int]:
    """Get sheet dimensions without loading all data.

    CLAUDE-PERFORMANCE: Uses read_only mode and doesn't iterate cells.

    Returns:
        Tuple of (max_row, max_column)
    """
    try:
        wb = load_workbook(excel_path, read_only=True)
        ws = wb[sheet_name]
        dimensions = (ws.max_row, ws.max_column)
        wb.close()
        return dimensions
    except Exception:
        logger.exception("Failed to get sheet dimensions", excel_path=str(excel_path), sheet=sheet_name)
        return (0, 0)


def detect_sheet_features(excel_path: Path, sheet_name: str, sample_rows: int = 100) -> dict[str, bool]:
    """Detect sheet features like formulas, pivot tables, etc.

    CLAUDE-KNOWLEDGE: Only samples first N rows for performance.

    Returns:
        Dict with feature flags (has_formulas, has_merged_cells, etc.)
    """
    features = {
        "has_formulas": False,
        "has_merged_cells": False,
        "has_pivot_tables": False,
        "formula_count": 0,
        "merged_ranges": 0,
    }

    try:
        wb = load_workbook(excel_path, read_only=True, data_only=False)
        ws = wb[sheet_name]

        # Check for formulas in sample
        for row in ws.iter_rows(min_row=1, max_row=min(sample_rows, ws.max_row), values_only=False):
            for cell in row:
                if cell.data_type == "f":  # Formula type
                    features["has_formulas"] = True
                    features["formula_count"] += 1

        # Check for merged cells
        if hasattr(ws, "merged_cells") and ws.merged_cells:
            features["has_merged_cells"] = True
            features["merged_ranges"] = len(ws.merged_cells.ranges)

        wb.close()
    except Exception as e:
        logger.exception("Failed to detect sheet features", error=str(e))

    return features
