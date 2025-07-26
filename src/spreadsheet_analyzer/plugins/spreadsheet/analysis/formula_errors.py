"""Excel formula error detection utilities using openpyxl."""

import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell import Cell
from structlog import get_logger

logger = get_logger(__name__)

# Excel error types
EXCEL_ERRORS = {
    "#DIV/0!": "Division by zero error",
    "#N/A": "Value not available error",
    "#NAME?": "Unrecognized name or formula error",
    "#NULL!": "Null intersection error",
    "#NUM!": "Invalid numeric value error",
    "#REF!": "Invalid cell reference error",
    "#VALUE!": "Wrong value type error",
    "#SPILL!": "Spill range is not blank error",
    "#CALC!": "Calculation error",
    "#GETTING_DATA": "Temporary loading state (Excel 365)",
}

# Pattern to match Excel errors in strings
ERROR_PATTERN = re.compile(r"#[A-Z]+[?!]|#GETTING_DATA")


def is_excel_error(value: Any) -> bool:
    """Check if a cell value is an Excel error.

    Args:
        value: Cell value to check

    Returns:
        True if the value is an Excel error
    """
    if value is None:
        return False

    # Convert to string for pattern matching
    str_value = str(value)

    # Check if it's a known error type
    return str_value in EXCEL_ERRORS or ERROR_PATTERN.match(str_value) is not None


def get_error_type(value: Any) -> str | None:
    """Get the Excel error type from a cell value.

    Args:
        value: Cell value to check

    Returns:
        Error type string (e.g., "#DIV/0!") or None if not an error
    """
    if not is_excel_error(value):
        return None

    str_value = str(value)

    # Return exact match if it's a known error
    if str_value in EXCEL_ERRORS:
        return str_value

    # Try to extract error pattern
    match = ERROR_PATTERN.match(str_value)
    if match:
        return match.group(0)

    return None


def scan_sheet_for_errors(sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, list[dict[str, Any]]]:
    """Scan a worksheet for formula errors.

    Args:
        sheet: Openpyxl worksheet to scan

    Returns:
        Dictionary with error types as keys and lists of error details as values
    """
    errors_found = {}

    # Scan all cells in the used range
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.value is not None and is_excel_error(cell.value):
                error_type = get_error_type(cell.value)
                if error_type:
                    if error_type not in errors_found:
                        errors_found[error_type] = []

                    # Collect error details
                    error_details = {
                        "cell": cell.coordinate,
                        "formula": cell.value if isinstance(cell, Cell) and cell.data_type == "f" else None,
                        "value": str(cell.value),
                        "row": cell.row,
                        "column": cell.column,
                        "has_formula": isinstance(cell, Cell) and cell.data_type == "f",
                    }

                    # Try to get the formula if available
                    if hasattr(cell, "_value") and cell._value != cell.value:
                        error_details["formula"] = cell._value

                    errors_found[error_type].append(error_details)

    return errors_found


def scan_workbook_for_errors(
    excel_path: Path | str, sheet_names: list[str] | None = None
) -> dict[str, dict[str, list[dict[str, Any]]]]:
    """Scan an entire workbook or specific sheets for formula errors.

    Args:
        excel_path: Path to Excel file
        sheet_names: Optional list of sheet names to scan (None = all sheets)

    Returns:
        Dictionary with sheet names as keys and error dictionaries as values
    """
    excel_path = Path(excel_path)

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    try:
        # Open workbook in read-only mode for efficiency
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        all_errors = {}

        # Determine which sheets to scan
        sheets_to_scan = sheet_names if sheet_names else wb.sheetnames

        for sheet_name in sheets_to_scan:
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found in workbook")
                continue

            sheet = wb[sheet_name]
            sheet_errors = scan_sheet_for_errors(sheet)

            if sheet_errors:
                all_errors[sheet_name] = sheet_errors
                logger.info(
                    f"Found errors in sheet '{sheet_name}'",
                    error_count=sum(len(errors) for errors in sheet_errors.values()),
                    error_types=list(sheet_errors.keys()),
                )

        wb.close()

    except Exception:
        logger.exception("Error scanning workbook for errors")
        raise
    else:
        return all_errors


def detect_inconsistent_formulas(sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, list[dict[str, Any]]]:
    """Detect cells whose formulas differ from the majority in their row/column.

    This helps identify 'wrong but valid' formulas that are often more problematic
    than outright errors in spreadsheet audits.

    Args:
        sheet: Openpyxl worksheet to analyze

    Returns:
        Dictionary with 'row_inconsistencies' and 'column_inconsistencies' as keys
    """
    from collections import Counter

    row_inconsistencies = []
    col_inconsistencies = []

    # Analyze formula consistency by row
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row), start=1):
        # Collect formulas in this row
        formula_cells = [(cell, cell.value) for cell in row if cell.data_type == "f" and cell.value]

        if len(formula_cells) > 1:
            # Normalize formulas (remove absolute references for comparison)
            normalized_formulas = []
            for cell, formula in formula_cells:
                # Simple normalization: remove $ signs for comparison
                normalized = str(formula).replace("$", "")
                normalized_formulas.append((cell, formula, normalized))

            # Find most common pattern
            formula_counts = Counter(norm for _, _, norm in normalized_formulas)
            if len(formula_counts) > 1:  # There are variations
                most_common, count = formula_counts.most_common(1)[0]

                # Report cells that deviate from the pattern
                for cell, original_formula, normalized in normalized_formulas:
                    if normalized != most_common:
                        row_inconsistencies.append(
                            {
                                "cell": cell.coordinate,
                                "row": row_idx,
                                "column": cell.column,
                                "formula": original_formula,
                                "expected_pattern": most_common,
                                "deviation_type": "row",
                                "similar_cells_count": count,
                            }
                        )

    # Analyze formula consistency by column
    for col_idx in range(1, sheet.max_column + 1):
        # Collect formulas in this column
        formula_cells = []
        for row_idx in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.data_type == "f" and cell.value:
                formula_cells.append((cell, cell.value))

        if len(formula_cells) > 1:
            # Normalize formulas for column comparison
            normalized_formulas = []
            for cell, formula in formula_cells:
                # For columns, also try to normalize row references
                normalized = str(formula).replace("$", "")
                # Simple pattern: replace row numbers with placeholder
                import re

                normalized = re.sub(r"\d+", "N", normalized)
                normalized_formulas.append((cell, formula, normalized))

            # Find most common pattern
            formula_counts = Counter(norm for _, _, norm in normalized_formulas)
            if len(formula_counts) > 1:  # There are variations
                most_common, count = formula_counts.most_common(1)[0]

                # Report cells that deviate from the pattern
                for cell, original_formula, normalized in normalized_formulas:
                    if normalized != most_common:
                        col_inconsistencies.append(
                            {
                                "cell": cell.coordinate,
                                "row": cell.row,
                                "column": col_idx,
                                "formula": original_formula,
                                "expected_pattern": most_common,
                                "deviation_type": "column",
                                "similar_cells_count": count,
                            }
                        )

    return {"row_inconsistencies": row_inconsistencies, "column_inconsistencies": col_inconsistencies}


def generate_error_analysis_code(excel_path: Path | str, sheet_name: str | None = None) -> str:
    """Generate Python code to analyze formula errors in a workbook.

    Args:
        excel_path: Path to Excel file
        sheet_name: Optional specific sheet to analyze

    Returns:
        Python code string for error analysis
    """
    code = f'''# Analyze Excel formula errors and consistency using openpyxl
import openpyxl
from pathlib import Path
import pandas as pd
from collections import Counter
import re

# Load workbook - need data_only=False to see formulas
excel_path = Path({str(excel_path)!r})
wb = openpyxl.load_workbook(excel_path, data_only=False)
wb_values = openpyxl.load_workbook(excel_path, data_only=True)  # For seeing evaluated values

# Excel error types to check
EXCEL_ERRORS = {{
    "#DIV/0!": "Division by zero",
    "#N/A": "Value not available",
    "#NAME?": "Unrecognized name",
    "#NULL!": "Null intersection",
    "#NUM!": "Invalid numeric value",
    "#REF!": "Invalid reference",
    "#VALUE!": "Wrong value type",
    "#SPILL!": "Spill range not blank",
}}

def check_cell_error(cell):
    """Check if a cell contains an error."""
    if cell.value is None:
        return None
    str_val = str(cell.value)
    for error_type in EXCEL_ERRORS:
        if error_type in str_val:
            return error_type
    return None

def analyze_formula_consistency(sheet):
    """Detect formula pattern inconsistencies."""
    inconsistencies = []

    # Check rows
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row), start=1):
        formulas = [(cell, cell.value) for cell in row if cell.data_type == 'f' and cell.value]
        if len(formulas) > 1:
            # Normalize by removing $ signs
            normalized = [(cell, str(formula).replace('$', '')) for cell, formula in formulas]
            counts = Counter(norm for _, norm in normalized)
            if len(counts) > 1:
                most_common, _ = counts.most_common(1)[0]
                for cell, norm in normalized:
                    if norm != most_common:
                        inconsistencies.append({{
                            'cell': cell.coordinate,
                            'type': 'row_deviation',
                            'formula': cell.value,
                            'row': row_idx
                        }})

    # Check columns
    for col_idx in range(1, sheet.max_column + 1):
        formulas = []
        for row_idx in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.data_type == 'f' and cell.value:
                formulas.append((cell, cell.value))

        if len(formulas) > 1:
            # Normalize by removing $ and replacing numbers
            normalized = []
            for cell, formula in formulas:
                norm = str(formula).replace('$', '')
                norm = re.sub(r'\\d+', 'N', norm)  # Replace row numbers
                normalized.append((cell, norm))

            counts = Counter(norm for _, norm in normalized)
            if len(counts) > 1:
                most_common, _ = counts.most_common(1)[0]
                for cell, norm in normalized:
                    if norm != most_common:
                        inconsistencies.append({{
                            'cell': cell.coordinate,
                            'type': 'column_deviation',
                            'formula': cell.value,
                            'column': col_idx
                        }})

    return inconsistencies

# Analyze errors and consistency
all_errors = []
all_inconsistencies = []
'''

    if sheet_name:
        code += f"""
# Analyze specific sheet
sheet = wb[{sheet_name!r}]
print(f"Analyzing sheet '{{sheet.title}}' for formula errors...")

for row in sheet.iter_rows():
    for cell in row:
        error_type = check_cell_error(cell)
        if error_type:
            all_errors.append({{
                'sheet': sheet.title,
                'cell': cell.coordinate,
                'error_type': error_type,
                'description': EXCEL_ERRORS[error_type],
                'formula': getattr(cell, '_value', None) if hasattr(cell, '_value') else None
            }})
"""
    else:
        code += """
# Analyze all sheets
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"Analyzing sheet '{sheet.title}' for formula errors...")

    for row in sheet.iter_rows():
        for cell in row:
            error_type = check_cell_error(cell)
            if error_type:
                all_errors.append({
                    'sheet': sheet.title,
                    'cell': cell.coordinate,
                    'error_type': error_type,
                    'description': EXCEL_ERRORS[error_type],
                    'formula': getattr(cell, '_value', None) if hasattr(cell, '_value') else None
                })
"""

    # Add formula consistency analysis
    if sheet_name:
        code += f"""
# Analyze formula consistency
print("\\nAnalyzing formula consistency in '{sheet_name}'...")
inconsistencies = analyze_formula_consistency(wb[{sheet_name!r}])
all_inconsistencies.extend(inconsistencies)
"""
    else:
        code += """
# Analyze formula consistency across all sheets
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\\nAnalyzing formula consistency in '{sheet.title}'...")
    sheet_inconsistencies = analyze_formula_consistency(sheet)
    for inc in sheet_inconsistencies:
        inc['sheet'] = sheet_name
    all_inconsistencies.extend(sheet_inconsistencies)
"""

    code += """
# Close workbook
wb.close()

# Display error results
if all_errors:
    error_df = pd.DataFrame(all_errors)
    print(f"\\nFound {len(all_errors)} formula errors:")
    print("\\nError Summary:")
    print(error_df['error_type'].value_counts())

    print("\\nDetailed Error List:")
    for _, error in error_df.iterrows():
        print(f"  {error['sheet']}!{error['cell']} - {error['error_type']}: {error['description']}")
        if error['formula']:
            print(f"    Formula: {error['formula']}")
else:
    print("\\nNo formula errors found in the workbook!")

# Display consistency results
if all_inconsistencies:
    inc_df = pd.DataFrame(all_inconsistencies)
    print(f"\\n\\nFound {len(all_inconsistencies)} formula inconsistencies:")

    # Group by type
    row_deviations = inc_df[inc_df['type'] == 'row_deviation']
    col_deviations = inc_df[inc_df['type'] == 'column_deviation']

    if len(row_deviations) > 0:
        print(f"\\nRow inconsistencies ({len(row_deviations)}):")
        for _, inc in row_deviations.head(10).iterrows():
            print(f"  {inc.get('sheet', '')}!{inc['cell']} (Row {inc['row']}): {inc['formula']}")

    if len(col_deviations) > 0:
        print(f"\\nColumn inconsistencies ({len(col_deviations)}):")
        for _, inc in col_deviations.head(10).iterrows():
            print(f"  {inc.get('sheet', '')}!{inc['cell']} (Col {inc['column']}): {inc['formula']}")

    if len(all_inconsistencies) > 20:
        print(f"\\n... and {len(all_inconsistencies) - 20} more inconsistencies")
else:
    print("\\n\\nNo formula inconsistencies detected - formulas follow consistent patterns!")

# Create comprehensive visualization
if (all_errors and len(all_errors) > 0) or (all_inconsistencies and len(all_inconsistencies) > 0):
    import matplotlib.pyplot as plt

    # Determine subplot layout
    has_errors = all_errors and len(all_errors) > 0
    has_inconsistencies = all_inconsistencies and len(all_inconsistencies) > 0

    if has_errors and has_inconsistencies:
        fig, axes = plt.subplots(2, 2, figsize=(14, 10))
        fig.suptitle('Excel Formula Analysis Results', fontsize=16)
    elif has_errors or has_inconsistencies:
        fig, axes = plt.subplots(1, 2, figsize=(12, 5))
        fig.suptitle('Excel Formula Analysis Results', fontsize=16)
        axes = axes.reshape(1, -1) if has_errors or has_inconsistencies else axes

    plot_idx = 0

    # Plot error distribution if errors exist
    if has_errors:
        error_df = pd.DataFrame(all_errors)
        ax = axes.flat[plot_idx] if has_errors and has_inconsistencies else axes[plot_idx]

        # Bar chart of error types
        error_counts = error_df['error_type'].value_counts()
        ax.bar(error_counts.index, error_counts.values, color='coral')
        ax.set_xlabel('Error Type')
        ax.set_ylabel('Count')
        ax.set_title('Formula Error Distribution')
        ax.tick_params(axis='x', rotation=45)
        plot_idx += 1

        # Pie chart by sheet
        if 'sheet' in error_df.columns:
            ax = axes.flat[plot_idx] if has_errors and has_inconsistencies else axes[plot_idx]
            sheet_counts = error_df['sheet'].value_counts()
            ax.pie(sheet_counts.values, labels=sheet_counts.index, autopct='%1.1f%%')
            ax.set_title('Errors by Sheet')
            plot_idx += 1

    # Plot inconsistency distribution if inconsistencies exist
    if has_inconsistencies:
        inc_df = pd.DataFrame(all_inconsistencies)
        ax = axes.flat[plot_idx] if has_errors and has_inconsistencies else axes[plot_idx]

        # Bar chart of inconsistency types
        type_counts = inc_df['type'].value_counts()
        colors = {'row_deviation': 'skyblue', 'column_deviation': 'lightgreen'}
        ax.bar(type_counts.index, type_counts.values,
               color=[colors.get(t, 'gray') for t in type_counts.index])
        ax.set_xlabel('Inconsistency Type')
        ax.set_ylabel('Count')
        ax.set_title('Formula Pattern Deviations')
        plot_idx += 1

        # Top affected areas
        if plot_idx < len(axes.flat):
            ax = axes.flat[plot_idx]
            # Show top 10 most affected rows/columns
            if 'row' in inc_df.columns:
                row_counts = inc_df[inc_df['type'] == 'row_deviation']['row'].value_counts().head(10)
                if len(row_counts) > 0:
                    ax.barh(row_counts.index.astype(str), row_counts.values, color='gold')
                    ax.set_xlabel('Number of Inconsistencies')
                    ax.set_ylabel('Row Number')
                    ax.set_title('Most Affected Rows')

    plt.tight_layout()
    plt.show()

# Summary statistics
print("\\n\\n=== FORMULA ANALYSIS SUMMARY ===")
print(f"Total Formula Errors: {len(all_errors)}")
print(f"Total Inconsistencies: {len(all_inconsistencies)}")
if all_errors or all_inconsistencies:
    print("\\nRecommendations:")
    if all_errors:
        print("- Fix formula errors immediately to prevent calculation issues")
    if all_inconsistencies:
        print("- Review inconsistent formulas - they may indicate copy/paste errors")
        print("- Consider using consistent formula patterns across rows/columns")
        print("- Verify that formula deviations are intentional")
"""

    return code
