"""Optimized notebook creation with memory-efficient Excel loading."""

import asyncio
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from structlog import get_logger

from spreadsheet_analyzer.notebook_llm.models import Cell, CellType, NotebookDocument

logger = get_logger(__name__)


async def create_optimized_sheet_notebook(
    excel_path: Path,
    sheet_name: str,
    deterministic_result: dict[str, Any] | None = None,
) -> NotebookDocument:
    """Create a NotebookDocument with optimized Excel loading.

    Key optimizations:
    1. Uses read_only mode for openpyxl
    2. Uses asyncio.to_thread for blocking I/O
    3. Uses na_filter=False for faster pandas loading
    4. Optimized memory usage patterns
    """
    cells = []

    # Cell 1: Overview
    overview_content = f"""# Excel Sheet Analysis: {sheet_name}

This notebook analyzes the sheet '{sheet_name}' from the Excel file '{excel_path.name}'.

**Analysis Date**: {pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")}
"""
    cells.append(
        Cell(id="overview", cell_type=CellType.MARKDOWN, source=overview_content, metadata={"cell_type": "overview"})
    )

    # Cell 2: Import and setup
    setup_code = """# Import required libraries
import pandas as pd
import numpy as np
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

# Excel file path
excel_file = Path(%s)
sheet_name = %s

print(f"Analyzing sheet '{sheet_name}' from {excel_file.name}")
""" % (repr(str(excel_path)), repr(sheet_name))

    cells.append(Cell(id="setup", cell_type=CellType.CODE, source=setup_code, metadata={"cell_type": "setup"}))

    # Cell 3: Optimized data loading
    load_code = """# Load the sheet data with optimized settings
try:
    # CLAUDE-PERFORMANCE: Use na_filter=False for faster loading
    # Read with all data types preserved
    df = pd.read_excel(
        excel_file,
        sheet_name=sheet_name,
        header=None,
        na_filter=False,  # Faster loading, handle NA values later if needed
        engine='openpyxl'  # Explicit engine selection
    )
    print(f"Successfully loaded sheet: {df.shape[0]} rows × {df.shape[1]} columns")

    # CLAUDE-KNOWLEDGE: openpyxl loads entire workbook into memory - use read_only mode for large files
    # Load with read_only mode to check formulas (memory efficient)
    from openpyxl import load_workbook
    wb = load_workbook(excel_file, read_only=True, data_only=False)
    ws = wb[sheet_name]

    # Get dimensions without loading all cells
    print(f"Sheet dimensions from openpyxl: {ws.max_row} rows × {ws.max_column} columns")

    # Check for formulas (limited scan for performance)
    has_formulas = False
    formula_count = 0
    for row in ws.iter_rows(min_row=1, max_row=min(100, ws.max_row), values_only=False):
        for cell in row:
            if cell.data_type == 'f':  # Formula type
                has_formulas = True
                formula_count += 1

    if has_formulas:
        print(f"\\nDetected {formula_count} formulas in first 100 rows")

    wb.close()  # Important: close read-only workbook

except Exception as e:
    print(f"Error loading sheet: {e}")
    df = None
"""

    cells.append(
        Cell(id="load_data", cell_type=CellType.CODE, source=load_code, metadata={"cell_type": "data_loading"})
    )

    # Cell 4: Memory-efficient exploration
    explore_code = """# Memory-efficient data exploration
if df is not None:
    print("\\n=== Data Overview ===")
    print(f"Shape: {df.shape}")

    # Calculate memory usage
    memory_mb = df.memory_usage(deep=True).sum() / 1024 / 1024
    print(f"Memory usage: {memory_mb:.2f} MB")

    if memory_mb > 100:
        print("⚠️  Large dataset detected. Consider sampling for analysis.")

    # Check for headers efficiently
    potential_header_row = None
    for i in range(min(5, len(df))):
        row = df.iloc[i]
        # Convert to string for consistent checking
        string_values = row.astype(str)
        # Check if mostly non-numeric strings
        non_numeric = sum(not val.replace('.', '', 1).replace('-', '', 1).isdigit()
                         for val in string_values if val and val != 'nan')
        if non_numeric > len(df.columns) * 0.5:  # More than 50% non-numeric
            potential_header_row = i
            break

    if potential_header_row is not None:
        print(f"\\nPotential header row detected at row {potential_header_row}")
        print("Headers:", list(df.iloc[potential_header_row]))

    # Display first few rows (limited for memory)
    print("\\n=== First 5 rows ===")
    print(df.head(5).to_string())

    # Basic statistics for numeric columns only
    numeric_cols = df.select_dtypes(include=[np.number]).columns
    if len(numeric_cols) > 0:
        print(f"\\n=== Numeric columns: {len(numeric_cols)} ===")
        print(df[numeric_cols].describe())
"""

    cells.append(
        Cell(id="explore", cell_type=CellType.CODE, source=explore_code, metadata={"cell_type": "exploration"})
    )

    # Cell 5: Add deterministic results if available
    if deterministic_result:
        det_content = f"""## Deterministic Analysis Results

### Summary
- **Execution Time**: {deterministic_result.get("execution_time", "N/A")}s
- **Success**: {deterministic_result.get("success", False)}

"""
        if deterministic_result.get("structure"):
            det_content += f"- **Total Sheets**: {len(deterministic_result['structure'].sheets)}\n"
        if deterministic_result.get("formulas"):
            det_content += f"- **Total Formulas**: {len(deterministic_result['formulas'].dependency_graph)}\n"
        if deterministic_result.get("security"):
            det_content += f"- **Security Risk**: {deterministic_result['security'].risk_level}\n"

        cells.append(
            Cell(
                id="deterministic_results",
                cell_type=CellType.MARKDOWN,
                source=det_content,
                metadata={"analysis_type": "deterministic"},
            )
        )

    # Cell 6: Memory management tips
    memory_tips = """## Memory Management Tips

For large Excel files:
1. Use `chunksize` parameter in pd.read_excel() to process in batches
2. Select specific columns: `usecols=['A', 'B', 'C']`
3. Sample data for exploration: `df.sample(n=10000)`
4. Use dtype specification to reduce memory: `dtype={'col': 'category'}`
5. Consider converting to Parquet for better performance

Example for chunked processing:
```python
# Process large file in chunks
chunk_size = 10000
for chunk in pd.read_excel(excel_file, sheet_name=sheet_name, chunksize=chunk_size):
    # Process each chunk
    process_chunk(chunk)
```
"""

    cells.append(
        Cell(
            id="memory_tips",
            cell_type=CellType.MARKDOWN,
            source=memory_tips,
            metadata={"cell_type": "documentation"},
        )
    )

    return NotebookDocument(cells=cells)


async def load_excel_async(
    excel_path: Path,
    sheet_name: str,
    *,
    read_only: bool = True,
    data_only: bool = True,
    na_filter: bool = False,
    nrows: int | None = None,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Load Excel file asynchronously with optimized settings.

    Args:
        excel_path: Path to Excel file
        sheet_name: Name of sheet to load
        read_only: Use read-only mode for openpyxl (memory efficient)
        data_only: Load cell values only (no formulas)
        na_filter: Filter NA values (False for better performance)
        nrows: Limit number of rows to read

    Returns:
        Tuple of (DataFrame, metadata dict)
    """

    async def _load_pandas():
        """Load Excel file with pandas in thread pool."""
        return await asyncio.to_thread(
            pd.read_excel,
            excel_path,
            sheet_name=sheet_name,
            header=None,
            na_filter=na_filter,
            nrows=nrows,
            engine="openpyxl",
        )

    async def _load_openpyxl():
        """Load Excel file with openpyxl for metadata."""

        def _get_metadata():
            wb = load_workbook(excel_path, read_only=read_only, data_only=data_only)
            try:
                ws = wb[sheet_name]
                metadata = {
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "sheet_names": wb.sheetnames,
                    "has_formulas": False,
                    "formula_count": 0,
                }

                # Quick formula check (limited rows for performance)
                if not data_only:
                    for row in ws.iter_rows(min_row=1, max_row=min(100, ws.max_row), values_only=False):
                        for cell in row:
                            if hasattr(cell, "data_type") and cell.data_type == "f":
                                metadata["has_formulas"] = True
                                metadata["formula_count"] += 1

                return metadata
            finally:
                wb.close()

        return await asyncio.to_thread(_get_metadata)

    # Load both concurrently
    df, metadata = await asyncio.gather(
        _load_pandas(),
        _load_openpyxl(),
    )

    return df, metadata
