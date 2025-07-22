"""Integration module for spreadsheet LLM analysis framework.

This module provides the SpreadsheetLLMAnalyzer class that ties together
the three-layer architecture:
1. Tool Registry (registry.py)
2. Notebook Executor (notebook_executor.py)
3. Analysis Strategies (analysis_strategies.py)
"""

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .analysis_strategies import (
    AnalysisStrategy,
    CellLevelAnalysisStrategy,
    DataQualityStrategy,
    FormulaAnalysisStrategy,
    StructureAnalysisStrategy,
)
from .notebook_executor import NotebookExecutor
from .registry import ToolRegistry


class SpreadsheetLLMAnalyzer:
    """Main integration class for LLM-powered spreadsheet analysis.

    This class provides a high-level interface for analyzing spreadsheets
    using different strategies, each powered by LLM queries executed through
    Jupyter notebooks.

    Architecture:
        - Tool Registry: Manages available analysis tools
        - Notebook Executor: Handles Jupyter kernel and execution
        - Analysis Strategies: Implements different analysis approaches
    """

    def __init__(self, workbook_path: Path | None = None):
        """Initialize the analyzer.

        Args:
            workbook_path: Optional path to Excel file to analyze
        """
        self.workbook_path = workbook_path
        self.workbook: Workbook | None = None

        # Initialize components
        self.registry = ToolRegistry()
        self.executor = NotebookExecutor(self.registry)

        # Available strategies
        self.strategies: dict[str, AnalysisStrategy] = {
            "structure": StructureAnalysisStrategy(self.executor),
            "formulas": FormulaAnalysisStrategy(self.executor),
            "data_quality": DataQualityStrategy(self.executor),
            "cell_level": CellLevelAnalysisStrategy(self.executor),
        }

        # Analysis results cache
        self.results: dict[str, Any] = {}

        # Load workbook if path provided
        if workbook_path:
            self.load_workbook(workbook_path)

    def load_workbook(self, path: Path) -> None:
        """Load an Excel workbook for analysis.

        Args:
            path: Path to the Excel file

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file is not a valid Excel file
        """
        if not path.exists():
            raise FileNotFoundError(f"Workbook not found: {path}")

        try:
            self.workbook = load_workbook(path, read_only=True, data_only=False)
            self.workbook_path = path

            # Register workbook tools
            self.registry.register_workbook_tools(self.workbook)

            # Clear cached results
            self.results.clear()

        except Exception as e:
            raise ValueError(f"Failed to load workbook: {e}")

    def analyze(self, strategy_name: str, context: dict[str, Any] | None = None, **kwargs) -> dict[str, Any]:
        """Run analysis using specified strategy.

        Args:
            strategy_name: Name of strategy to use
            context: Optional context for analysis
            **kwargs: Additional arguments for strategy

        Returns:
            Analysis results

        Raises:
            ValueError: If strategy not found or workbook not loaded
        """
        if not self.workbook:
            raise ValueError("No workbook loaded")

        if strategy_name not in self.strategies:
            raise ValueError(f"Unknown strategy: {strategy_name}. Available: {list(self.strategies.keys())}")

        # Prepare context
        if context is None:
            context = {}

        # Add workbook info to context
        context.update(
            {
                "workbook_path": str(self.workbook_path),
                "sheet_names": self.workbook.sheetnames,
                "active_sheet": self.workbook.active.title if self.workbook.active else None,
            }
        )

        # Run strategy
        strategy = self.strategies[strategy_name]
        result = strategy.analyze(self.workbook, context, **kwargs)

        # Cache results
        self.results[strategy_name] = result

        return result

    def analyze_all(self, exclude: list[str] | None = None) -> dict[str, dict[str, Any]]:
        """Run all available analysis strategies.

        Args:
            exclude: List of strategy names to exclude

        Returns:
            Dict mapping strategy names to results
        """
        if exclude is None:
            exclude = []

        results = {}
        for name, strategy in self.strategies.items():
            if name not in exclude:
                try:
                    results[name] = self.analyze(name)
                except Exception as e:
                    results[name] = {"error": str(e)}

        return results

    def get_recommendations(self) -> list[dict[str, Any]]:
        """Get actionable recommendations based on analysis results.

        Returns:
            List of recommendations with priority and details
        """
        recommendations = []

        # Check if we have analysis results
        if not self.results:
            return [
                {
                    "priority": "info",
                    "category": "analysis",
                    "message": "No analysis performed yet",
                    "action": "Run analyze() or analyze_all() first",
                }
            ]

        # Extract recommendations from each strategy result
        for strategy_name, result in self.results.items():
            if isinstance(result, dict) and "recommendations" in result:
                for rec in result["recommendations"]:
                    rec["source"] = strategy_name
                    recommendations.append(rec)

        # Sort by priority
        priority_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}
        recommendations.sort(key=lambda x: priority_order.get(x.get("priority", "info"), 5))

        return recommendations

    def get_summary(self) -> dict[str, Any]:
        """Get a summary of all analysis results.

        Returns:
            Summary dict with key findings and metrics
        """
        if not self.results:
            return {"status": "no_analysis", "message": "No analysis performed"}

        summary = {
            "workbook": str(self.workbook_path),
            "analyses_performed": list(self.results.keys()),
            "key_findings": [],
            "metrics": {},
            "recommendations_count": len(self.get_recommendations()),
        }

        # Extract key findings from each analysis
        for strategy_name, result in self.results.items():
            if isinstance(result, dict):
                # Add metrics
                if "metrics" in result:
                    summary["metrics"][strategy_name] = result["metrics"]

                # Add key findings
                if "summary" in result:
                    summary["key_findings"].append({"source": strategy_name, "finding": result["summary"]})

        return summary

    def export_results(self, output_path: Path, format: str = "json") -> None:
        """Export analysis results to file.

        Args:
            output_path: Path for output file
            format: Output format (json, markdown)
        """
        if format == "json":
            with open(output_path, "w") as f:
                json.dump(
                    {
                        "summary": self.get_summary(),
                        "results": self.results,
                        "recommendations": self.get_recommendations(),
                    },
                    f,
                    indent=2,
                )

        elif format == "markdown":
            with open(output_path, "w") as f:
                f.write("# Spreadsheet Analysis Report\n\n")

                # Summary
                summary = self.get_summary()
                f.write(f"**Workbook:** {summary['workbook']}\n\n")
                f.write(f"**Analyses Performed:** {', '.join(summary['analyses_performed'])}\n\n")

                # Key Findings
                f.write("## Key Findings\n\n")
                for finding in summary["key_findings"]:
                    f.write(f"- **{finding['source']}:** {finding['finding']}\n")
                f.write("\n")

                # Recommendations
                f.write("## Recommendations\n\n")
                for rec in self.get_recommendations():
                    f.write(f"### {rec.get('priority', 'info').upper()}: {rec.get('category', 'general')}\n")
                    f.write(f"{rec.get('message', '')}\n")
                    if "action" in rec:
                        f.write(f"**Action:** {rec['action']}\n")
                    f.write("\n")

        else:
            raise ValueError(f"Unsupported format: {format}")

    def cleanup(self) -> None:
        """Clean up resources."""
        self.executor.cleanup()
        if self.workbook:
            self.workbook.close()


# Convenience functions
def analyze_spreadsheet(
    workbook_path: str | Path, strategies: list[str] | None = None, output_path: Path | None = None
) -> dict[str, Any]:
    """Convenience function for quick spreadsheet analysis.

    Args:
        workbook_path: Path to Excel file
        strategies: List of strategies to run (None = all)
        output_path: Optional path to save results

    Returns:
        Analysis results summary
    """
    analyzer = SpreadsheetLLMAnalyzer()

    try:
        # Load workbook
        analyzer.load_workbook(Path(workbook_path))

        # Run analysis
        if strategies:
            results = {}
            for strategy in strategies:
                results[strategy] = analyzer.analyze(strategy)
        else:
            results = analyzer.analyze_all()

        # Export if requested
        if output_path:
            analyzer.export_results(output_path)

        # Return summary
        return analyzer.get_summary()

    finally:
        analyzer.cleanup()
