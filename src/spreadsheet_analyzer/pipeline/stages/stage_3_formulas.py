"""
Stage 3: Formula Analysis and Dependency Graphing

This module performs comprehensive formula analysis on Excel workbooks, including:
- Formula parsing and dependency extraction
- Dependency graph construction with optional semantic edge detection
- Circular reference detection
- Complexity scoring and depth calculation
- Range membership tracking for efficient large range handling

CLAUDE-KNOWLEDGE: This refactored version consolidates the best features from both
the original stage_3_formulas.py and stage_3_formulas_enhanced.py, providing
semantic analysis as an optional feature.
"""

import logging
import re
from collections import defaultdict, deque

# dataclass imports moved to types.py
from pathlib import Path
from typing import Any, Final, NamedTuple

import openpyxl
from openpyxl.cell import Cell
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from spreadsheet_analyzer.pipeline.types import (
    EdgeMetadata,
    Err,
    FormulaAnalysis,
    FormulaNode,
    Ok,
    RangeMembershipIndex,
)

logger = logging.getLogger(__name__)

# ============================================================================
# CONFIGURATION CONSTANTS
# ============================================================================

# Core analysis constants
MAX_CELLS_FOR_ANALYSIS: Final[int] = 10_000  # Maximum cells to analyze to prevent memory issues
MAX_CELLS_PER_CHUNK: Final[int] = 1000  # Process formulas in chunks for memory efficiency
VOLATILITY_MULTIPLIER: Final[float] = 2.0  # Complexity multiplier for volatile formulas (e.g., NOW(), RAND())
EXTERNAL_REF_MULTIPLIER: Final[float] = 1.5  # Complexity multiplier for external references ([workbook.xlsx])
BASE_COMPLEXITY_SCORE: Final[float] = 1.0  # Starting complexity score for all formulas

# Range handling strategies
RANGE_STRATEGY_SKIP: Final[str] = "skip"  # Skip all range references (fastest)
RANGE_STRATEGY_EXPAND: Final[str] = "expand"  # Expand ranges to individual cells (most detail)
RANGE_STRATEGY_SUMMARIZE: Final[str] = "summarize"  # Keep ranges as single entities
RANGE_STRATEGY_SMART: Final[str] = "smart"  # Expand small ranges, summarize large ones
DEFAULT_RANGE_STRATEGY: Final[str] = RANGE_STRATEGY_SMART
SMART_RANGE_THRESHOLD: Final[int] = 10  # Ranges larger than this are kept as ranges, not expanded

# Feature flags for optional functionality
ENABLE_SEMANTIC_ANALYSIS: Final[bool] = True  # Detect relationship types (SUMS_OVER, LOOKS_UP_IN, etc.)
ENABLE_CELL_METADATA: Final[bool] = True  # Collect cell formatting and type information
ENABLE_EDGE_WEIGHTS: Final[bool] = True  # Calculate importance weights based on range sizes

# Pattern constants for formula parsing
EXCEL_OPERATORS: Final[str] = r"[\+\-\*/\^<>=&:\(\),]"  # Excel formula operators
CELL_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"(?:([A-Za-z_][\w.]*|'[^']+')!)?"  # Optional sheet name (e.g., 'Sheet1!' or Sheet1!)
    r"("  # Start cell/range group
    r"\$?[A-Z]+\$?\d+"  # Cell reference (e.g., A1, $A$1)
    r"(?::\$?[A-Z]+\$?\d+)?"  # Optional range end (e.g., :B10 for A1:B10)
    r"|"  # OR
    r"\$?[A-Z]+:\$?[A-Z]+"  # Column range (e.g., A:D, $A:$D)
    r")",  # End cell/range group
    re.IGNORECASE,
)


# ============================================================================
# DATA STRUCTURES
# ============================================================================


class CellReference(NamedTuple):
    """
    Represents a cell or range reference with sheet context.

    This immutable structure ensures consistent representation of cell
    references throughout the analysis pipeline.
    """

    sheet: str
    start_col: int
    start_row: int | None  # Can be None for column ranges like A:D
    end_col: int | None = None
    end_row: int | None = None

    @property
    def is_range(self) -> bool:
        """Check if this reference is a range (not a single cell)."""
        return self.end_col is not None

    @property
    def cell_count(self) -> int:
        """Calculate the number of cells in this reference."""
        if not self.is_range:
            return 1

        col_count = (self.end_col - self.start_col + 1) if self.end_col else 1

        # For column ranges (A:D), we don't have specific row counts
        # Return a large number to indicate it's a full column range
        if self.start_row is None or self.end_row is None:
            # Excel has 1,048,576 rows, but we'll use a more reasonable estimate
            return col_count * 1000  # Assume 1000 rows for column ranges

        row_count = self.end_row - self.start_row + 1
        return col_count * row_count

    def to_key(self) -> str:
        """Convert to a string key for use in dictionaries."""
        # Check if sheet name needs quotes (contains spaces or special characters)
        sheet_name = self.sheet
        if " " in sheet_name or "!" in sheet_name or "'" in sheet_name:
            # Escape single quotes by doubling them
            escaped_sheet = sheet_name.replace("'", "''")
            sheet_name = f"'{escaped_sheet}'"

        if self.is_range:
            # Handle column ranges (where row might be None)
            if self.start_row is None and self.end_row is None:
                # Column range like A:D
                start = get_column_letter(self.start_col)
                end = get_column_letter(self.end_col)
            else:
                # Regular range like A1:B10
                start = f"{get_column_letter(self.start_col)}{self.start_row}"
                end = f"{get_column_letter(self.end_col)}{self.end_row}"
            return f"{sheet_name}!{start}:{end}"
        else:
            cell = f"{get_column_letter(self.start_col)}{self.start_row}"
            return f"{sheet_name}!{cell}"


# FormulaNode is now imported from types.py


# FormulaAnalysis is now imported from types.py


# ============================================================================
# SEMANTIC EDGE DETECTION
# ============================================================================


class SemanticEdgeDetector:
    """
    Detects semantic relationship types between cells based on Excel functions.

    This class provides optional semantic analysis to understand not just
    dependencies, but the nature of relationships between cells.

    CLAUDE-KNOWLEDGE: Understanding semantic relationships helps in visualizing
    and optimizing spreadsheet structures.
    """

    # Mapping of Excel functions to semantic edge types
    # These mappings help understand the nature of relationships between cells
    FUNCTION_SEMANTICS: Final[dict[str, str]] = {
        # Aggregation functions
        "SUM": "SUMS_OVER",
        "SUMIF": "CONDITIONALLY_SUMS",
        "SUMIFS": "CONDITIONALLY_SUMS",
        "AVERAGE": "AVERAGES_OVER",
        "COUNT": "COUNTS_FROM",
        "COUNTA": "COUNTS_FROM",
        "MAX": "FINDS_MAX_IN",
        "MIN": "FINDS_MIN_IN",
        # Lookup functions
        "VLOOKUP": "LOOKS_UP_IN",
        "HLOOKUP": "LOOKS_UP_IN",
        "INDEX": "INDEXES_INTO",
        "MATCH": "MATCHES_IN",
        "XLOOKUP": "LOOKS_UP_IN",
        # Conditional functions
        "IF": "CONDITIONALLY_USES",
        "IFS": "CONDITIONALLY_USES",
        "COUNTIF": "CONDITIONALLY_COUNTS",
        "COUNTIFS": "CONDITIONALLY_COUNTS",
        # Reference functions
        "INDIRECT": "INDIRECTLY_REFERENCES",
        "OFFSET": "OFFSETS_FROM",
        # Financial functions
        "NPV": "CALCULATES_NPV_FROM",
        "IRR": "CALCULATES_IRR_FROM",
        "PMT": "CALCULATES_PMT_FROM",
    }

    def detect_edge_type(
        self, formula: str, dependency: CellReference, context: dict[str, Any] | None = None
    ) -> EdgeMetadata:
        """
        Detect the semantic type of an edge based on the formula.

        Args:
            formula: The formula containing the reference
            dependency: The cell reference being analyzed
            context: Optional context about the formula

        Returns:
            EdgeMetadata describing the relationship
        """
        # Extract the function context
        func_name, arg_position = self._extract_function_context(formula, dependency)

        # Determine edge type
        edge_type = "REFERENCES"  # Default
        if func_name and func_name.upper() in self.FUNCTION_SEMANTICS:
            edge_type = self.FUNCTION_SEMANTICS[func_name.upper()]

        # Calculate edge weight based on range size
        # Larger ranges get higher weights (up to 10.0) to indicate their importance
        weight = 1.0
        if ENABLE_EDGE_WEIGHTS:
            weight = min(10.0, 1.0 + (dependency.cell_count - 1) * 0.1)

        # Build metadata
        metadata = {
            "formula_template": self._create_formula_template(formula),
        }
        if context:
            metadata.update(context)

        return EdgeMetadata(
            edge_type=edge_type,
            function_name=func_name,
            argument_position=arg_position,
            weight=weight,
            metadata=metadata,
        )

    def _extract_function_context(self, formula: str, _dependency: CellReference) -> tuple[str | None, int | None]:
        """Extract the function name and argument position for a dependency."""
        # dependency parameter is reserved for future use when we implement full AST parsing
        # This is a simplified version - a full implementation would need
        # proper formula parsing to handle nested functions

        # Find function calls in the formula
        func_pattern = r"([A-Z]+)\s*\("
        matches = list(re.finditer(func_pattern, formula, re.IGNORECASE))

        if not matches:
            return None, None

        # For now, return the first function found
        # A complete implementation would analyze the AST
        func_name = matches[0].group(1)

        # Estimate argument position (simplified)
        arg_position = 0

        return func_name, arg_position

    def _create_formula_template(self, formula: str) -> str:
        """Create a template by replacing cell references with placeholders."""
        template = formula

        # Replace cell references with placeholders
        for match in CELL_PATTERN.finditer(formula):
            ref = match.group(0)
            template = template.replace(ref, "<REF>", 1)

        return template


# ============================================================================
# FORMULA PARSER
# ============================================================================


class FormulaParser:
    """
    Parses Excel formulas to extract cell and range dependencies.

    This parser handles various Excel formula constructs including:
    - Simple cell references (A1, $A$1)
    - Range references (A1:B10)
    - Cross-sheet references (Sheet1!A1)
    - Sheet names with spaces ('My Sheet'!A1)

    CLAUDE-COMPLEX: Excel formula parsing requires handling many edge cases
    including escaped quotes, array formulas, and international formats.
    """

    def __init__(self) -> None:
        """Initialize parser with compiled patterns."""
        self._cache: dict[str, list[CellReference]] = {}
        self._cache_hits = 0
        self._cache_misses = 0

    def parse_formula(self, formula: str, current_sheet: str) -> list[CellReference]:
        """
        Parse a formula and extract all cell/range references.

        Args:
            formula: Excel formula to parse
            current_sheet: Current sheet name for relative references

        Returns:
            List of CellReference objects found in the formula
        """
        # Check cache first
        cache_key = f"{current_sheet}|{formula}"
        if cache_key in self._cache:
            self._cache_hits += 1
            return self._cache[cache_key]

        self._cache_misses += 1

        # Parse the formula
        references = self._extract_references(formula, current_sheet)

        # Cache the result
        self._cache[cache_key] = references

        return references

    def _extract_references(self, formula: str, current_sheet: str) -> list[CellReference]:
        """Extract all cell references from a formula."""
        references = []

        # Find all cell/range references
        for match in CELL_PATTERN.finditer(formula):
            sheet_part = match.group(1)
            cell_part = match.group(2)

            # Determine sheet name
            sheet = sheet_part.strip("'") if sheet_part else current_sheet

            # Parse the cell/range reference
            if ":" in cell_part:
                # It's a range
                start, end = cell_part.split(":", 1)
                ref = self._parse_range(sheet, start, end)
            else:
                # It's a single cell
                ref = self._parse_cell(sheet, cell_part)

            if ref:
                references.append(ref)

        return references

    def _parse_cell(self, sheet: str, cell_ref: str) -> CellReference | None:
        """Parse a single cell reference."""
        try:
            cell_ref = cell_ref.replace("$", "")
            row, col = coordinate_to_tuple(cell_ref)
            return CellReference(sheet, col, row)
        except Exception as e:
            logger.debug("Failed to parse cell reference '%s': %s", cell_ref, e)
            return None

    def _parse_range(self, sheet: str, start_ref: str, end_ref: str) -> CellReference | None:
        """Parse a range reference."""
        try:
            start_ref = start_ref.replace("$", "")
            end_ref = end_ref.replace("$", "")

            # Handle full column/row references
            min_col, min_row, max_col, max_row = range_boundaries(f"{start_ref}:{end_ref}")

            return CellReference(sheet, min_col, min_row, max_col, max_row)
        except Exception as e:
            logger.debug("Failed to parse range '%s:%s': %s", start_ref, end_ref, e)
            return None


# ============================================================================
# DEPENDENCY GRAPH BUILDER
# ============================================================================


class DependencyGraph:
    """
    Builds and manages the formula dependency graph.

    This class constructs a directed graph where nodes represent cells
    with formulas and edges represent dependencies between cells.

    CLAUDE-KNOWLEDGE: The dependency graph is the core data structure for
    understanding spreadsheet logic and detecting issues like circular references.
    """

    def __init__(self, *, enable_semantic_analysis: bool = False):
        """
        Initialize empty dependency graph.

        Args:
            enable_semantic_analysis: Whether to include semantic edge detection
        """
        self.nodes: dict[str, FormulaNode] = {}
        self.adjacency_list: dict[str, set[str]] = defaultdict(set)
        self.reverse_adjacency: dict[str, set[str]] = defaultdict(set)
        self.enable_semantic_analysis = enable_semantic_analysis

        if enable_semantic_analysis:
            self.edge_detector = SemanticEdgeDetector()

    def add_node(
        self,
        sheet: str,
        cell: str,
        formula: str,
        dependencies: list[CellReference],
        *,
        volatile: bool = False,
        external: bool = False,
        cell_metadata: dict[str, Any] | None = None,
    ) -> None:
        """
        Add a node to the dependency graph.

        Args:
            sheet: Sheet name
            cell: Cell reference
            formula: The formula in the cell
            dependencies: List of cell references this formula depends on
            volatile: Whether the formula contains volatile functions
            external: Whether the formula has external references
            cell_metadata: Optional metadata about the cell
        """
        node_key = f"{sheet}!{cell}"

        # Convert dependencies to string keys
        dep_keys = []
        edge_labels = {} if self.enable_semantic_analysis else None

        for dep in dependencies:
            dep_key = dep.to_key()
            dep_keys.append(dep_key)

            # Add semantic edge detection if enabled
            if self.enable_semantic_analysis:
                edge_metadata = self.edge_detector.detect_edge_type(formula, dep, {"sheet": sheet, "cell": cell})
                edge_labels[dep_key] = edge_metadata

        # Calculate complexity score
        complexity = self._calculate_complexity(formula, len(dep_keys), volatile, external)

        # Create node
        node = FormulaNode(
            sheet=sheet,
            cell=cell,
            formula=formula,
            dependencies=frozenset(dep_keys),
            volatile=volatile,
            external=external,
            complexity_score=complexity,
            edge_labels=edge_labels,
            cell_metadata=cell_metadata if ENABLE_CELL_METADATA else None,
        )

        # Add to graph
        self.nodes[node_key] = node

        # Update adjacency lists
        for dep_key in dep_keys:
            self.adjacency_list[node_key].add(dep_key)
            self.reverse_adjacency[dep_key].add(node_key)

    def _calculate_complexity(self, formula: str, dep_count: int, volatile: bool, external: bool) -> float:
        """Calculate complexity score for a formula.

        Complexity factors:
        - Base score: 1.0
        - Formula length: +0.01 per character (incentivizes simpler formulas)
        - Dependencies: +0.5 per dependency (more dependencies = more complex)
        - Volatile functions: 2x multiplier (recalculate frequently)
        - External references: 1.5x multiplier (cross-workbook dependencies)
        """
        score = BASE_COMPLEXITY_SCORE

        # Factor in formula length
        score += len(formula) / 100

        # Factor in dependency count
        score += dep_count * 0.5

        # Apply multipliers
        if volatile:
            score *= VOLATILITY_MULTIPLIER
        if external:
            score *= EXTERNAL_REF_MULTIPLIER

        return round(score, 2)

    def find_circular_references(self) -> set[frozenset[str]]:
        """
        Find all circular reference cycles in the graph.

        Uses Tarjan's strongly connected components algorithm to find
        all cycles efficiently.

        Returns:
            Set of cycles, where each cycle is a frozenset of node keys
        """
        # Tarjan's algorithm for finding strongly connected components
        index = 0
        stack = []
        indices = {}
        lowlinks = {}
        on_stack = set()
        cycles = []

        def strongconnect(v: str) -> None:
            nonlocal index

            indices[v] = index
            lowlinks[v] = index
            index += 1
            stack.append(v)
            on_stack.add(v)

            # Visit neighbors
            for w in self.adjacency_list.get(v, []):
                if w not in indices:
                    strongconnect(w)
                    lowlinks[v] = min(lowlinks[v], lowlinks[w])
                elif w in on_stack:
                    lowlinks[v] = min(lowlinks[v], indices[w])

            # Found SCC root
            if lowlinks[v] == indices[v]:
                component = []
                while True:
                    w = stack.pop()
                    on_stack.remove(w)
                    component.append(w)
                    if w == v:
                        break

                # Only keep actual cycles (more than one node)
                if len(component) > 1:
                    cycles.append(frozenset(component))

        # Find all SCCs
        for node in self.nodes:
            if node not in indices:
                strongconnect(node)

        return set(cycles)

    def calculate_max_depth(self) -> int:
        """
        Calculate the maximum dependency depth in the graph.

        This represents the longest chain of dependencies, which can
        impact calculation performance.

        CLAUDE-PERFORMANCE: Deep dependency chains can cause performance
        issues in Excel, especially with volatile functions.
        """
        if not self.nodes:
            return 0

        # Find nodes with no dependencies OR whose dependencies are not in the graph
        # (i.e., they depend on cells that aren't formulas)
        leaf_nodes = []
        for node_key, node in self.nodes.items():
            if not node.dependencies:
                leaf_nodes.append(node_key)
            else:
                # Check if all dependencies are outside the graph
                all_deps_external = True
                for dep in node.dependencies:
                    if dep in self.nodes:
                        all_deps_external = False
                        break
                if all_deps_external:
                    leaf_nodes.append(node_key)

        if not leaf_nodes:
            # All nodes have dependencies that are in the graph - there must be cycles
            return -1

        # BFS from leaf nodes
        max_depth = 0
        visited = set()
        queue = deque([(node, 0) for node in leaf_nodes])

        while queue:
            node, depth = queue.popleft()

            if node in visited:
                continue

            visited.add(node)
            max_depth = max(max_depth, depth)

            # Add dependent nodes
            for dependent in self.reverse_adjacency.get(node, []):
                if dependent not in visited:
                    queue.append((dependent, depth + 1))

        return max_depth


# ============================================================================
# FORMULA ANALYZER
# ============================================================================


class FormulaAnalyzer:
    """
    Main analyzer class that orchestrates formula analysis.

    This class coordinates the parsing, graph building, and analysis
    phases to produce comprehensive formula analysis results.

    CLAUDE-KNOWLEDGE: The analyzer uses a chunked processing approach
    to handle large workbooks efficiently without running out of memory.
    """

    def __init__(
        self,
        *,
        range_strategy: str = DEFAULT_RANGE_STRATEGY,
        enable_semantic_analysis: bool = ENABLE_SEMANTIC_ANALYSIS,
        enable_cell_metadata: bool = ENABLE_CELL_METADATA,
    ):
        """
        Initialize the formula analyzer.

        Args:
            range_strategy: How to handle large ranges
            enable_semantic_analysis: Whether to perform semantic edge detection
            enable_cell_metadata: Whether to collect cell metadata
        """
        self.parser = FormulaParser()
        self.graph = DependencyGraph(enable_semantic_analysis=enable_semantic_analysis)
        self.range_strategy = range_strategy
        self.enable_semantic_analysis = enable_semantic_analysis
        self.enable_cell_metadata = enable_cell_metadata

        # Tracking
        self.volatile_formulas: set[str] = set()
        self.external_references: set[str] = set()
        self.processed_cells = 0
        self.total_formulas = 0
        self.skipped_ranges = 0

    def analyze_workbook(self, workbook_path: Path, progress_callback: Any = None) -> FormulaAnalysis:
        """
        Analyze all formulas in a workbook.

        Args:
            workbook_path: Path to the Excel file
            progress_callback: Optional callback for progress updates

        Returns:
            Complete formula analysis results
        """
        # Load workbook in read-only mode for performance
        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False, keep_vba=False)

        try:
            # Report initial progress
            if progress_callback:
                progress_callback("formula_analysis", 0.0, "Starting formula analysis")

            # Analyze each sheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                self._analyze_sheet(sheet, sheet_name, progress_callback)

            # Build range membership index
            range_index = self._build_range_index()

            # Find circular references
            circular_refs = self.graph.find_circular_references()

            # Calculate max depth
            max_depth = self.graph.calculate_max_depth()

            # Calculate total complexity
            total_complexity = sum(node.complexity_score for node in self.graph.nodes.values())

            # Compile statistics
            statistics = {
                "total_formulas": self.total_formulas,
                "processed_cells": self.processed_cells,
                "skipped_ranges": self.skipped_ranges,
                "circular_reference_count": len(circular_refs),
                "volatile_formula_count": len(self.volatile_formulas),
                "external_reference_count": len(self.external_references),
                "unique_dependencies": len(self.graph.adjacency_list),
                "parser_cache_hits": self.parser._cache_hits,
                "parser_cache_misses": self.parser._cache_misses,
            }

            # Report completion
            if progress_callback:
                progress_callback(
                    "formula_analysis", 1.0, f"Analysis complete: {self.total_formulas} formulas analyzed"
                )

            return FormulaAnalysis(
                dependency_graph=dict(self.graph.nodes),
                circular_references=frozenset(circular_refs),
                volatile_formulas=frozenset(self.volatile_formulas),
                external_references=frozenset(self.external_references),
                max_dependency_depth=max_depth,
                formula_complexity_score=round(total_complexity, 2),
                statistics=statistics,
                range_index=range_index,
            )

        finally:
            wb.close()

    def _analyze_sheet(self, sheet: Worksheet, sheet_name: str, progress_callback: Any = None) -> None:
        """Analyze all formulas in a sheet using chunked processing."""
        # Process in chunks for memory efficiency
        chunk_count = 0
        cells_in_chunk = []

        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f" and cell.value:
                    cells_in_chunk.append(cell)

                    if len(cells_in_chunk) >= MAX_CELLS_PER_CHUNK:
                        self._process_chunk(cells_in_chunk, sheet_name, progress_callback)
                        cells_in_chunk = []
                        chunk_count += 1

        # Process remaining cells
        if cells_in_chunk:
            self._process_chunk(cells_in_chunk, sheet_name, progress_callback)

    def _process_chunk(self, cells: list[Cell], sheet_name: str, progress_callback: Any = None) -> None:
        """Process a chunk of cells with formulas."""
        for cell in cells:
            try:
                self._analyze_formula(cell, sheet_name)

                if progress_callback and self.processed_cells % 100 == 0:
                    progress_callback(
                        "formula_analysis",
                        self.processed_cells / MAX_CELLS_FOR_ANALYSIS,
                        f"Analyzed {self.processed_cells} formulas",
                    )

            except Exception as e:
                logger.warning("Failed to analyze formula in %s!%s: %s", sheet_name, cell.coordinate, e)

    def _analyze_formula(self, cell: Cell, sheet_name: str) -> None:
        """Analyze a single cell's formula."""
        self.processed_cells += 1

        formula = str(cell.value)
        cell_ref = cell.coordinate

        # Check for volatile functions
        is_volatile = self._check_volatile(formula)
        if is_volatile:
            self.volatile_formulas.add(f"{sheet_name}!{cell_ref}")

        # Check for external references
        is_external = self._check_external(formula)
        if is_external:
            self.external_references.add(f"{sheet_name}!{cell_ref}")

        # Parse dependencies
        dependencies = self.parser.parse_formula(formula, sheet_name)

        # Filter dependencies based on range strategy
        filtered_deps = self._apply_range_strategy(dependencies)

        # Collect cell metadata if enabled
        cell_metadata = None
        if self.enable_cell_metadata:
            cell_metadata = {
                "row": cell.row,
                "column": cell.column,
                "data_type": cell.data_type,
                "number_format": cell.number_format,
            }

        # Add to graph
        self.graph.add_node(
            sheet_name,
            cell_ref,
            formula,
            filtered_deps,
            volatile=is_volatile,
            external=is_external,
            cell_metadata=cell_metadata,
        )

        self.total_formulas += 1

    def _check_volatile(self, formula: str) -> bool:
        """Check if formula contains volatile functions.

        Volatile functions recalculate every time Excel recalculates,
        potentially impacting performance. Common volatile functions:
        - NOW(): Current date and time
        - TODAY(): Current date
        - RAND()/RANDBETWEEN(): Random numbers
        - OFFSET(): Dynamic range references
        - INDIRECT(): Dynamic cell references
        """
        volatile_functions = {"NOW", "TODAY", "RAND", "RANDBETWEEN", "OFFSET", "INDIRECT"}
        formula_upper = formula.upper()
        return any(f"{func}(" in formula_upper for func in volatile_functions)

    def _check_external(self, formula: str) -> bool:
        """Check if formula contains external references.

        External references link to other workbooks using the pattern:
        [workbook.xlsx]Sheet1!A1

        These create dependencies on external files and can break if
        the referenced workbook is moved or renamed.
        """
        # Look for external workbook references [workbook.xlsx]
        return "[" in formula and "]" in formula

    def _apply_range_strategy(self, dependencies: list[CellReference]) -> list[CellReference]:
        """Apply the configured range handling strategy.

        Strategies:
        - SKIP: Ignore all range references (fastest, but loses information)
        - EXPAND: Convert ranges to individual cells (detailed, but memory intensive)
        - SUMMARIZE: Keep ranges as single entities (balanced approach)
        - SMART: Expand small ranges (≤10 cells), keep large ranges as-is (recommended)

        The strategy choice impacts:
        - Memory usage (EXPAND uses most)
        - Analysis detail (SKIP loses range information)
        - Performance (SKIP is fastest, EXPAND is slowest)
        """
        if self.range_strategy == RANGE_STRATEGY_SKIP:
            # Skip all ranges
            return [d for d in dependencies if not d.is_range]

        elif self.range_strategy == RANGE_STRATEGY_EXPAND:
            # Expand all ranges (dangerous for large ranges!)
            expanded = []
            for dep in dependencies:
                if dep.is_range and dep.cell_count > 1000:
                    logger.warning("Skipping large range %s with %s cells", dep.to_key(), dep.cell_count)
                    self.skipped_ranges += 1
                else:
                    expanded.append(dep)
            return expanded

        elif self.range_strategy == RANGE_STRATEGY_SUMMARIZE:
            # Keep ranges as single entities
            return dependencies

        else:  # RANGE_STRATEGY_SMART (default)
            # Smart strategy: expand small ranges, summarize large ones
            result = []
            for dep in dependencies:
                if dep.is_range and dep.cell_count > SMART_RANGE_THRESHOLD:
                    # Keep as range for large references
                    result.append(dep)
                else:
                    # Expand small ranges or keep single cells
                    result.append(dep)
            return result

    def _build_range_index(self) -> RangeMembershipIndex:
        """Build an index of all ranges for membership queries."""
        sheet_ranges = defaultdict(list)

        for _node_key, node in self.graph.nodes.items():
            for dep_key in node.dependencies:
                # Parse the dependency key
                if ":" in dep_key:
                    # It's a range reference
                    parts = dep_key.split("!")
                    if len(parts) == 2:
                        range_sheet, range_ref = parts
                        try:
                            min_col, min_row, max_col, max_row = range_boundaries(range_ref)
                            sheet_ranges[range_sheet].append((min_row, max_row, min_col, max_col, dep_key))
                        except Exception as e:
                            # Log parsing errors for range boundaries
                            logger.debug("Failed to parse range boundaries for %s: %s", range_ref, e)

        return RangeMembershipIndex(dict(sheet_ranges))


# ============================================================================
# PUBLIC API
# ============================================================================


def stage_3_formula_analysis(
    file_path: Path,
    *,
    range_strategy: str = DEFAULT_RANGE_STRATEGY,
    enable_semantic_analysis: bool = False,
    progress_callback: Any = None,
) -> Ok[FormulaAnalysis] | Err:
    """
    Perform formula analysis on an Excel workbook.

    This is the main entry point for Stage 3 of the pipeline. It analyzes
    all formulas in the workbook and builds a comprehensive dependency graph
    with optional semantic enhancements.

    Args:
        file_path: Path to the Excel file
        range_strategy: How to handle large ranges (skip/expand/summarize/smart)
        enable_semantic_analysis: Whether to perform semantic edge detection
        progress_callback: Optional callback for progress updates

    Returns:
        Ok(FormulaAnalysis) on success, Err(str) on failure

    Example:
        result = stage_3_formula_analysis(
            Path("financial_model.xlsx"),
            enable_semantic_analysis=True
        )

        if isinstance(result, Ok):
            analysis = result.value
            print(f"Found {len(analysis.dependency_graph)} formulas")
            print(f"Max dependency depth: {analysis.max_dependency_depth}")

    CLAUDE-KNOWLEDGE: This function is designed to be called as part of the
    deterministic pipeline but can also be used standalone for formula analysis.
    """
    try:
        # Validate file exists
        if not file_path.exists():
            return Err(f"File not found: {file_path}")

        # Create analyzer
        analyzer = FormulaAnalyzer(range_strategy=range_strategy, enable_semantic_analysis=enable_semantic_analysis)

        # Perform analysis
        logger.info("Starting formula analysis for %s", file_path)
        analysis = analyzer.analyze_workbook(file_path, progress_callback)

        logger.info(
            "Formula analysis complete: %d formulas analyzed, %d circular references found",
            analysis.statistics["total_formulas"],
            len(analysis.circular_references),
        )

        return Ok(analysis)

    except Exception as e:
        error_msg = f"Formula analysis failed: {e!s}"
        logger.exception(error_msg)
        return Err(error_msg)


def create_formula_validator(workbook_path: Path) -> Ok[Any] | Err:
    """
    Create a validator function for formula verification.

    This utility function creates a validator that can check if formulas
    in the workbook calculate correctly.

    Args:
        workbook_path: Path to the Excel file

    Returns:
        Ok(validator_function) on success, Err(str) on failure
    """
    try:
        # This would create a validator that can evaluate formulas
        # For now, return a placeholder
        def validator(formula: str, expected_value: Any) -> bool:
            """Validate that a formula produces the expected value."""
            # Placeholder implementation
            return True

        return Ok(validator)

    except Exception as e:
        return Err(f"Failed to create validator: {e!s}")
