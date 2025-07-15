"""
Stage 2: Structural Mapping (Hybrid FP/OOP).

This module implements structural analysis of Excel files using a hybrid approach:
- Functional programming for pure transformations and analysis
- Object-oriented for managing complex workbook state during traversal
"""

import logging
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path
from typing import Final

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from spreadsheet_analyzer.pipeline.types import Err, Ok, Result, SheetStructure, WorkbookStructure

logger = logging.getLogger(__name__)

# ==================== Constants ====================

# Complexity scoring thresholds

SMALL_SHEET_COUNT: Final[int] = 3
MEDIUM_SHEET_COUNT: Final[int] = 10
LARGE_SHEET_COUNT: Final[int] = 50

# Feature detection sampling limits
FEATURE_SAMPLE_ROWS: Final[int] = 100
FEATURE_SAMPLE_COLS: Final[int] = 100

SMALL_CELL_COUNT: Final[int] = 1000
MEDIUM_CELL_COUNT: Final[int] = 10000
LARGE_CELL_COUNT: Final[int] = 100000

NAMED_RANGE_THRESHOLD: Final[int] = 10

# ==================== Data Classes ====================


@dataclass(frozen=True)
class ComplexityMetrics:
    """Metrics for calculating complexity score."""

    sheet_count: int
    total_cells: int
    total_formulas: int
    named_range_count: int
    has_vba: bool
    has_external_links: bool


# ==================== Pure Functions for Analysis ====================


def get_cell_data_type(cell) -> str:
    """
    Determine the data type of a cell value.

    CLAUDE-KNOWLEDGE: openpyxl cell types:
    - 'n' = numeric
    - 's' = string
    - 'b' = boolean
    - 'd' = date
    - 'e' = error
    - 'f' = formula
    """
    if cell.value is None:
        return "empty"

    # Map data types to descriptions
    type_mapping = {
        "f": "formula",
        "n": "number",
        "s": "string",
        "b": "boolean",
        "d": "date",
        "e": "error",
    }

    return type_mapping.get(cell.data_type, "unknown")


def calculate_used_range(worksheet: Worksheet) -> tuple[str, int, int]:
    """
    Calculate the actual used range of a worksheet.

    CLAUDE-PERFORMANCE: We iterate only through used cells
    instead of the entire grid.
    """
    min_row = worksheet.min_row
    max_row = worksheet.max_row
    min_col = worksheet.min_column
    max_col = worksheet.max_column

    # Check if sheet is effectively empty (default 1,1 dimensions)
    if min_row == max_row == min_col == max_col == 1:
        # Check if the single cell has any content
        cell_value = worksheet.cell(1, 1).value
        if cell_value is None:
            return "A1", 0, 0

    # Convert to Excel notation
    start_cell = f"{get_column_letter(min_col)}{min_row}"
    end_cell = f"{get_column_letter(max_col)}{max_row}"
    used_range = f"{start_cell}:{end_cell}"

    # Calculate dimensions
    row_count = max_row - min_row + 1
    col_count = max_col - min_col + 1

    return used_range, row_count, col_count


def analyze_sheet_features(worksheet: Worksheet) -> dict[str, bool]:
    """
    Analyze sheet features using pure functions.

    Returns dict of feature flags.
    """
    features = {
        "has_data": False,
        "has_formulas": False,
        "has_charts": False,
        "has_pivot_tables": False,
    }

    # Check for data and formulas by sampling cells
    # Use the used range to avoid iterating over empty cells
    max_row = min(worksheet.max_row or 1, FEATURE_SAMPLE_ROWS)
    max_col = min(worksheet.max_column or 1, FEATURE_SAMPLE_COLS)

    for row in worksheet.iter_rows(max_row=max_row, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                features["has_data"] = True
                # Check if it's a formula cell
                if hasattr(cell, "data_type") and cell.data_type == "f":
                    features["has_formulas"] = True
                # Early exit if we found both
                if features["has_data"] and features["has_formulas"]:
                    break
        if features["has_data"] and features["has_formulas"]:
            break

    # Check for charts
    if hasattr(worksheet, "_charts") and worksheet._charts:  # noqa: SLF001
        features["has_charts"] = True

    # Check for pivot tables
    if hasattr(worksheet, "_pivots") and worksheet._pivots:  # noqa: SLF001
        features["has_pivot_tables"] = True

    return features


def count_cells_and_formulas(worksheet: Worksheet) -> tuple[int, int]:
    """
    Count non-empty cells and formulas in worksheet.

    CLAUDE-PERFORMANCE: We limit iteration to the used range to avoid
    processing empty cells beyond the data area.

    CLAUDE-GOTCHA: In openpyxl, when data_only=False, formulas are stored
    in cell.value and identified by cell.data_type == 'f'
    """
    cell_count = 0
    formula_count = 0

    # Get the actual used range to avoid iterating over empty cells
    min_row = worksheet.min_row
    max_row = worksheet.max_row
    min_col = worksheet.min_column
    max_col = worksheet.max_column

    # Skip if sheet is empty
    if not all([min_row, max_row, min_col, max_col]):
        return 0, 0

    # Iterate only through used area
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                cell_count += 1
                # Check if it's a formula cell
                if hasattr(cell, "data_type") and cell.data_type == "f":
                    formula_count += 1

    return cell_count, formula_count


def extract_named_ranges(workbook: Workbook) -> tuple[str, ...]:
    """
    Extract all named ranges from workbook.

    CLAUDE-KNOWLEDGE: Named ranges can be workbook-scoped or sheet-scoped.
    """
    named_ranges: list[str] = []

    try:
        if hasattr(workbook, "defined_names"):
            # Try to iterate over defined names
            named_ranges.extend(str(name) for name in workbook.defined_names)
    except (AttributeError, ValueError, TypeError):
        # If there's any error accessing defined names, just return empty
        logger.debug("Unable to access defined names from workbook")

    return tuple(sorted(set(named_ranges)))


def calculate_complexity_score(metrics: ComplexityMetrics) -> int:
    """
    Calculate workbook complexity score (0-100).

    Pure function that considers multiple factors.
    """
    score = 0

    # Sheet complexity (0-20 points)
    if metrics.sheet_count <= SMALL_SHEET_COUNT:
        score += 5
    elif metrics.sheet_count <= MEDIUM_SHEET_COUNT:
        score += 10
    elif metrics.sheet_count <= LARGE_SHEET_COUNT:
        score += 15
    else:
        score += 20

    # Cell complexity (0-20 points)
    if metrics.total_cells <= SMALL_CELL_COUNT:
        score += 5
    elif metrics.total_cells <= MEDIUM_CELL_COUNT:
        score += 10
    elif metrics.total_cells <= LARGE_CELL_COUNT:
        score += 15
    else:
        score += 20

    # Formula complexity (0-30 points)
    formula_ratio = metrics.total_formulas / max(metrics.total_cells, 1)
    score += min(30, int(formula_ratio * 100))

    # Feature complexity (0-30 points)
    if metrics.named_range_count > NAMED_RANGE_THRESHOLD:
        score += 10
    elif metrics.named_range_count > 0:
        score += 5

    if metrics.has_vba:
        score += 10

    if metrics.has_external_links:
        score += 10

    return min(100, score)


# ==================== Hybrid Structural Mapper Class ====================


class StructuralMapper:
    """
    Hybrid mapper that maintains state during traversal but uses
    pure functions for analysis.

    CLAUDE-COMPLEX: We need state to efficiently traverse the workbook
    structure, but all analysis is done through pure functions.
    CLAUDE-IMPORTANT: This loads the entire workbook into memory -
    use openpyxl's read_only=True for large files.
    """

    def __init__(self, workbook: Workbook):
        """Initialize mapper with workbook."""
        self.workbook = workbook
        self._sheet_cache: dict[str, SheetStructure] = {}

    def analyze_sheet(self, sheet_name: str) -> SheetStructure:
        """
        Analyze a single sheet, using cache if available.

        CLAUDE-PERFORMANCE: Caching prevents re-analysis of sheets
        when referenced multiple times.
        """
        if sheet_name in self._sheet_cache:
            return self._sheet_cache[sheet_name]

        try:
            worksheet = self.workbook[sheet_name]

            # CLAUDE-KNOWLEDGE: Use pure functions for analysis to ensure
            # reproducible results and easier testing
            used_range, row_count, col_count = calculate_used_range(worksheet)
            features = analyze_sheet_features(worksheet)
            cell_count, formula_count = count_cells_and_formulas(worksheet)

            # Create immutable structure
            structure = SheetStructure(
                name=sheet_name,
                row_count=row_count,
                column_count=col_count,
                used_range=used_range,
                has_data=features["has_data"],
                has_formulas=features["has_formulas"],
                has_charts=features["has_charts"],
                has_pivot_tables=features["has_pivot_tables"],
                cell_count=cell_count,
                formula_count=formula_count,
            )

            # Cache result
            self._sheet_cache[sheet_name] = structure
        except (KeyError, AttributeError, ValueError, TypeError) as e:
            logger.warning("Failed to analyze sheet %s: %s", sheet_name, str(e))
            # Return minimal structure for failed sheets
            return SheetStructure(
                name=sheet_name,
                row_count=0,
                column_count=0,
                used_range="",
                has_data=False,
                has_formulas=False,
                has_charts=False,
                has_pivot_tables=False,
                cell_count=0,
                formula_count=0,
            )
        else:
            return structure

    def analyze_all_sheets(self) -> list[SheetStructure]:
        """Analyze all sheets in workbook."""
        structures = []

        for sheet_name in self.workbook.sheetnames:
            structure = self.analyze_sheet(sheet_name)
            structures.append(structure)

        return structures

    def build_workbook_structure(self, *, has_vba: bool = False, has_external_links: bool = False) -> WorkbookStructure:
        """
        Build complete workbook structure using analyzed sheets.
        """
        # Analyze all sheets
        sheets = self.analyze_all_sheets()

        # Calculate totals using pure functions
        total_cells = sum(sheet.cell_count for sheet in sheets)
        total_formulas = sum(sheet.formula_count for sheet in sheets)

        # Extract named ranges
        named_ranges = extract_named_ranges(self.workbook)

        # Calculate complexity
        metrics = ComplexityMetrics(
            sheet_count=len(sheets),
            total_cells=total_cells,
            total_formulas=total_formulas,
            named_range_count=len(named_ranges),
            has_vba=has_vba,
            has_external_links=has_external_links,
        )
        complexity_score = calculate_complexity_score(metrics)

        # Create immutable structure
        return WorkbookStructure(
            sheets=tuple(sheets),
            total_cells=total_cells,
            total_formulas=total_formulas,
            named_ranges=named_ranges,
            has_vba_project=has_vba,
            has_external_links=has_external_links,
            complexity_score=complexity_score,
        )


# ==================== Main Stage Function ====================


def stage_2_structural_mapping(
    file_path: Path, *, has_vba: bool = False, has_external_links: bool = False, read_only: bool = True
) -> Result:
    """
    Perform structural analysis using hybrid approach.

    Args:
        file_path: Path to Excel file
        has_vba: Whether file has VBA (from security scan)
        has_external_links: Whether file has external links (from security scan)
        read_only: Whether to open in read-only mode

    Returns:
        Ok(WorkbookStructure) if analysis succeeds
        Err(error_message) if analysis fails
    """
    try:
        # Open workbook
        # CLAUDE-PERFORMANCE: read_only mode is much faster for large files
        # and uses less memory
        # CLAUDE-IMPORTANT: We need data_only=False to detect formulas
        workbook = openpyxl.load_workbook(
            filename=str(file_path),
            read_only=read_only,
            keep_vba=False,  # Don't load VBA to avoid security issues
            data_only=False,  # Need formulas for structural analysis
            keep_links=False,  # Don't update external links
        )

        try:
            # Create mapper and analyze
            mapper = StructuralMapper(workbook)
            structure = mapper.build_workbook_structure(has_vba=has_vba, has_external_links=has_external_links)

            return Ok(structure)

        finally:
            # Always close workbook
            workbook.close()

    except (OSError, ValueError, TypeError, MemoryError) as e:
        return Err(f"Structural mapping failed: {e!s}", {"exception": str(e)})


# ==================== Utility Functions ====================


def create_structure_validator(
    max_sheets: int = 100, max_cells: int = 1_000_000, max_complexity: int = 80
) -> Callable[[Path], list[str]]:
    """
    Create a structure validator with specific limits.
    """

    def validator(file_path: Path) -> list[str]:
        """Validate file structure and return issues."""
        issues = []

        # Run structural analysis
        result = stage_2_structural_mapping(file_path)

        if isinstance(result, Err):
            issues.append(f"Structural analysis failed: {result.error}")
            return issues

        structure = result.value

        # Check limits
        if structure.sheet_count > max_sheets:
            issues.append(f"Too many sheets: {structure.sheet_count} (limit: {max_sheets})")

        if structure.total_cells > max_cells:
            issues.append(f"Too many cells: {structure.total_cells} (limit: {max_cells})")

        if structure.complexity_score > max_complexity:
            issues.append(f"File too complex: score {structure.complexity_score} (limit: {max_complexity})")

        # Check for empty workbook
        if structure.total_cells == 0:
            issues.append("Workbook contains no data")

        return issues

    return validator


def create_sheet_filter(min_cells: int = 10, *, require_formulas: bool = False) -> Callable[[SheetStructure], bool]:
    """
    Create a filter function for sheets based on criteria.

    Demonstrates functional composition.
    """

    def filter_sheet(sheet: SheetStructure) -> bool:
        """Check if sheet meets criteria."""
        if sheet.cell_count < min_cells:
            return False

        return not (require_formulas and not sheet.has_formulas)

    return filter_sheet
