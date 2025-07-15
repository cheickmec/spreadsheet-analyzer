"""
Stage 4: Content Intelligence (Functional Programming).

This module implements intelligent content analysis using pure functions
to detect patterns, assess data quality, and generate insights.
"""

import re
from collections import Counter
from collections.abc import Callable
from datetime import datetime
from pathlib import Path
from typing import Any, Final

import openpyxl

from spreadsheet_analyzer.pipeline.types import ContentAnalysis, ContentInsight, DataPattern, Err, Ok, Result

# ==================== Constants ====================

# Pattern detection thresholds
DATE_PATTERN_THRESHOLD: Final[float] = 0.5
EMAIL_PATTERN_THRESHOLD: Final[float] = 0.7
PHONE_PATTERN_THRESHOLD: Final[float] = 0.7
PHONE_MIN_LENGTH: Final[int] = 7
ID_CODE_CONSISTENCY_THRESHOLD: Final[float] = 0.8
ALPHANUMERIC_THRESHOLD: Final[float] = 0.7
CURRENCY_THRESHOLD: Final[float] = 0.3
NUMERIC_THRESHOLD: Final[float] = 0.8

# Quality scoring thresholds
LOW_COMPLETENESS_THRESHOLD: Final[float] = 0.5
CRITICAL_COMPLETENESS_THRESHOLD: Final[float] = 0.2
AVERAGE_QUALITY_THRESHOLD: Final[int] = 60
PROBLEM_SHEET_THRESHOLD: Final[int] = 50
PATTERN_DETECTION_THRESHOLD: Final[float] = 0.5
SHEET_ISSUES_THRESHOLD: Final[int] = 70

# Quality description thresholds
EXCELLENT_QUALITY_THRESHOLD: Final[int] = 90
GOOD_QUALITY_THRESHOLD: Final[int] = 70
FAIR_QUALITY_THRESHOLD: Final[int] = 50

# Excel date range
EXCEL_DATE_MAX: Final[int] = 100000

# ==================== Pattern Detection Functions ====================


def detect_date_patterns(values: list[Any]) -> DataPattern | None:
    """
    Detect date/time patterns in data.

    CLAUDE-KNOWLEDGE: Excel stores dates as numbers, making
    pattern detection complex.
    """
    date_count = 0
    date_formats: Counter[str] = Counter()

    for value in values:
        if value is None:
            continue

        # Check if it's a datetime object
        if isinstance(value, datetime):
            date_count += 1
            date_formats["datetime"] += 1
        # Check if it's a number that could be a date (Excel date serial)
        elif isinstance(value, int | float) and 1 < value < EXCEL_DATE_MAX:
            # Excel dates are typically between 1 (1900-01-01) and ~45000 (2023)
            date_count += 1
            date_formats["serial"] += 1
        # Check string patterns
        elif isinstance(value, str):
            # Common date patterns
            patterns = [
                (r"\d{4}-\d{2}-\d{2}", "ISO"),
                (r"\d{2}/\d{2}/\d{4}", "US"),
                (r"\d{2}\.\d{2}\.\d{4}", "EU"),
                (r"\d{1,2}-\w{3}-\d{4}", "Text"),
            ]
            for pattern, format_name in patterns:
                if re.match(pattern, value):
                    date_count += 1
                    date_formats[format_name] += 1
                    break

    if date_count > len(values) * DATE_PATTERN_THRESHOLD:  # More than 50% dates
        confidence = date_count / len(values)
        most_common_format = date_formats.most_common(1)[0][0] if date_formats else "mixed"

        return DataPattern(
            pattern_type="DATE_COLUMN",
            confidence=confidence,
            locations=(),  # Will be filled by caller
            description=f"Date column with {most_common_format} format",
        )

    return None


def _extract_currency_value(value: str) -> float | None:
    """Extract numeric value from currency string."""
    try:
        return float(re.sub(r"[^\d.-]", "", value))
    except ValueError:
        return None


def _extract_percentage_value(value: str) -> float | None:
    """Extract numeric value from percentage string."""
    try:
        return float(value.rstrip("%")) / 100
    except ValueError:
        return None


def _analyze_string_value(value: str) -> tuple[bool, bool, float | None]:
    """Analyze string value for currency, percentage, or numeric content.

    Returns: (is_currency, is_percentage, numeric_value)
    """
    currency_symbols = ["$", "€", "£", "¥"]

    if any(symbol in value for symbol in currency_symbols):
        return True, False, _extract_currency_value(value)
    if "%" in value:
        return False, True, _extract_percentage_value(value)
    return False, False, None


def _count_numeric_patterns(values: list[Any]) -> tuple[list[float], int, int]:
    """Count and extract numeric patterns from values.

    Returns: (numeric_values, currency_count, percentage_count)
    """
    numeric_values = []
    currency_count = 0
    percentage_count = 0

    for value in values:
        if value is None:
            continue

        if isinstance(value, int | float):
            numeric_values.append(value)
        elif isinstance(value, str):
            is_currency, is_percentage, numeric_val = _analyze_string_value(value)

            if is_currency:
                currency_count += 1
                if numeric_val is not None:
                    numeric_values.append(numeric_val)
            elif is_percentage:
                percentage_count += 1
                if numeric_val is not None:
                    numeric_values.append(numeric_val)

    return numeric_values, currency_count, percentage_count


def detect_numeric_patterns(values: list[Any]) -> DataPattern | None:
    """
    Detect numeric patterns (currency, percentage, etc).

    CLAUDE-KNOWLEDGE: We look for consistent formatting and ranges
    to identify pattern types.
    CLAUDE-GOTCHA: Excel formatting can hide actual values -
    $1.23 might be stored as 1.23 with currency formatting.
    """
    numeric_values, currency_count, percentage_count = _count_numeric_patterns(values)

    if not numeric_values:
        return None

    # Determine pattern type
    total_values = len(values)
    threshold = total_values * CURRENCY_THRESHOLD

    if currency_count > threshold:
        return DataPattern(
            pattern_type="CURRENCY", confidence=0.9, locations=(), description="Currency values detected"
        )
    if percentage_count > threshold:
        return DataPattern(
            pattern_type="PERCENTAGE", confidence=0.9, locations=(), description="Percentage values detected"
        )
    if len(numeric_values) > len(values) * NUMERIC_THRESHOLD:
        # Check if integers or decimals
        all_integers = all(isinstance(v, int) or v.is_integer() for v in numeric_values)
        if all_integers:
            return DataPattern(
                pattern_type="INTEGER_SEQUENCE", confidence=0.8, locations=(), description="Integer numeric column"
            )
        return DataPattern(
            pattern_type="DECIMAL_VALUES", confidence=0.8, locations=(), description="Decimal numeric column"
        )

    return None


def detect_text_patterns(values: list[Any]) -> DataPattern | None:
    """
    Detect patterns in text data (emails, codes, etc).
    """
    text_values = [str(v) for v in values if v is not None]
    if not text_values:
        return None

    # Email pattern
    email_count = sum(1 for v in text_values if re.match(r"^[\w\.-]+@[\w\.-]+\.\w+$", v))
    if email_count > len(text_values) * EMAIL_PATTERN_THRESHOLD:
        return DataPattern(
            pattern_type="EMAIL_ADDRESSES",
            confidence=email_count / len(text_values),
            locations=(),
            description="Email address column",
        )

    # Phone pattern
    phone_count = sum(1 for v in text_values if re.match(r"^[\d\s\-\(\)]+$", v) and len(v) >= PHONE_MIN_LENGTH)
    if phone_count > len(text_values) * PHONE_PATTERN_THRESHOLD:
        return DataPattern(
            pattern_type="PHONE_NUMBERS",
            confidence=phone_count / len(text_values),
            locations=(),
            description="Phone number column",
        )

    # ID/Code pattern (alphanumeric with consistent length)
    lengths = Counter(len(v) for v in text_values)
    most_common_length = lengths.most_common(1)[0][0] if lengths else 0
    consistent_length = sum(1 for v in text_values if len(v) == most_common_length)

    if consistent_length > len(text_values) * ID_CODE_CONSISTENCY_THRESHOLD:
        alphanumeric = sum(1 for v in text_values if v.isalnum())
        if alphanumeric > len(text_values) * ALPHANUMERIC_THRESHOLD:
            return DataPattern(
                pattern_type="ID_CODES",
                confidence=0.8,
                locations=(),
                description=f"ID/Code column with length {most_common_length}",
            )

    return None


# ==================== Data Quality Assessment Functions ====================


def assess_completeness(values: list[Any]) -> float:
    """
    Assess data completeness (non-null ratio).
    """
    if not values:
        return 0.0

    non_null = sum(1 for v in values if v is not None and str(v).strip())
    return non_null / len(values)


def assess_uniqueness(values: list[Any]) -> float:
    """
    Assess uniqueness ratio of values.
    """
    non_null_values = [v for v in values if v is not None]
    if not non_null_values:
        return 0.0

    unique_values = set(non_null_values)
    return len(unique_values) / len(non_null_values)


def assess_consistency(values: list[Any]) -> float:
    """
    Assess data type consistency.
    """
    if not values:
        return 0.0

    type_counts = Counter(type(v).__name__ for v in values if v is not None)
    if not type_counts:
        return 0.0

    most_common_type_count = type_counts.most_common(1)[0][1]
    total_non_null = sum(type_counts.values())

    return most_common_type_count / total_non_null


def calculate_data_quality_score(completeness: float, consistency: float, validity: float = 1.0) -> int:
    """
    Calculate overall data quality score (0-100).
    """
    # Weighted average
    score = (
        completeness * 40  # 40% weight on completeness
        + consistency * 40  # 40% weight on consistency
        + validity * 20  # 20% weight on validity
    )

    return int(min(100, max(0, score)))


# ==================== Insight Generation Functions ====================


def generate_completeness_insights(sheet_completeness: dict[str, dict[str, float]]) -> list[ContentInsight]:
    """
    Generate insights about data completeness.
    """
    insights = []

    for sheet_name, column_scores in sheet_completeness.items():
        # Find columns with low completeness
        incomplete_columns = [
            (col, score) for col, score in column_scores.items() if score < LOW_COMPLETENESS_THRESHOLD
        ]

        if incomplete_columns:
            severity = (
                "HIGH" if any(score < CRITICAL_COMPLETENESS_THRESHOLD for _, score in incomplete_columns) else "MEDIUM"
            )
            insights.append(
                ContentInsight(
                    insight_type="DATA_COMPLETENESS",
                    title=f"Incomplete data in {sheet_name}",
                    description=f"Found {len(incomplete_columns)} columns with less than 50% data completeness",
                    severity=severity,
                    affected_areas=tuple(f"{sheet_name}!{col}" for col, _ in incomplete_columns),
                    recommendation="Review and fill missing data or document why values are missing",
                )
            )

    return insights


def generate_pattern_insights(patterns: list[DataPattern]) -> list[ContentInsight]:
    """
    Generate insights from detected patterns.
    """
    insights = []

    # Group patterns by type
    pattern_groups: dict[str, list[DataPattern]] = {}
    for pattern in patterns:
        pattern_groups.setdefault(pattern.pattern_type, []).append(pattern)

    # Generate insights for each pattern type
    for pattern_type, group in pattern_groups.items():
        if pattern_type == "EMAIL_ADDRESSES":
            insights.append(
                ContentInsight(
                    insight_type="PII_DETECTED",
                    title="Personal information detected",
                    description=f"Found {len(group)} columns containing email addresses",
                    severity="HIGH",
                    affected_areas=tuple(p.locations[0] for p in group if p.locations),
                    recommendation="Ensure proper data protection measures are in place",
                )
            )
        elif pattern_type == "CURRENCY":
            insights.append(
                ContentInsight(
                    insight_type="FINANCIAL_DATA",
                    title="Financial data detected",
                    description=f"Found {len(group)} columns with currency values",
                    severity="MEDIUM",
                    affected_areas=tuple(p.locations[0] for p in group if p.locations),
                    recommendation="Verify currency formatting consistency",
                )
            )

    return insights


def generate_quality_insights(quality_scores: dict[str, int]) -> list[ContentInsight]:
    """
    Generate insights about overall data quality.
    """
    insights = []

    # Calculate average quality
    if quality_scores:
        avg_quality = sum(quality_scores.values()) / len(quality_scores)

        if avg_quality < AVERAGE_QUALITY_THRESHOLD:
            insights.append(
                ContentInsight(
                    insight_type="DATA_QUALITY",
                    title="Low overall data quality",
                    description=f"Average data quality score is {int(avg_quality)}%",
                    severity="HIGH",
                    affected_areas=tuple(
                        sheet for sheet, score in quality_scores.items() if score < AVERAGE_QUALITY_THRESHOLD
                    ),
                    recommendation="Review data entry processes and validation rules",
                )
            )

        # Find specific problem sheets
        problem_sheets = [(sheet, score) for sheet, score in quality_scores.items() if score < PROBLEM_SHEET_THRESHOLD]

        for sheet, score in problem_sheets:
            insights.append(
                ContentInsight(
                    insight_type="DATA_QUALITY",
                    title=f"Poor data quality in {sheet}",
                    description=f"Data quality score is only {score}%",
                    severity="HIGH",
                    affected_areas=(sheet,),
                    recommendation="Focus on improving data consistency and completeness",
                )
            )

    return insights


# ==================== Column Analysis Functions ====================


def analyze_column(values: list[Any]) -> dict[str, Any]:
    """
    Comprehensive analysis of a single column.
    """
    # Basic metrics
    completeness = assess_completeness(values)
    uniqueness = assess_uniqueness(values)
    consistency = assess_consistency(values)

    # Pattern detection
    pattern = None
    if completeness > PATTERN_DETECTION_THRESHOLD:  # Only detect patterns if enough data
        pattern = detect_date_patterns(values) or detect_numeric_patterns(values) or detect_text_patterns(values)

    return {
        "completeness": completeness,
        "uniqueness": uniqueness,
        "consistency": consistency,
        "pattern": pattern,
        "row_count": len(values),
        "non_null_count": sum(1 for v in values if v is not None),
    }


# ==================== Main Content Analysis Function ====================


def analyze_sheet_content(sheet, sheet_name: str, sample_size: int = 1000) -> tuple[dict[str, Any], list[DataPattern]]:
    """
    Analyze content of a single sheet.

    Args:
        sheet: The worksheet to analyze
        sheet_name: Name of the sheet
        sample_size: Maximum rows to sample per sheet
    """
    column_analyses = {}
    patterns = []

    # Analyze each column
    for column in sheet.iter_cols(max_row=sample_size):  # Sample first sample_size rows
        if not column or not column[0].value:
            continue

        # Get column header
        header = str(column[0].value)
        column_letter = column[0].column_letter

        # Get column values (excluding header)
        values = [cell.value for cell in column[1:]]

        # Analyze column
        analysis = analyze_column(values)
        column_analyses[header] = analysis

        # Add pattern location if found
        if analysis["pattern"]:
            pattern = analysis["pattern"]
            patterns.append(
                DataPattern(
                    pattern_type=pattern.pattern_type,
                    confidence=pattern.confidence,
                    locations=(f"{sheet_name}!{column_letter}",),
                    description=pattern.description,
                )
            )

    return column_analyses, patterns


# ==================== Main Stage Function ====================


def stage_4_content_intelligence(file_path: Path, sample_size: int = 1000) -> Result:
    """
    Perform intelligent content analysis.

    Args:
        file_path: Path to Excel file
        sample_size: Maximum rows to sample per sheet

    Returns:
        Ok(ContentAnalysis) if analysis succeeds
        Err(error_message) if analysis fails
    """
    try:
        # Open workbook
        workbook = openpyxl.load_workbook(filename=str(file_path), read_only=True, data_only=True, keep_links=False)

        try:
            all_patterns = []
            all_insights = []
            sheet_quality_scores = {}
            sheet_completeness = {}
            key_metrics = {
                "total_sheets_analyzed": 0,
                "total_patterns_found": 0,
                "average_completeness": 0,
                "sheets_with_issues": 0,
            }

            # Analyze each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Skip empty sheets
                if sheet.max_row == 0 or sheet.max_column == 0:
                    continue

                key_metrics["total_sheets_analyzed"] += 1

                # Analyze content
                column_analyses, patterns = analyze_sheet_content(sheet, sheet_name, sample_size)
                all_patterns.extend(patterns)

                # Calculate sheet quality
                if column_analyses:
                    completeness_scores = {col: analysis["completeness"] for col, analysis in column_analyses.items()}
                    sheet_completeness[sheet_name] = completeness_scores

                    avg_completeness = sum(completeness_scores.values()) / len(completeness_scores)
                    avg_consistency = sum(analysis["consistency"] for analysis in column_analyses.values()) / len(
                        column_analyses
                    )

                    quality_score = calculate_data_quality_score(avg_completeness, avg_consistency)
                    sheet_quality_scores[sheet_name] = quality_score

                    if quality_score < SHEET_ISSUES_THRESHOLD:
                        key_metrics["sheets_with_issues"] += 1

            # Generate insights
            all_insights.extend(generate_completeness_insights(sheet_completeness))
            all_insights.extend(generate_pattern_insights(all_patterns))
            all_insights.extend(generate_quality_insights(sheet_quality_scores))

            # Calculate overall metrics
            key_metrics["total_patterns_found"] = len(all_patterns)
            if sheet_quality_scores:
                key_metrics["average_data_quality"] = int(
                    sum(sheet_quality_scores.values()) / len(sheet_quality_scores)
                )

            # Generate summary
            summary = generate_summary(
                sheet_count=len(sheet_quality_scores),
                pattern_count=len(all_patterns),
                insight_count=len(all_insights),
                avg_quality=key_metrics.get("average_data_quality", 0),
            )

            # Create analysis result
            return Ok(
                ContentAnalysis(
                    data_patterns=tuple(all_patterns),
                    insights=tuple(all_insights),
                    data_quality_score=key_metrics.get("average_data_quality", 0),
                    summary=summary,
                    key_metrics=key_metrics,
                )
            )

        finally:
            workbook.close()

    except (OSError, ValueError, TypeError, MemoryError) as e:
        return Err(f"Content analysis failed: {e!s}", {"exception": str(e)})


# ==================== Utility Functions ====================


def generate_summary(sheet_count: int, pattern_count: int, insight_count: int, avg_quality: int) -> str:
    """
    Generate human-readable summary of analysis.
    """
    quality_desc = (
        "excellent"
        if avg_quality >= EXCELLENT_QUALITY_THRESHOLD
        else "good"
        if avg_quality >= GOOD_QUALITY_THRESHOLD
        else "fair"
        if avg_quality >= FAIR_QUALITY_THRESHOLD
        else "poor"
    )

    summary_parts = [
        f"Analyzed {sheet_count} sheets",
        f"found {pattern_count} data patterns",
        f"generated {insight_count} insights.",
        f"Overall data quality is {quality_desc} ({avg_quality}%).",
    ]

    return " ".join(summary_parts)


def create_content_validator(
    min_quality_score: int = 60, required_patterns: set[str] | None = None, forbidden_patterns: set[str] | None = None
) -> Callable[[Path], list[str]]:
    """
    Create a content validator with specific requirements.
    """

    def validator(file_path: Path) -> list[str]:
        """Validate content and return issues."""
        issues = []

        # Run content analysis
        result = stage_4_content_intelligence(file_path)

        if isinstance(result, Err):
            issues.append(f"Content analysis failed: {result.error}")
            return issues

        analysis = result.value

        # Check quality score
        if analysis.data_quality_score < min_quality_score:
            issues.append(f"Data quality too low: {analysis.data_quality_score}% (minimum: {min_quality_score}%)")

        # Check required patterns
        if required_patterns:
            found_types = {p.pattern_type for p in analysis.data_patterns}
            missing = required_patterns - found_types
            if missing:
                issues.append(f"Missing required patterns: {', '.join(missing)}")

        # Check forbidden patterns
        if forbidden_patterns:
            found_types = {p.pattern_type for p in analysis.data_patterns}
            forbidden_found = forbidden_patterns & found_types
            if forbidden_found:
                issues.append(f"Forbidden patterns found: {', '.join(forbidden_found)}")

        # Check critical insights
        critical_insights = [i for i in analysis.insights if i.severity == "HIGH"]
        issues.extend(f"Critical issue: {insight.title}" for insight in critical_insights)

        return issues

    return validator
