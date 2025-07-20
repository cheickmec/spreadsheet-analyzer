"""
Excel-aware DataFrame wrapper for maintaining coordinate mapping.

This module provides a DataFrame wrapper that preserves Excel coordinate
information when loading data with offsets, enabling agents to correctly
map between Pandas indices and Excel references.
"""

import re
from typing import Any

import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter

from spreadsheet_analyzer.graph_db.query_interface import EnhancedQueryInterface, GraphQueryInterface
from spreadsheet_analyzer.graph_db.range_membership import RangeMembershipIndex


class ExcelAwareDataFrame:
    """
    DataFrame wrapper that maintains Excel coordinate mapping.

    CLAUDE-KNOWLEDGE: When loading Excel data into Pandas with skiprows
    or usecols, the mapping between Pandas indices and Excel coordinates
    is lost. This wrapper preserves that mapping.
    """

    def __init__(
        self,
        df: pd.DataFrame,
        metadata: dict[str, Any],
        graph_interface: GraphQueryInterface | None = None,
        enhanced_interface: EnhancedQueryInterface | None = None,
        range_index: RangeMembershipIndex | None = None,
    ):
        """
        Initialize with DataFrame and Excel metadata.

        Args:
            df: Pandas DataFrame with the data
            metadata: Dictionary containing:
                - sheet: Sheet name
                - start_row: First Excel row (1-based)
                - start_col: First Excel column letter
                - skiprows: Number of rows skipped
                - usecols: Columns specification used
            graph_interface: Optional graph query interface
            enhanced_interface: Optional enhanced query interface with range support
            range_index: Optional range membership index for intersection queries
        """
        self.df = df
        self.sheet_name = metadata["sheet"]
        self.start_row = metadata["start_row"]  # Excel row (1-based)
        self.start_col = metadata["start_col"]  # Excel column letter
        self.skiprows = metadata.get("skiprows", 0)
        self.usecols = metadata.get("usecols")
        self.graph_interface = graph_interface
        self.enhanced_interface = enhanced_interface
        self.range_index = range_index

        # Build column mapping
        self._build_column_mapping()

    def _build_column_mapping(self):
        """Build mapping from DataFrame column index to Excel column."""
        self.col_mapping = {}

        if self.usecols is None:
            # All columns from start_col
            start_idx = column_index_from_string(self.start_col) - 1
            for i, _col in enumerate(self.df.columns):
                excel_col = get_column_letter(start_idx + i + 1)
                self.col_mapping[i] = excel_col
        elif isinstance(self.usecols, str):
            # Column range like "B:F"
            if ":" in self.usecols:
                start, end = self.usecols.split(":")
                start_idx = column_index_from_string(start) - 1
                for i in range(len(self.df.columns)):
                    excel_col = get_column_letter(start_idx + i + 1)
                    self.col_mapping[i] = excel_col
            else:
                # Single column
                self.col_mapping[0] = self.usecols
        elif isinstance(self.usecols, list):
            # List of column names or indices
            for i, col_spec in enumerate(self.usecols):
                if isinstance(col_spec, str) and re.match(r"^[A-Z]+$", col_spec):
                    # Excel column letter
                    self.col_mapping[i] = col_spec
                elif isinstance(col_spec, int):
                    # Column index
                    self.col_mapping[i] = get_column_letter(col_spec + 1)
                else:
                    # Column name - try to infer
                    self.col_mapping[i] = get_column_letter(i + column_index_from_string(self.start_col))

    def to_excel_ref(self, row: int, col: int | str) -> str:
        """
        Convert Pandas indices to Excel reference.

        Args:
            row: Pandas row index (0-based)
            col: Pandas column index (0-based) or column name

        Returns:
            Full Excel reference (e.g., "Sheet1!B5")
        """
        # Handle column specification
        if isinstance(col, str):
            # Column name - find its position
            try:
                col_idx = self.df.columns.get_loc(col)
            except KeyError as e:
                raise ValueError(f"Column '{col}' not found in DataFrame") from e
        else:
            col_idx = col

        # Calculate Excel row (accounting for skiprows and header)
        excel_row = self.start_row + row

        # Get Excel column
        excel_col = self.col_mapping.get(col_idx)
        if not excel_col:
            # Fallback calculation
            excel_col = get_column_letter(column_index_from_string(self.start_col) + col_idx)

        return f"{self.sheet_name}!{excel_col}{excel_row}"

    def from_excel_ref(self, ref: str) -> tuple[int, int]:
        """
        Convert Excel reference to Pandas indices.

        Args:
            ref: Excel reference (e.g., "B5" or "Sheet1!B5")

        Returns:
            Tuple of (row_index, column_index) in Pandas
        """
        # Parse reference
        if "!" in ref:
            sheet, cell_ref = ref.split("!", 1)
            if sheet != self.sheet_name:
                raise ValueError(f"Reference is from different sheet: {sheet}")
        else:
            cell_ref = ref

        # Extract column and row
        match = re.match(r"^([A-Z]+)(\d+)$", cell_ref)
        if not match:
            raise ValueError(f"Invalid cell reference: {cell_ref}")

        excel_col, excel_row = match.groups()
        excel_row = int(excel_row)

        # Calculate Pandas row index
        row_idx = excel_row - self.start_row
        if row_idx < 0 or row_idx >= len(self.df):
            raise ValueError(f"Row {excel_row} is outside DataFrame range")

        # Calculate Pandas column index
        col_idx = None
        for idx, mapped_col in self.col_mapping.items():
            if mapped_col == excel_col:
                col_idx = idx
                break

        if col_idx is None:
            raise ValueError(f"Column {excel_col} not found in mapping")

        return (row_idx, col_idx)

    def query_dependencies(self, row: int, col: int | str) -> list[dict[str, Any]]:
        """
        Query graph database for dependencies of a cell.

        Args:
            row: Pandas row index
            col: Pandas column index or name

        Returns:
            List of dependencies from graph database
        """
        if not self.graph_interface:
            raise ValueError("No graph interface configured")

        excel_ref = self.to_excel_ref(row, col)
        return self.graph_interface.get_dependencies(excel_ref)

    def query_dependents(self, row: int, col: int | str) -> list[dict[str, Any]]:
        """
        Query graph database for cells that depend on this cell.

        Args:
            row: Pandas row index
            col: Pandas column index or name

        Returns:
            List of dependent cells from graph database
        """
        if not self.graph_interface:
            raise ValueError("No graph interface configured")

        excel_ref = self.to_excel_ref(row, col)
        return self.graph_interface.get_dependents(excel_ref)

    def get_cell_info(self, row: int, col: int | str) -> dict[str, Any]:
        """
        Get comprehensive information about a cell.

        Args:
            row: Pandas row index
            col: Pandas column index or name

        Returns:
            Dictionary with cell value, Excel ref, and graph info
        """
        excel_ref = self.to_excel_ref(row, col)

        info = {
            "excel_ref": excel_ref,
            "pandas_loc": (row, col),
            "value": self.df.iloc[row, col if isinstance(col, int) else self.df.columns.get_loc(col)],
        }

        if self.graph_interface:
            # Add dependency information
            deps = self.query_dependencies(row, col)
            info["dependencies"] = deps
            info["dependency_count"] = len(deps)

        return info

    def is_cell_in_formula_range(self, row: int, col: int | str) -> bool:
        """
        Check if a cell is part of any formula range.

        Args:
            row: Pandas row index
            col: Pandas column index or name

        Returns:
            True if cell is part of a formula range, False otherwise
        """
        if not self.range_index:
            return False

        excel_ref = self.to_excel_ref(row, col)
        # Extract just the cell reference part
        _, cell_ref = excel_ref.split("!", 1)

        return self.range_index.is_cell_in_any_range(self.sheet_name, cell_ref)

    def get_formula_ranges_containing_cell(self, row: int, col: int | str) -> list[str]:
        """
        Get all formula ranges that contain a specific cell.

        Args:
            row: Pandas row index
            col: Pandas column index or name

        Returns:
            List of formula keys that reference ranges containing this cell
        """
        if not self.range_index:
            return []

        excel_ref = self.to_excel_ref(row, col)
        _, cell_ref = excel_ref.split("!", 1)

        return self.range_index.get_ranges_containing_cell(self.sheet_name, cell_ref)

    def find_range_intersections(
        self, start_row: int, end_row: int, start_col: int | str, end_col: int | str
    ) -> list[tuple[str, str]]:
        """
        Find all formula ranges that intersect with a DataFrame range.

        Args:
            start_row: Starting Pandas row index
            end_row: Ending Pandas row index (inclusive)
            start_col: Starting column (index or name)
            end_col: Ending column (index or name)

        Returns:
            List of tuples (formula_key, range_ref) that intersect
        """
        if not self.range_index:
            return []

        # Convert to Excel references
        start_ref = self.to_excel_ref(start_row, start_col)
        end_ref = self.to_excel_ref(end_row, end_col)

        # Extract cell parts
        _, start_cell = start_ref.split("!", 1)
        _, end_cell = end_ref.split("!", 1)

        return self.range_index.get_ranges_intersecting_range(self.sheet_name, start_cell, end_cell)

    def highlight_formula_ranges(self) -> pd.DataFrame:
        """
        Create a DataFrame mask showing which cells are part of formula ranges.

        Returns:
            Boolean DataFrame with True for cells in formula ranges
        """
        if not self.range_index:
            return pd.DataFrame(False, index=self.df.index, columns=self.df.columns)

        mask = pd.DataFrame(False, index=self.df.index, columns=self.df.columns)

        for row_idx in range(len(self.df)):
            for col_idx in range(len(self.df.columns)):
                if self.is_cell_in_formula_range(row_idx, col_idx):
                    mask.iloc[row_idx, col_idx] = True

        return mask

    def get_range_dependencies_for_selection(
        self, rows: slice | list[int] | None = None, cols: slice | list[str | int] | None = None
    ) -> dict[str, Any]:
        """
        Get dependency information for a selection of cells including range memberships.

        Args:
            rows: Row selection (slice, list, or None for all)
            cols: Column selection (slice, list, or None for all)

        Returns:
            Dictionary with dependency statistics for the selection
        """
        if not self.enhanced_interface:
            return {"error": "No enhanced interface configured"}

        # Default to all rows/cols if not specified
        if rows is None:
            rows = slice(None)
        if cols is None:
            cols = slice(None)

        # Get actual row/col indices
        row_indices = self.df.index[rows] if isinstance(rows, slice) else rows
        col_indices = self.df.columns[cols] if isinstance(cols, slice) else cols

        stats = {
            "cells_with_formulas": 0,
            "cells_in_ranges": 0,
            "unique_ranges": set(),
            "total_dependencies": 0,
            "total_dependents": 0,
        }

        for row in row_indices:
            for col in col_indices:
                excel_ref = self.to_excel_ref(row, col)
                _, cell_ref = excel_ref.split("!", 1)

                # Query dependencies
                result = self.enhanced_interface.get_cell_dependencies(self.sheet_name, cell_ref, include_ranges=True)

                if result.has_formula:
                    stats["cells_with_formulas"] += 1

                if result.is_in_ranges:
                    stats["cells_in_ranges"] += 1
                    stats["unique_ranges"].update(result.is_in_ranges)

                stats["total_dependencies"] += result.total_dependencies
                stats["total_dependents"] += result.total_dependents

        # Convert set to list for JSON serialization
        stats["unique_ranges"] = list(stats["unique_ranges"])
        stats["selection_size"] = len(row_indices) * len(col_indices)

        return stats

    # Delegate DataFrame methods
    def __getattr__(self, name):
        """Delegate attribute access to underlying DataFrame."""
        return getattr(self.df, name)

    def __getitem__(self, key):
        """Delegate item access to underlying DataFrame."""
        return self.df[key]

    def __repr__(self):
        """String representation."""
        return f"ExcelAwareDataFrame(sheet='{self.sheet_name}', shape={self.df.shape}, start={self.start_col}{self.start_row})"


def load_excel_with_mapping(
    file_path: str,
    sheet_name: str | int = 0,
    skiprows: int | None = None,
    usecols: str | list | None = None,
    graph_interface: GraphQueryInterface | None = None,
    enhanced_interface: EnhancedQueryInterface | None = None,
    range_index: RangeMembershipIndex | None = None,
    **kwargs,
) -> ExcelAwareDataFrame:
    """
    Load Excel file with coordinate mapping preserved.

    This function loads an Excel file into a DataFrame while maintaining
    the mapping between Pandas indices and Excel coordinates.

    Args:
        file_path: Path to Excel file
        sheet_name: Sheet to load (default first sheet)
        skiprows: Number of rows to skip
        usecols: Columns to load (e.g., "B:F" or ["B", "C", "D"])
        graph_interface: Optional graph query interface
        enhanced_interface: Optional enhanced query interface with range support
        range_index: Optional range membership index for intersection queries
        **kwargs: Additional arguments passed to pd.read_excel

    Returns:
        ExcelAwareDataFrame with coordinate mapping and range tracking

    Example:
        >>> df = load_excel_with_mapping(
        ...     "data.xlsx",
        ...     sheet_name="Sales",
        ...     skiprows=4,
        ...     usecols="B:F",
        ...     enhanced_interface=query_interface,
        ...     range_index=formula_analysis.range_membership_index
        ... )
        >>> df.to_excel_ref(0, 0)  # Returns "Sales!B5"
        >>> df.is_cell_in_formula_range(10, 2)  # Check if cell is in any range
    """
    # Load the DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=skiprows, usecols=usecols, **kwargs)

    # Determine start position
    start_row = (skiprows or 0) + 1  # +1 for 1-based Excel rows
    if kwargs.get("header") is None or kwargs.get("header") == 0:
        start_row += 1  # Account for header row

    # Determine start column
    if usecols is None:
        start_col = "A"
    elif isinstance(usecols, str):
        # Extract start column from range or single column
        start_col = usecols.split(":")[0] if ":" in usecols else usecols
    elif isinstance(usecols, list) and usecols:
        # Use first column in list
        first_col = usecols[0]
        if isinstance(first_col, str) and re.match(r"^[A-Z]+$", first_col):
            start_col = first_col
        elif isinstance(first_col, int):
            start_col = get_column_letter(first_col + 1)
        else:
            start_col = "A"  # Fallback
    else:
        start_col = "A"

    # Get actual sheet name if integer was provided
    if isinstance(sheet_name, int):
        xl_file = pd.ExcelFile(file_path)
        sheet_name = xl_file.sheet_names[sheet_name]
        xl_file.close()

    # Create metadata
    metadata = {
        "sheet": sheet_name,
        "start_row": start_row,
        "start_col": start_col,
        "skiprows": skiprows,
        "usecols": usecols,
    }

    return ExcelAwareDataFrame(df, metadata, graph_interface, enhanced_interface, range_index)
