"""Excel-specific tool implementations.

This module provides functional tools for working with Excel files
including reading cells, analyzing formulas, and extracting data.

CLAUDE-KNOWLEDGE: These tools wrap openpyxl functionality in a
functional interface for safe, predictable Excel operations.
"""

from pathlib import Path
from typing import Any

import openpyxl
from pydantic import BaseModel, Field

from ...core.errors import ToolError
from ...core.types import Result, err, ok
from ..types import FunctionalTool, create_tool


# Input schemas for Excel tools
class CellReaderInput(BaseModel):
    """Input for reading a single cell."""

    file_path: str = Field(description="Path to Excel file")
    sheet_name: str = Field(description="Name of the sheet")
    cell_reference: str = Field(description="Cell reference (e.g., 'A1')")


class RangeReaderInput(BaseModel):
    """Input for reading a range of cells."""

    file_path: str = Field(description="Path to Excel file")
    sheet_name: str = Field(description="Name of the sheet")
    range_reference: str = Field(description="Range reference (e.g., 'A1:B10')")
    include_formulas: bool = Field(default=False, description="Include formula strings")


class SheetReaderInput(BaseModel):
    """Input for reading an entire sheet."""

    file_path: str = Field(description="Path to Excel file")
    sheet_name: str = Field(description="Name of the sheet")
    max_rows: int | None = Field(default=None, description="Maximum rows to read")
    max_cols: int | None = Field(default=None, description="Maximum columns to read")


class WorkbookReaderInput(BaseModel):
    """Input for reading workbook metadata."""

    file_path: str = Field(description="Path to Excel file")


class FormulaAnalyzerInput(BaseModel):
    """Input for analyzing formulas."""

    file_path: str = Field(description="Path to Excel file")
    sheet_name: str = Field(description="Name of the sheet")
    include_dependencies: bool = Field(default=True, description="Include formula dependencies")


# Tool implementations
def create_cell_reader_tool() -> FunctionalTool:
    """Create a tool for reading individual cells."""

    def execute(args: CellReaderInput) -> Result[dict[str, Any], ToolError]:
        """Read a single cell from Excel."""
        try:
            workbook = openpyxl.load_workbook(args.file_path, read_only=True, data_only=True)

            if args.sheet_name not in workbook.sheetnames:
                return err(
                    ToolError(
                        f"Sheet '{args.sheet_name}' not found",
                        tool_name="cell_reader",
                        details={"available_sheets": workbook.sheetnames},
                    )
                )

            sheet = workbook[args.sheet_name]
            cell = sheet[args.cell_reference]

            # Load with formulas if needed
            formula = None
            if hasattr(cell, "value") and isinstance(cell.value, str) and cell.value.startswith("="):
                # Reload to get formula
                wb_formula = openpyxl.load_workbook(args.file_path, read_only=True, data_only=False)
                formula = wb_formula[args.sheet_name][args.cell_reference].value
                wb_formula.close()

            result = {
                "reference": args.cell_reference,
                "value": cell.value,
                "data_type": type(cell.value).__name__,
                "formula": formula,
                "row": cell.row,
                "column": cell.column,
                "coordinate": cell.coordinate,
            }

            workbook.close()
            return ok(result)

        except Exception as e:
            return err(ToolError(f"Failed to read cell: {e}", tool_name="cell_reader", cause=e))

    return create_tool(
        name="read_cell",
        description="Read a single cell from an Excel file",
        args_schema=CellReaderInput,
        execute_fn=execute,
        category="excel",
        return_type=dict,
        tags=["excel", "read", "cell"],
    )


def create_range_reader_tool() -> FunctionalTool:
    """Create a tool for reading cell ranges."""

    def execute(args: RangeReaderInput) -> Result[list[list[Any]], ToolError]:
        """Read a range of cells from Excel."""
        try:
            workbook = openpyxl.load_workbook(args.file_path, read_only=True, data_only=not args.include_formulas)

            if args.sheet_name not in workbook.sheetnames:
                return err(
                    ToolError(
                        f"Sheet '{args.sheet_name}' not found",
                        tool_name="range_reader",
                        details={"available_sheets": workbook.sheetnames},
                    )
                )

            sheet = workbook[args.sheet_name]

            # Parse the range
            cells = sheet[args.range_reference]

            # Convert to list of lists
            result = []
            for row in cells:
                row_data = []
                for cell in row:
                    if (
                        args.include_formulas
                        and hasattr(cell, "value")
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        row_data.append({"value": cell.value, "formula": cell.value, "coordinate": cell.coordinate})
                    else:
                        row_data.append(cell.value)
                result.append(row_data)

            workbook.close()
            return ok(result)

        except Exception as e:
            return err(ToolError(f"Failed to read range: {e}", tool_name="range_reader", cause=e))

    return create_tool(
        name="read_range",
        description="Read a range of cells from an Excel file",
        args_schema=RangeReaderInput,
        execute_fn=execute,
        category="excel",
        return_type=list,
        tags=["excel", "read", "range"],
    )


def create_sheet_reader_tool() -> FunctionalTool:
    """Create a tool for reading entire sheets."""

    def execute(args: SheetReaderInput) -> Result[dict[str, Any], ToolError]:
        """Read an entire sheet from Excel."""
        try:
            workbook = openpyxl.load_workbook(args.file_path, read_only=True, data_only=True)

            if args.sheet_name not in workbook.sheetnames:
                return err(
                    ToolError(
                        f"Sheet '{args.sheet_name}' not found",
                        tool_name="sheet_reader",
                        details={"available_sheets": workbook.sheetnames},
                    )
                )

            sheet = workbook[args.sheet_name]

            # Determine bounds
            max_row = min(sheet.max_row, args.max_rows) if args.max_rows else sheet.max_row
            max_col = min(sheet.max_column, args.max_cols) if args.max_cols else sheet.max_column

            # Read all data
            data = []
            for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                row_data = [cell.value for cell in row]
                data.append(row_data)

            result = {
                "sheet_name": args.sheet_name,
                "data": data,
                "dimensions": {"rows": max_row, "columns": max_col, "total_cells": max_row * max_col},
                "has_more_data": {
                    "rows": sheet.max_row > max_row if args.max_rows else False,
                    "columns": sheet.max_column > max_col if args.max_cols else False,
                },
            }

            workbook.close()
            return ok(result)

        except Exception as e:
            return err(ToolError(f"Failed to read sheet: {e}", tool_name="sheet_reader", cause=e))

    return create_tool(
        name="read_sheet",
        description="Read an entire sheet from an Excel file",
        args_schema=SheetReaderInput,
        execute_fn=execute,
        category="excel",
        return_type=dict,
        tags=["excel", "read", "sheet"],
    )


def create_workbook_reader_tool() -> FunctionalTool:
    """Create a tool for reading workbook metadata."""

    def execute(args: WorkbookReaderInput) -> Result[dict[str, Any], ToolError]:
        """Read workbook metadata."""
        try:
            workbook = openpyxl.load_workbook(args.file_path, read_only=True)

            # Collect metadata
            sheets_info = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheets_info.append(
                    {
                        "name": sheet_name,
                        "max_row": sheet.max_row,
                        "max_column": sheet.max_column,
                        "cell_count": sheet.max_row * sheet.max_column,
                    }
                )

            result = {
                "file_path": args.file_path,
                "file_name": Path(args.file_path).name,
                "sheet_count": len(workbook.sheetnames),
                "sheet_names": workbook.sheetnames,
                "sheets": sheets_info,
                "properties": {
                    "created": str(workbook.properties.created) if workbook.properties.created else None,
                    "modified": str(workbook.properties.modified) if workbook.properties.modified else None,
                    "creator": workbook.properties.creator,
                    "title": workbook.properties.title,
                },
            }

            workbook.close()
            return ok(result)

        except Exception as e:
            return err(ToolError(f"Failed to read workbook: {e}", tool_name="workbook_reader", cause=e))

    return create_tool(
        name="read_workbook",
        description="Read workbook metadata from an Excel file",
        args_schema=WorkbookReaderInput,
        execute_fn=execute,
        category="excel",
        return_type=dict,
        tags=["excel", "read", "workbook", "metadata"],
    )


def create_formula_analyzer_tool() -> FunctionalTool:
    """Create a tool for analyzing formulas in a sheet."""

    def execute(args: FormulaAnalyzerInput) -> Result[dict[str, Any], ToolError]:
        """Analyze formulas in an Excel sheet."""
        try:
            # Load with formulas
            workbook = openpyxl.load_workbook(args.file_path, read_only=True, data_only=False)

            if args.sheet_name not in workbook.sheetnames:
                return err(
                    ToolError(
                        f"Sheet '{args.sheet_name}' not found",
                        tool_name="formula_analyzer",
                        details={"available_sheets": workbook.sheetnames},
                    )
                )

            sheet = workbook[args.sheet_name]

            formulas = []
            formula_cells = []

            # Find all formulas
            for row in sheet.iter_rows():
                for cell in row:
                    if hasattr(cell, "value") and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_info = {
                            "cell": cell.coordinate,
                            "formula": cell.value,
                            "row": cell.row,
                            "column": cell.column,
                        }

                        # Extract references if requested
                        if args.include_dependencies:
                            import re

                            # Simple regex to find cell references
                            refs = re.findall(r"[A-Z]+\d+", cell.value)
                            formula_info["references"] = list(set(refs))

                        formulas.append(formula_info)
                        formula_cells.append(cell.coordinate)

            # Analyze patterns
            function_usage = {}
            for formula in formulas:
                # Extract functions
                import re

                functions = re.findall(r"([A-Z]+)\(", formula["formula"])
                for func in functions:
                    function_usage[func] = function_usage.get(func, 0) + 1

            result = {
                "sheet_name": args.sheet_name,
                "formula_count": len(formulas),
                "formula_cells": formula_cells,
                "formulas": formulas,
                "function_usage": function_usage,
                "statistics": {
                    "total_formulas": len(formulas),
                    "unique_functions": len(function_usage),
                    "cells_with_formulas": len(formula_cells),
                    "formula_density": len(formulas) / (sheet.max_row * sheet.max_column)
                    if sheet.max_row * sheet.max_column > 0
                    else 0,
                },
            }

            workbook.close()
            return ok(result)

        except Exception as e:
            return err(ToolError(f"Failed to analyze formulas: {e}", tool_name="formula_analyzer", cause=e))

    return create_tool(
        name="analyze_formulas",
        description="Analyze formulas in an Excel sheet",
        args_schema=FormulaAnalyzerInput,
        execute_fn=execute,
        category="excel",
        return_type=dict,
        tags=["excel", "analyze", "formula"],
    )
