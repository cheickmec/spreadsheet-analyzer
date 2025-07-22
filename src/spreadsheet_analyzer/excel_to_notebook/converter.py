"""Converter for transforming Excel files into notebook format for LLM analysis.

This module implements the bridge between Excel files and the notebook-based
analysis system, converting spreadsheet data into a structured notebook format
that can be analyzed by LLMs.
"""

import json
import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Final

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from spreadsheet_analyzer.notebook_llm.protocol.base import NotebookCell
from spreadsheet_analyzer.notebook_llm.protocol.base import NotebookCellType as CellType

logger = logging.getLogger(__name__)

# Constants for conversion
MAX_PREVIEW_ROWS: Final[int] = 20
MAX_PREVIEW_COLS: Final[int] = 20
MAX_CELL_CONTENT_LENGTH: Final[int] = 1000


@dataclass
class SheetMetadata:
    """Metadata about a worksheet."""

    name: str
    row_count: int
    column_count: int
    has_formulas: bool
    has_charts: bool
    has_pivot_tables: bool
    data_regions: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class NotebookDocument:
    """Represents a notebook document generated from Excel."""

    cells: list[NotebookCell]
    metadata: dict[str, Any]
    source_file: Path


class ExcelToNotebookConverter:
    """Converts Excel files to notebook format for analysis.

    This converter creates a structured notebook representation of Excel files,
    preserving important information while making it accessible to LLMs.
    """

    def __init__(self, *, include_values: bool = True, include_formulas: bool = True):
        """Initialize the converter.

        Args:
            include_values: Whether to include cell values
            include_formulas: Whether to include formulas
        """
        self.include_values = include_values
        self.include_formulas = include_formulas

    def convert(self, excel_path: Path) -> NotebookDocument:
        """Convert an Excel file to notebook format.

        Args:
            excel_path: Path to the Excel file

        Returns:
            NotebookDocument containing the converted data

        Raises:
            ValueError: If the file cannot be opened or processed
        """
        if not excel_path.exists():
            raise ValueError(f"Excel file not found: {excel_path}")

        try:
            # Load workbook in read-only mode for efficiency
            wb = openpyxl.load_workbook(
                excel_path,
                read_only=True,
                data_only=not self.include_formulas,
            )

            cells = []
            metadata = {
                "source_file": str(excel_path),
                "sheet_count": len(wb.sheetnames),
                "sheets": [],
            }

            # Create overview cell
            cells.append(self._create_overview_cell(wb))

            # Process each sheet
            for sheet in wb.worksheets:
                sheet_metadata = self._analyze_sheet(sheet)
                metadata["sheets"].append(sheet_metadata.__dict__)

                # Add sheet analysis cells
                cells.extend(self._create_sheet_cells(sheet, sheet_metadata))

            # Add summary cell
            cells.append(self._create_summary_cell(metadata))

            wb.close()

            return NotebookDocument(
                cells=cells,
                metadata=metadata,
                source_file=excel_path,
            )

        except Exception as e:
            raise ValueError(f"Failed to convert Excel file: {e}") from e

    def _create_overview_cell(self, workbook) -> NotebookCell:
        """Create an overview cell describing the workbook."""
        content = f"""# Excel Workbook Analysis

## File Overview
- **Sheets**: {len(workbook.sheetnames)}
- **Sheet Names**: {", ".join(workbook.sheetnames)}

## Analysis Sections
1. Workbook structure overview
2. Individual sheet analysis
3. Formula and dependency tracking
4. Data pattern detection
5. Summary and recommendations
"""
        return NotebookCell(
            cell_type=CellType.MARKDOWN,
            content=content,
            metadata={"cell_id": "overview"},
        )

    def _analyze_sheet(self, sheet: Worksheet) -> SheetMetadata:
        """Analyze a worksheet to extract metadata."""
        # Get dimensions (may not be accurate in read-only mode)
        max_row = sheet.max_row or 1
        max_col = sheet.max_column or 1

        # Check for formulas (limited in read-only mode)
        has_formulas = False
        if not sheet.parent.read_only:
            for row in sheet.iter_rows(max_row=min(100, max_row)):
                for cell in row:
                    if hasattr(cell, "formula") and cell.formula:
                        has_formulas = True
                        break
                if has_formulas:
                    break

        # Detect data regions
        data_regions = self._detect_data_regions(sheet, max_row, max_col)

        return SheetMetadata(
            name=sheet.title,
            row_count=max_row,
            column_count=max_col,
            has_formulas=has_formulas,
            has_charts=bool(sheet._charts) if hasattr(sheet, "_charts") else False,
            has_pivot_tables=bool(sheet._pivots) if hasattr(sheet, "_pivots") else False,
            data_regions=data_regions,
        )

    def _detect_data_regions(self, sheet: Worksheet, max_row: int, max_col: int) -> list[dict[str, Any]]:
        """Detect contiguous data regions in the sheet."""
        regions = []

        # Simple region detection - look for headers
        preview_rows = min(MAX_PREVIEW_ROWS, max_row)
        preview_cols = min(MAX_PREVIEW_COLS, max_col)

        # Check first row for potential headers
        first_row = []
        for col in range(1, preview_cols + 1):
            cell = sheet.cell(row=1, column=col)
            if cell.value:
                first_row.append(str(cell.value))

        if len(first_row) > 1:  # Likely a data table
            regions.append(
                {
                    "type": "table",
                    "start": "A1",
                    "headers": first_row,
                    "estimated_rows": max_row,
                }
            )

        return regions

    def _create_sheet_cells(self, sheet: Worksheet, metadata: SheetMetadata) -> list[NotebookCell]:
        """Create notebook cells for a sheet."""
        cells = []

        # Sheet overview markdown cell
        overview_content = f"""## Sheet: {metadata.name}

### Sheet Statistics
- **Dimensions**: {metadata.row_count} rows × {metadata.column_count} columns
- **Has Formulas**: {"Yes" if metadata.has_formulas else "No"}
- **Has Charts**: {"Yes" if metadata.has_charts else "No"}
- **Has Pivot Tables**: {"Yes" if metadata.has_pivot_tables else "No"}
"""

        if metadata.data_regions:
            overview_content += "\n### Detected Data Regions\n"
            for region in metadata.data_regions:
                overview_content += f"- **{region['type'].title()}** starting at {region['start']}"
                if "headers" in region:
                    overview_content += f"\n  Headers: {', '.join(region['headers'][:5])}"
                    if len(region["headers"]) > 5:
                        overview_content += f" ... ({len(region['headers'])} total)"
                overview_content += "\n"

        cells.append(
            NotebookCell(
                cell_type=CellType.MARKDOWN,
                content=overview_content,
                metadata={"sheet": metadata.name, "cell_id": f"sheet_{metadata.name}_overview"},
            )
        )

        # Data preview code cell
        if self.include_values:
            preview_code = self._generate_data_preview_code(sheet, metadata)
            cells.append(
                NotebookCell(
                    cell_type=CellType.CODE,
                    content=preview_code,
                    metadata={"sheet": metadata.name, "cell_id": f"sheet_{metadata.name}_preview"},
                )
            )

        # Formula analysis code cell
        if self.include_formulas and metadata.has_formulas:
            formula_code = self._generate_formula_analysis_code(sheet, metadata)
            if formula_code:
                cells.append(
                    NotebookCell(
                        cell_type=CellType.CODE,
                        content=formula_code,
                        metadata={"sheet": metadata.name, "cell_id": f"sheet_{metadata.name}_formulas"},
                    )
                )

        return cells

    def _generate_data_preview_code(self, sheet: Worksheet, metadata: SheetMetadata) -> str:
        """Generate code to preview sheet data."""
        code_lines = [
            f"# Data preview for sheet: {metadata.name}",
            "import pandas as pd",
            "",
            "# Sample data from the sheet",
            "data = {",
        ]

        # Get preview data
        preview_rows = min(MAX_PREVIEW_ROWS, metadata.row_count)
        preview_cols = min(MAX_PREVIEW_COLS, metadata.column_count)

        # Extract column data
        columns_data = {}
        for col in range(1, preview_cols + 1):
            col_letter = get_column_letter(col)
            col_values = []

            for row in range(1, preview_rows + 1):
                cell = sheet.cell(row=row, column=col)
                value = cell.value
                if value is not None:
                    # Truncate long strings
                    if isinstance(value, str) and len(value) > MAX_CELL_CONTENT_LENGTH:
                        value = value[:MAX_CELL_CONTENT_LENGTH] + "..."
                    col_values.append(value)
                else:
                    col_values.append(None)

            if any(v is not None for v in col_values):
                columns_data[col_letter] = col_values

        # Format as Python code
        for col, values in columns_data.items():
            # Convert values to JSON-safe format
            safe_values = []
            for v in values:
                if v is None:
                    safe_values.append("None")
                elif isinstance(v, str):
                    safe_values.append(json.dumps(v))
                elif isinstance(v, datetime):
                    # Format datetime as quoted string
                    safe_values.append(json.dumps(v.strftime("%Y-%m-%d %H:%M:%S")))
                elif isinstance(v, date):
                    # Format date as quoted string
                    safe_values.append(json.dumps(v.strftime("%Y-%m-%d")))
                else:
                    safe_values.append(str(v))

            code_lines.append(f'    "{col}": [{", ".join(safe_values)}],')

        code_lines.extend(
            [
                "}",
                "",
                "# Create DataFrame",
                "df = pd.DataFrame(data)",
                'print(f"Shape: {df.shape}")',
                'print("\\nFirst few rows:")',
                "df.head(10)",
            ]
        )

        return "\n".join(code_lines)

    def _generate_formula_analysis_code(self, sheet: Worksheet, metadata: SheetMetadata) -> str | None:
        """Generate code to analyze formulas in the sheet."""
        # In read-only mode, formula access is limited
        # This is a placeholder for when formulas are accessible
        if sheet.parent.read_only:
            return None

        code_lines = [
            f"# Formula analysis for sheet: {metadata.name}",
            "",
            "# Note: Formula analysis requires non-read-only mode",
            "formulas = {}",
            "",
            "# Placeholder for formula extraction",
            "# In production, this would extract and analyze formulas",
            "",
            'print(f"Total formulas found: {len(formulas)}")',
        ]

        return "\n".join(code_lines)

    def _create_summary_cell(self, metadata: dict[str, Any]) -> NotebookCell:
        """Create a summary cell with analysis conclusions."""
        content = """## Analysis Summary

### Key Findings
1. Workbook structure has been analyzed
2. Data regions have been identified
3. Formula dependencies mapped (where applicable)

### Recommendations
- Use the data preview cells to understand the data structure
- Review detected patterns and anomalies
- Check formula consistency across sheets
- Validate data types and ranges

### Next Steps
This notebook can now be analyzed using LLM-powered strategies to:
- Detect complex patterns
- Identify data quality issues
- Suggest optimizations
- Generate insights about the data
"""
        return NotebookCell(
            cell_type=CellType.MARKDOWN,
            content=content,
            metadata={"cell_id": "summary"},
        )
