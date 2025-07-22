"""Analyze a single Excel sheet with its own dedicated agent and notebook.

This implements the correct architecture:
- One notebook per sheet
- One agent per notebook
- Reads Excel file into pandas (no inlined data)
- Saves to proper folder structure: analysis_results/[excel_file]/[sheet_name].ipynb
"""

import argparse
import asyncio
import json
import os
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from spreadsheet_analyzer.notebook_llm.llm_providers import get_provider
from spreadsheet_analyzer.notebook_llm.llm_providers.base import Message, Role
from spreadsheet_analyzer.notebook_llm.nap.protocols import Cell, CellType, NotebookDocument
from spreadsheet_analyzer.notebook_llm.strategies import get_strategy
from spreadsheet_analyzer.notebook_llm.strategies.base import AnalysisFocus, AnalysisTask, ResponseFormat
from spreadsheet_analyzer.pipeline.pipeline import DeterministicPipeline


def create_sheet_notebook(
    excel_path: Path, sheet_name: str, deterministic_results: dict | None = None
) -> NotebookDocument:
    """Create a notebook for analyzing a single sheet.

    Args:
        excel_path: Path to Excel file
        sheet_name: Name of sheet to analyze
        deterministic_results: Optional results from deterministic analysis

    Returns:
        NotebookDocument with initial cells for sheet analysis
    """
    cells = []

    # Cell 1: Overview
    overview_content = f"""# Excel Sheet Analysis: {sheet_name}

## File Information
- **File**: {excel_path.name}
- **Sheet**: {sheet_name}
- **Analysis Date**: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

## Analysis Plan
1. Load sheet data into pandas DataFrame
2. Explore sheet structure and statistics
3. Analyze formulas and dependencies
4. Identify patterns and anomalies
5. Generate insights and recommendations
"""
    cells.append(
        Cell(id="overview", cell_type=CellType.MARKDOWN, source=overview_content, metadata={"cell_type": "overview"})
    )

    # Cell 2: Import and setup
    setup_code = """# Import required libraries
import pandas as pd
import numpy as np
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

# Excel file path
excel_file = Path(%s)
sheet_name = %s

print(f"Analyzing sheet '{sheet_name}' from {excel_file.name}")
""" % (repr(str(excel_path)), repr(sheet_name))

    cells.append(Cell(id="setup", cell_type=CellType.CODE, source=setup_code, metadata={"cell_type": "setup"}))

    # Cell 3: Load data
    load_code = """# Load the sheet data
try:
    # Read with all data types preserved
    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
    print(f"Successfully loaded sheet: {df.shape[0]} rows × {df.shape[1]} columns")

    # Also load to check formulas
    from openpyxl import load_workbook
    wb = load_workbook(excel_file, data_only=False)
    ws = wb[sheet_name]
    print(f"Sheet dimensions from openpyxl: {ws.max_row} rows × {ws.max_column} columns")

except Exception as e:
    print(f"Error loading sheet: {e}")
    df = None
"""

    cells.append(
        Cell(id="load_data", cell_type=CellType.CODE, source=load_code, metadata={"cell_type": "data_loading"})
    )

    # Cell 4: Initial exploration
    explore_code = """# Initial data exploration
if df is not None:
    print("\\n=== Data Overview ===")
    print(f"Shape: {df.shape}")
    print(f"Memory usage: {df.memory_usage().sum() / 1024:.1f} KB")

    # Check for headers
    potential_header_row = None
    for i in range(min(5, len(df))):
        row = df.iloc[i]
        non_null = row.notna().sum()
        if non_null > len(df.columns) * 0.5:  # More than 50% non-null
            string_count = sum(isinstance(val, str) for val in row if pd.notna(val))
            if string_count > len(df.columns) * 0.5:  # More than 50% strings
                potential_header_row = i
                break

    if potential_header_row is not None:
        print(f"\\nPotential header row detected at row {potential_header_row}")
        print("Headers:", list(df.iloc[potential_header_row]))

    # Display first few rows
    print("\\n=== First 10 rows ===")
    display(df.head(10))
"""

    cells.append(
        Cell(id="explore", cell_type=CellType.CODE, source=explore_code, metadata={"cell_type": "exploration"})
    )

    # Cell 5: Add deterministic results if available
    if deterministic_results:
        det_content = f"""## Deterministic Analysis Results

### Summary
- **Execution Time**: {deterministic_results.get("execution_time", "N/A")}s
- **Success**: {deterministic_results.get("success", False)}

"""
        if deterministic_results.get("structure"):
            det_content += f"- **Total Sheets**: {len(deterministic_results['structure'].sheets)}\n"
        if deterministic_results.get("formulas"):
            det_content += f"- **Total Formulas**: {len(deterministic_results['formulas'].dependency_graph)}\n"
        if deterministic_results.get("security"):
            det_content += f"- **Security Risk**: {deterministic_results['security'].risk_level}\n"

        cells.append(
            Cell(
                id="deterministic_results",
                cell_type=CellType.MARKDOWN,
                source=det_content,
                metadata={"analysis_type": "deterministic"},
            )
        )

    # Create notebook document
    return NotebookDocument(
        id=f"sheet_analysis_{sheet_name}",
        cells=cells,
        metadata={"excel_file": str(excel_path), "sheet_name": sheet_name},
        kernel_spec={"name": "python3", "display_name": "Python 3"},
        language_info={"name": "python", "version": "3.12"},
    )


def save_notebook(notebook: NotebookDocument, output_path: Path) -> None:
    """Save notebook in Jupyter format with proper line formatting."""

    def format_source(source: str) -> list[str]:
        """Format source content with proper newlines."""
        if not source:
            return []

        lines = source.split("\n")
        formatted = []
        for i, line in enumerate(lines):
            if i < len(lines) - 1:
                formatted.append(line + "\n")
            else:
                formatted.append(line)
        return formatted

    jupyter_nb = {
        "cells": [
            {
                "cell_type": cell.cell_type.value,
                "metadata": cell.metadata,
                "source": format_source(cell.source),
                "outputs": cell.outputs if hasattr(cell, "outputs") else [],
            }
            for cell in notebook.cells
        ],
        "metadata": {**notebook.metadata, "kernelspec": notebook.kernel_spec, "language_info": notebook.language_info},
        "nbformat": 4,
        "nbformat_minor": 5,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(jupyter_nb, f, indent=2)


async def analyze_sheet_with_llm(
    notebook: NotebookDocument, provider, strategy_name: str = "hierarchical"
) -> tuple[NotebookDocument, dict]:
    """Analyze a sheet using LLM with specified strategy.

    Args:
        notebook: Initial notebook with sheet data
        provider: LLM provider instance
        strategy_name: Strategy to use for analysis

    Returns:
        Updated notebook and analysis results
    """
    strategy = get_strategy(strategy_name)

    # Prepare context
    context = strategy.prepare_context(notebook, AnalysisFocus.GENERAL, token_budget=4000)

    print(f"\nUsing {strategy_name} strategy:")
    print(f"  - Compression: {context.compression_method}")
    print(f"  - Token count: {context.token_count}")
    print(f"  - Cells selected: {len(context.cells)}")

    # Create analysis task
    task = AnalysisTask(
        name="sheet_analysis",
        description=f"Analyze the Excel sheet '{notebook.metadata.get('sheet_name', 'Unknown')}' for structure, patterns, and insights",
        focus=AnalysisFocus.GENERAL,
        expected_format=ResponseFormat.STRUCTURED,
    )

    # Format prompt
    prompt = strategy.format_prompt(context, task)

    # Get LLM response
    messages = [
        Message(
            role=Role.SYSTEM,
            content="You are an expert Excel analyst. Analyze the provided sheet data and notebook context to identify patterns, anomalies, and insights.",
        ),
        Message(role=Role.USER, content=prompt),
    ]

    print("Sending request to LLM...")
    response = provider.complete(messages, temperature=0.1)

    # Add analysis results to notebook
    analysis_cell = Cell(
        id=f"llm_analysis_{strategy_name}",
        cell_type=CellType.MARKDOWN,
        source=f"""## LLM Analysis ({strategy_name.replace("_", " ").title()})

{response.content}
""",
        metadata={"analysis_type": "llm", "strategy": strategy_name},
    )
    notebook.cells.append(analysis_cell)

    results = {
        "strategy": strategy_name,
        "context": context.to_dict(),
        "response": response.content,
        "usage": response.usage,
    }

    return notebook, results


def list_sheets(excel_file: Path) -> list[str]:
    """List all sheets in an Excel file."""
    try:
        wb = load_workbook(excel_file, read_only=True)
        return wb.sheetnames
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return []


def create_parser() -> argparse.ArgumentParser:
    """Create command-line argument parser."""
    parser = argparse.ArgumentParser(
        description="Analyze a single Excel sheet using LLM-powered analysis",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Analyze a specific sheet
  %(prog)s path/to/file.xlsx "Sheet Name"

  # List all sheets in a file
  %(prog)s path/to/file.xlsx --list-sheets

  # Use a different LLM model
  %(prog)s path/to/file.xlsx "Sheet1" --model claude-opus-4-20250514

  # Use a different analysis strategy
  %(prog)s path/to/file.xlsx "Sheet1" --strategy detailed
""",
    )

    parser.add_argument("excel_file", type=Path, help="Path to Excel file to analyze")

    parser.add_argument(
        "sheet_name", nargs="?", help="Name of sheet to analyze (required unless --list-sheets is used)"
    )

    parser.add_argument(
        "--list-sheets", "-l", action="store_true", help="List all available sheets in the Excel file and exit"
    )

    parser.add_argument(
        "--model", "-m", default="claude-sonnet-4-20250514", help="LLM model to use (default: claude-sonnet-4-20250514)"
    )

    parser.add_argument(
        "--strategy",
        "-s",
        default="hierarchical",
        choices=["basic", "hierarchical", "detailed"],
        help="Analysis strategy to use (default: hierarchical)",
    )

    parser.add_argument(
        "--output-dir",
        "-o",
        type=Path,
        default=Path("analysis_results"),
        help="Output directory for results (default: analysis_results)",
    )

    parser.add_argument("--skip-deterministic", action="store_true", help="Skip deterministic analysis phase")

    parser.add_argument("--verbose", "-v", action="store_true", help="Enable verbose output")

    return parser


def main():
    """Main function to analyze a single sheet."""
    parser = create_parser()
    args = parser.parse_args()

    # Validate Excel file exists
    if not args.excel_file.exists():
        print(f"ERROR: Excel file not found: {args.excel_file}")
        sys.exit(1)

    # List sheets if requested
    if args.list_sheets:
        sheets = list_sheets(args.excel_file)
        if sheets:
            print(f"Sheets in {args.excel_file.name}:")
            for i, sheet in enumerate(sheets, 1):
                print(f"  {i}. {sheet}")
        else:
            print("No sheets found or error reading file.")
        sys.exit(0)

    # Validate sheet name is provided
    if not args.sheet_name:
        print("ERROR: Sheet name is required (use --list-sheets to see available sheets)")
        parser.print_help()
        sys.exit(1)

    # Check if sheet exists
    sheets = list_sheets(args.excel_file)
    if args.sheet_name not in sheets:
        print(f"ERROR: Sheet '{args.sheet_name}' not found in {args.excel_file.name}")
        print("\nAvailable sheets:")
        for sheet in sheets:
            print(f"  - {sheet}")
        sys.exit(1)

    # Check if Anthropic API key is set
    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("ERROR: ANTHROPIC_API_KEY environment variable not set!")
        print("Please set it with: export ANTHROPIC_API_KEY='your-api-key'")
        sys.exit(1)

    print("=== Analyzing Single Sheet ===")
    print(f"File: {args.excel_file.name}")
    print(f"Target Sheet: {args.sheet_name}")
    print(f"Model: {args.model}")
    print(f"Strategy: {args.strategy}")
    print(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

    # Step 1: Run deterministic analysis on the whole file
    det_dict = None
    if not args.skip_deterministic:
        print("Step 1: Running deterministic analysis...")
        pipeline = DeterministicPipeline()
        det_result = pipeline.run(args.excel_file)

        det_dict = {
            "success": det_result.success,
            "execution_time": det_result.execution_time,
            "errors": det_result.errors,
        }
        if det_result.structure:
            det_dict["structure"] = det_result.structure
        if det_result.formulas:
            det_dict["formulas"] = det_result.formulas
        if det_result.security:
            det_dict["security"] = det_result.security

        print(f"✓ Deterministic analysis complete ({det_result.execution_time:.2f}s)\n")
    else:
        print("Step 1: Skipping deterministic analysis (--skip-deterministic)\n")

    # Step 2: Create notebook for the target sheet
    print(f"Step 2: Creating notebook for sheet '{args.sheet_name}'...")
    notebook = create_sheet_notebook(args.excel_file, args.sheet_name, det_dict)
    print(f"✓ Created notebook with {len(notebook.cells)} initial cells\n")

    # Step 3: Run LLM analysis
    print("Step 3: Running LLM analysis...")
    try:
        provider = get_provider("anthropic", model=args.model)
        print(f"✓ Initialized {provider.model_name}")

        if args.verbose:
            print(f"  - Strategy: {args.strategy}")
            print("  - Token budget: 4000")

        # Run analysis
        notebook, results = asyncio.run(analyze_sheet_with_llm(notebook, provider, args.strategy))

        print("✓ LLM analysis complete")
        if isinstance(results.get("usage"), dict):
            print(f"  - Tokens used: {results['usage'].get('total_tokens', 'N/A')}")

    except Exception as e:
        print(f"ERROR in LLM analysis: {e}")
        if args.verbose:
            import traceback

            traceback.print_exc()
        results = {"error": str(e)}

    # Step 4: Save results
    print("\nStep 4: Saving results...")

    # Create proper folder structure
    excel_name = args.excel_file.stem
    output_dir = args.output_dir / excel_name
    output_dir.mkdir(parents=True, exist_ok=True)

    # Save notebook
    sheet_filename = args.sheet_name.replace(" ", "_").lower()
    notebook_path = output_dir / f"{sheet_filename}.ipynb"
    save_notebook(notebook, notebook_path)
    print(f"✓ Notebook saved to: {notebook_path}")

    # Save analysis metadata
    metadata_path = output_dir / f"{sheet_filename}_metadata.json"
    metadata = {
        "excel_file": str(args.excel_file),
        "sheet_name": args.sheet_name,
        "analysis_date": datetime.now().isoformat(),
        "model": args.model,
        "strategy": args.strategy,
        "llm_results": results if "error" not in results else {"error": results["error"]},
    }
    with open(metadata_path, "w") as f:
        json.dump(metadata, f, indent=2)
    print(f"✓ Metadata saved to: {metadata_path}")

    print("\n=== Analysis Complete ===")
    print(f"Results saved in: {output_dir}/")
    print(f"Open {notebook_path.name} in Jupyter to view the analysis.")


if __name__ == "__main__":
    main()
