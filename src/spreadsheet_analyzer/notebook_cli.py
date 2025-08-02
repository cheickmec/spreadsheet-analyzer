#!/usr/bin/env python3
"""
Notebook Tools CLI

Automated Excel analysis using LLM function calling with the notebook tools interface.
"""

import argparse
import asyncio
import json
import logging
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from structlog import get_logger

from spreadsheet_analyzer.graph_db.query_interface import (
    create_enhanced_query_interface,
)
from spreadsheet_analyzer.notebook_session import notebook_session

# Import observability components
from spreadsheet_analyzer.observability import (
    PhoenixConfig,
    get_cost_tracker,
    initialize_cost_tracker,
    initialize_phoenix,
    instrument_all,
)
from spreadsheet_analyzer.pipeline import DeterministicPipeline
from spreadsheet_analyzer.pipeline.types import (
    ContentAnalysis,
    FormulaAnalysis,
    IntegrityResult,
    PipelineResult,
    SecurityReport,
    WorkbookStructure,
)

logger = get_logger(__name__)


class StructuredFileNameGenerator:
    """Generate structured file names for notebooks and logs with all relevant parameters."""

    def __init__(
        self,
        excel_file: Path,
        model: str,
        sheet_index: int,
        sheet_name: str | None = None,
        max_rounds: int = 5,
        session_id: str | None = None,
    ):
        """
        Initialize the file name generator.

        Args:
            excel_file: Path to the Excel file being analyzed
            model: LLM model name being used
            sheet_index: Index of the sheet being analyzed
            sheet_name: Name of the sheet being analyzed (optional)
            max_rounds: Maximum number of analysis rounds
            session_id: Custom session ID (optional)
        """
        self.excel_file = excel_file
        self.model = self._sanitize_model_name(model)
        self.sheet_index = sheet_index
        self.sheet_name = self._sanitize_sheet_name(sheet_name) if sheet_name else None
        self.max_rounds = max_rounds
        self.session_id = session_id

    def _sanitize_model_name(self, model: str) -> str:
        """Sanitize model name for use in file names."""
        # Remove version suffixes and special characters
        model_clean = model.replace("-", "_").replace(".", "_")
        # Extract the main model name (e.g., "claude_3_5_sonnet" from "claude-3-5-sonnet-20241022")
        if "claude" in model_clean.lower():
            # Extract Claude model variant
            if "opus" in model_clean.lower():
                return "claude_opus"
            elif "sonnet" in model_clean.lower():
                return "claude_sonnet"
            elif "haiku" in model_clean.lower():
                return "claude_haiku"
            else:
                return "claude"
        elif "gpt" in model_clean.lower():
            # Extract GPT model variant
            if "4" in model_clean:
                return "gpt4"
            elif "3" in model_clean:
                return "gpt3"
            else:
                return "gpt"
        elif any(name in model_clean.lower() for name in ["ollama", "mistral", "llama", "mixtral", "codellama"]):
            # Extract Ollama/local model variant
            if "mistral" in model_clean.lower():
                return "ollama_mistral"
            elif "llama" in model_clean.lower():
                return "ollama_llama"
            elif "mixtral" in model_clean.lower():
                return "ollama_mixtral"
            elif "codellama" in model_clean.lower():
                return "ollama_codellama"
            else:
                return "ollama"
        else:
            # For other models, use a simplified version
            return model_clean.split("_")[0] if "_" in model_clean else model_clean

    def _sanitize_sheet_name(self, sheet_name: str) -> str:
        """Sanitize sheet name for use in file names."""
        # Remove or replace characters that are problematic in file names
        # Replace spaces, special characters with underscores
        sanitized = re.sub(r"[^\w\-_.]", "_", sheet_name)
        # Remove multiple consecutive underscores
        sanitized = re.sub(r"_+", "_", sanitized)
        # Remove leading/trailing underscores
        sanitized = sanitized.strip("_")
        # Limit length
        return sanitized[:50] if len(sanitized) > 50 else sanitized

    def generate_notebook_name(self, include_timestamp: bool = False) -> str:
        """
        Generate a structured notebook file name.

        Args:
            include_timestamp: Whether to include timestamp in notebook name

        Returns:
            Structured notebook file name
        """
        # Build the base name with all parameters
        parts = [
            self.excel_file.stem,  # Excel file name
            f"sheet{self.sheet_index}",  # Sheet index
        ]

        # Add sheet name if available
        if self.sheet_name:
            parts.append(self.sheet_name)

        # Add model name
        parts.append(self.model)

        # Add max rounds
        parts.append(f"r{self.max_rounds}")

        # Add timestamp if requested
        if include_timestamp:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            parts.append(timestamp)

        # Join parts and add extension
        return f"{'_'.join(parts)}.ipynb"

    def get_cost_tracking_path(self, output_dir: Path | None = None) -> Path:
        """Get the path for the cost tracking file."""
        parts = [
            self.excel_file.stem,
            f"sheet{self.sheet_index}",
        ]
        if self.sheet_name:
            parts.append(self.sheet_name)
        parts.append(self.model)
        parts.append(f"r{self.max_rounds}")
        parts.append("cost_tracking")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        parts.append(timestamp)

        cost_name = f"{'_'.join(parts)}.json"

        if output_dir:
            output_dir.mkdir(parents=True, exist_ok=True)
            return output_dir / cost_name
        else:
            # Use logs directory with date organization
            date_str = datetime.now().strftime("%Y%m%d")
            logs_dir = Path("logs") / date_str
            logs_dir.mkdir(parents=True, exist_ok=True)
            return logs_dir / cost_name

    def generate_log_name(self, include_timestamp: bool = True) -> str:
        """
        Generate a structured log file name.

        Args:
            include_timestamp: Whether to include timestamp in log name (default True)

        Returns:
            Structured log file name
        """
        # Build the base name with all parameters
        parts = [
            self.excel_file.stem,  # Excel file name
            f"sheet{self.sheet_index}",  # Sheet index
        ]

        # Add sheet name if available
        if self.sheet_name:
            parts.append(self.sheet_name)

        # Add model name
        parts.append(self.model)

        # Add max rounds
        parts.append(f"r{self.max_rounds}")

        # Add log type
        parts.append("llm_messages")

        # Add timestamp (default for logs)
        if include_timestamp:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            parts.append(timestamp)

        # Join parts and add extension
        return f"{'_'.join(parts)}.log"

    def generate_session_id(self) -> str:
        """Generate a structured session ID with timestamp.
        Returns:
            Structured session ID matching log file naming convention
        """
        parts = [
            self.excel_file.stem,  # Excel file name
            f"sheet{self.sheet_index}",  # Sheet index
        ]

        # Add sheet name if available
        if self.sheet_name:
            parts.append(self.sheet_name)

        # Add model name
        parts.append(self.model)

        # Add max rounds
        parts.append(f"r{self.max_rounds}")

        # Add session suffix
        parts.append("analysis_session")

        # Add timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        parts.append(timestamp)

        return "_".join(parts)


class PipelineResultsToMarkdown:
    """Convert pipeline results to well-formatted markdown cells."""

    @staticmethod
    def create_header(pipeline_result: PipelineResult) -> str:
        """Create the main header markdown."""
        return f"""# 📊 Excel Analysis Report

**File:** `{pipeline_result.context.file_path.name}`
"""

    @staticmethod
    def integrity_to_markdown(integrity: IntegrityResult) -> str:
        """Convert integrity results to markdown."""
        trust_stars = "⭐" * integrity.trust_tier

        md = f"""## 🔒 File Integrity Analysis

**File Hash:** `{integrity.file_hash[:16]}...`
**File Size:** {integrity.metadata.size_mb:.2f} MB
**MIME Type:** {integrity.metadata.mime_type}
**Excel Format:** {"✅ Valid Excel" if integrity.is_excel else "❌ Not Excel"}
**Trust Level:** {trust_stars} ({integrity.trust_tier}/5)
**Processing Class:** {integrity.processing_class}
"""

        if not integrity.validation_passed:
            md += "\n### ⚠️ Validation Status: Failed\n"

        return md

    @staticmethod
    def security_to_markdown(security: SecurityReport) -> str:
        """Convert security results to markdown."""
        # Only show security section if risk is MEDIUM or higher
        if security.risk_level in ("LOW", "NONE"):
            return ""

        risk_emoji = {"CRITICAL": "🔴", "HIGH": "🟠", "MEDIUM": "🟡"}

        md = f"""## 🛡️ Security Analysis

**Risk Level:** {risk_emoji.get(security.risk_level, "❓")} {security.risk_level}
"""

        if security.has_macros:
            md += "⚠️ **Contains VBA Macros**\n"
        if security.has_external_links:
            md += "⚠️ **Contains External Links**\n"

        # Only show non-hidden sheet threats
        relevant_threats = [t for t in security.threats if t.threat_type not in ("HIDDEN_SHEET", "VERY_HIDDEN_SHEET")]

        if relevant_threats:
            md += "\n### Detected Threats:\n"
            for threat in relevant_threats[:5]:
                md += f"- **{threat.threat_type}**: {threat.description}\n"
            if len(relevant_threats) > 5:
                md += f"- ...and {len(relevant_threats) - 5} more threats\n"

        return md

    @staticmethod
    def structure_to_markdown(structure: WorkbookStructure, security: SecurityReport | None = None) -> str:
        """Convert structure results to markdown."""
        # Extract hidden sheet names from security threats
        hidden_sheets = set()
        if security and security.threats:
            for threat in security.threats:
                if threat.threat_type in ("HIDDEN_SHEET", "VERY_HIDDEN_SHEET"):
                    sheet_name = threat.details.get("sheet_name")
                    if sheet_name:
                        hidden_sheets.add(sheet_name)

        md = f"""## 📋 Structural Analysis

**Total Sheets:** {structure.sheet_count}
**Total Cells with Data:** {structure.total_cells:,}
**Named Ranges:** {len(structure.named_ranges)}
**Complexity Score:** {structure.complexity_score}/100

### Sheet Details:

| Sheet Name | Hidden | Rows | Columns | Formulas |
|------------|--------|------|---------|----------|
"""

        for sheet in structure.sheets:
            is_hidden = sheet.name in hidden_sheets
            md += f"| {sheet.name} | {is_hidden} | {sheet.row_count:,} | {sheet.column_count} | {sheet.formula_count:,} |\n"

        return md

    @staticmethod
    def formulas_to_markdown(formulas: FormulaAnalysis) -> str:
        """Convert formula analysis to markdown."""
        md = f"""## 🔗 Formula Analysis

**Max Dependency Depth:** {formulas.max_dependency_depth} levels
**Formula Complexity Score:** {formulas.formula_complexity_score}/100
**Circular References:** {"⚠️ Yes" if formulas.has_circular_references else "✅ No"}
**Volatile Formulas:** {len(formulas.volatile_formulas)}
**External References:** {len(formulas.external_references)}
"""

        if formulas.circular_references:
            md += "\n### ⚠️ Circular References Found:\n"
            for i, chain in enumerate(list(formulas.circular_references)[:3], 1):
                chain_str = " → ".join(f"`{cell}`" for cell in list(chain)[:5])
                if len(chain) > 5:
                    chain_str += " → ..."
                md += f"{i}. {chain_str}\n"
            if len(formulas.circular_references) > 3:
                md += f"...and {len(formulas.circular_references) - 3} more circular reference chains\n"

        if formulas.volatile_formulas:
            md += "\n### 🔄 Volatile Formulas (recalculate on every change):\n"
            for cell in list(formulas.volatile_formulas)[:5]:
                md += f"- `{cell}`\n"
            if len(formulas.volatile_formulas) > 5:
                md += f"- ...and {len(formulas.volatile_formulas) - 5} more volatile formulas\n"

        return md

    @staticmethod
    def content_to_markdown(content: ContentAnalysis) -> str:
        """Convert content analysis to markdown."""
        quality_emoji = "🟢" if content.data_quality_score >= 80 else "🟡" if content.data_quality_score >= 60 else "🔴"

        md = f"""## 📊 Content Analysis

**Data Quality Score:** {quality_emoji} {content.data_quality_score}/100
"""

        # Filter out "incomplete data" insights which are usually false positives in Excel
        relevant_insights = [
            i
            for i in content.insights
            if "incomplete data" not in i.title.lower()
            and "missing data" not in i.title.lower()
            and "data completeness" not in i.description.lower()
        ]

        if relevant_insights:
            md += "\n### 💡 Key Insights:\n"
            for insight in relevant_insights[:5]:
                severity_emoji = "🔴" if insight.severity == "HIGH" else "🟡" if insight.severity == "MEDIUM" else "🟢"
                md += f"\n**{insight.title}** {severity_emoji}\n"
                md += f"{insight.description}\n"
                if insight.recommendation:
                    md += f"- **Recommendation:** {insight.recommendation}\n"
            if len(relevant_insights) > 5:
                md += f"\n...and {len(relevant_insights) - 5} more insights\n"

        if content.data_patterns:
            md += "\n### 🔍 Data Patterns:\n"
            for pattern in content.data_patterns[:3]:
                confidence = int(pattern.confidence * 100)
                md += f"- **{pattern.pattern_type}** (Confidence: {confidence}%): {pattern.description}\n"

        return md

    @staticmethod
    def create_summary(pipeline_result: PipelineResult) -> str:
        """Create a summary markdown section."""
        md = "## 📌 Analysis Summary\n\n"

        if pipeline_result.errors:
            md += "### ❌ Errors Encountered:\n"
            for error in pipeline_result.errors:
                md += f"- {error}\n"
            md += "\n"

        md += "### ✅ Completed Analysis Stages:\n"
        if pipeline_result.integrity:
            md += "- ✓ File Integrity Check\n"
        if pipeline_result.security:
            md += "- ✓ Security Scan\n"
        if pipeline_result.structure:
            md += "- ✓ Structural Analysis\n"
        if pipeline_result.formulas:
            md += "- ✓ Formula Analysis\n"
        if pipeline_result.content:
            md += "- ✓ Content Intelligence\n"

        md += "\n---\n\n*This analysis was generated using the deterministic pipeline. Additional interactive analysis can be performed using the cells below.*"

        return md


async def track_llm_usage(response: Any, model: str) -> None:
    """
    Track token usage from LLM response.

    Args:
        response: LLM response object
        model: Model name
    """
    try:
        # Extract usage metadata from response
        usage_metadata = None

        # Try different ways to get usage data based on provider
        if hasattr(response, "usage_metadata"):
            usage_metadata = response.usage_metadata
        elif hasattr(response, "usage"):
            usage_metadata = response.usage
        elif hasattr(response, "response_metadata"):
            usage_metadata = response.response_metadata.get("usage", {})

        if usage_metadata:
            input_tokens = (
                usage_metadata.get("input_tokens", 0)
                or usage_metadata.get("prompt_tokens", 0)
                or usage_metadata.get("total_tokens", 0) // 2  # Rough estimate
            )
            output_tokens = (
                usage_metadata.get("output_tokens", 0)
                or usage_metadata.get("completion_tokens", 0)
                or usage_metadata.get("total_tokens", 0) // 2  # Rough estimate
            )

            if input_tokens > 0 or output_tokens > 0:
                cost_tracker = get_cost_tracker()
                cost_tracker.track_usage(
                    model=model,
                    input_tokens=input_tokens,
                    output_tokens=output_tokens,
                    metadata={"source": "notebook_cli"},
                )

    except Exception as e:
        logger.debug(f"Failed to track LLM usage: {e}")


class NotebookCLI:
    """CLI interface for automated Excel analysis with LLM integration."""

    def __init__(self):
        self.parser = self._create_parser()

    def _create_parser(self):
        parser = argparse.ArgumentParser(
            description="Automated Excel analysis using LLM function calling with Phoenix observability.",
            formatter_class=argparse.RawDescriptionHelpFormatter,
            epilog="""
Examples:
  # Basic analysis with default Claude model
  %(prog)s data.xlsx

  # Use GPT-4 instead of Claude
  %(prog)s data.xlsx --model gpt-4 --sheet-index 1

  # Use a specific Claude model
  %(prog)s data.xlsx --model claude-3-opus-20240229

  # Custom output location and session
  %(prog)s data.xlsx --output-dir results --session-id analysis-001

  # Configure Phoenix observability
  %(prog)s data.xlsx --phoenix-mode docker --phoenix-host localhost

  # Set cost limit
  %(prog)s data.xlsx --cost-limit 5.0
""",
        )

        parser.add_argument("excel_file", type=Path, help="Path to the Excel file to analyze.")
        parser.add_argument(
            "--model",
            type=str,
            default="claude-3-5-sonnet-20241022",
            help="LLM model to use. Default: Claude 3.5 Sonnet. Examples: 'gpt-4', 'claude-3-opus-20240229'",
        )
        parser.add_argument(
            "--api-key",
            type=str,
            default=None,
            help="API key for the LLM. Defaults to environment variable (ANTHROPIC_API_KEY or OPENAI_API_KEY).",
        )
        parser.add_argument(
            "--session-id",
            type=str,
            default=None,
            help="Unique ID for the notebook session. Defaults to a name derived from the Excel file.",
        )
        parser.add_argument(
            "--notebook-path",
            type=Path,
            default=None,
            help="Path to save/load the notebook. Defaults to notebook_{session_id}.ipynb.",
        )
        parser.add_argument(
            "--max-rounds",
            type=int,
            default=5,
            help="Maximum number of analysis rounds (i.e., LLM calls).",
        )
        parser.add_argument(
            "--sheet-index",
            type=int,
            default=0,
            help="Index of the sheet to analyze (0-based). Default is 0 (first sheet).",
        )
        parser.add_argument(
            "-o",
            "--output-dir",
            type=Path,
            help="Output directory for results (default: analysis_results/YYYYMMDD/)",
        )
        parser.add_argument("--verbose", action="store_true", help="Enable verbose logging.")

        # Phoenix observability options
        phoenix_group = parser.add_argument_group("Phoenix Observability")
        phoenix_group.add_argument(
            "--phoenix-mode",
            choices=["local", "cloud", "docker", "none"],
            default="docker",
            help="Phoenix deployment mode (default: docker)",
        )
        phoenix_group.add_argument(
            "--phoenix-host",
            default="localhost",
            help="Phoenix host for docker mode (default: localhost)",
        )
        phoenix_group.add_argument(
            "--phoenix-port",
            type=int,
            default=6006,
            help="Phoenix port for local mode (default: 6006)",
        )
        phoenix_group.add_argument(
            "--phoenix-api-key",
            help="Phoenix API key for cloud mode (or set PHOENIX_API_KEY env var)",
        )
        phoenix_group.add_argument(
            "--phoenix-project",
            default="spreadsheet-analyzer",
            help="Phoenix project name (default: spreadsheet-analyzer)",
        )

        # Cost tracking options
        cost_group = parser.add_argument_group("Cost Tracking")
        cost_group.add_argument(
            "--cost-limit",
            type=float,
            help="Set spending limit in USD",
        )
        cost_group.add_argument(
            "--track-costs",
            action="store_true",
            default=True,
            help="Enable cost tracking (default: True)",
        )

        return parser

    async def run_analysis(self, args):
        """Run the automated analysis loop."""
        # Resolve the excel file path to be absolute
        excel_path = args.excel_file.resolve()

        # Initialize Phoenix observability
        tracer_provider = None
        if args.phoenix_mode != "none":
            phoenix_config = PhoenixConfig(
                mode=args.phoenix_mode,
                host=args.phoenix_host,
                port=args.phoenix_port,
                api_key=args.phoenix_api_key or os.getenv("PHOENIX_API_KEY"),
                project_name=args.phoenix_project,
            )
            tracer_provider = initialize_phoenix(phoenix_config)

            if tracer_provider:
                # Instrument all providers
                results = instrument_all(tracer_provider)
                logger.info("Phoenix instrumentation complete", results=results)
            else:
                logger.warning("Phoenix initialization failed, continuing without observability")

        # Get sheet name from Excel file if possible
        sheet_name = None
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            if args.sheet_index < len(workbook.sheetnames):
                sheet_name = workbook.sheetnames[args.sheet_index]
            workbook.close()
        except Exception as e:
            logger.warning(f"Could not read sheet name from Excel file: {e}")

        # Create file name generator with all parameters
        file_name_generator = StructuredFileNameGenerator(
            excel_file=excel_path,
            model=args.model,
            sheet_index=args.sheet_index,
            sheet_name=sheet_name,
            max_rounds=args.max_rounds,
            session_id=args.session_id,
        )

        # Generate structured file names
        notebook_name = file_name_generator.generate_notebook_name(include_timestamp=False)
        notebook_path = args.notebook_path or excel_path.parent / notebook_name
        session_id = args.session_id or file_name_generator.generate_session_id()
        llm_log_name = file_name_generator.generate_log_name(include_timestamp=True)
        llm_log_path = excel_path.parent / llm_log_name

        # Set up LLM message logging to file

        llm_logger = logging.getLogger("llm_messages")
        llm_logger.setLevel(logging.DEBUG)
        llm_file_handler = logging.FileHandler(llm_log_path, mode="w")
        llm_file_handler.setLevel(logging.DEBUG)
        llm_formatter = logging.Formatter("%(asctime)s - %(message)s")
        llm_file_handler.setFormatter(llm_formatter)
        llm_logger.addHandler(llm_file_handler)
        llm_logger.info(f"Starting LLM message logging for: {excel_path}")

        # Initialize cost tracking
        cost_tracking_path = file_name_generator.get_cost_tracking_path(args.output_dir)
        if args.track_costs:
            cost_tracker = initialize_cost_tracker(cost_limit=args.cost_limit, save_path=cost_tracking_path)
            logger.info(f"💰 Cost tracking enabled: {cost_tracking_path}")
            if args.cost_limit:
                logger.info(f"💰 Cost limit set: ${args.cost_limit:.2f}")

        # Check if file exists
        if not excel_path.exists():
            logger.error(f"Excel file not found: {excel_path}")
            return

        logger.info(f"Starting analysis of: {excel_path}")
        logger.info(f"Session ID: {session_id}")
        logger.info(f"Notebook Path: {notebook_path}")
        logger.info(f"LLM message log: {llm_log_path}")

        # Step 1: Run deterministic pipeline
        logger.info("Running deterministic pipeline analysis...")
        pipeline_result = None
        query_interface = None
        formula_cache_path = None

        try:
            # Run the pipeline with progress tracking
            pipeline = DeterministicPipeline()
            pipeline_result = await asyncio.to_thread(pipeline.run, excel_path)

            if pipeline_result.success:
                logger.info(f"✅ Pipeline analysis completed in {pipeline_result.execution_time:.2f} seconds")

                # Save formula analysis to pickle cache if available
                if pipeline_result.formulas:
                    import pickle

                    cache_dir = Path(".pipeline_cache")
                    cache_dir.mkdir(exist_ok=True)

                    # Create a unique filename based on the Excel file
                    cache_filename = f"{excel_path.stem}_formula_analysis.pkl"
                    formula_cache_path = cache_dir / cache_filename

                    with formula_cache_path.open("wb") as f:
                        pickle.dump(pipeline_result.formulas, f)

                    logger.info(f"Saved formula analysis to cache: {formula_cache_path}")

                    # Create query interface
                    query_interface = create_enhanced_query_interface(pipeline_result.formulas)
                    logger.info("Created enhanced query interface for formula analysis")
            else:
                logger.error("Pipeline analysis failed:")
                for error in pipeline_result.errors:
                    logger.error(f"  - {error}")

        except Exception:
            logger.exception("Error running deterministic pipeline")
            # Continue anyway - we can still create a notebook

        # Step 2: Create notebook and add pipeline results
        async with notebook_session(session_id, notebook_path) as session:
            # Add query interface to session if available
            session.query_interface = query_interface

            # Register the session with the global session manager so tools can access it
            from spreadsheet_analyzer.notebook_llm_interface import get_session_manager

            session_manager = get_session_manager()
            session_manager._sessions["default_session"] = session

            # Add pipeline results as markdown cells if available
            if pipeline_result:
                logger.info("Adding pipeline results to notebook...")
                toolkit = session.toolkit

                # Add header
                header_md = PipelineResultsToMarkdown.create_header(pipeline_result)
                result = await toolkit.render_markdown(header_md)
                if result.is_err():
                    logger.warning(f"Failed to add header: {result.err_value}")

                # Skip integrity analysis - not useful for LLM
                # if pipeline_result.integrity:
                #     integrity_md = PipelineResultsToMarkdown.integrity_to_markdown(pipeline_result.integrity)
                #     result = await toolkit.render_markdown(integrity_md)
                #     if result.is_err():
                #         logger.warning(f"Failed to add integrity analysis: {result.err_value}")

                # Add security analysis
                if pipeline_result.security:
                    security_md = PipelineResultsToMarkdown.security_to_markdown(pipeline_result.security)
                    result = await toolkit.render_markdown(security_md)
                    if result.is_err():
                        logger.warning(f"Failed to add security analysis: {result.err_value}")

                # Add structure analysis
                if pipeline_result.structure:
                    structure_md = PipelineResultsToMarkdown.structure_to_markdown(
                        pipeline_result.structure, pipeline_result.security
                    )
                    result = await toolkit.render_markdown(structure_md)
                    if result.is_err():
                        logger.warning(f"Failed to add structure analysis: {result.err_value}")

                # Add formula analysis
                if pipeline_result.formulas:
                    formulas_md = PipelineResultsToMarkdown.formulas_to_markdown(pipeline_result.formulas)
                    result = await toolkit.render_markdown(formulas_md)
                    if result.is_err():
                        logger.warning(f"Failed to add formula analysis: {result.err_value}")

                # Skip content analysis - data quality scores are misleading for Excel
                # if pipeline_result.content:
                #     content_md = PipelineResultsToMarkdown.content_to_markdown(pipeline_result.content)
                #     result = await toolkit.render_markdown(content_md)
                #     if result.is_err():
                #         logger.warning(f"Failed to add content analysis: {result.err_value}")

                # Skip summary - not useful for LLM
                # summary_md = PipelineResultsToMarkdown.create_summary(pipeline_result)
                # result = await toolkit.render_markdown(summary_md)
                # if result.is_err():
                #     logger.warning(f"Failed to add summary: {result.err_value}")

                logger.info("Pipeline results added to notebook")

            # Add a basic data loading cell
            logger.info("Adding data loading cell...")

            # Calculate relative path from repo root
            try:
                repo_root = Path.cwd()
                relative_path = excel_path.relative_to(repo_root)
                path_str = f'"{relative_path}"'
            except ValueError:
                # If file is outside repo, use absolute path
                path_str = f'r"{excel_path}"'

            load_data_code = f"""import pandas as pd
from pathlib import Path

# Load the Excel file (path relative to repo root)
excel_path = Path({path_str})
sheet_index = {args.sheet_index}

print(f"Loading data from: {{excel_path}}")
print(f"Loading sheet at index: {{sheet_index}}")

try:
    # First, get information about all sheets
    xl_file = pd.ExcelFile(excel_path)
    sheet_names = xl_file.sheet_names
    print(f"\\nAvailable sheets: {{sheet_names}}")

    if sheet_index >= len(sheet_names):
        print(f"\\nError: Sheet index {{sheet_index}} is out of range. File has {{len(sheet_names)}} sheets.")
        print("Please use --sheet-index with a value between 0 and {{}}".format(len(sheet_names)-1))
        df = None
    else:
        # Load the specified sheet
        selected_sheet = sheet_names[sheet_index]
        print(f"\\nLoading sheet: '{{selected_sheet}}'")

        df = pd.read_excel(excel_path, sheet_name=sheet_index)
        print(f"\\nLoaded data with shape: {{df.shape}}")
        print(f"Columns: {{list(df.columns)}}")

        # Display first few rows
        print("\\nFirst 5 rows:")
        display(df.head())

        # Basic info
        print("\\nData types:")
        print(df.dtypes)

except Exception as e:
    print(f"Error loading Excel file: {{e}}")
    df = None
"""

            result = await toolkit.execute_code(load_data_code)
            if result.is_ok():
                logger.info(f"✅ Data loading cell executed: {result.ok_value.cell_id}")
            else:
                logger.error(f"❌ Failed to execute data loading cell: {result.err_value}")

            # Add query interface loading cell if formula cache exists
            if formula_cache_path and formula_cache_path.exists():
                logger.info("Adding query interface loading cell...")

                # Calculate relative path for the cache file
                try:
                    repo_root = Path.cwd()
                    relative_cache_path = formula_cache_path.relative_to(repo_root)
                    cache_path_str = f'"{relative_cache_path}"'
                except ValueError:
                    # If cache is outside repo, use absolute path
                    cache_path_str = f'r"{formula_cache_path}"'

                query_interface_code = f'''# Query interface for formula dependency analysis
# The pipeline has already analyzed all formulas and cached the results

import pickle
from pathlib import Path
from spreadsheet_analyzer.graph_db.query_interface import create_enhanced_query_interface

# Load cached formula analysis
cache_file = Path({cache_path_str})
with open(cache_file, 'rb') as f:
    formula_analysis = pickle.load(f)

query_interface = create_enhanced_query_interface(formula_analysis)

# Convenience functions for graph queries
def get_cell_dependencies(sheet, cell_ref):
    """Get complete dependency information for a specific cell."""
    result = query_interface.get_cell_dependencies(sheet, cell_ref)
    print(f"\\nCell {{sheet}}!{{cell_ref}}:")
    print(f"  Has formula: {{result.has_formula}}")
    if result.formula:
        print(f"  Formula: {{result.formula}}")
    if result.direct_dependencies:
        print(f"  Direct dependencies: {{', '.join(result.direct_dependencies[:5])}}")
        if len(result.direct_dependencies) > 5:
            print(f"    ...and {{len(result.direct_dependencies) - 5}} more")
    if result.direct_dependents:
        print(f"  Cells that depend on this: {{', '.join(result.direct_dependents[:5])}}")
        if len(result.direct_dependents) > 5:
            print(f"    ...and {{len(result.direct_dependents) - 5}} more")
    return result

def find_cells_affecting_range(sheet, start_cell, end_cell):
    """Find all cells that affect any cell within the specified range."""
    result = query_interface.find_cells_affecting_range(sheet, start_cell, end_cell)
    print(f"\\nCells affecting range {{sheet}}!{{start_cell}}:{{end_cell}}:")
    for cell, deps in list(result.items())[:5]:
        print(f"  {{cell}} depends on: {{', '.join(deps[:3])}}")
        if len(deps) > 3:
            print(f"    ...and {{len(deps) - 3}} more")
    if len(result) > 5:
        print(f"  ...and {{len(result) - 5}} more cells")
    return result

def get_formula_statistics():
    """Get comprehensive statistics about formulas in the workbook."""
    stats = query_interface.get_formula_statistics_with_ranges()
    print("\\nFormula Statistics:")
    print(f"  Total formulas: {{stats['total_formulas']:,}}")
    print(f"  Formulas with dependencies: {{stats['formulas_with_dependencies']:,}}")
    print(f"  Unique cells referenced: {{stats['unique_cells_referenced']:,}}")
    print(f"  Max dependency depth: {{stats['max_dependency_depth']}} levels")
    print(f"  Circular references: {{stats['circular_reference_chains']}}")
    print(f"  Formula complexity score: {{stats['complexity_score']}}/100")
    return stats

def find_empty_cells_in_formula_ranges(sheet):
    """Find empty cells that are part of formula ranges."""
    result = query_interface.find_empty_cells_in_formula_ranges(sheet)
    print(f"\\nEmpty cells in formula ranges for sheet '{{sheet}}':")
    if result:
        print(f"  Found {{len(result)}} empty cells")
        # Group by rows for display
        rows = {{}}
        for cell in list(result)[:20]:
            row_num = ''.join(filter(str.isdigit, cell))
            if row_num not in rows:
                rows[row_num] = []
            rows[row_num].append(cell)
        for row, cells in list(rows.items())[:5]:
            print(f"  Row {{row}}: {{', '.join(cells)}}")
        if len(result) > 20:
            print(f"  ...and {{len(result) - 20}} more")
    else:
        print("  No empty cells found in formula ranges")
    return result
'''

                result = await toolkit.execute_code(query_interface_code)
                if result.is_ok():
                    logger.info(f"✅ Query interface loading cell executed: {result.ok_value.cell_id}")
                else:
                    logger.error(f"❌ Failed to execute query interface loading cell: {result.err_value}")

            # Add graph query interface tools if formula analysis succeeded
            if query_interface and pipeline_result and pipeline_result.formulas:
                logger.info("Adding graph query interface tools...")

                # Add markdown documentation for graph queries
                graph_tools_doc = """## 🔍 Formula Analysis Tools

You have TWO approaches available for formula analysis:

### 1️⃣ Graph-Based Dependency Analysis (Recommended for Complex Files)
The deterministic pipeline has analyzed all formulas and created a dependency graph. These tools are robust and handle complex Excel files:

- **get_cell_dependencies** - Analyze what a cell depends on and what depends on it
- **find_cells_affecting_range** - Find all cells that affect a specific range
- **find_empty_cells_in_formula_ranges** - Find gaps in data that formulas reference
- **get_formula_statistics** - Get overall statistics about formulas
- **find_circular_references** - Find all circular reference chains

### 2️⃣ Formulas Library for Advanced Formula Evaluation (Recommended)
Robust formula evaluation using the 'formulas' library that handles complex Excel files:

- **load_excel_with_formulas** - Load Excel file for formula evaluation
- **evaluate_cell** - Get calculated cell values and formulas
- **set_cell_and_recalculate** - What-if analysis with recalculation
- **get_cell_dependencies_formulas** - Track formula dependencies
- **export_formulas_model** - Export model to JSON
- **get_formulas_help** - Get detailed help

✅ **Recommended**: The formulas library handles complex Excel files much better than other alternatives.

### Usage:
All tools are available through the tool-calling interface. Use graph-based analysis for quick dependency queries, and the formulas library for accurate formula evaluation and what-if analysis.
"""

                result = await toolkit.render_markdown(graph_tools_doc)
                if result.is_err():
                    logger.warning(f"Failed to add graph tools documentation: {result.err_value}")

                logger.info("Graph query tools are available via tool-calling interface")

                # Add marker to delineate LLM analysis region
                marker_result = await toolkit.render_markdown("## --- LLM Analysis Start ---")
                if marker_result.is_err():
                    logger.warning(f"Failed to add LLM analysis marker: {marker_result.err_value}")
                else:
                    logger.info("Added LLM analysis boundary marker")

            # Step 3: Setup LLM interaction (if requested)
            if args.max_rounds > 0:
                logger.info("Setting up LLM interaction...")

                # Import LangChain components
                try:
                    from langchain_anthropic import ChatAnthropic
                    from langchain_core.messages import HumanMessage, SystemMessage, ToolMessage
                    from langchain_ollama import ChatOllama
                    from langchain_openai import ChatOpenAI
                except ImportError:
                    logger.exception("Failed to import LangChain components")
                    logger.info("Please install langchain-anthropic, langchain-openai, and langchain-ollama packages")
                else:
                    # Get notebook tools
                    from spreadsheet_analyzer.notebook_llm_interface import get_notebook_tools

                    tools = get_notebook_tools()

                    # Initialize LLM based on model selection
                    try:
                        if "claude" in args.model.lower():
                            api_key = args.api_key or os.getenv("ANTHROPIC_API_KEY")
                            if not api_key:
                                logger.error("No API key provided. Set ANTHROPIC_API_KEY or use --api-key")
                                llm = None
                            else:
                                llm = ChatAnthropic(
                                    model_name=args.model,
                                    api_key=api_key,
                                    max_tokens=4096,
                                )
                        elif "gpt" in args.model.lower():
                            api_key = args.api_key or os.getenv("OPENAI_API_KEY")
                            if not api_key:
                                logger.error("No API key provided. Set OPENAI_API_KEY or use --api-key")
                                llm = None
                            else:
                                llm = ChatOpenAI(
                                    model_name=args.model,
                                    api_key=api_key,
                                    temperature=0,
                                )
                        elif any(
                            name in args.model.lower()
                            for name in [
                                "ollama",
                                "mistral",
                                "llama",
                                "mixtral",
                                "codellama",
                                "qwen",
                                "deepseek",
                                "command",
                                "phi",
                            ]
                        ):
                            # Ollama models - no tool support validation
                            logger.info(f"Using Ollama model: {args.model}")
                            # Extract the model name (remove "ollama:" prefix if present)
                            model_name = (
                                args.model.replace("ollama:", "") if args.model.startswith("ollama:") else args.model
                            )

                            # Directly create ChatOllama instance without any checks
                            llm = ChatOllama(
                                model=model_name,
                                base_url=os.getenv("OLLAMA_BASE_URL", "http://localhost:11434"),
                                temperature=0,
                            )
                        else:
                            logger.error(f"Unsupported model: {args.model}")
                            llm = None

                        if llm:
                            # Bind tools to LLM - let it fail loudly if not supported
                            llm_with_tools = llm.bind_tools(tools)

                            # Get current notebook state in py:percent format
                            try:
                                notebook_state = session.toolkit.export_to_percent_format()
                            except Exception:
                                logger.exception("Failed to export notebook state")
                                import traceback

                                traceback.print_exc()
                                notebook_state = "# Failed to export notebook state"

                            # Create initial prompt with notebook context
                            sheet_info = f" (sheet index {args.sheet_index})" if args.sheet_index != 0 else ""
                            initial_prompt = f"""I've loaded the Excel file '{excel_path.name}'{sheet_info} into a Jupyter notebook.

## Current Notebook State:
```python
{notebook_state}
```

The notebook already contains:
- Pipeline analysis results (Security, Structure, Formula Analysis)
- Data loaded into DataFrame 'df' with initial exploration showing shape and first rows
{"- Query interface for formula dependencies (graph-based analysis)" if formula_cache_path and formula_cache_path.exists() else ""}
- Formula analysis tools: graph-based for dependencies, formulas library for evaluation

Please continue the analysis from where it left off. **DO NOT re-execute cells that already have output.**

You can:
1. Execute NEW Python code to explore the data further
2. Use pandas operations to analyze patterns
3. Create visualizations if helpful
4. {"Query the formula dependency graph using graph-based tools" if formula_cache_path and formula_cache_path.exists() else "Look for data quality issues"}
5. Use formulas library tools for formula evaluation and what-if analysis
6. Provide insights and recommendations

Focus on deeper analysis that builds upon what's already been done.

IMPORTANT: Track your progress against the completion criteria and create a final summary when done."""

                            # Import session tracking
                            from spreadsheet_analyzer.observability import add_session_metadata, phoenix_session

                            messages = [
                                SystemMessage(
                                    f"""You are an autonomous data analyst AI conducting comprehensive spreadsheet analysis.

CRITICAL CONSTRAINT - NO VISUAL ACCESS:
- Cannot see images, plots, charts, or visualizations
- Must extract insights through textual methods only
- All analysis based on numerical summaries and textual descriptions

CONTEXT:
- Analyzing Excel file: {excel_path.name}
- Sheet index: {args.sheet_index}
- Sheet name: {sheet_name}

CURRENT NOTEBOOK STATE:
```python
{notebook_state}
```

AUTONOMOUS ANALYSIS PROTOCOL:
1. Conduct systematic analysis without seeking user approval
2. Back all assumptions with evidence from the data
3. Reach solid conclusions based on thorough investigation
4. Document reasoning in code comments
5. Use available tools to execute code and explore data
6. Focus on actionable insights and recommendations

MULTI-TABLE DETECTION - CRITICAL FIRST STEP:
Before analyzing data, ALWAYS check if the sheet contains multiple tables using BOTH mechanical AND semantic analysis:

1. **Initial Structure Scan**
   ```python
   # First, examine the raw structure with MORE context
   print(f"Sheet dimensions: {{df.shape}}")
   print("\n--- First 30 rows ---")
   print(df.head(30))

   # Look at ALL columns including unnamed ones
   print("\n--- Column overview ---")
   for col in df.columns:
       non_null = df[col].notna().sum()
       print(f"{{col}}: {{non_null}} non-null values, dtype: {{df[col].dtype}}")
   ```

2. **Mechanical Detection** (empty rows/columns)
   ```python
   # Check for empty row patterns that separate tables
   empty_rows = df.isnull().all(axis=1)
   empty_row_groups = empty_rows.groupby((~empty_rows).cumsum()).sum()
   print(f"\nEmpty row blocks: {{empty_row_groups[empty_row_groups > 0].to_dict()}}")

   # Check for empty columns (potential horizontal separators)
   empty_cols = df.isnull().all(axis=0)
   print(f"Empty columns: {{list(df.columns[empty_cols])}}")
   ```

3. **Semantic Detection** (USE HUMAN JUDGMENT)
   ```python
   # Ask yourself: What does each row represent?
   # Example: If row 1 = "Product ABC, $50" and row 100 = "John Smith, Developer"
   # These are clearly different entity types!

   # Check for semantic shifts in data
   print("\n--- Checking for semantic table boundaries ---")

   # 1. Analyze column groupings - do they describe the same type of thing?
   # Example: ['Customer', 'Address', 'Phone'] vs ['Date', 'Amount', 'Transaction ID']

   # 2. Look for granularity changes
   # Example: 5 summary rows followed by 1000 detail rows

   # 3. Check if column names suggest different purposes
   # Example: Financial columns vs HR columns in same sheet
   ```

4. **Common Multi-Table Patterns to Recognize**
   - **Master/Detail**: Header info (few rows) + Line items (many rows)
   - **Summary/Breakdown**: Totals followed by individual components
   - **Different Domains**: Unrelated business data side-by-side
   - **Time Periods**: Current data adjacent to historical data

   Ask: "Would these naturally be separate tables in a database?"

5. **Decision Framework**
   Even WITHOUT empty rows/columns, declare multiple tables if:
   - Rows represent fundamentally different entity types
   - Column sets serve different business purposes
   - There's a clear shift in data granularity
   - A business analyst would logically separate them

6. **Multi-Table Handling Strategy**
   If multiple tables detected:
   - Document the table boundaries and what each represents
   - Analyze each table's purpose separately
   - Use `.iloc[start:end, start_col:end_col]` for extraction
   - Focus on the most relevant table(s) for insights

COMPLETION PROTOCOL - CRITICAL:
- FIRST: Always perform multi-table detection using the empty row analysis code
- Complete ALL analysis steps autonomously without asking for user input
- NEVER ask "Would you like me to..." or "Let me know if..." or "Do you need..."
- When analysis is complete, provide a final summary and STOP
- If errors occur, implement workarounds or fix code, re-run it and continue analysis
- End with definitive conclusions
- DO NOT offer to perform additional analysis - just complete what's needed

TEXTUAL DATA EXPLORATION TECHNIQUES:
- `.iloc[start:end]` or `.loc[condition]` to examine specific data regions
- `.sample(n)` for random sampling
- `.groupby()` for categorical analysis
- `.value_counts()` for frequency distributions
- `.describe()` for statistical summaries
- `.corr()` for correlation analysis
- `.isnull().sum()` for missing data analysis

TEXTUAL VISUALIZATION ALTERNATIVES (NO IMAGES):
**Distributions & Patterns:**
- `.value_counts().head(10)` to show frequency distribution
- `.describe()` to show quartiles, mean, std, min/max
- `.quantile([0.1, 0.25, 0.5, 0.75, 0.9])` for detailed percentiles
- `.hist(bins=20).value_counts()` for histogram-like data

**Trends & Relationships:**
- `.groupby().agg(['mean', 'std', 'count'])` to show patterns by category
- `.corr().round(3)` to show correlation matrix numerically
- `.pivot_table()` to show cross-tabulations
- `.rolling(window=5).mean()` for moving averages

**Outliers & Anomalies:**
- IQR method: Q1, Q3 = df.quantile([0.25, 0.75]); IQR = Q3 - Q1
- `.quantile([0.01, 0.99])` to show extreme values
- `(df > df.quantile(0.99)) | (df < df.quantile(0.01))` to identify outliers
- `.std()` and z-score calculations for statistical outliers

**Missing Data Patterns:**
- `.isnull().sum()` for column-wise missing counts
- `.isnull().sum(axis=1)` for row-wise missing patterns
- `.isnull().groupby(df['category']).sum()` for missing by category

**Data Quality Assessment:**
- `.dtypes` to check data types
- `.nunique()` to check cardinality
- `.duplicated().sum()` to find duplicates
- `.apply(lambda x: x.astype(str).str.len().max())` for string length analysis

ERROR VALIDATION REQUIREMENTS:
- **Data Type Validation**: Check for mixed data types in columns
- **Range Validation**: Identify values outside expected ranges
- **Formula Verification**: If formulas exist, verify calculations manually
- **Consistency Checks**: Look for inconsistent naming, formatting, or values
- **Missing Data Patterns**: Analyze if missing data follows patterns
- **Duplicate Detection**: Check for duplicate rows or suspicious duplicates
- **Business Logic Validation**: Verify data makes business sense
- **Cross-Reference Validation**: Check relationships between columns

EVIDENCE-BASED ANALYSIS:
- Never assume - always verify with data
- Show calculations - don't just state conclusions
- Provide confidence levels - indicate uncertainty
- Cross-validate findings - use multiple methods
- Document assumptions - clearly state what you're assuming

OUTPUT REQUIREMENTS:
- All outputs truncated at 1000 characters
- Use ONLY textual summaries and numerical descriptions
- Provide specific data examples to support conclusions
- Include error detection findings in analysis
- Reach definitive, evidence-based conclusions
- Describe patterns, trends, and relationships in words

BEST PRACTICES:
- Include reasoning in code comments
- Document analysis approach and findings
- Build upon existing analysis
- Provide clear, actionable recommendations
- Always validate findings with multiple approaches
- Use descriptive statistics to paint a picture of the data

ANALYSIS COMPLETION CRITERIA:
Mark analysis as COMPLETE when ALL of the following are achieved:
1. ✓ Multi-table detection performed (empty row analysis to identify table boundaries)
2. ✓ Data quality assessment completed (missing data, duplicates, anomalies)
3. ✓ Statistical analysis performed (distributions, correlations, patterns)
4. ✓ Business logic validated (calculations, relationships, consistency)
5. ✓ Key findings documented in markdown cells
6. ✓ Actionable recommendations provided
7. ✓ Final comprehensive analysis report created in markdown cell with title "## 📊 Analysis Complete"

When these criteria are met, create a final comprehensive analysis report in a markdown cell:

**Report Structure Required:**
# 📊 Analysis Complete

## Executive Summary
- Brief overview of the analysis performed
- Most important findings in 2-3 sentences

## Data Overview
- Dataset characteristics (size, timeframe, scope)
- Multi-table detection results
- Data quality summary

## Key Findings
1. **Finding 1**: [Detailed description with supporting evidence]
2. **Finding 2**: [Detailed description with supporting evidence]
3. **Finding 3**: [Detailed description with supporting evidence]
(Include 3-5 major findings)

## Data Quality Issues
- Missing data patterns
- Anomalies detected
- Validation concerns

## Statistical Insights
- Key distributions and patterns
- Significant correlations
- Trend analysis results

## Business Implications
- What these findings mean for business operations
- Risk factors identified
- Opportunities discovered

## Recommendations
1. **Immediate Actions**: [What should be done right away]
2. **Short-term Improvements**: [1-3 month timeline]
3. **Long-term Considerations**: [Strategic recommendations]

## Technical Notes
- Analysis methodology used
- Assumptions made
- Limitations of the analysis

Then STOP the analysis - do not ask for further instructions."""
                                ),
                                HumanMessage(content=initial_prompt),
                            ]

                            # Wrap analysis in session tracking
                            with phoenix_session(
                                session_id=session_id,
                                user_id=None,  # Could be set from args or env
                            ):
                                # Add session metadata
                                add_session_metadata(
                                    session_id,
                                    {
                                        "excel_file": excel_path.name,
                                        "sheet_index": args.sheet_index,
                                        "sheet_name": sheet_name or f"Sheet {args.sheet_index}",
                                        "model": args.model,
                                        "max_rounds": args.max_rounds,
                                        "cost_limit": args.cost_limit if args.track_costs else None,
                                    },
                                )

                                for round_num in range(1, args.max_rounds + 1):
                                    logger.info(f"Starting analysis round {round_num}/{args.max_rounds}")

                                    # Log round start with clear visual delimiter
                                    llm_logger.info(f"\n{'🔄' * 40}")
                                    llm_logger.info(f"{'🔄' * 15} ROUND {round_num} - Starting Analysis {'🔄' * 15}")
                                    llm_logger.info(f"{'🔄' * 40}\n")

                                    # Log messages being sent to LLM
                                    llm_logger.info(f"{'═' * 20} Messages to LLM {'═' * 20}")
                                    for i, msg in enumerate(messages):
                                        msg_type = type(msg).__name__
                                        msg_content = getattr(msg, "content", str(msg))
                                        llm_logger.info(f"\nMessage {i + 1} ({msg_type}):\n{msg_content}")
                                        if hasattr(msg, "tool_calls") and msg.tool_calls:
                                            llm_logger.info(f"Tool calls: {json.dumps(msg.tool_calls, indent=2)}")

                                    if args.verbose:
                                        logger.info("Sending messages to LLM:", messages=messages)

                                    try:
                                        response = await llm_with_tools.ainvoke(messages)

                                        # Track token usage
                                        await track_llm_usage(response, args.model)

                                        # Log response from LLM
                                        llm_logger.info(f"\n{'═' * 20} LLM Response {'═' * 20}")
                                        llm_logger.info(f"Response type: {type(response).__name__}")
                                        llm_logger.info(f"Content: {response.content}")
                                        if hasattr(response, "tool_calls") and response.tool_calls:
                                            llm_logger.info(f"Tool calls: {json.dumps(response.tool_calls, indent=2)}")

                                    except Exception:
                                        logger.exception("Model API call failed")
                                        break

                                    if args.verbose:
                                        logger.info("Received response from LLM:", response=response)

                                    # Process tool calls
                                    tool_output_messages = []
                                    if response.tool_calls:
                                        llm_logger.info(f"\n{'═' * 20} Tool Executions {'═' * 20}")
                                        # Add the AI response with tool calls to the conversation first
                                        messages.append(response)

                                        for tool_call in response.tool_calls:
                                            tool_name = tool_call.get("name")
                                            tool_args = tool_call.get("args")
                                            logger.info(f"LLM called tool: {tool_name}", args=tool_args)

                                            # Dynamically call the tool function
                                            tool_func = next((t for t in tools if t.name == tool_name), None)
                                            if tool_func:
                                                try:
                                                    tool_output = await tool_func.ainvoke(tool_args)
                                                    logger.info(f"Tool output: {tool_output}")
                                                except Exception as tool_error:
                                                    # Log the error but continue analysis
                                                    logger.exception("Tool execution failed")
                                                    tool_output = (
                                                        f"Tool execution failed: {tool_error!s}. "
                                                        f"Continuing analysis with alternative approach. "
                                                        f"The analysis will proceed without this specific operation."
                                                    )
                                                    # Don't break the analysis loop - let the LLM adapt

                                                # Log tool call and output
                                                llm_logger.info(f"\nTOOL CALL: {tool_name}")
                                                llm_logger.info(f"Arguments: {json.dumps(tool_args, indent=2)}")
                                                llm_logger.info(f"Output: {tool_output}")

                                                tool_output_messages.append(
                                                    ToolMessage(content=str(tool_output), tool_call_id=tool_call["id"])
                                                )
                                            else:
                                                logger.warning(f"LLM tried to call unknown tool: {tool_name}")

                                        # Add tool results to conversation
                                        messages.extend(tool_output_messages)

                                    elif response.content:
                                        logger.info(f"LLM response: {response.content}")

                                        # Check if the response contains patterns indicating it's asking for user input
                                        forbidden_patterns = [
                                            "would you like me to",
                                            "let me know if",
                                            "do you need",
                                            "should i proceed",
                                            "would you prefer",
                                            "shall i continue",
                                            "feel free to ask",
                                            "if you'd like",
                                            "please let me know",
                                        ]

                                        response_lower = response.content.lower()
                                        if any(pattern in response_lower for pattern in forbidden_patterns):
                                            logger.warning(
                                                "LLM attempted to ask for user input - enforcing autonomous completion"
                                            )
                                            # Add a system message to remind the LLM to complete autonomously
                                            messages.append(response)
                                            messages.append(
                                                SystemMessage(
                                                    content="""
    REMINDER: You must complete the analysis autonomously.
    - Create a final comprehensive analysis report in markdown with "## 📊 Analysis Complete"
    - Follow the required report structure (Executive Summary, Data Overview, Key Findings, etc.)
    - Include all sections: findings, data quality, statistical insights, business implications, recommendations
    - Then STOP - do not ask for further instructions
    Complete the analysis now."""
                                                )
                                            )
                                            continue  # Continue to next round instead of breaking

                                        # Check if analysis is complete
                                        if (
                                            "analysis complete" in response_lower
                                            or "📊 analysis complete" in response_lower
                                        ):
                                            logger.info("Analysis marked as complete by LLM")
                                            # Log round completion
                                            llm_logger.info(f"\n{'✅' * 40}")
                                            llm_logger.info(f"{'✅' * 15} ROUND {round_num} Complete {'✅' * 15}")
                                            llm_logger.info(f"{'✅' * 40}\n")
                                            break

                                        # If no tool calls and not asking for input, the LLM is done
                                        # Log round completion
                                        llm_logger.info(f"\n{'✅' * 40}")
                                        llm_logger.info(f"{'✅' * 15} ROUND {round_num} Complete {'✅' * 15}")
                                        llm_logger.info(f"{'✅' * 40}\n")
                                        break
                                    else:
                                        logger.warning("LLM response was empty.")
                                        # Log round completion
                                        llm_logger.info(f"\n{'✅' * 40}")
                                        llm_logger.info(f"{'✅' * 15} ROUND {round_num} Complete {'✅' * 15}")
                                        llm_logger.info(f"{'✅' * 40}\n")
                                        break

                    except Exception:
                        logger.exception("Failed to initialize LLM")

            # Log cost summary if tracking enabled
            if args.track_costs:
                cost_summary = cost_tracker.get_summary()
                logger.info("\n💰 Cost Summary:")
                logger.info(f"  Total Cost: ${cost_summary['total_cost_usd']:.4f}")
                logger.info(f"  Total Tokens: {cost_summary['total_tokens']['total']:,}")
                if cost_summary["cost_by_model"]:
                    logger.info("  Cost by Model:")
                    for model, cost in cost_summary["cost_by_model"].items():
                        logger.info(f"    {model}: ${cost:.4f}")
                if args.cost_limit:
                    logger.info(
                        f"  Budget Status: {'✅ Within' if cost_summary['within_budget'] else '❌ Exceeded'} limit (${args.cost_limit:.2f})"
                    )

            # Ensure notebook is saved at the end
            logger.info("Analysis complete. Saving notebook...")
            save_result = session.toolkit.save_notebook(notebook_path, overwrite=True)
            if save_result.is_ok():
                logger.info(f"✅ Notebook saved successfully to: {save_result.ok_value}")
            else:
                logger.error(f"❌ Failed to save notebook: {save_result.err_value}")

            logger.info("Analysis session completed.")

    def run(self):
        """Parse arguments and run the analysis."""
        args = self.parser.parse_args()
        # Setup basic logging
        log_level = logging.INFO if args.verbose else logging.WARNING
        logging.basicConfig(level=log_level, stream=sys.stdout)

        # Give a more specific logger name
        global logger
        logger = get_logger(f"notebook_cli.{args.model}")

        asyncio.run(self.run_analysis(args))


if __name__ == "__main__":
    NotebookCLI().run()
