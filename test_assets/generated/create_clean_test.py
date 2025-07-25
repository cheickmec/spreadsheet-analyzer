"""Create a clean test Excel file without namespace issues."""

from pathlib import Path

import openpyxl

# Create test directory if it doesn't exist
test_dir = Path(__file__).parent
test_dir.mkdir(exist_ok=True)

# Create a simple workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Data"

# Add simple data without any external references
ws["A1"] = "Name"
ws["B1"] = "Value"
ws["C1"] = "Status"

# Add data rows
data = [
    ("Item 1", 100, "Active"),
    ("Item 2", 200, "Active"),
    ("Item 3", 150, "Pending"),
    ("Item 4", 75, "Inactive"),
]

for i, (name, value, status) in enumerate(data, start=2):
    ws[f"A{i}"] = name
    ws[f"B{i}"] = value
    ws[f"C{i}"] = status

# Add a simple sum formula
ws["B6"] = "=SUM(B2:B5)"
ws["A6"] = "Total"

# Save the file
output_file = test_dir / "clean_test.xlsx"
wb.save(output_file)

print(f"Clean test file created: {output_file}")
