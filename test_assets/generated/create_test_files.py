#!/usr/bin/env python3
"""Create comprehensive test data files for CLI testing.

This script generates realistic test files that cover different spreadsheet
analysis scenarios as outlined in CLI_REDESIGN_PLAN.md.
"""

import random
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def create_simple_sales_xlsx() -> None:
    """Create simple_sales.xlsx with multi-sheet sales data."""
    print("Creating simple_sales.xlsx...")

    # Sheet 1: Monthly Sales Data
    monthly_data = {
        "Month": ["Jan 2024", "Feb 2024", "Mar 2024", "Apr 2024", "May 2024", "Jun 2024"],
        "Product_A_Sales": [15000, 18000, 22000, 19000, 25000, 23000],
        "Product_B_Sales": [12000, 14000, 16000, 18000, 20000, 22000],
        "Product_C_Sales": [8000, 9500, 11000, 12500, 14000, 15500],
        "Total_Sales": [35000, 41500, 49000, 49500, 59000, 60500],
        "Returns": [1750, 2075, 2450, 2475, 2950, 3025],
        "Net_Sales": [33250, 39425, 46550, 47025, 56050, 57475],
    }

    # Sheet 2: Regional Breakdown
    regional_data = {
        "Region": ["North", "South", "East", "West", "Central"],
        "Q1_Sales": [85000, 92000, 78000, 88000, 67000],
        "Q2_Sales": [91000, 98000, 82000, 94000, 72000],
        "Sales_Growth": [7.1, 6.5, 5.1, 6.8, 7.5],
        "Market_Share": [23.2, 25.1, 21.3, 24.0, 18.3],
        "Target_Achievement": [105.2, 98.7, 112.4, 101.3, 89.6],
    }

    # Sheet 3: Product Performance
    product_data = {
        "Product_ID": ["PRD-001", "PRD-002", "PRD-003", "PRD-004", "PRD-005"],
        "Product_Name": ["Widget Pro", "Super Gadget", "Basic Tool", "Premium Kit", "Standard Set"],
        "Unit_Price": [49.99, 129.99, 19.99, 299.99, 89.99],
        "Units_Sold": [1200, 450, 2100, 180, 680],
        "Revenue": [59988, 58496, 41979, 53998, 61193],
        "Cost_Per_Unit": [25.00, 75.00, 12.00, 180.00, 50.00],
        "Profit_Margin": [49.9, 42.3, 40.0, 40.0, 44.4],
    }

    # Create workbook with multiple sheets
    wb = Workbook()

    # Monthly Sales sheet
    ws1 = wb.active
    ws1.title = "Monthly Sales"
    df1 = pd.DataFrame(monthly_data)
    for r in dataframe_to_rows(df1, index=False, header=True):
        ws1.append(r)

    # Regional Breakdown sheet
    ws2 = wb.create_sheet("Regional Breakdown")
    df2 = pd.DataFrame(regional_data)
    for r in dataframe_to_rows(df2, index=False, header=True):
        ws2.append(r)

    # Product Performance sheet
    ws3 = wb.create_sheet("Product Performance")
    df3 = pd.DataFrame(product_data)
    for r in dataframe_to_rows(df3, index=False, header=True):
        ws3.append(r)

    # Add some basic formatting
    for ws in [ws1, ws2, ws3]:
        # Header row formatting
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

    wb.save("simple_sales.xlsx")
    print("✅ Created simple_sales.xlsx")


def create_financial_model_xlsx() -> None:
    """Create financial_model.xlsx with complex formulas."""
    print("Creating financial_model.xlsx...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Model"

    # Headers
    headers = ["Item", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Revenue assumptions
    ws.cell(row=2, column=1, value="Revenue")
    ws.cell(row=3, column=1, value="Base Revenue")
    ws.cell(row=3, column=2, value=1000000)  # Year 1 base

    # Growth rate assumption
    ws.cell(row=4, column=1, value="Growth Rate")
    ws.cell(row=4, column=2, value=0.15)  # 15% growth

    # Calculate revenue for subsequent years using formulas
    for col in range(3, 7):  # Years 2-5
        # Revenue grows by growth rate each year
        ws.cell(row=3, column=col, value=f"=B3*(1+$B$4)^{col - 2}")

    # Cost structure
    ws.cell(row=6, column=1, value="Costs")
    ws.cell(row=7, column=1, value="Variable Costs (% of Revenue)")
    ws.cell(row=7, column=2, value=0.40)  # 40% of revenue

    ws.cell(row=8, column=1, value="Fixed Costs")
    ws.cell(row=8, column=2, value=200000)  # Year 1 fixed costs

    ws.cell(row=9, column=1, value="Fixed Cost Inflation")
    ws.cell(row=9, column=2, value=0.03)  # 3% annual inflation

    # Calculate variable costs (formulas)
    for col in range(2, 7):  # Years 1-5
        ws.cell(row=10, column=1, value="Total Variable Costs")
        ws.cell(row=10, column=col, value="=B3*$B$7" if col == 2 else f"={chr(65 + col)}3*$B$7")

    # Calculate fixed costs with inflation (formulas)
    for col in range(2, 7):  # Years 1-5
        ws.cell(row=11, column=1, value="Total Fixed Costs")
        if col == 2:
            ws.cell(row=11, column=col, value="=B8")
        else:
            ws.cell(row=11, column=col, value=f"=B8*(1+$B$9)^{col - 2}")

    # Total costs
    for col in range(2, 7):  # Years 1-5
        ws.cell(row=12, column=1, value="Total Costs")
        ws.cell(row=12, column=col, value=f"={chr(65 + col)}10+{chr(65 + col)}11")

    # EBITDA calculation
    for col in range(2, 7):  # Years 1-5
        ws.cell(row=14, column=1, value="EBITDA")
        ws.cell(row=14, column=col, value=f"={chr(65 + col)}3-{chr(65 + col)}12")

    # Depreciation
    ws.cell(row=15, column=1, value="Depreciation")
    for col in range(2, 7):
        ws.cell(row=15, column=col, value=50000)  # Fixed depreciation

    # EBIT calculation
    for col in range(2, 7):  # Years 1-5
        ws.cell(row=16, column=1, value="EBIT")
        ws.cell(row=16, column=col, value=f"={chr(65 + col)}14-{chr(65 + col)}15")

    # Tax rate
    ws.cell(row=17, column=1, value="Tax Rate")
    ws.cell(row=17, column=2, value=0.25)  # 25% tax rate

    # Tax calculation
    for col in range(2, 7):  # Years 1-5
        ws.cell(row=18, column=1, value="Taxes")
        ws.cell(row=18, column=col, value=f"=MAX(0,{chr(65 + col)}16*$B$17)")

    # Net Income
    for col in range(2, 7):  # Years 1-5
        ws.cell(row=19, column=1, value="Net Income")
        ws.cell(row=19, column=col, value=f"={chr(65 + col)}16-{chr(65 + col)}18")

    # Cash Flow calculation
    for col in range(2, 7):  # Years 1-5
        ws.cell(row=21, column=1, value="Operating Cash Flow")
        ws.cell(row=21, column=col, value=f"={chr(65 + col)}19+{chr(65 + col)}15")

    # NPV calculation assumptions
    ws.cell(row=23, column=1, value="Discount Rate")
    ws.cell(row=23, column=2, value=0.10)  # 10% discount rate

    # Present value calculations
    for col in range(2, 7):  # Years 1-5
        ws.cell(row=24, column=1, value="PV of Cash Flow")
        ws.cell(row=24, column=col, value=f"={chr(65 + col)}21/(1+$B$23)^{col - 1}")

    # Total NPV
    ws.cell(row=25, column=1, value="Total NPV")
    ws.cell(row=25, column=2, value="=SUM(B24:F24)")

    wb.save("financial_model.xlsx")
    print("✅ Created financial_model.xlsx")


def create_inventory_tracking_csv() -> None:
    """Create inventory_tracking.csv with typical inventory data."""
    print("Creating inventory_tracking.csv...")

    # Generate realistic inventory data
    np.random.seed(42)  # For reproducible data

    categories = ["Electronics", "Clothing", "Home & Garden", "Sports", "Books"]
    statuses = ["In Stock", "Low Stock", "Out of Stock", "Discontinued"]
    suppliers = ["Supplier A", "Supplier B", "Supplier C", "Supplier D"]

    data = []
    for i in range(500):  # 500 inventory items
        item_id = f"INV-{i + 1:04d}"
        category = random.choice(categories)

        # Generate correlated data
        base_price = random.uniform(5, 500)
        quantity = random.randint(0, 1000)

        # Create some realistic patterns
        if quantity == 0:
            status = "Out of Stock"
        elif quantity < 20:
            status = "Low Stock"
        elif random.random() < 0.05:
            status = "Discontinued"
        else:
            status = "In Stock"

        # Last restocked date
        days_ago = random.randint(1, 365)
        last_restocked = (datetime.now() - timedelta(days=days_ago)).strftime("%Y-%m-%d")

        # Generate some data quality issues intentionally
        if random.random() < 0.02:  # 2% missing data
            supplier = None
        else:
            supplier = random.choice(suppliers)

        if random.random() < 0.01:  # 1% negative quantities (data error)
            quantity = -quantity

        data.append(
            {
                "Item_ID": item_id,
                "Item_Name": f"Product {i + 1}",
                "Category": category,
                "Quantity_On_Hand": quantity,
                "Unit_Price": round(base_price, 2),
                "Total_Value": round(quantity * base_price, 2),
                "Supplier": supplier,
                "Last_Restocked": last_restocked,
                "Status": status,
                "Reorder_Point": random.randint(10, 50),
                "Max_Stock_Level": random.randint(100, 1000),
            }
        )

    df = pd.DataFrame(data)
    df.to_csv("inventory_tracking.csv", index=False)
    print("✅ Created inventory_tracking.csv")


def create_employee_records_xlsx() -> None:
    """Create employee_records.xlsx with data quality issues."""
    print("Creating employee_records.xlsx...")

    np.random.seed(42)

    departments = ["Sales", "Engineering", "Marketing", "HR", "Finance", "Operations"]
    titles = ["Manager", "Senior", "Junior", "Lead", "Director", "Associate"]
    locations = ["New York", "San Francisco", "Chicago", "Austin", "Remote"]

    data = []
    for i in range(200):  # 200 employee records
        emp_id = f"EMP-{i + 1:04d}"

        # Generate name with potential duplicates/issues
        first_names = ["John", "Jane", "Mike", "Sarah", "David", "Lisa", "Chris", "Amy"]
        last_names = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller"]

        first_name = random.choice(first_names)
        last_name = random.choice(last_names)

        # Introduce some data quality issues
        if random.random() < 0.03:  # 3% name formatting issues
            first_name = first_name.lower()  # inconsistent case
        if random.random() < 0.02:  # 2% extra spaces
            last_name = f" {last_name} "

        # Email generation with some issues
        email = f"{first_name.lower().strip()}.{last_name.lower().strip()}@company.com"
        if random.random() < 0.01:  # 1% malformed emails
            email = email.replace("@", "")

        # Salary with some outliers and missing data
        base_salary = random.randint(40000, 150000)
        if random.random() < 0.02:  # 2% unrealistic salaries
            base_salary = random.choice([999999, 1, 0])
        if random.random() < 0.01:  # 1% missing salaries
            base_salary = None

        # Hire date with some inconsistencies
        days_ago = random.randint(30, 3650)  # 30 days to 10 years ago
        hire_date = datetime.now() - timedelta(days=days_ago)

        if random.random() < 0.01:  # 1% future hire dates (data error)
            hire_date = datetime.now() + timedelta(days=random.randint(1, 365))

        # Department and title
        department = random.choice(departments)
        title = f"{random.choice(titles)} {department.rstrip('s')}"

        # Performance rating with some missing values
        performance = random.choice([1, 2, 3, 4, 5])
        if random.random() < 0.05:  # 5% missing performance data
            performance = None

        data.append(
            {
                "Employee_ID": emp_id,
                "First_Name": first_name,
                "Last_Name": last_name,
                "Email": email,
                "Department": department,
                "Job_Title": title,
                "Hire_Date": hire_date.strftime("%Y-%m-%d"),
                "Salary": base_salary,
                "Location": random.choice(locations),
                "Performance_Rating": performance,
                "Is_Active": random.choice([True, False]) if random.random() < 0.1 else True,
                "Manager_ID": f"EMP-{random.randint(1, 50):04d}" if random.random() < 0.8 else None,
            }
        )

    # Create DataFrame and add some duplicate records (data quality issue)
    df = pd.DataFrame(data)

    # Add 5 duplicate records
    duplicates = df.sample(5).copy()
    df = pd.concat([df, duplicates], ignore_index=True)

    # Save to Excel
    df.to_excel("employee_records.xlsx", index=False)
    print("✅ Created employee_records.xlsx")


def main() -> None:
    """Create all test data files."""
    # Files are created in the current directory (test_assets/generated/)

    print("🏗️  Creating comprehensive test data files...")
    print()

    try:
        create_simple_sales_xlsx()
        create_financial_model_xlsx()
        create_inventory_tracking_csv()
        create_employee_records_xlsx()

        print()
        print("✅ All test data files created successfully!")
        print()
        print("📊 Test files created:")
        print("  - simple_sales.xlsx (multi-sheet sales data)")
        print("  - financial_model.xlsx (complex formulas)")
        print("  - inventory_tracking.csv (CSV data)")
        print("  - employee_records.xlsx (data quality issues)")

    except Exception as e:
        print(f"❌ Error creating test files: {e}")
        raise


if __name__ == "__main__":
    main()
