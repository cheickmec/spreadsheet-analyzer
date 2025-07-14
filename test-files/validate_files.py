#!/usr/bin/env python3
"""
Validation script for downloaded Excel test files.
Tests that files can be opened and basic structure can be analyzed.
"""

import json
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed. Run: uv add openpyxl")
    sys.exit(1)


def validate_excel_file(file_path: Path) -> dict[str, Any]:
    """
    Validate an Excel file and extract basic structure information.

    Args:
        file_path: Path to the Excel file

    Returns:
        Dictionary with validation results and basic file info
    """
    result = {
        "file_name": file_path.name,
        "file_size_kb": round(file_path.stat().st_size / 1024, 1),
        "valid": False,
        "error": None,
        "sheets": [],
        "total_sheets": 0,
        "has_formulas": False,
        "has_charts": False,
        "has_named_ranges": False,
        "has_macros": False,
    }

    try:
        # Attempt to load the workbook
        wb = load_workbook(file_path, read_only=True, keep_vba=True, data_only=False)
        result["valid"] = True
        result["total_sheets"] = len(wb.sheetnames)
        result["sheets"] = wb.sheetnames[:10]  # First 10 sheet names

        # Check for named ranges
        if wb.defined_names:
            result["has_named_ranges"] = True

        # Check for macros (VBA)
        if hasattr(wb, "vba_archive") and wb.vba_archive:
            result["has_macros"] = True

        # Sample first sheet for formulas and charts
        if wb.sheetnames:
            ws = wb[wb.sheetnames[0]]

            # Check for formulas (sample first 100 cells)
            cell_count = 0
            cell_sample_limit = 100
            for row in ws.iter_rows(max_row=10, max_col=10):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        result["has_formulas"] = True
                        break
                    cell_count += 1
                    if cell_count > cell_sample_limit:
                        break
                if result["has_formulas"]:
                    break

            # Check for charts (using public API if available)
            # Note: Chart detection may not work in read-only mode
            try:
                if hasattr(ws, "_charts") and ws._charts:  # noqa: SLF001
                    result["has_charts"] = True
            except AttributeError:
                pass

        wb.close()

    except (OSError, ValueError) as e:
        result["error"] = str(e)
        result["valid"] = False

    return result


def main():
    """Main validation function."""
    print("🔍 Validating downloaded Excel test files...\n")

    # Find all Excel files in test-files directory
    test_files_dir = Path(__file__).parent
    excel_files: list[Path] = []

    # Scan all subdirectories for Excel files
    for category_dir in test_files_dir.iterdir():
        if category_dir.is_dir() and not category_dir.name.startswith("."):
            excel_files.extend(category_dir.glob("*.xl*"))

    if not excel_files:
        print("❌ No Excel files found in test-files directory")
        return

    print(f"Found {len(excel_files)} Excel files to validate:\n")

    validation_results = []
    valid_count = 0

    for file_path in excel_files:
        print(f"Validating: {file_path.relative_to(test_files_dir)}")
        result = validate_excel_file(file_path)
        validation_results.append(result)

        if result["valid"]:
            valid_count += 1
            print(f"  ✅ Valid - {result['total_sheets']} sheets, {result['file_size_kb']}KB")
            if result["has_formulas"]:
                print("     📋 Contains formulas")
            if result["has_charts"]:
                print("     📊 Contains charts")
            if result["has_named_ranges"]:
                print("     🏷️  Contains named ranges")
            if result["has_macros"]:
                print("     🔧 Contains VBA macros")
        else:
            print(f"  ❌ Invalid - {result['error']}")
        print()

    # Summary
    print("📊 Validation Summary:")
    print(f"   Total files: {len(excel_files)}")
    print(f"   Valid files: {valid_count}")
    print(f"   Invalid files: {len(excel_files) - valid_count}")
    print(f"   Success rate: {(valid_count / len(excel_files) * 100):.1f}%")

    # Save validation results
    results_file = test_files_dir / "validation_results.json"
    with results_file.open("w") as f:
        json.dump(
            {
                "validation_date": "2024-07-14",
                "total_files": len(excel_files),
                "valid_files": valid_count,
                "success_rate": round(valid_count / len(excel_files) * 100, 1),
                "files": validation_results,
            },
            f,
            indent=2,
        )

    print(f"\n💾 Detailed results saved to: {results_file}")


if __name__ == "__main__":
    main()
