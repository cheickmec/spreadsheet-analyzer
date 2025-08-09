#!/usr/bin/env python3
"""
Sheet-Cartographer Agent (OpenAI Implementation) - Version 0.11a

LLM-first decide-or-split conflict resolution:
- Removes rule-based relabels in S3.5
- Adds formulas-view peek, numeric consistency evidence, and LLM-driven keep/relabel/split

Provider: OpenAI
Models Supported: gpt-4o, gpt-4o-mini, gpt-4-turbo, gpt-3.5-turbo
Author: Sheet-Cartographer v0.11a (OpenAI)
Date: 2025-08-09
"""

import argparse
import json
import sys
import time
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Final, Literal

import numpy as np
from scipy import stats

# Add parent directory for utils import
sys.path.append(str(Path(__file__).parent))
# Import Excel and OpenAI libraries
import openpyxl
from openai import OpenAI
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.worksheet import Worksheet

# Import new semantic classification system
from semantic_classification import DataCharacteristics, StructuralType
from utils import ExperimentLogger

# =========================
# Constants & thresholds
# =========================
DEFAULT_MODEL: Final[str] = "gpt-4o-mini"
MAX_WINDOW_HEIGHT: Final[int] = 50
MAX_WINDOW_WIDTH: Final[int] = 50
MAX_CELL_META_CALLS: Final[int] = 140
MAX_RANGE_AREA: Final[int] = 10000
MAX_LLM_TOKENS: Final[int] = 32000
MAX_RUNTIME_SECONDS: Final[int] = 30
DENSITY_THRESHOLD: Final[float] = 0.05
HIGH_DENSITY_THRESHOLD: Final[float] = 0.8
BLANK_TAIL_THRESHOLD: Final[float] = 0.3
HOMOGENEITY_THRESHOLD: Final[float] = 0.8

# Heuristic thresholds (now used only to *trigger* LLM review; not to override)
HDR_MAX_FORMULA_DENSITY: Final[float] = 0.05
HDR_MAX_NUMERIC_RATIO: Final[float] = 0.30
AGG_MIN_NUMERIC_RATIO: Final[float] = 0.30
AGG_MIN_FORMULA_DENSITY: Final[float] = 0.08
AGG_KEYWORDS = ("total", "subtotal", "grand total", "kpi", "sum", "balance")

# Confidence presets
FACT_TABLE_CONFIDENCE: Final[float] = 0.75
HEADER_BANNER_CONFIDENCE: Final[float] = 0.70
KPI_SUMMARY_CONFIDENCE: Final[float] = 0.75
PIVOT_CACHE_CONFIDENCE: Final[float] = 0.60
CHART_ANCHOR_CONFIDENCE: Final[float] = 1.00
SUMMARY_CONFIDENCE: Final[float] = 0.80
UNKNOWN_CONFIDENCE: Final[float] = 0.20

# Type mapping from string to enum
TYPE_MAP: Final[dict[str, StructuralType]] = {
    "DataTable": StructuralType.DATA_TABLE,
    "HeaderZone": StructuralType.HEADER_ZONE,
    "AggregationZone": StructuralType.AGGREGATION_ZONE,
    "MetadataZone": StructuralType.METADATA_ZONE,
    "VisualizationAnchor": StructuralType.VISUALIZATION_ANCHOR,
    "FormInputZone": StructuralType.FORM_INPUT_ZONE,
    "NavigationZone": StructuralType.NAVIGATION_ZONE,
    "UnstructuredText": StructuralType.UNSTRUCTURED_TEXT,
    "EmptyPadding": StructuralType.EMPTY_PADDING,
}


@dataclass
class SheetInfo:
    rows: int
    cols: int
    used_range: str
    hidden_rows: list[int] = field(default_factory=list)
    hidden_cols: list[int] = field(default_factory=list)


@dataclass
class CellMeta:
    value_type: Literal["num", "text", "date", "blank", "error"]
    is_formula: bool
    is_merged: bool
    style_flags: list[str] = field(default_factory=list)


@dataclass
class Block:
    id: str
    range: str
    structural_type: StructuralType
    semantic_description: str
    confidence: float
    classification_reasoning: str = ""
    data_characteristics: DataCharacteristics = field(default_factory=DataCharacteristics)
    suggested_operations: list[str] = field(default_factory=list)
    named_range: str | None = None


@dataclass
class ChartObject:
    id: str
    anchor: str
    type: str
    linked_block: str | None = None


class ExcelToolHandler:
    """Handles tool execution for Excel operations."""

    def __init__(self, worksheet: Worksheet, logger: ExperimentLogger, excel_path: Path):
        self.worksheet = worksheet
        self.logger = logger
        self.excel_path = excel_path
        self.tool_call_count = defaultdict(int)
        self.total_tokens_used = 0

    def execute_tool(self, tool_name: str, arguments: dict[str, Any]) -> Any:
        self.tool_call_count[tool_name] += 1
        self.logger.main.debug(f"🔧 Executing tool: {tool_name} with args: {arguments}")

        if tool_name == "sheet_info":
            return self._sheet_info()
        elif tool_name == "window_grab":
            return self._window_grab(**arguments)
        elif tool_name == "range_peek":
            return self._range_peek(**arguments)
        elif tool_name == "range_formulas":
            return self._range_formulas(**arguments)
        elif tool_name == "cell_meta":
            return self._cell_meta(**arguments)
        elif tool_name == "merged_list":
            return self._merged_list()
        elif tool_name == "names_list":
            return self._names_list()
        elif tool_name == "object_scan":
            return self._object_scan(**arguments)
        else:
            raise ValueError(f"Unknown tool: {tool_name}")

    def _sheet_info(self) -> dict[str, Any]:
        """Get sheet dimensions and metadata. Guard read-only limitations."""
        min_row = self.worksheet.min_row or 1
        max_row = self.worksheet.max_row or 1
        min_col = self.worksheet.min_column or 1
        max_col = self.worksheet.max_column or 1
        used_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

        # ReadOnlyWorksheet lacks row_dimensions/column_dimensions; guard it.
        hidden_rows = []
        hidden_cols = []
        try:
            if hasattr(self.worksheet, "row_dimensions") and isinstance(self.worksheet.row_dimensions, dict):
                hidden_rows = [idx for idx, rd in self.worksheet.row_dimensions.items() if getattr(rd, "hidden", False)]
        except Exception as e:
            self.logger.main.debug(f"Hidden rows unavailable in read-only mode: {e}")

        try:
            if hasattr(self.worksheet, "column_dimensions") and isinstance(self.worksheet.column_dimensions, dict):
                hidden_cols = [
                    column_index_from_string(col_letter)
                    for col_letter, cd in self.worksheet.column_dimensions.items()
                    if getattr(cd, "hidden", False)
                ]
        except Exception as e:
            self.logger.main.debug(f"Hidden cols unavailable in read-only mode: {e}")

        return {
            "rows": max_row,
            "cols": max_col,
            "usedRange": used_range,
            "hiddenRows": hidden_rows,
            "hiddenCols": hidden_cols,
        }

    def _safe_str(self, value: Any, max_chars: int = 25) -> str:
        if value is None:
            return ""
        s = str(value)
        if len(s) <= max_chars:
            return s
        if len(s) <= 80:
            return s[:max_chars] + "↯"
        return f"{s[:max_chars]}↯{s[-10:]}"

    def _window_grab(self, top: int, left: int, height: int, width: int, format: str = "markdown") -> str:
        height = min(height, MAX_WINDOW_HEIGHT)
        width = min(width, MAX_WINDOW_WIDTH)

        lines = [f"# viewport: top={top} left={left} height={height} width={width}"]

        col_headers = ["# cols:"]
        for col in range(left, left + width):
            col_headers.append(f"{get_column_letter(col):>8}")
        lines.append(" | ".join(col_headers))

        for row in range(top, top + height):
            row_data = [f"{row:3d}"]
            for col in range(left, left + width):
                cell = self.worksheet.cell(row=row, column=col)
                str_value = self._safe_str(cell.value)
                if str_value == "":
                    str_value = "␀"
                row_data.append(f"{str_value:>8}")
            lines.append(" | ".join(row_data))

        return "\n".join(lines)

    def _range_peek(self, range: str, format: str = "markdown") -> str:
        parts = range.split(":")
        if len(parts) != 2:
            raise ValueError(f"Invalid range: {range}")

        start_col, start_row = coordinate_from_string(parts[0])
        end_col, end_row = coordinate_from_string(parts[1])
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)

        area = (end_row - start_row + 1) * (end_col_idx - start_col_idx + 1)
        if area > MAX_RANGE_AREA:
            raise ValueError(f"Range area {area} exceeds maximum {MAX_RANGE_AREA}")

        return self._window_grab(
            top=start_row,
            left=start_col_idx,
            height=end_row - start_row + 1,
            width=end_col_idx - start_col_idx + 1,
            format=format,
        )

    def _ensure_formula_workbook(self):
        """Open a separate handle with data_only=False for formula access."""
        if not hasattr(self, "_formula_workbook"):
            # Using non-read-only to access full features and formulas
            self._formula_workbook = openpyxl.load_workbook(self.excel_path, read_only=False, data_only=False)
        return self._formula_workbook

    def _range_formulas(self, range: str, format: str = "markdown") -> str:
        """Peek a specific range but rendering formulas where present."""
        parts = range.split(":")
        if len(parts) != 2:
            raise ValueError(f"Invalid range: {range}")

        start_col, start_row = coordinate_from_string(parts[0])
        end_col, end_row = coordinate_from_string(parts[1])
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)

        area = (end_row - start_row + 1) * (end_col_idx - start_col_idx + 1)
        if area > MAX_RANGE_AREA:
            raise ValueError(f"Range area {area} exceeds maximum {MAX_RANGE_AREA}")

        # Prepare header
        height = min(end_row - start_row + 1, MAX_WINDOW_HEIGHT)
        width = min(end_col_idx - start_col_idx + 1, MAX_WINDOW_WIDTH)
        lines = [f"# viewport(formulas): top={start_row} left={start_col_idx} height={height} width={width}"]
        col_headers = ["# cols:"]
        for col in range(start_col_idx, start_col_idx + width):
            col_headers.append(f"{get_column_letter(col):>8}")
        lines.append(" | ".join(col_headers))

        # Access formulas workbook
        try:
            fwb = self._ensure_formula_workbook()
            sheet_index = self.worksheet.parent.worksheets.index(self.worksheet)
            fws = fwb.worksheets[sheet_index]
        except Exception as e:
            self.logger.main.debug(f"Failed to open formula workbook: {e}")
            # Fallback to values view
            return self._range_peek(range, format=format)

        # Render rows
        for r in range(start_row, start_row + height):
            row_data = [f"{r:3d}"]
            for c in range(start_col_idx, start_col_idx + width):
                fcell = fws.cell(row=r, column=c)
                val = fcell.value
                s: str = ""
                if isinstance(val, str) and val.startswith("="):
                    s = val[:80] + ("↯" if len(val) > 80 else "")
                else:
                    s = self._safe_str(val)
                if s == "":
                    s = "␀"
                row_data.append(f"{s:>8}")
            lines.append(" | ".join(row_data))

        return "\n".join(lines)

    def _cell_meta(self, row: int, col: int) -> dict[str, Any]:
        if self.tool_call_count["cell_meta"] > MAX_CELL_META_CALLS:
            raise ValueError(f"Exceeded maximum cell_meta calls ({MAX_CELL_META_CALLS})")

        cell = self.worksheet.cell(row=row, column=col)

        value_type = "blank"
        if cell.value is not None:
            if isinstance(cell.value, (int, float)):
                value_type = "num"
            elif isinstance(cell.value, str):
                if cell.value.startswith("="):
                    value_type = "num"
                else:
                    value_type = "text"
            elif hasattr(cell.value, "date"):
                value_type = "date"
            elif cell.value == "#ERROR":
                value_type = "error"

        is_formula = False
        try:
            fwb = self._ensure_formula_workbook()
            sheet_index = self.worksheet.parent.worksheets.index(self.worksheet)
            formula_worksheet = fwb.worksheets[sheet_index]
            fcell = formula_worksheet.cell(row=row, column=col)
            if hasattr(fcell, "value") and isinstance(fcell.value, str) and fcell.value.startswith("="):
                is_formula = True
        except Exception as e:
            self.logger.main.debug(f"Formula detection fallback for {row},{col}: {e}")
            if isinstance(cell.value, str) and cell.value.startswith("="):
                is_formula = True

        is_merged = False
        if hasattr(self.worksheet, "merged_cells"):
            for merged_range in self.worksheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    is_merged = True
                    break

        style_flags = []
        if cell.font and cell.font.bold:
            style_flags.append("bold")
        if cell.alignment and cell.alignment.horizontal == "center":
            style_flags.append("center")

        return {"valueType": value_type, "isFormula": is_formula, "isMerged": is_merged, "styleFlags": style_flags}

    def _merged_list(self) -> list[dict[str, str]]:
        merged = []
        if hasattr(self.worksheet, "merged_cells"):
            for merged_range in self.worksheet.merged_cells.ranges:
                merged.append({"range": str(merged_range)})
                if len(merged) >= 2000:
                    break
        return merged

    def _names_list(self) -> list[dict[str, str]]:
        """Use DefinedName.destinations for robust named range resolution."""
        named: list[dict[str, str]] = []
        try:
            dn_table = self.worksheet.parent.defined_names
            for name in dn_table:
                dn = dn_table[name]
                for sheet_name, coord in dn.destinations:
                    named.append({"name": name, "sheet": sheet_name, "range": coord})
                    if len(named) >= 1000:
                        return named
        except Exception as e:
            self.logger.main.debug(f"Could not retrieve named ranges: {e}")
        return named

    def _object_scan(self, type: str) -> list[dict[str, Any]]:
        objects = []
        if type == "chart":
            try:
                if hasattr(self.worksheet, "_charts"):
                    for chart in self.worksheet._charts:
                        anchor = "A1"
                        if hasattr(chart, "anchor"):
                            if hasattr(chart.anchor, "_from"):
                                fm = chart.anchor._from
                                if hasattr(fm, "col") and hasattr(fm, "row"):
                                    col_letter = get_column_letter(fm.col + 1)
                                    anchor = f"{col_letter}{fm.row + 1}"
                            elif hasattr(chart.anchor, "ref"):
                                anchor = (
                                    str(chart.anchor.ref).split(":")[0]
                                    if ":" in str(chart.anchor.ref)
                                    else str(chart.anchor.ref)
                                )
                        objects.append(
                            {
                                "type": "chart",
                                "anchor": anchor,
                                "title": chart.title if hasattr(chart, "title") else "Untitled Chart",
                            }
                        )
                        if len(objects) >= 50:
                            break
            except Exception as e:
                self.logger.main.debug(f"Error scanning for charts: {e}")
        elif type == "pivot":
            try:
                if hasattr(self.worksheet.parent, "pivots"):
                    for pivot in self.worksheet.parent.pivots:
                        objects.append(
                            {
                                "type": "pivot",
                                "anchor": "A1",
                                "name": pivot.name if hasattr(pivot, "name") else "Pivot Table",
                            }
                        )
                        if len(objects) >= 20:
                            break
            except Exception as e:
                self.logger.main.debug(f"Error scanning for pivots: {e}")
        return objects

    def calculate_layout_homogeneity(self, range_text: str) -> float:
        lines = range_text.split("\n")[2:]
        if not lines:
            return 1.0
        data_matrix = []
        for line in lines:
            cells = line.split("|")[1:]
            if cells:
                data_matrix.append([cell.strip() for cell in cells])
        if not data_matrix:
            return 1.0

        first_row = data_matrix[0] if data_matrix else []
        if first_row:
            consecutive_count = 1
            max_consecutive = 1
            for i in range(1, len(first_row)):
                if first_row[i] == first_row[i - 1] and first_row[i] not in ["", "␀"]:
                    consecutive_count += 1
                    max_consecutive = max(max_consecutive, consecutive_count)
                else:
                    consecutive_count = 1
            if max_consecutive > len(first_row) * 0.5:
                return 0.3

        row_densities = []
        for row in data_matrix:
            non_empty = sum(1 for cell in row if cell and cell != "␀")
            density = non_empty / len(row) if row else 0
            row_densities.append(density)

        num_cols = max(len(row) for row in data_matrix) if data_matrix else 0
        col_densities = []
        for col_idx in range(num_cols):
            non_empty = 0
            total = 0
            for row in data_matrix:
                if col_idx < len(row):
                    total += 1
                    if row[col_idx] and row[col_idx] != "␀":
                        non_empty += 1
            density = non_empty / total if total > 0 else 0
            col_densities.append(density)

        row_homogeneity = 1.0
        if len(set(row_densities)) > 1:
            bins = np.linspace(0, 1, 11)
            digitized = np.digitize(row_densities, bins)
            _, counts = np.unique(digitized, return_counts=True)
            probabilities = counts / len(digitized)
            entropy = stats.entropy(probabilities, base=2)
            max_entropy = np.log2(len(bins))
            row_homogeneity = 1.0 - (entropy / max_entropy if max_entropy > 0 else 0)

        col_homogeneity = 1.0
        if len(set(col_densities)) > 1:
            bins = np.linspace(0, 1, 11)
            digitized = np.digitize(col_densities, bins)
            _, counts = np.unique(digitized, return_counts=True)
            probabilities = counts / len(digitized)
            entropy = stats.entropy(probabilities, base=2)
            max_entropy = np.log2(len(bins))
            col_homogeneity = 1.0 - (entropy / max_entropy if max_entropy > 0 else 0)

        homogeneity = 0.7 * row_homogeneity + 0.3 * col_homogeneity
        return homogeneity

    def close(self):
        try:
            if hasattr(self, "_formula_workbook"):
                self._formula_workbook.close()
                del self._formula_workbook
        except Exception:
            pass

    def __del__(self):
        self.close()


class SheetCartographer:
    """Main agent that orchestrates sheet mapping."""

    def __init__(self, excel_path: Path, sheet_index: int, model: str, logger: ExperimentLogger):
        self.excel_path = excel_path
        self.sheet_index = sheet_index
        self.model = model
        self.logger = logger
        self.start_time = time.time()

        self.logger.main.info(f"📂 Loading Excel file: {excel_path}")
        self.workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        self.worksheet = self.workbook.worksheets[sheet_index]
        self.logger.main.info(f"📊 Loaded sheet {sheet_index}: {self.worksheet.title}")

        self.tool_handler = ExcelToolHandler(self.worksheet, logger, self.excel_path)
        self.client = OpenAI()  # Requires OPENAI_API_KEY
        self.total_tokens = 0

        self.blocks: list[Block] = []
        self.objects: list[ChartObject] = []
        self.unresolved: list[dict[str, str]] = []

        self.sheet_metadata: dict[str, Any] = {}

        # lightweight cache to avoid repeated peeks during conflict resolution
        self._peek_cache: dict[str, str] = {}
        self._peek_formula_cache: dict[str, str] = {}

    # --------------------------
    # Run pipeline
    # --------------------------
    def run(self) -> dict[str, Any]:
        self.logger.main.info("🚀 Starting Sheet-Cartographer analysis")

        try:
            sheet_info = self._orient()
            density_grid = self._sweep(sheet_info)
            candidate_blocks = self._cluster(density_grid, sheet_info)
            self._probe(candidate_blocks)

            # LLM-first Conflict resolution pass
            self._resolve_conflicts()

            self._overlay()
            cartographer_map = self._emit()

            elapsed = time.time() - self.start_time
            self.logger.log_metrics(
                {
                    "runtime_seconds": elapsed,
                    "total_tokens": self.total_tokens,
                    "blocks_detected": len(self.blocks),
                    "tool_calls": dict(self.tool_handler.tool_call_count),
                    "token_efficiency": self.total_tokens / max(1, (sheet_info["rows"] * sheet_info["cols"])),
                }
            )
            return cartographer_map

        except Exception as e:
            self.logger.error.error(f"Cartographer failed: {e}", exc_info=True)
            raise

    # --------------------------
    # S0: Orient
    # --------------------------
    def _orient(self) -> dict[str, Any]:
        self.logger.main.info("📍 S0: ORIENT - Getting sheet dimensions")
        sheet_info = self.tool_handler.execute_tool("sheet_info", {})
        self.logger.main.info(f"   Sheet size: {sheet_info['rows']}x{sheet_info['cols']} ({sheet_info['usedRange']})")

        self.sheet_metadata = sheet_info
        self.sheet_metadata["sheet_name"] = self.worksheet.title

        self.system_metadata = self._create_system_metadata(sheet_info)
        self.logger.main.debug(f"System metadata for LLM:\n{self.system_metadata}")
        return sheet_info

    def _create_system_metadata(self, sheet_info: dict[str, Any]) -> str:
        rows = sheet_info["rows"]
        cols = sheet_info["cols"]
        used_range = sheet_info["usedRange"]
        max_col_letter = get_column_letter(cols)

        parts = [
            "You are the Sheet-Cartographer analyzing spreadsheet structure.",
            "",
            "SHEET METADATA:",
            f"• Sheet name: {self.worksheet.title}",
            f"• Dimensions: {rows} rows × {cols} columns (A–{max_col_letter})",
            f"• Used range: {used_range}",
        ]
        if sheet_info.get("hiddenRows"):
            parts.append(f"• Hidden rows: {sheet_info['hiddenRows']}")
        if sheet_info.get("hiddenCols"):
            parts.append(f"• Hidden columns: {sheet_info['hiddenCols']}")
        parts.extend(
            [
                "",
                "Always treat any viewport or range as a sub-region of this full grid.",
                "This helps you understand where data likely continues beyond the current view.",
            ]
        )
        return "\n".join(parts)

    # --------------------------
    # S1: Sweep
    # --------------------------
    def _sweep(self, sheet_info: dict[str, Any]) -> list[list[float]]:
        self.logger.main.info("🔍 S1: SWEEP - Scanning sheet density")
        rows = sheet_info["rows"]
        cols = sheet_info["cols"]

        window_height = 40
        window_width = 20
        self.sweep_step_h = window_height // 2
        self.sweep_step_w = window_width // 2

        density_grid: list[list[float]] = []

        for row_start in range(1, rows + 1, self.sweep_step_h):
            if time.time() - self.start_time > MAX_RUNTIME_SECONDS:
                windows_processed = sum(len(r) for r in density_grid)
                total_windows = ((rows + self.sweep_step_h - 1) // self.sweep_step_h) * (
                    (cols + self.sweep_step_w - 1) // self.sweep_step_w
                )
                self.logger.main.warning(
                    f"Runtime cap hit during sweep; processed {windows_processed}/{total_windows} windows"
                )
                break

            density_row = []
            for col_start in range(1, cols + 1, self.sweep_step_w):
                current_height = min(window_height, rows - row_start + 1)
                current_width = min(window_width, cols - col_start + 1)

                window_text = self.tool_handler.execute_tool(
                    "window_grab",
                    {
                        "top": row_start,
                        "left": col_start,
                        "height": current_height,
                        "width": current_width,
                        "format": "markdown",
                    },
                )

                lines = window_text.split("\n")[2:]
                non_empty = 0
                total = 0
                for line in lines:
                    cells = line.split("|")[1:]
                    for cell in cells:
                        total += 1
                        if (ct := cell.strip()) and ct not in ("", "␀"):
                            non_empty += 1

                density = non_empty / total if total > 0 else 0
                density_row.append(density)

                self.logger.main.debug(
                    f"   Window [{row_start}:{min(row_start + current_height - 1, rows)}, "
                    f"{col_start}:{min(col_start + current_width - 1, cols)}] "
                    f"non_empty: {non_empty}/{total}, density: {density:.3f}"
                )

                if density > HIGH_DENSITY_THRESHOLD:
                    self.logger.main.debug(f"   Dense window detected at [{row_start}, {col_start}]")

            density_grid.append(density_row)

        self.logger.main.info(
            f"   Density grid size: {len(density_grid)}x{len(density_grid[0]) if density_grid else 0}"
        )
        return density_grid

    # --------------------------
    # S2: Cluster
    # --------------------------
    def _cluster(self, density_grid: list[list[float]], sheet_info: dict[str, Any]) -> list[dict[str, Any]]:
        self.logger.main.info("🔗 S2: CLUSTER - Merging dense regions")
        candidate_blocks = []
        visited = set()

        for i, row in enumerate(density_grid):
            for j, density in enumerate(row):
                if density > DENSITY_THRESHOLD and (i, j) not in visited:
                    cluster = self._flood_fill(density_grid, i, j, visited)
                    min_i = min(c[0] for c in cluster)
                    max_i = max(c[0] for c in cluster)
                    min_j = min(c[1] for c in cluster)
                    max_j = max(c[1] for c in cluster)

                    start_row = min_i * self.sweep_step_h + 1
                    end_row = min((max_i + 2) * self.sweep_step_h, sheet_info["rows"])
                    start_col = min_j * self.sweep_step_w + 1
                    end_col = min((max_j + 2) * self.sweep_step_w, sheet_info["cols"])

                    block_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
                    candidate_blocks.append(
                        {"range": block_range, "density": sum(density_grid[c[0]][c[1]] for c in cluster) / len(cluster)}
                    )

        self.logger.main.info(f"   Found {len(candidate_blocks)} candidate blocks")
        return candidate_blocks

    def _flood_fill(self, grid: list[list[float]], i: int, j: int, visited: set) -> list[tuple[int, int]]:
        cluster = []
        stack = [(i, j)]
        while stack:
            ci, cj = stack.pop()
            if (ci, cj) in visited:
                continue
            if ci < 0 or ci >= len(grid) or cj < 0 or cj >= len(grid[0]):
                continue
            if grid[ci][cj] <= DENSITY_THRESHOLD:
                continue
            visited.add((ci, cj))
            cluster.append((ci, cj))
            stack.extend([(ci - 1, cj), (ci + 1, cj), (ci, cj - 1), (ci, cj + 1)])
        return cluster

    # --------------------------
    # S3: Probe & classify
    # --------------------------
    def _split_until_homogeneous(self, range_str: str, depth: int = 0, max_depth: int = 3) -> list[dict[str, Any]]:
        if depth >= max_depth:
            return [{"range": range_str, "homogeneity": 0.5, "split": False}]

        parts = range_str.split(":")
        if len(parts) != 2:
            return [{"range": range_str, "homogeneity": 0.5, "split": False}]

        start_col, start_row = coordinate_from_string(parts[0])
        end_col, end_row = coordinate_from_string(parts[1])
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)

        area = (end_row - start_row + 1) * (end_col_idx - start_col_idx + 1)
        if area > MAX_RANGE_AREA:
            self.logger.main.debug(f"   Area {area} exceeds max, forcing split")
            mid_row = (start_row + end_row) // 2
            upper_range = f"{start_col}{start_row}:{end_col}{mid_row}"
            lower_range = f"{start_col}{mid_row + 1}:{end_col}{end_row}"
            return self._split_until_homogeneous(upper_range, depth + 1, max_depth) + self._split_until_homogeneous(
                lower_range, depth + 1, max_depth
            )

        range_text = self.tool_handler.execute_tool("range_peek", {"range": range_str, "format": "markdown"})
        homogeneity = self.tool_handler.calculate_layout_homogeneity(range_text)
        self.logger.main.debug(f"   {'  ' * depth}Range {range_str}: homogeneity={homogeneity:.3f}")

        if homogeneity >= HOMOGENEITY_THRESHOLD:
            return [{"range": range_str, "homogeneity": homogeneity, "split": False, "text": range_text}]

        total_rows = end_row - start_row + 1
        total_cols = end_col_idx - start_col_idx + 1
        sub_ranges: list[dict[str, Any]] = []

        if total_rows > total_cols * 2:
            mid_row = (start_row + end_row) // 2
            upper = f"{start_col}{start_row}:{end_col}{mid_row}"
            lower = f"{start_col}{mid_row + 1}:{end_col}{end_row}"
            self.logger.main.debug(f"   {'  ' * depth}Splitting horizontally at row {mid_row}")
            sub_ranges.extend(self._split_until_homogeneous(upper, depth + 1, max_depth))
            sub_ranges.extend(self._split_until_homogeneous(lower, depth + 1, max_depth))
        elif total_cols > total_rows * 2:
            mid_col_idx = (start_col_idx + end_col_idx) // 2
            mid_col = get_column_letter(mid_col_idx)
            left = f"{start_col}{start_row}:{mid_col}{end_row}"
            right = f"{get_column_letter(mid_col_idx + 1)}{start_row}:{end_col}{end_row}"
            self.logger.main.debug(f"   {'  ' * depth}Splitting vertically at column {mid_col}")
            sub_ranges.extend(self._split_until_homogeneous(left, depth + 1, max_depth))
            sub_ranges.extend(self._split_until_homogeneous(right, depth + 1, max_depth))
        else:
            mid_row = (start_row + end_row) // 2
            mid_col_idx = (start_col_idx + end_col_idx) // 2
            mid_col = get_column_letter(mid_col_idx)
            self.logger.main.debug(f"   {'  ' * depth}Splitting into quadrants at {mid_col}{mid_row}")
            sub_ranges.extend(
                self._split_until_homogeneous(f"{start_col}{start_row}:{mid_col}{mid_row}", depth + 1, max_depth)
            )
            sub_ranges.extend(
                self._split_until_homogeneous(
                    f"{get_column_letter(mid_col_idx + 1)}{start_row}:{end_col}{mid_row}", depth + 1, max_depth
                )
            )
            sub_ranges.extend(
                self._split_until_homogeneous(f"{start_col}{mid_row + 1}:{mid_col}{end_row}", depth + 1, max_depth)
            )
            sub_ranges.extend(
                self._split_until_homogeneous(
                    f"{get_column_letter(mid_col_idx + 1)}{mid_row + 1}:{end_col}{end_row}", depth + 1, max_depth
                )
            )

        return sub_ranges

    def _probe(self, candidate_blocks: list[dict[str, Any]]) -> None:
        self.logger.main.info("🔬 S3: PROBE - Segmenting and classifying blocks")
        block_id = 1

        for idx, candidate in enumerate(candidate_blocks):
            if time.time() - self.start_time > MAX_RUNTIME_SECONDS:
                self.logger.main.warning(
                    f"Runtime cap hit during probe; processed {idx}/{len(candidate_blocks)} regions"
                )
                break

            self.logger.main.info(f"   Processing candidate {idx + 1}/{len(candidate_blocks)}: {candidate['range']}")
            sub_regions = self._split_until_homogeneous(candidate["range"])
            self.logger.main.info(f"   Split into {len(sub_regions)} homogeneous regions")

            for sub_region in sub_regions:
                if "text" in sub_region:
                    lines = sub_region["text"].split("\n")[2:]
                    total_cells = sum(len(line.split("|")[1:]) for line in lines)
                    non_empty = sum(
                        1 for line in lines for cell in line.split("|")[1:] if (c := cell.strip()) and c != "␀"
                    )
                    density = non_empty / total_cells if total_cells > 0 else 0
                    if density < 0.05:
                        self.logger.main.debug(
                            f"   Skipping sparse region {sub_region['range']} (density: {density:.3f})"
                        )
                        continue

                if "text" not in sub_region:
                    sub_region["text"] = self.tool_handler.execute_tool(
                        "range_peek", {"range": sub_region["range"], "format": "markdown"}
                    )

                text_lower = sub_region["text"].lower()
                has_embedded_totals = any(
                    k in text_lower
                    for k in [
                        "total revenue",
                        "total revenues",
                        "total expense",
                        "total expenses",
                        "total income",
                        "total cost",
                        "summary",
                        "subtotal",
                        "grand total",
                    ]
                )

                result = self._classify_block_with_llm(sub_region["text"], sub_region["range"])
                (
                    block_type,
                    semantic_desc,
                    confidence,
                    reasoning,
                    open_questions,
                    peek_requests,
                    suggested_ops,
                    agg_subtype,
                ) = result

                if has_embedded_totals and confidence > 0.5:
                    self.logger.main.info(
                        f"   Embedded totals detected in {sub_region['range']} - lowering confidence to trigger split"
                    )
                    confidence = 0.4

                refinement_count = 0
                max_refinements = 3
                peeked_ranges = set()
                prev_confidence = confidence
                prev_block_type = block_type
                all_peeked_text: list[str] = []
                running_ops = list(suggested_ops)

                while (open_questions or peek_requests) and refinement_count < max_refinements:
                    self.logger.main.debug(f"   Refinement {refinement_count + 1} for {sub_region['range']}")
                    additional_context_parts = []

                    if peek_requests:
                        for peek_req in peek_requests:
                            peek_range = peek_req.get("range", "")
                            peek_reason = peek_req.get("reason", "")
                            if peek_range in peeked_ranges:
                                continue
                            peeked_ranges.add(peek_range)
                            try:
                                peek_text = self.tool_handler.execute_tool(
                                    "range_peek", {"range": peek_range, "format": "markdown"}
                                )
                                additional_context_parts.append(
                                    f"Peek at {peek_range} (reason: {peek_reason}):\n{peek_text}"
                                )
                                all_peeked_text.append(peek_text)
                            except Exception as e:
                                self.logger.main.debug(f"   Peek failed at {peek_range}: {e}")

                    if open_questions:
                        heuristic_context = self._address_open_questions(
                            open_questions, sub_region["range"], sub_region["text"]
                        )
                        if heuristic_context and heuristic_context != "No additional context gathered":
                            additional_context_parts.append(heuristic_context)

                    additional_context = "\n\n".join(additional_context_parts) if additional_context_parts else None
                    result = self._classify_block_with_llm(
                        sub_region["text"], sub_region["range"], additional_context=additional_context
                    )
                    (
                        new_type,
                        new_semantic,
                        new_conf,
                        new_reasoning,
                        open_questions,
                        peek_requests,
                        suggested_ops,
                        new_agg_subtype,
                    ) = result

                    if abs(new_conf - prev_confidence) < 0.05 and new_type == prev_block_type and refinement_count > 0:
                        self.logger.main.info(
                            f"   Confidence plateau ({prev_confidence:.2f} → {new_conf:.2f}), forcing split"
                        )
                        new_conf = 0.4
                        break

                    if all_peeked_text:
                        combined_text = sub_region["text"] + "\n" + "\n".join(all_peeked_text)
                        new_homog = self.tool_handler.calculate_layout_homogeneity(combined_text)
                        if new_homog < HOMOGENEITY_THRESHOLD:
                            self.logger.main.info(f"   Homogeneity ↓ to {new_homog:.2f} after peeks -> re-split")
                            new_conf = 0.4
                            break

                    block_type = new_type
                    semantic_desc = new_semantic
                    confidence = new_conf
                    reasoning = new_reasoning
                    running_ops = list(dict.fromkeys(running_ops + suggested_ops))
                    if new_agg_subtype and new_agg_subtype != "none":
                        agg_subtype = new_agg_subtype
                    prev_confidence = new_conf
                    prev_block_type = new_type
                    refinement_count += 1

                if confidence < 0.5 or has_embedded_totals:
                    self.logger.main.info(
                        f"   Low confidence or mixed semantics -> deeper split of {sub_region['range']}"
                    )
                    deeper_regions = self._force_split_region(sub_region["range"])
                    for deep_region in deeper_regions:
                        if "text" not in deep_region:
                            deep_region["text"] = self.tool_handler.execute_tool(
                                "range_peek", {"range": deep_region["range"], "format": "markdown"}
                            )
                        deep_result = self._classify_block_with_llm(deep_region["text"], deep_region["range"])
                        (
                            deep_type,
                            deep_sem,
                            deep_conf,
                            deep_reasoning,
                            _open_q,
                            _peek_reqs,
                            deep_ops,
                            _deep_subtype,
                        ) = deep_result
                        struct_type = TYPE_MAP.get(deep_type, StructuralType.UNSTRUCTURED_TEXT)

                        deep_data_chars = self._infer_characteristics(deep_region["text"], deep_region["range"])
                        if struct_type == StructuralType.HEADER_ZONE:
                            deep_data_chars.has_headers = True
                            deep_data_chars.header_rows = max(deep_data_chars.header_rows, 1)
                        deep_default_ops = self._get_default_operations(struct_type)
                        deep_final_ops = list(dict.fromkeys(deep_ops + deep_default_ops))

                        self.blocks.append(
                            Block(
                                id=f"blk_{block_id:02d}",
                                range=deep_region["range"],
                                structural_type=struct_type,
                                semantic_description=deep_sem or "Split region from mixed semantics",
                                confidence=deep_conf,
                                classification_reasoning=deep_reasoning or "Forced split due to mixed semantics",
                                data_characteristics=deep_data_chars,
                                suggested_operations=deep_final_ops,
                            )
                        )
                        self.logger.main.info(
                            f"   Block {block_id}: {deep_region['range']} -> {deep_type} (conf: {deep_conf:.2f})"
                        )
                        block_id += 1
                else:
                    struct_type = TYPE_MAP.get(block_type, StructuralType.UNSTRUCTURED_TEXT)
                    data_chars = self._infer_characteristics(sub_region["text"], sub_region["range"])
                    if agg_subtype and agg_subtype != "none":
                        data_chars.aggregation_zone_subtype = agg_subtype
                    if struct_type == StructuralType.HEADER_ZONE:
                        data_chars.has_headers = True
                        data_chars.header_rows = max(data_chars.header_rows, 1)
                    default_ops = self._get_default_operations(struct_type)
                    final_ops = list(dict.fromkeys(running_ops + default_ops))

                    self.blocks.append(
                        Block(
                            id=f"blk_{block_id:02d}",
                            range=sub_region["range"],
                            structural_type=struct_type,
                            semantic_description=semantic_desc or f"Region classified as {block_type}",
                            confidence=confidence,
                            classification_reasoning=reasoning,
                            data_characteristics=data_chars,
                            suggested_operations=final_ops,
                        )
                    )
                    self.logger.main.info(
                        f"   Block {block_id}: {sub_region['range']} -> {block_type} (conf: {confidence:.2f}, homog: {sub_region.get('homogeneity', 0):.2f})"
                    )
                    block_id += 1

    # --------------------------
    # S3.5: Conflict Resolution (LLM-first decide-or-split)
    # --------------------------
    def _resolve_conflicts(self) -> None:
        """Ask the LLM to KEEP, RELABEL, or SPLIT when signals suggest uncertainty."""
        self.logger.main.info("🧭 S3.5: RESOLVE - LLM-first decide-or-split")

        new_blocks: list[Block] = []
        next_id = 1 + max([int(b.id.split("_")[1]) for b in self.blocks if b.id.startswith("blk_")] + [0])

        for block in self.blocks:
            rng = block.range
            text = self._peek_cache.get(rng) or self.tool_handler.execute_tool(
                "range_peek", {"range": rng, "format": "markdown"}
            )
            self._peek_cache[rng] = text

            ftext = self._peek_formula_cache.get(rng) or self.tool_handler.execute_tool(
                "range_formulas", {"range": rng, "format": "markdown"}
            )
            self._peek_formula_cache[rng] = ftext

            ev = self._collect_block_evidence(block, text, ftext)

            # decide when to involve the LLM (no auto overrides!)
            needs_review = False
            if block.structural_type in (StructuralType.HEADER_ZONE, StructuralType.AGGREGATION_ZONE):
                if (
                    (0.04 <= ev["formula_density"] <= 0.12)
                    or (0.20 <= ev["numeric_ratio"] <= 0.45)
                    or ev["has_agg_keywords"]
                    or ev["numeric_consistency"].get("checked_cols", 0) > 0
                ):
                    needs_review = True

            if not needs_review:
                new_blocks.append(block)
                continue

            decision, payload = self._decide_or_split_with_llm(block, text, ftext, ev)

            if decision == "keep":
                self.logger.main.info(f"   LLM: keep {block.id} {block.range} as {block.structural_type}")
                # May boost/adjust confidence slightly based on LLM confidence
                llm_conf = float(payload.get("confidence", block.confidence))
                block.confidence = max(block.confidence, min(1.0, llm_conf))
                block.classification_reasoning = (
                    f"{block.classification_reasoning}\n[LLM] {payload.get('reasoning', '')}".strip()
                )
                new_blocks.append(block)

            elif decision == "relabel":
                new_label = payload.get("new_label")
                agg_subtype = payload.get("aggregation_subtype", "none")
                conf = float(payload.get("confidence", SUMMARY_CONFIDENCE))
                reason = payload.get("reasoning", "LLM relabel")
                self._apply_decision(
                    block,
                    new_label,
                    ev,
                    new_confidence=conf,
                    reason=reason,
                    agg_subtype=agg_subtype,
                    source="arbitration",
                )
                new_blocks.append(block)
                self.logger.main.info(f"   LLM: relabel {block.id} {block.range} -> {new_label} (conf {conf:.2f})")

            elif decision == "split":
                slices: list[dict] = payload.get("slices", [])
                applied = self._validate_and_apply_split(block, slices, ev, next_id)
                if applied:
                    created = applied["created"]
                    next_id += len(created)
                    new_blocks.extend(created)
                    self.logger.main.info(
                        f"   LLM: split {block.id} into {len(created)} blocks: {[b.range for b in created]}"
                    )
                else:
                    # fallback if split was invalid: keep original unchanged but append reason
                    block.classification_reasoning = (
                        f"{block.classification_reasoning}\n[LLM] Split invalid; keeping.".strip()
                    )
                    new_blocks.append(block)
            else:
                # Arbitration failed or unknown -> keep
                new_blocks.append(block)

        self.blocks = new_blocks

    def _validate_and_apply_split(self, block: Block, slices: list[dict], ev: dict, next_id: int) -> dict | None:
        """Ensure slices are within the original range, non-overlapping, and well-typed; then create new blocks."""
        if not slices:
            return None
        try:
            # parse original range
            (s_col, s_row) = coordinate_from_string(block.range.split(":")[0])
            (e_col, e_row) = coordinate_from_string(block.range.split(":")[1])
            sidx, eidx = column_index_from_string(s_col), column_index_from_string(e_col)

            def _norm(r: str) -> tuple[int, int, int, int]:
                (a_col, a_row) = coordinate_from_string(r.split(":")[0])
                (b_col, b_row) = coordinate_from_string(r.split(":")[1])
                return (column_index_from_string(a_col), a_row, column_index_from_string(b_col), b_row)

            # bounds + overlap checks
            boxes = []
            for sl in slices:
                if "range" not in sl or "label" not in sl:
                    return None
                x1, y1, x2, y2 = _norm(sl["range"])
                if x1 < sidx or x2 > eidx or y1 < s_row or y2 > e_row:
                    return None
                if x1 > x2 or y1 > y2:
                    return None
                boxes.append((x1, y1, x2, y2, sl))

            # check overlap
            def _overlap(b1, b2):
                return not (b1[2] < b2[0] or b2[2] < b1[0] or b1[3] < b2[1] or b2[3] < b1[1])

            for i in range(len(boxes)):
                for j in range(i + 1, len(boxes)):
                    if _overlap(boxes[i], boxes[j]):
                        return None

            # build new blocks
            created: list[Block] = []
            for _, _, _, _, sl in sorted(boxes, key=lambda t: (t[1], t[0])):
                rng = sl["range"]
                lbl = sl["label"]
                sub = sl.get("aggregation_subtype", "none")
                # extract views and infer
                text = self._peek_cache.get(rng) or self.tool_handler.execute_tool(
                    "range_peek", {"range": rng, "format": "markdown"}
                )
                self._peek_cache[rng] = text
                data_chars = self._infer_characteristics(text, rng)
                if lbl == "AggregationZone" and sub and sub != "none":
                    data_chars.aggregation_zone_subtype = sub
                if lbl == "HeaderZone":
                    data_chars.has_headers = True
                    data_chars.header_rows = max(data_chars.header_rows, 1)
                struct_type = TYPE_MAP.get(lbl, block.structural_type)
                default_ops = self._get_default_operations(struct_type)

                created.append(
                    Block(
                        id=f"blk_{next_id:02d}",
                        range=rng,
                        structural_type=struct_type,
                        semantic_description=sl.get("semantic_description", f"Slice {lbl} from {block.id}"),
                        confidence=float(sl.get("confidence", SUMMARY_CONFIDENCE)),
                        classification_reasoning=sl.get("reasoning", "LLM split decision"),
                        data_characteristics=data_chars,
                        suggested_operations=default_ops,
                    )
                )
                next_id += 1

            return {"created": created}
        except Exception as e:
            self.logger.main.debug(f"Split validation failed: {e}")
            return None

    def _apply_decision(
        self,
        block: Block,
        decision: str,
        ev: dict,
        *,
        new_confidence: float | None = None,
        reason: str | None = None,
        agg_subtype: str | None = None,
        source: str = "deterministic",
    ) -> None:
        struct_type = TYPE_MAP.get(decision, block.structural_type)
        block.structural_type = struct_type

        # Update data characteristics
        dc = block.data_characteristics
        if struct_type == StructuralType.HEADER_ZONE:
            dc.has_headers = True
            dc.header_rows = max(dc.header_rows, 1)
            dc.aggregation_zone_subtype = None
        elif struct_type == StructuralType.AGGREGATION_ZONE:
            dc.has_headers = False
            if agg_subtype and agg_subtype != "none":
                dc.aggregation_zone_subtype = agg_subtype
            else:
                if "grand total" in ev["keywords"]:
                    dc.aggregation_zone_subtype = "grand_total"
                elif "subtotal" in ev["keywords"]:
                    dc.aggregation_zone_subtype = "subtotal"
                else:
                    dc.aggregation_zone_subtype = "kpi"

        # Confidence and reasoning
        if new_confidence is not None:
            block.confidence = new_confidence
        else:
            block.confidence = (
                SUMMARY_CONFIDENCE if struct_type == StructuralType.AGGREGATION_ZONE else HEADER_BANNER_CONFIDENCE
            )

        src_tag = "LLM arbitration" if source != "deterministic" else "Rule-based"
        add_reason = (
            reason
            or f"{src_tag} using formula_density={ev['formula_density']:.2f}, numeric_ratio={ev['numeric_ratio']:.2f}, keywords={list(ev['keywords'])}"
        )
        block.classification_reasoning = f"{block.classification_reasoning}\n[{src_tag}] {add_reason}".strip()

        # Suggested operations refresh
        default_ops = self._get_default_operations(struct_type)
        block.suggested_operations = list(dict.fromkeys((block.suggested_operations or []) + default_ops))

        # Semantic tweak
        if struct_type == StructuralType.AGGREGATION_ZONE and "total" in ev["keywords"]:
            if "grand total" in ev["keywords"]:
                block.semantic_description = block.semantic_description or "Grand total summary section"
            elif "subtotal" in ev["keywords"]:
                block.semantic_description = block.semantic_description or "Subtotal summary section"
            else:
                block.semantic_description = block.semantic_description or "KPI/total summary section"
        elif struct_type == StructuralType.HEADER_ZONE:
            block.semantic_description = block.semantic_description or "Column headers for the data that follows"

    def _neighbor_row_peek(self, range_str: str) -> dict[str, str]:
        """Peek one row above/below same width as range, best-effort."""
        result = {"above": "", "below": ""}
        try:
            start_col, start_row = coordinate_from_string(range_str.split(":")[0])
            end_col, end_row = coordinate_from_string(range_str.split(":")[1])
            # above
            if start_row > 1:
                above = f"{start_col}{start_row - 1}:{end_col}{start_row - 1}"
                result["above"] = self.tool_handler.execute_tool("range_peek", {"range": above, "format": "markdown"})
            # below
            total_rows = self.sheet_metadata.get("rows", 0)
            if end_row < total_rows:
                below = f"{start_col}{end_row + 1}:{end_col}{end_row + 1}"
                result["below"] = self.tool_handler.execute_tool("range_peek", {"range": below, "format": "markdown"})
        except Exception as e:
            self.logger.main.debug(f"Neighbor row peek failed: {e}")
        return result

    def _collect_block_evidence(self, block: Block, text: str, formulas_text: str) -> dict:
        """Compute robust, sheet-aware evidence for conflict decisions."""
        rng = block.range
        dc = self._infer_characteristics(text, rng)

        # sample for formula detection beyond _infer_characteristics
        parts = rng.split(":")
        start_col, start_row = coordinate_from_string(parts[0])
        end_col, end_row = coordinate_from_string(parts[1])
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)

        samples = []
        for col in range(start_col_idx, min(start_col_idx + 2, end_col_idx + 1)):
            samples.append((start_row, col))
        for col in range(max(start_col_idx, end_col_idx - 1), end_col_idx + 1):
            samples.append((end_row, col))
        for row in range(start_row, min(start_row + 2, end_row + 1)):
            samples.append((row, start_col_idx))
        for row in range(max(start_row, end_row - 1), end_row + 1):
            samples.append((row, end_col_idx))
        samples = samples[:8]

        formula_hits = 0
        checked = 0
        for r, c in samples:
            try:
                meta = self.tool_handler.execute_tool("cell_meta", {"row": r, "col": c})
                checked += 1
                if meta.get("isFormula"):
                    formula_hits += 1
            except Exception:
                pass

        sampled_formula_density = (formula_hits / checked) if checked else 0.0
        formula_density = max(dc.formula_density or 0.0, sampled_formula_density)

        # Keywords & position
        text_lower = text.lower()
        keywords_present = set(k for k in AGG_KEYWORDS if k in text_lower)
        position = self._position_flags(rng)

        # Neighbor context & formulas view
        neighbors = self._neighbor_row_peek(rng)
        formulas_view = (formulas_text or "")[:2000]

        # Numeric consistency check
        numeric_consistency = self._numeric_consistency_check(rng)

        return {
            "numeric_ratio": dc.numeric_ratio or 0.0,
            "formula_density": formula_density,
            "has_agg_keywords": any(k in keywords_present for k in ("total", "subtotal", "grand total")),
            "keywords": keywords_present,
            "position": position,
            "neighbors": neighbors,
            "formulas_view": formulas_view,
            "numeric_consistency": numeric_consistency,
        }

    def _position_flags(self, range_str: str) -> dict[str, bool]:
        parts = range_str.split(":")
        start_col, start_row = coordinate_from_string(parts[0])
        end_col, end_row = coordinate_from_string(parts[1])
        total_rows = self.sheet_metadata.get("rows", 1)
        total_cols = self.sheet_metadata.get("cols", 1)
        sidx = column_index_from_string(start_col)
        eidx = column_index_from_string(end_col)
        return {
            "near_top": start_row <= 5,
            "near_bottom": end_row >= total_rows - 5,
            "left_side": sidx <= 3,
            "right_side": eidx >= total_cols - 3,
            "full_width": (eidx - sidx + 1) == total_cols,
            "full_height": (end_row - start_row + 1) == total_rows,
        }

    def _to_float(self, val) -> tuple[bool, float]:
        """Try to coerce to float; return (ok, number)."""
        if val is None:
            return False, 0.0
        if isinstance(val, (int, float)):
            try:
                return True, float(val)
            except Exception:
                return False, 0.0
        s = str(val).strip()
        if s == "" or s == "␀":
            return False, 0.0
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        s = s.replace("$", "").replace("€", "").replace(",", "")
        try:
            return True, float(s)
        except Exception:
            return False, 0.0

    def _numeric_consistency_check(self, range_str: str) -> dict:
        """Check if numbers in a suspected header/aggregation row match sums of the data below (typed totals welcome)."""
        try:
            (start_col, start_row) = coordinate_from_string(range_str.split(":")[0])
            (end_col, end_row) = coordinate_from_string(range_str.split(":")[1])
            sidx = column_index_from_string(start_col)
            eidx = column_index_from_string(end_col)
            total_rows = self.sheet_metadata.get("rows", 0)

            single_row = end_row == start_row
            if not single_row:
                return {"checked_cols": 0, "matches": 0, "details": []}

            header_nums: dict[int, float] = {}
            for c in range(sidx, eidx + 1):
                ok, num = self._to_float(self.worksheet.cell(row=start_row, column=c).value)
                if ok:
                    header_nums[c] = num
            if not header_nums:
                return {"checked_cols": 0, "matches": 0, "details": []}

            data_start = end_row + 1
            max_scan_rows = min(200, max(0, total_rows - end_row))
            blank_streak = 0
            sums: dict[int, float] = dict.fromkeys(header_nums.keys(), 0.0)
            scanned_rows = 0

            for r in range(data_start, data_start + max_scan_rows):
                scanned_rows += 1
                all_blank = True
                for c in range(sidx, eidx + 1):
                    v = self.worksheet.cell(row=r, column=c).value
                    ok, num = self._to_float(v)
                    if ok:
                        all_blank = False
                        if c in sums:
                            sums[c] += num
                if all_blank:
                    blank_streak += 1
                    if blank_streak >= 2:
                        break
                else:
                    blank_streak = 0

            details = []
            matches = 0
            for c, hv in header_nums.items():
                s = sums.get(c, 0.0)
                abs_tol = 0.5
                rel_tol = 0.02 if abs(hv) >= 10 else 0.05
                rel_err = abs(s - hv) / (abs(hv) if hv != 0 else 1.0)
                ok_match = (abs(s - hv) <= abs_tol) or (rel_err <= rel_tol)
                if ok_match:
                    matches += 1
                details.append(
                    {
                        "col": get_column_letter(c),
                        "header_value": hv,
                        "sum_below": s,
                        "rel_err": rel_err,
                        "match": ok_match,
                    }
                )

            return {
                "checked_cols": len(header_nums),
                "matches": matches,
                "details": details,
                "rows_scanned": scanned_rows,
            }
        except Exception as e:
            self.logger.main.debug(f"Numeric consistency check failed for {range_str}: {e}")
            return {"checked_cols": 0, "matches": 0, "details": [], "error": str(e)}

    def _decide_or_split_with_llm(
        self, block: Block, values_text: str, formulas_text: str, evidence: dict
    ) -> tuple[str, dict]:
        """Ask the LLM to keep, relabel, or split (with slices)."""
        system_prompt = f"""{self.system_metadata}

You are deciding if a block is labelled correctly or should be split.
Outcomes:
- keep: leave the label as-is.
- relabel: change to HeaderZone or AggregationZone (with optional subtype).
- split: return 2-4 non-overlapping slices with explicit ranges and labels.

Guidelines:
- 'HeaderZone' is mostly text, rarely formulas; usually top of a table.
- 'AggregationZone' is numbers and often formulas; keywords TOTAL/SUBTOTAL/GRAND TOTAL are strong signals.
- Typed totals (numbers without formulas) still count as Aggregation if they summarize the region.
- Prefer 'split' when a single row/region mixes headers and totals/KPIs.
Return concise reasoning (no chain-of-thought)."""

        payload = {
            "range": block.range,
            "current_label": str(block.structural_type),
            "values_view": values_text[:1600],
            "formulas_view": (formulas_text or "")[:1600],
            "neighbors": {
                "above": evidence.get("neighbors", {}).get("above", "")[:600],
                "below": evidence.get("neighbors", {}).get("below", "")[:600],
            },
            "evidence": {
                "numeric_ratio": round(evidence["numeric_ratio"], 4),
                "formula_density": round(evidence["formula_density"], 4),
                "has_agg_keywords": bool(evidence["has_agg_keywords"]),
                "keywords": sorted(list(evidence["keywords"])),
                "position": evidence["position"],
                "numeric_consistency": evidence.get("numeric_consistency", {}),
            },
        }

        tools = [
            {
                "type": "function",
                "function": {
                    "name": "decide_or_split",
                    "description": "Keep, relabel, or split a block into slices",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "decision": {"type": "string", "enum": ["keep", "relabel", "split"]},
                            "confidence": {"type": "number"},
                            "reasoning": {"type": "string"},
                            "new_label": {
                                "type": "string",
                                "enum": ["HeaderZone", "AggregationZone"],
                                "description": "Required if decision=relabel",
                            },
                            "aggregation_subtype": {
                                "type": "string",
                                "enum": ["kpi", "subtotal", "grand_total", "none"],
                            },
                            "slices": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "range": {"type": "string"},
                                        "label": {
                                            "type": "string",
                                            "enum": ["HeaderZone", "AggregationZone", "DataTable", "MetadataZone"],
                                        },
                                        "aggregation_subtype": {
                                            "type": "string",
                                            "enum": ["kpi", "subtotal", "grand_total", "none"],
                                        },
                                        "reasoning": {"type": "string"},
                                        "semantic_description": {"type": "string"},
                                        "confidence": {"type": "number"},
                                    },
                                    "required": ["range", "label"],
                                },
                                "minItems": 2,
                                "maxItems": 4,
                            },
                        },
                        "required": ["decision", "confidence", "reasoning"],
                    },
                },
            }
        ]

        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": json.dumps(payload, ensure_ascii=False)},
        ]

        try:
            if self.total_tokens > MAX_LLM_TOKENS:
                raise RuntimeError(f"Token budget exceeded: {self.total_tokens} > {MAX_LLM_TOKENS}")

            resp = self.client.chat.completions.create(
                model=self.model,
                messages=messages,
                tools=tools,
                tool_choice={"type": "function", "function": {"name": "decide_or_split"}},
                temperature=0.1,
                max_tokens=350,
            )

            actual_model = getattr(resp, "model", None)
            if actual_model and not hasattr(self, "_model_updated"):
                self.logger.update_model_actual(actual_model)
                self._model_updated = True

            if hasattr(resp, "usage"):
                self.total_tokens += resp.usage.total_tokens
                self.logger.log_llm_interaction(
                    model=self.model,
                    prompt=messages[-1]["content"],
                    response=str(resp.choices[0].message),
                    tokens={
                        "input": resp.usage.prompt_tokens,
                        "output": resp.usage.completion_tokens,
                        "total": resp.usage.total_tokens,
                    },
                    actual_model=actual_model,
                )

            if resp.choices[0].message.tool_calls:
                args = json.loads(resp.choices[0].message.tool_calls[0].function.arguments)
                decision = args.get("decision", "keep")
                return decision, args
        except Exception as e:
            self.logger.error.error(f"LLM decide_or_split failed: {e}")

        return "keep", {"confidence": block.confidence, "reasoning": "Arbitration failed; keeping"}

    # --------------------------
    # LLM calls (classification)
    # --------------------------
    def _classify_block_with_llm(
        self, range_text: str, range_str: str, additional_context: str | None = None
    ) -> tuple[str, str, float, str, list[str], list[dict], list[str], str]:
        tools = [
            {
                "type": "function",
                "function": {
                    "name": "classify_block",
                    "description": "Classify the type of spreadsheet block with open questions for refinement",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "block_type": {
                                "type": "string",
                                "enum": [
                                    "DataTable",
                                    "HeaderZone",
                                    "AggregationZone",
                                    "MetadataZone",
                                    "VisualizationAnchor",
                                    "FormInputZone",
                                    "NavigationZone",
                                    "UnstructuredText",
                                    "EmptyPadding",
                                ],
                                "description": "Structural type based on layout pattern",
                            },
                            "confidence": {"type": "number", "description": "Confidence between 0 and 1"},
                            "semantic_description": {
                                "type": "string",
                                "description": "One-sentence semantic description",
                            },
                            "reasoning": {"type": "string", "description": "Brief explanation (no chain-of-thought)"},
                            "aggregation_subtype": {
                                "type": "string",
                                "enum": ["kpi", "subtotal", "grand_total", "none"],
                            },
                            "suggested_operations": {"type": "array", "items": {"type": "string"}},
                            "open_questions": {"type": "array", "items": {"type": "string"}},
                            "peek_requests": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "properties": {"range": {"type": "string"}, "reason": {"type": "string"}},
                                    "required": ["range", "reason"],
                                },
                                "maxItems": 3,
                            },
                        },
                        "required": ["block_type", "semantic_description", "confidence", "reasoning"],
                    },
                },
            }
        ]

        system_prompt = f"""{self.system_metadata}

CLASSIFICATION INSTRUCTIONS:

STEP 1 - Choose ONE structural type:
- DataTable: Regular rows/columns with consistent schema (>70% Data cells)
- HeaderZone: Column/row headers referenced by formulas (rarely numeric or formulaic themselves)
- AggregationZone: Summaries, totals, KPIs (specify subtype: kpi/subtotal/grand_total)
- MetadataZone: Titles, descriptions not referenced by formulas
- VisualizationAnchor: Charts, sparklines, conditional formatting
- FormInputZone: Data entry areas with validation/dropdowns
- NavigationZone: Links, buttons, menus (need HYPERLINK evidence)
- UnstructuredText / EmptyPadding: Free text / whitespace

STEP 2 - Semantic description: business meaning, not just structure.
STEP 3 - Provide confidence, reasoning, suggested operations.

IMPORTANT:
- If mixed semantics, return confidence < 0.5 so we split further.
- Consider the block's position within the full grid (top, bottom, sides).
"""

        breadcrumb = self._create_range_breadcrumb(range_str)
        user_content = f"{breadcrumb}\n\nClassify this spreadsheet block at range {range_str}:\n\n{range_text}"
        if additional_context:
            user_content += f"\n\nAdditional context from refinement:\n{additional_context}"

        messages = [{"role": "system", "content": system_prompt}, {"role": "user", "content": user_content}]

        try:
            if self.total_tokens > MAX_LLM_TOKENS:
                self.logger.error.error(f"Token budget exceeded: {self.total_tokens} > {MAX_LLM_TOKENS}")
                raise RuntimeError(f"Token budget exceeded: {self.total_tokens} > {MAX_LLM_TOKENS}")

            response = self.client.chat.completions.create(
                model=self.model,
                messages=messages,
                tools=tools,
                tool_choice={"type": "function", "function": {"name": "classify_block"}},
                temperature=0.1,
                max_tokens=500,
            )

            actual_model = getattr(response, "model", None)
            if actual_model and not hasattr(self, "_model_updated"):
                self.logger.update_model_actual(actual_model)
                self._model_updated = True

            if hasattr(response, "usage"):
                tokens_used = response.usage.total_tokens
                self.total_tokens += tokens_used
                self.logger.log_llm_interaction(
                    model=self.model,
                    prompt=messages[-1]["content"],
                    response=str(response.choices[0].message),
                    tokens={
                        "input": response.usage.prompt_tokens,
                        "output": response.usage.completion_tokens,
                        "total": tokens_used,
                    },
                    actual_model=actual_model,
                )

            if response.choices[0].message.tool_calls:
                tool_call = response.choices[0].message.tool_calls[0]
                args = json.loads(tool_call.function.arguments)
                return (
                    args["block_type"],
                    args.get("semantic_description", ""),
                    args["confidence"],
                    args.get("reasoning", ""),
                    args.get("open_questions", []),
                    args.get("peek_requests", []),
                    args.get("suggested_operations", []),
                    args.get("aggregation_subtype", "none"),
                )

        except Exception as e:
            self.logger.error.error(f"LLM classification failed: {e}")

        return (
            "UnstructuredText",
            "Unable to classify region",
            UNKNOWN_CONFIDENCE,
            "Classification failed",
            [],
            [],
            [],
            "none",
        )

    def _force_split_region(self, range_str: str) -> list[dict[str, Any]]:
        parts = range_str.split(":")
        if len(parts) != 2:
            return [{"range": range_str, "homogeneity": 0.5}]
        start_col, start_row = coordinate_from_string(parts[0])
        end_col, end_row = coordinate_from_string(parts[1])
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        total_rows = end_row - start_row + 1
        total_cols = end_col_idx - start_col_idx + 1

        sub_regions: list[dict[str, Any]] = []

        if total_rows >= 2:
            try:
                first_row_range = f"{start_col}{start_row}:{end_col}{start_row}"
                second_row_range = f"{start_col}{start_row + 1}:{end_col}{start_row + 1}"
                fr_text = self.tool_handler.execute_tool("range_peek", {"range": first_row_range, "format": "markdown"})
                sr_text = self.tool_handler.execute_tool(
                    "range_peek", {"range": second_row_range, "format": "markdown"}
                )

                fr_cells, sr_cells = [], []
                for line in fr_text.split("\n")[2:]:
                    fr_cells.extend([c.strip() for c in line.split("|")[1:]])
                for line in sr_text.split("\n")[2:]:
                    sr_cells.extend([c.strip() for c in line.split("|")[1:]])

                fr_lower = " ".join(fr_cells).lower()
                totals_count = sum(1 for k in ["total", "sum", "revenue", "expense"] if k in fr_lower)

                def is_num(x: str) -> bool:
                    if not x or x == "␀":
                        return False
                    cleaned = x.strip().replace(",", "").replace("$", "").replace("%", "")
                    if cleaned.startswith("(") and cleaned.endswith(")"):
                        cleaned = "-" + cleaned[1:-1]
                    try:
                        float(cleaned)
                        return True
                    except:
                        return False

                fr_num = sum(1 for c in fr_cells if is_num(c))
                sr_num = sum(1 for c in sr_cells if is_num(c))

                should_split_first = False
                if totals_count >= 2 or (fr_num > len(fr_cells) * 0.5 and sr_num < len(sr_cells) * 0.3):
                    should_split_first = True

                if should_split_first:
                    header_range = f"{start_col}{start_row}:{end_col}{start_row}"
                    rest_range = f"{start_col}{start_row + 1}:{end_col}{end_row}"
                    sub_regions.append({"range": header_range, "homogeneity": 0.9})
                    sub_regions.append({"range": rest_range, "homogeneity": 0.8})
                    return sub_regions

            except Exception as e:
                self.logger.main.debug(f"   First-row analysis error: {e}")

        if total_cols >= 3:
            if total_cols == 9:
                split_col = start_col_idx + 5
            else:
                split_col = start_col_idx + (total_cols * 2 // 3)
            left_range = f"{start_col}{start_row}:{get_column_letter(split_col)}{end_row}"
            right_range = f"{get_column_letter(split_col + 1)}{start_row}:{end_col}{end_row}"
            self.logger.main.debug(f"   Force split columns at {get_column_letter(split_col)}")
            sub_regions.append({"range": left_range, "homogeneity": 0.8})
            sub_regions.append({"range": right_range, "homogeneity": 0.8})
        elif total_rows >= 2:
            mid_row = (start_row + end_row) // 2
            upper_range = f"{start_col}{start_row}:{end_col}{mid_row}"
            lower_range = f"{start_col}{mid_row + 1}:{end_col}{end_row}"
            self.logger.main.debug(f"   Force splitting rows at {mid_row}")
            sub_regions.append({"range": upper_range, "homogeneity": 0.8})
            sub_regions.append({"range": lower_range, "homogeneity": 0.8})
        else:
            sub_regions.append({"range": range_str, "homogeneity": 0.5})

        return sub_regions

    def _address_open_questions(self, questions: list[str], range_str: str, range_text: str) -> str:
        """Heuristically answer the model's open questions (headers? totals? sparse tail?).
        Returns a short, human-readable summary string or 'No additional context gathered'."""
        context_parts: list[str] = []

        # Parse range safely
        parts = range_str.split(":")
        if len(parts) != 2:
            return "No additional context gathered"
        start_col, start_row = coordinate_from_string(parts[0])
        end_col, end_row = coordinate_from_string(parts[1])
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)

        for q in questions or []:
            ql = (q or "").lower()

            # 1) Header-y? Check bold on first row cells
            if "header" in ql or "row 1" in ql or "row one" in ql:
                try:
                    style_sample = []
                    sample_end = min(start_col_idx + 3, end_col_idx + 1)
                    for c in range(start_col_idx, sample_end):
                        meta = self.tool_handler.execute_tool("cell_meta", {"row": start_row, "col": c})
                        style_sample.append("bold" if "bold" in meta.get("styleFlags", []) else "normal")
                    if any(s == "bold" for s in style_sample):
                        context_parts.append(f"Row {start_row} has bold labels → likely headers")
                    else:
                        context_parts.append(f"Row {start_row} not bolded → less likely headers")
                except Exception as e:
                    self.logger.main.debug(f"_address_open_questions header check failed: {e}")

            # 2) Totals? Look for formulas near the bottom of the region
            if "total" in ql or "sum" in ql or "subtotal" in ql:
                try:
                    has_formulas = False
                    probe_rows = range(max(end_row - 2, start_row), end_row + 1)
                    probe_cols = range(start_col_idx, min(start_col_idx + 3, end_col_idx + 1))
                    for r in probe_rows:
                        for c in probe_cols:
                            meta = self.tool_handler.execute_tool("cell_meta", {"row": r, "col": c})
                            if meta.get("isFormula"):
                                has_formulas = True
                                break
                        if has_formulas:
                            break
                    if has_formulas:
                        context_parts.append(
                            f"Bottom rows of {range_str} contain formulas → likely summary/aggregation"
                        )
                    # also check if literal 'total' text appears in the values view
                    if "total" in (range_text or "").lower():
                        context_parts.append("Text includes 'total' keyword")
                except Exception as e:
                    self.logger.main.debug(f"_address_open_questions totals check failed: {e}")

            # 3) Sparse tail? Check if the bottom quarter is mostly blank
            if any(k in ql for k in ("sparse", "empty", "blank")):
                try:
                    lines = (range_text or "").split("\n")[2:]  # skip header lines from window renderer
                    quarter = max(1, len(lines) // 4)
                    tail = lines[-quarter:] if lines else []
                    blank_rows = 0
                    for line in tail:
                        cells = [c.strip() for c in line.split("|")[1:]]
                        if all(c in ("", "␀") for c in cells):
                            blank_rows += 1
                    if tail and blank_rows / len(tail) > 0.5:
                        context_parts.append("Bottom portion is mostly blank → likely end of data")
                    elif tail:
                        context_parts.append("Bottom portion still has data → not just padding")
                except Exception as e:
                    self.logger.main.debug(f"_address_open_questions sparsity check failed: {e}")

        return " | ".join(context_parts) if context_parts else "No additional context gathered"

    def _create_range_breadcrumb(self, range_str: str) -> str:
        parts = range_str.split(":")
        if len(parts) != 2:
            return f"📍 Context: Examining range {range_str} within full sheet ({self.sheet_metadata.get('rows', '?')}×{self.sheet_metadata.get('cols', '?')})"
        start_col, start_row = coordinate_from_string(parts[0])
        end_col, end_row = coordinate_from_string(parts[1])
        total_rows = self.sheet_metadata.get("rows", 1)
        total_cols = self.sheet_metadata.get("cols", 1)
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        position_hints = []
        position_hints.append(
            "near top"
            if start_row <= 5
            else ("near bottom" if end_row >= total_rows - 5 else f"middle region (row {start_row}/{total_rows})")
        )
        position_hints.append(
            "left side" if start_col_idx <= 3 else ("right side" if end_col_idx >= total_cols - 3 else "center columns")
        )
        if end_col_idx - start_col_idx + 1 == total_cols:
            position_hints.append("full width")
        if end_row - start_row + 1 == total_rows:
            position_hints.append("full height")
        position_str = ", ".join(position_hints)
        return f"📍 Context: Examining {range_str} ({position_str}) within full sheet ({total_rows}×{total_cols})"

    # --------------------------
    # S4: Overlay
    # --------------------------
    def _overlay(self) -> None:
        self.logger.main.info("🎨 S4: OVERLAY - Adding special elements")

        merged = self.tool_handler.execute_tool("merged_list", {})
        self.logger.main.info(f"   Found {len(merged)} merged ranges")

        named = self.tool_handler.execute_tool("names_list", {})
        self.logger.main.info(f"   Found {len(named)} named ranges")
        current_sheet_name = self.worksheet.title
        for name_info in named:
            if name_info.get("sheet") != current_sheet_name:
                continue
            range_str = name_info.get("range", "")
            if not range_str:
                continue
            for block in self.blocks:
                if self._ranges_overlap(range_str, block.range):
                    block.named_range = name_info["name"]
                    break

        charts = self.tool_handler.execute_tool("object_scan", {"type": "chart"})
        for idx, chart in enumerate(charts):
            self.objects.append(
                ChartObject(id=f"cht_{idx + 1:02d}", anchor=chart.get("anchor", ""), type="chart", linked_block=None)
            )

    def _ranges_overlap(self, range1: str, range2: str) -> bool:
        try:
            clean_range1 = self._strip_sheet_prefix(range1)
            clean_range2 = self._strip_sheet_prefix(range2)

            if ":" in clean_range1:
                p1 = clean_range1.split(":")
                s1_col, s1_row = coordinate_from_string(p1[0])
                e1_col, e1_row = coordinate_from_string(p1[1])
            else:
                s1_col, s1_row = coordinate_from_string(clean_range1)
                e1_col, e1_row = s1_col, s1_row

            if ":" in clean_range2:
                p2 = clean_range2.split(":")
                s2_col, s2_row = coordinate_from_string(p2[0])
                e2_col, e2_row = coordinate_from_string(p2[1])
            else:
                s2_col, s2_row = coordinate_from_string(clean_range2)
                e2_col, e2_row = s2_col, s2_row

            s1_cidx, e1_cidx = column_index_from_string(s1_col), column_index_from_string(e1_col)
            s2_cidx, e2_cidx = column_index_from_string(s2_col), column_index_from_string(e2_col)

            rows_overlap = not (e1_row < s2_row or e2_row < s1_row)
            cols_overlap = not (e1_cidx < s2_cidx or e2_cidx < s1_cidx)
            return rows_overlap and cols_overlap
        except Exception as e:
            self.logger.main.debug(f"Error checking range overlap: {e}")
            return False

    def _strip_sheet_prefix(self, range_str: str) -> str:
        if not range_str:
            return range_str
        if "!" in range_str:
            range_str = range_str.split("!", 1)[1]
        return range_str.replace("$", "")

    # --------------------------
    # S5: Emit
    # --------------------------
    def _emit(self) -> dict[str, Any]:
        self.logger.main.info("📤 S5: EMIT - Generating cartographer map")
        cartographer_map = {
            "sheet": self.worksheet.title,
            "analyzedAt": datetime.now().isoformat() + "Z",
            "blocks": [asdict(block) for block in self.blocks],
            "objects": [asdict(obj) for obj in self.objects],
            "unresolved": self.unresolved,
        }
        self.logger.save_results(cartographer_map)
        return cartographer_map

    # --------------------------
    # Parsing / inference utils
    # --------------------------
    def _infer_characteristics(self, viewport_text: str, range_str: str) -> DataCharacteristics:
        chars = DataCharacteristics()

        parts = range_str.split(":")
        if len(parts) == 2:
            start_col, start_row = coordinate_from_string(parts[0])
            end_col, end_row = coordinate_from_string(parts[1])
            start_col_idx = column_index_from_string(start_col)
            end_col_idx = column_index_from_string(end_col)

            sample_coords = []
            for col in range(start_col_idx, min(start_col_idx + 2, end_col_idx + 1)):
                sample_coords.append((start_row, col))
            for col in range(max(start_col_idx, end_col_idx - 1), end_col_idx + 1):
                sample_coords.append((end_row, col))

            formula_hits = 0
            samples_taken = 0
            for row, col in sample_coords[:4]:
                try:
                    meta = self.tool_handler.execute_tool("cell_meta", {"row": row, "col": col})
                    samples_taken += 1
                    if meta.get("isFormula"):
                        formula_hits += 1
                except Exception:
                    pass
            if samples_taken > 0:
                chars.formula_density = formula_hits / samples_taken

        lines = viewport_text.split("\n")
        if len(lines) < 3:
            return chars

        data_lines = lines[2:]
        if not data_lines:
            return chars

        rows = []
        for line in data_lines:
            parts = line.split("|")
            if len(parts) > 1:
                cells = [cell.strip() for cell in parts[1:]]
                while cells and cells[-1] == "":
                    cells.pop()
                if cells:
                    rows.append(cells)
        if not rows:
            return chars

        first_row = rows[0]
        has_numeric_first_row = any(self._is_numeric(cell) for cell in first_row if cell and cell != "␀")
        has_text_first_row = any(not self._is_numeric(cell) and cell and cell != "␀" for cell in first_row)
        if len(rows) > 1 and has_text_first_row and not has_numeric_first_row:
            chars.has_headers = True
            chars.header_rows = 1

        total_cells = 0
        numeric_cells = 0
        date_cells = 0
        formula_cells = 0
        empty_cells = 0

        for row in rows[chars.header_rows :]:
            for cell in row:
                if not cell or cell == "␀":
                    empty_cells += 1
                    continue
                total_cells += 1
                if cell.startswith("="):
                    formula_cells += 1
                    numeric_cells += 1
                elif self._is_numeric(cell):
                    numeric_cells += 1
                elif self._is_date(cell):
                    date_cells += 1

        if total_cells > 0:
            chars.numeric_ratio = numeric_cells / total_cells
            sampled_formula_density = chars.formula_density or 0.0
            computed_formula_density = formula_cells / total_cells
            chars.formula_density = max(sampled_formula_density, computed_formula_density)
            chars.has_formulas = chars.formula_density > 0
            chars.data_density = total_cells / (total_cells + empty_cells) if (total_cells + empty_cells) > 0 else 0

            if chars.numeric_ratio > 0.5:
                has_currency = any(
                    "$" in str(cell) or "€" in str(cell) or "," in str(cell) for row in rows for cell in row
                )
                if has_currency or chars.formula_density > 0.2:
                    chars.primary_data_type = "financial"
                else:
                    chars.primary_data_type = "numeric"
            elif date_cells / total_cells > 0.2:
                chars.primary_data_type = "temporal"
                chars.is_time_series = True
            else:
                chars.primary_data_type = "categorical"

        if rows:
            last_row = rows[-1]
            if any("total" in str(cell).lower() for cell in last_row):
                chars.has_totals_row = True

        return chars

    def _is_numeric(self, value: str) -> bool:
        if not value or value == "␀":
            return False
        cleaned = value.strip().replace(",", "").replace("$", "").replace("%", "")
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = "-" + cleaned[1:-1]
        try:
            float(cleaned)
            return True
        except Exception:
            return False

    def _is_date(self, value: str) -> bool:
        if not value or value == "␀":
            return False
        date_patterns = [
            r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}",
            r"\d{4}[/-]\d{1,2}[/-]\d{1,2}",
            r"[A-Za-z]{3,9}\s+\d{1,2},?\s+\d{4}",
        ]
        import re

        return any(re.match(pattern, value.strip()) for pattern in date_patterns)

    def _get_default_operations(self, structural_type: StructuralType) -> list[str]:
        defaults = {
            StructuralType.DATA_TABLE: ["add_filters", "check_duplicates", "profile_columns"],
            StructuralType.HEADER_ZONE: ["extract_column_names", "validate_references"],
            StructuralType.AGGREGATION_ZONE: ["verify_calculations", "extract_summary_values"],
            StructuralType.METADATA_ZONE: ["extract_metadata", "parse_report_info"],
            StructuralType.VISUALIZATION_ANCHOR: ["extract_chart_data", "analyze_visualization"],
            StructuralType.FORM_INPUT_ZONE: ["validate_inputs", "extract_form_schema"],
            StructuralType.NAVIGATION_ZONE: ["extract_links", "map_navigation"],
            StructuralType.UNSTRUCTURED_TEXT: ["extract_text", "parse_notes"],
            StructuralType.EMPTY_PADDING: [],
        }
        return defaults.get(structural_type, [])

    # --------------------------
    # CLI
    # --------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Sheet-Cartographer Agent (OpenAI) - Map Excel worksheet topology using OpenAI models"
    )
    parser.add_argument("excel_path", type=Path, help="Path to Excel file")
    parser.add_argument("--sheet-index", type=int, default=0, help="Sheet index to analyze (default: 0)")
    parser.add_argument(
        "--model", type=str, default=DEFAULT_MODEL, help=f"OpenAI model to use (default: {DEFAULT_MODEL})"
    )
    parser.add_argument("--verbose", action="store_true", help="Enable verbose logging")
    parser.add_argument(
        "--log-level",
        type=str,
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        help="Logging level (default: INFO)",
    )
    args = parser.parse_args()

    if not args.excel_path.exists():
        print(f"❌ Error: File not found: {args.excel_path}")
        sys.exit(1)

    logger = ExperimentLogger(
        module_path=__file__, model_name=args.model, excel_path=args.excel_path, sheet_index=args.sheet_index
    )

    logger.main.info("🎯 CONFIGURATION:")
    logger.main.info(f"   Excel file: {args.excel_path}")
    logger.main.info(f"   Sheet index: {args.sheet_index}")
    logger.main.info(f"   Model: {args.model}")
    logger.main.info(f"   Verbose: {args.verbose}")

    cartographer = None
    try:
        cartographer = SheetCartographer(
            excel_path=args.excel_path, sheet_index=args.sheet_index, model=args.model, logger=logger
        )
        result = cartographer.run()

        print("\n✅ ANALYSIS COMPLETE")
        print(f"📊 Detected {len(result['blocks'])} blocks")
        print(f"🎯 Detected {len(result['objects'])} objects")

        if args.verbose:
            print("\n📋 BLOCKS DETECTED:")
            for block in result["blocks"]:
                block_type = block.get("structural_type", block.get("type", "Unknown"))
                if isinstance(block_type, dict):
                    block_type = block_type.get("value", str(block_type))
                semantic = block.get("semantic_description", "")
                print(f"  - {block['id']}: {block_type} at {block['range']} (conf: {block['confidence']:.2f})")
                if semantic:
                    print(f"    → {semantic}")

        print(f"\n📁 Results saved to: {logger.outputs_dir}")

    except Exception as e:
        logger.error.error(f"Fatal error: {e}", exc_info=True)
        print(f"\n❌ FAILED: {e}")
        sys.exit(1)

    finally:
        if cartographer and hasattr(cartographer, "tool_handler"):
            cartographer.tool_handler.close()
        logger.finalize()


if __name__ == "__main__":
    main()
