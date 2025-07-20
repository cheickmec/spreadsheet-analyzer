"""Create a simple test Excel file for CLI testing."""

from pathlib import Path

import openpyxl

# Create test directory if it doesn't exist
test_dir = Path(__file__).parent
test_dir.mkdir(exist_ok=True)

# Create a simple workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Summary"

# Add some data
ws["A1"] = "Product"
ws["B1"] = "Quantity"
ws["C1"] = "Price"
ws["D1"] = "Total"

# Add data rows
products = [
    ("Widget A", 10, 15.99),
    ("Widget B", 5, 22.50),
    ("Widget C", 7, 18.75),
    ("Widget D", 3, 45.00),
]

for i, (product, qty, price) in enumerate(products, start=2):
    ws[f"A{i}"] = product
    ws[f"B{i}"] = qty
    ws[f"C{i}"] = price
    ws[f"D{i}"] = f"=B{i}*C{i}"  # Formula

# Add a sum formula
ws["D6"] = "=SUM(D2:D5)"
ws["A6"] = "Total"

# Create another sheet with more complex formulas
ws2 = wb.create_sheet("Details")
ws2["A1"] = "Reference"
ws2["B1"] = "Value"
ws2["C1"] = "Calculation"

# Add some cross-sheet references
ws2["A2"] = "Total Sales"
ws2["B2"] = "=Summary!D6"
ws2["C2"] = "=B2*1.1"  # 10% markup

# Add some potentially problematic formulas
ws2["A4"] = "Circular Ref Test"
ws2["B4"] = "=C4+1"
ws2["C4"] = "=B4-1"  # Creates circular reference

# Save the file
output_file = test_dir / "simple_test.xlsx"
wb.save(output_file)

print(f"Test file created: {output_file}")
