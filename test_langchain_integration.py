#!/usr/bin/env python3
"""Test script for LangChain integration.

This script tests the complete LangChain/LangGraph workflow
with a sample Excel file.
"""

import asyncio
import os
import sys
from pathlib import Path

# Add src to Python path
sys.path.insert(0, str(Path(__file__).parent / "src"))

from spreadsheet_analyzer.notebook_llm.llm_providers.langchain_integration import (
    analyze_sheet_with_langchain,
)


async def test_langchain_integration():
    """Test the LangChain/LangGraph integration."""

    # Check for sample Excel file
    excel_files = list(Path().glob("*.xlsx"))
    if not excel_files:
        print("ERROR: No Excel files found in current directory")
        print("Please provide a sample Excel file for testing")
        return 1

    excel_file = excel_files[0]
    print(f"Using Excel file: {excel_file}")

    # Get sheet names
    from openpyxl import load_workbook

    wb = load_workbook(excel_file, read_only=True)
    sheets = wb.sheetnames
    wb.close()

    if not sheets:
        print("ERROR: No sheets found in Excel file")
        return 1

    sheet_name = sheets[0]
    print(f"Using sheet: {sheet_name}")

    # Check API key
    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("ERROR: ANTHROPIC_API_KEY not set")
        return 1

    print("\n=== Testing LangChain Integration ===\n")

    try:
        # Run the analysis
        final_state = await analyze_sheet_with_langchain(
            excel_path=excel_file,
            sheet_name=sheet_name,
            skip_deterministic=False,
            provider="anthropic",
            model="claude-3-5-sonnet-20241022",
            temperature=0.1,
            enable_tracing=False,
        )

        # Check results
        print("\n=== Results ===")
        print(f"Excel Path: {final_state.get('excel_path')}")
        print(f"Sheet Name: {final_state.get('sheet_name')}")

        if final_state.get("execution_errors"):
            print("\nExecution Errors:")
            for error in final_state["execution_errors"]:
                print(f"  - {error}")

        if final_state.get("deterministic"):
            print("\nDeterministic Analysis: ✓")

        if final_state.get("notebook_json"):
            print("Notebook Created: ✓")

        if final_state.get("llm_response"):
            print("LLM Analysis: ✓")
            print(f"LLM Response Length: {len(final_state['llm_response'])} chars")

        if final_state.get("notebook_final"):
            print("LLM Code Executed: ✓")

        if final_state.get("output_path"):
            print(f"\nOutput saved to: {final_state['output_path']}")

        if final_state.get("token_usage"):
            usage = final_state["token_usage"]
            print("\nToken Usage:")
            print(f"  - Prompt: {usage.get('prompt_tokens', 'N/A')}")
            print(f"  - Completion: {usage.get('completion_tokens', 'N/A')}")
            print(f"  - Total: {usage.get('total_tokens', 'N/A')}")

        print("\n✅ Test completed successfully!")
        return 0

    except Exception as e:
        print(f"\n❌ Test failed: {e}")
        import traceback

        traceback.print_exc()
        return 1


if __name__ == "__main__":
    exit_code = asyncio.run(test_langchain_integration())
    sys.exit(exit_code)
