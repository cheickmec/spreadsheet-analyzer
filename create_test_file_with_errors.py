#!/usr/bin/env python
"""Create a test Excel file with various formula errors for testing."""

import openpyxl
from pathlib import Path


def create_test_file_with_errors():
    """Create an Excel file with various formula errors."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formula Errors Test"
    
    print("Creating test Excel file with formula errors...")
    
    # Headers
    ws['A1'] = "Error Type"
    ws['B1'] = "Formula"
    ws['C1'] = "Result"
    ws['D1'] = "Description"
    
    # Division by zero error
    ws['A2'] = "#DIV/0!"
    ws['B2'] = "=10/0"
    ws['C2'] = "=10/0"  # This will show #DIV/0!
    ws['D2'] = "Division by zero"
    
    # Reference error
    ws['A3'] = "#REF!"
    ws['B3'] = "=A1:A10"  # Then we'll delete some cells
    ws['C3'] = "#REF!"
    ws['D3'] = "Invalid cell reference"
    
    # Name error
    ws['A4'] = "#NAME?"
    ws['B4'] = "=UNKNOWNFUNCTION(A1)"
    ws['C4'] = "=UNKNOWNFUNCTION(A1)"  # This will show #NAME?
    ws['D4'] = "Unknown function name"
    
    # Value error
    ws['A5'] = "#VALUE!"
    ws['B5'] = '=VALUE("ABC")'
    ws['C5'] = '=VALUE("ABC")'  # This will show #VALUE!
    ws['D5'] = "Wrong value type"
    
    # N/A error
    ws['A6'] = "#N/A"
    ws['B6'] = '=VLOOKUP("NotFound",A1:D10,2,FALSE)'
    ws['C6'] = '=VLOOKUP("NotFound",A1:D10,2,FALSE)'  # This will show #N/A
    ws['D6'] = "Value not available"
    
    # Create another sheet with more complex errors
    ws2 = wb.create_sheet("Complex Errors")
    
    # Header
    ws2['A1'] = "Product"
    ws2['B1'] = "Price"
    ws2['C1'] = "Quantity"
    ws2['D1'] = "Total"
    ws2['E1'] = "Tax"
    ws2['F1'] = "Final"
    
    # Data with errors
    ws2['A2'] = "Product A"
    ws2['B2'] = 100
    ws2['C2'] = 0
    ws2['D2'] = "=B2/C2"  # Division by zero
    ws2['E2'] = "=D2*0.1"  # Error propagation
    ws2['F2'] = "=D2+E2"  # Error propagation
    
    ws2['A3'] = "Product B"
    ws2['B3'] = "Text"  # Text instead of number
    ws2['C3'] = 5
    ws2['D3'] = "=B3*C3"  # Value error
    ws2['E3'] = "=D3*0.1"
    ws2['F3'] = "=D3+E3"
    
    ws2['A4'] = "Product C"
    ws2['B4'] = 50
    ws2['C4'] = 10
    ws2['D4'] = "=B4*C4"  # This should work
    ws2['E4'] = "=VLOOKUP(A4,G1:H10,2,0)"  # N/A error (range doesn't exist)
    ws2['F4'] = "=D4+E4"
    
    # Create a third sheet with circular reference attempt
    ws3 = wb.create_sheet("Circular Reference")
    ws3['A1'] = "Circular Reference Example"
    ws3['A2'] = "Value 1"
    ws3['B2'] = "=C2+10"
    ws3['A3'] = "Value 2"
    ws3['B3'] = "=B2*2"
    ws3['A4'] = "Value 3"
    # Note: Most Excel versions prevent saving true circular references
    # but we can simulate the concept
    
    # Save the file
    output_path = Path("test_assets/generated/formula_errors_test.xlsx")
    output_path.parent.mkdir(exist_ok=True)
    wb.save(output_path)
    wb.close()
    
    print(f"Test file created: {output_path}")
    print("\nThe file contains:")
    print("- Sheet 1: Basic formula errors (#DIV/0!, #REF!, #NAME?, #VALUE!, #N/A)")
    print("- Sheet 2: Complex errors with error propagation")
    print("- Sheet 3: Circular reference example")
    
    return output_path


if __name__ == "__main__":
    file_path = create_test_file_with_errors()
    
    # Now test the error detection on this file
    print("\n" + "=" * 60)
    print("Testing error detection on the created file...")
    print("=" * 60)
    
    from spreadsheet_analyzer.plugins.spreadsheet.analysis.formula_errors import scan_workbook_for_errors
    
    errors = scan_workbook_for_errors(file_path)
    
    if errors:
        print(f"\nDetected errors in {len(errors)} sheet(s):")
        for sheet_name, sheet_errors in errors.items():
            print(f"\nSheet: {sheet_name}")
            for error_type, error_list in sheet_errors.items():
                print(f"  {error_type}: {len(error_list)} occurrence(s)")
                for error in error_list:
                    print(f"    - Cell {error['cell']}: {error.get('formula', 'No formula')}")
    else:
        print("\nNo errors detected (this might mean the errors weren't created properly)")